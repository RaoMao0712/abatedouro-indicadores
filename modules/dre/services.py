"""Servicos da DRE Gerencial."""

import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import repositories as repository
from modules.custos.services import CATEGORIAS_CUSTOS
from modules.financeiro.services import categoria_impacta_resultado_operacional


def formatar_numero_br(valor, casas=2):
    try:
        numero = float(valor or 0)
    except Exception:
        numero = 0
    texto = f"{numero:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_moeda_br(valor):
    return f"R$ {formatar_numero_br(valor, 2)}"


def formatar_percentual_br(valor):
    return f"{formatar_numero_br(valor, 2)}%"


def preparar_grafico_despesas_operacionais(linhas_custos, receita_bruta, limite=6):
    """
    Prepara os dados visuais do gráfico de rosca da DRE.
    O tamanho das fatias usa participação dentro das despesas operacionais.
    O rótulo exibido usa % da receita, para manter leitura gerencial.
    """
    itens_validos = [
        {
            "categoria": item["categoria"],
            "valor": float(item["valor"] or 0),
            "percentual_receita": float(item["percentual"] or 0)
        }
        for item in linhas_custos
        if float(item["valor"] or 0) > 0
    ]

    itens_validos = sorted(itens_validos, key=lambda item: item["valor"], reverse=True)
    total_despesas = sum(item["valor"] for item in itens_validos)

    if total_despesas <= 0:
        return {
            "itens": [],
            "gradiente": "#e5e7eb 0deg 360deg",
            "total": 0,
            "percentual_receita_total": 0
        }

    cores = [
        "#2563eb",
        "#16a34a",
        "#f97316",
        "#8b5cf6",
        "#0891b2",
        "#64748b",
        "#dc2626"
    ]

    limite_principal = max(1, limite - 1)
    principais = itens_validos[:limite_principal]
    restantes = itens_validos[limite_principal:]

    itens_grafico = []

    for item in principais:
        itens_grafico.append(item)

    if restantes:
        valor_outras = sum(item["valor"] for item in restantes)
        percentual_receita_outras = sum(item["percentual_receita"] for item in restantes)
        itens_grafico.append({
            "categoria": f"Outras categorias ({len(restantes)})",
            "valor": valor_outras,
            "percentual_receita": percentual_receita_outras
        })

    segmentos = []
    angulo_atual = 0
    itens_saida = []

    for indice, item in enumerate(itens_grafico):
        fatia = item["valor"] / total_despesas
        graus = fatia * 360
        inicio = angulo_atual
        fim = 360 if indice == len(itens_grafico) - 1 else angulo_atual + graus
        meio = (inicio + fim) / 2
        cor = cores[indice % len(cores)]
        segmentos.append(f"{cor} {inicio:.2f}deg {fim:.2f}deg")

        # Coordenadas dos rótulos ao redor da rosca.
        # Regra visual validada: só exibimos rótulos externos para categorias
        # com pelo menos 5% da receita. As demais continuam na tabela lateral.
        #
        # Também limitamos a distância dos rótulos para impedir invasão da tabela
        # lateral e evitar balões excessivamente afastados do gráfico.
        import math
        rad = math.radians(meio - 90)
        x = 50 + (32 * math.cos(rad))
        y = 50 + (31 * math.sin(rad))
        x = max(28, min(72, x))
        y = max(22, min(78, y))
        alinhamento = "right" if x < 50 else "left"
        mostrar_rotulo = float(item["percentual_receita"] or 0) >= 5

        itens_saida.append({
            "categoria": item["categoria"],
            "valor": round(item["valor"], 2),
            "valor_formatado": formatar_moeda_br(item["valor"]),
            "percentual_receita": round(item["percentual_receita"], 2),
            "percentual_receita_formatado": formatar_percentual_br(item["percentual_receita"]),
            "percentual_despesa": round(fatia * 100, 2),
            "percentual_despesa_formatado": formatar_percentual_br(fatia * 100),
            "cor": cor,
            "x": round(x, 2),
            "y": round(y, 2),
            "alinhamento": alinhamento,
            "mostrar_rotulo": mostrar_rotulo
        })

        angulo_atual = fim

    return {
        "itens": itens_saida,
        "gradiente": ", ".join(segmentos),
        "total": round(total_despesas, 2),
        "total_formatado": formatar_moeda_br(total_despesas),
        "percentual_receita_total": round((total_despesas / receita_bruta * 100) if receita_bruta > 0 else 0, 2),
        "percentual_receita_total_formatado": formatar_percentual_br((total_despesas / receita_bruta * 100) if receita_bruta > 0 else 0)
    }


def preparar_linhas_custos_executivas(linhas_custos, limite=6):
    itens = [item for item in linhas_custos if float(item["valor"] or 0) > 0]
    itens = sorted(itens, key=lambda item: float(item["valor"] or 0), reverse=True)

    if len(itens) <= limite:
        return itens

    principais = itens[:limite - 1]
    restantes = itens[limite - 1:]
    principais.append({
        "categoria": f"Outras categorias ({len(restantes)})",
        "valor": round(sum(float(item["valor"] or 0) for item in restantes), 2),
        "percentual": round(sum(float(item["percentual"] or 0) for item in restantes), 2)
    })
    return principais

def valor_linha_venda(item, campo, padrao=0):
    try:
        valor = item[campo]
    except Exception:
        valor = padrao

    if valor is None:
        return padrao

    return float(valor or 0)


def normalizar_venda_para_dre(item):
    sku = item["sku"]
    quantidade = valor_linha_venda(item, "quantidade")
    unidade = (item["unidade"] or "").lower()
    quantidade_unidades = valor_linha_venda(item, "quantidade_unidades")
    quantidade_kg = valor_linha_venda(item, "quantidade_kg")
    receita = valor_linha_venda(item, "receita")

    # Compatibilidade com registros antigos: antes a Galinha Cortada era lançada só em kg.
    if sku == "Galinha Cortada":
        if quantidade_kg <= 0 and unidade == "kg":
            quantidade_kg = quantidade

        if quantidade_unidades <= 0 and unidade in ["unidades", "unidade", "aves", "ave"]:
            quantidade_unidades = quantidade
    else:
        if quantidade_unidades <= 0:
            quantidade_unidades = quantidade

        quantidade_kg = 0

    return {
        "sku": sku,
        "receita": receita,
        "quantidade": quantidade,
        "unidade": unidade,
        "quantidade_unidades": quantidade_unidades,
        "quantidade_kg": quantidade_kg
    }


def buscar_dados_dre_gerencial(competencia):
    ano, mes = competencia.split("-")
    ultimo_dia = calendar.monthrange(int(ano), int(mes))[1]
    data_inicio = f"{competencia}-01"
    data_fim = f"{competencia}-{ultimo_dia:02d}"

    # CMV temporariamente congelado: vendas_diarias ainda fornece quantidades/SKUs
    # para o calculo do CMV, mas nao alimenta mais Receita Bruta.
    vendas_linhas = [
        normalizar_venda_para_dre(item)
        for item in repository.buscar_vendas_periodo(data_inicio, data_fim)
    ]

    vendas_por_sku_dict = {}
    for item in vendas_linhas:
        sku = item["sku"]

        if sku not in vendas_por_sku_dict:
            vendas_por_sku_dict[sku] = {
                "receita": 0,
                "quantidade": 0,
                "quantidade_unidades": 0,
                "quantidade_kg": 0
            }

        vendas_por_sku_dict[sku]["receita"] += item["receita"]
        vendas_por_sku_dict[sku]["quantidade_unidades"] += item["quantidade_unidades"]
        vendas_por_sku_dict[sku]["quantidade_kg"] += item["quantidade_kg"]

        if sku == "Galinha Cortada":
            vendas_por_sku_dict[sku]["quantidade"] += item["quantidade_kg"]
        else:
            vendas_por_sku_dict[sku]["quantidade"] += item["quantidade_unidades"]

    receita_bruta_movimentacoes = repository.buscar_receita_bruta_movimentacoes(data_inicio, data_fim)
    receita_bruta = receita_bruta_movimentacoes
    deducoes_receita = repository.buscar_deducoes_receita_movimentacoes(data_inicio, data_fim)
    receita_operacional_liquida = receita_bruta - deducoes_receita

    vendas_por_sku = []

    for sku, venda in vendas_por_sku_dict.items():
        quantidade_base = venda["quantidade_kg"] if sku == "Galinha Cortada" else venda["quantidade_unidades"]
        unidade_base = "kg" if sku == "Galinha Cortada" else "unidades"
        preco_medio = venda["receita"] / quantidade_base if quantidade_base > 0 else 0

        vendas_por_sku.append({
            "sku": sku,
            "receita": round(venda["receita"], 2),
            "quantidade": round(quantidade_base, 2),
            "unidade": unidade_base,
            "quantidade_unidades": round(venda["quantidade_unidades"], 2),
            "quantidade_kg": round(venda["quantidade_kg"], 2),
            "preco_medio": round(preco_medio, 4)
        })

    parametros = repository.buscar_parametros_custos_por_sku()

    cmv_por_sku = []
    cmv_total = 0

    for sku, venda in vendas_por_sku_dict.items():
        parametros_sku = parametros.get(sku)

        custo_ave = 0
        custo_embalagem = 0

        if parametros_sku:
            custo_ave = float(parametros_sku["custo_ave"] or 0)
            custo_embalagem = float(parametros_sku["custo_embalagem"] or 0)

        quantidade_unidades = float(venda["quantidade_unidades"] or 0)
        quantidade_kg = float(venda["quantidade_kg"] or 0)

        # Regra atual validada:
        # Galinha Cortada: (ave viva + embalagem) x bandejas vendidas.
        # CMV por kg = CMV total / kg vendidos.
        # Galinha Inteira: 1 x 1 por unidade vendida.
        custo_materia_prima_unitario = custo_ave
        custo_embalagem_unitario = custo_embalagem
        quantidade_cmv = quantidade_unidades

        custo_materia_prima = quantidade_cmv * custo_materia_prima_unitario
        custo_embalagens = quantidade_cmv * custo_embalagem_unitario
        cmv_sku = custo_materia_prima + custo_embalagens
        cmv_total += cmv_sku

        cmv_por_kg = 0
        if sku == "Galinha Cortada" and quantidade_kg > 0:
            cmv_por_kg = cmv_sku / quantidade_kg

        cmv_por_unidade = 0
        if quantidade_cmv > 0:
            cmv_por_unidade = cmv_sku / quantidade_cmv

        cmv_por_sku.append({
            "sku": sku,
            "quantidade_vendida": round(quantidade_kg if sku == "Galinha Cortada" else quantidade_unidades, 2),
            "quantidade_unidades": round(quantidade_unidades, 2),
            "quantidade_kg": round(quantidade_kg, 2),
            "custo_materia_prima_unitario": round(custo_materia_prima_unitario, 4),
            "custo_embalagem_unitario": round(custo_embalagem_unitario, 4),
            "materia_prima": round(custo_materia_prima, 2),
            "embalagem": round(custo_embalagens, 2),
            "cmv": round(cmv_sku, 2),
            "cmv_por_kg": round(cmv_por_kg, 4),
            "cmv_por_unidade": round(cmv_por_unidade, 4),
            "observacao_calculo": "CMV por bandeja vendida; CMV/kg calculado pelos kg vendidos." if sku == "Galinha Cortada" else "CMV 1 x 1 por unidade vendida."
        })

    custos_raw = repository.buscar_custos_operacionais_movimentacoes_por_categoria(competencia)

    categorias = CATEGORIAS_CUSTOS
    custos = {
        categoria: 0
        for categoria in categorias
    }

    for item in custos_raw:
        categoria = item["categoria"]
        valor = float(item["total"] or 0)
        if not categoria_impacta_resultado_operacional(categoria):
            continue
        # Compatibilidade: categorias novas vindas de Movimentacoes entram na DRE
        # mesmo que ainda nao existam na lista historica do modulo de Custos.
        custos[categoria] = custos.get(categoria, 0) + valor

    custos_operacionais_total = sum(custos.values())
    margem_bruta = receita_operacional_liquida - cmv_total
    resultado_operacional = margem_bruta - custos_operacionais_total
    resultado_nao_operacional = repository.buscar_resultado_nao_operacional_movimentacoes(data_inicio, data_fim)
    resultado_gerencial_periodo = resultado_operacional + resultado_nao_operacional

    def perc(valor):
        if receita_bruta > 0:
            return (valor / receita_bruta) * 100

        return 0

    linhas_custos = [
        {
            "categoria": categoria,
            "valor": round(valor, 2),
            "percentual": round(perc(valor), 2)
        }
        for categoria, valor in custos.items()
    ]

    linhas_custos_executivas = preparar_linhas_custos_executivas(linhas_custos)
    despesas_grafico = preparar_grafico_despesas_operacionais(linhas_custos, receita_bruta)

    return {
        "receita_bruta": round(receita_bruta, 2),
        "deducoes_receita": round(deducoes_receita, 2),
        "receita_operacional_liquida": round(receita_operacional_liquida, 2),
        "vendas_por_sku": vendas_por_sku,
        "cmv_total": round(cmv_total, 2),
        "cmv_percentual": round(perc(cmv_total), 2),
        "cmv_por_sku": cmv_por_sku,
        "margem_bruta": round(margem_bruta, 2),
        "margem_bruta_percentual": round(perc(margem_bruta), 2),
        "custos_operacionais_total": round(custos_operacionais_total, 2),
        "custos_operacionais_percentual": round(perc(custos_operacionais_total), 2),
        "linhas_custos": linhas_custos,
        "linhas_custos_executivas": linhas_custos_executivas,
        "despesas_grafico": despesas_grafico,
        "resultado_operacional": round(resultado_operacional, 2),
        "resultado_nao_operacional": round(resultado_nao_operacional, 2),
        "resultado_gerencial_periodo": round(resultado_gerencial_periodo, 2),
        "margem_operacional_percentual": round(perc(resultado_operacional), 2)
    }


def buscar_resumo_dre_gerencial(competencia):
    ano, mes = competencia.split("-")
    ultimo_dia = calendar.monthrange(int(ano), int(mes))[1]
    data_inicio = f"{competencia}-01"
    data_fim = f"{competencia}-{ultimo_dia:02d}"

    vendas_linhas = [
        normalizar_venda_para_dre(item)
        for item in repository.buscar_vendas_periodo(data_inicio, data_fim)
    ]
    vendas_por_sku_dict = {}
    for item in vendas_linhas:
        sku = item["sku"]
        venda = vendas_por_sku_dict.setdefault(sku, {"quantidade_unidades": 0, "quantidade_kg": 0})
        venda["quantidade_unidades"] += item["quantidade_unidades"]
        venda["quantidade_kg"] += item["quantidade_kg"]

    receita_bruta = repository.buscar_receita_bruta_movimentacoes(data_inicio, data_fim)
    deducoes_receita = repository.buscar_deducoes_receita_movimentacoes(data_inicio, data_fim)
    receita_operacional_liquida = receita_bruta - deducoes_receita

    parametros = repository.buscar_parametros_custos_por_sku()
    cmv_total = 0
    for sku, venda in vendas_por_sku_dict.items():
        parametros_sku = parametros.get(sku)
        custo_ave = float(parametros_sku["custo_ave"] or 0) if parametros_sku else 0
        custo_embalagem = float(parametros_sku["custo_embalagem"] or 0) if parametros_sku else 0
        quantidade_cmv = float(venda["quantidade_unidades"] or 0)
        cmv_total += quantidade_cmv * custo_ave + quantidade_cmv * custo_embalagem

    custos_raw = repository.buscar_custos_operacionais_movimentacoes_por_categoria(competencia)
    custos_operacionais_total = 0
    linhas_custos = []
    for item in custos_raw:
        categoria = item["categoria"]
        valor = float(item["total"] or 0)
        if not categoria_impacta_resultado_operacional(categoria):
            continue
        custos_operacionais_total += valor
        linhas_custos.append({"categoria": categoria, "valor": round(valor, 2)})

    margem_bruta = receita_operacional_liquida - cmv_total
    resultado_operacional = margem_bruta - custos_operacionais_total
    resultado_nao_operacional = repository.buscar_resultado_nao_operacional_movimentacoes(data_inicio, data_fim)
    resultado_gerencial_periodo = resultado_operacional + resultado_nao_operacional

    return {
        "receita_bruta": round(receita_bruta, 2),
        "deducoes_receita": round(deducoes_receita, 2),
        "receita_operacional_liquida": round(receita_operacional_liquida, 2),
        "custos_operacionais_total": round(custos_operacionais_total, 2),
        "resultado_operacional": round(resultado_operacional, 2),
        "resultado_nao_operacional": round(resultado_nao_operacional, 2),
        "resultado_gerencial_periodo": round(resultado_gerencial_periodo, 2),
        "linhas_custos": linhas_custos,
    }


def gerar_excel_dre_gerencial(competencia, dados):
    wb = Workbook()
    ws = wb.active
    ws.title = "DRE Gerencial"

    azul = "1F3B4D"
    laranja = "F97316"
    cinza = "F8FAFC"
    branco = "FFFFFF"
    azul_resultado = "2563EB"
    vermelho = "DC2626"

    fill_topo = PatternFill("solid", fgColor=azul)
    fill_laranja = PatternFill("solid", fgColor=laranja)
    fill_cinza = PatternFill("solid", fgColor=cinza)
    fill_resultado = PatternFill(
        "solid",
        fgColor=azul_resultado if dados["resultado_operacional"] >= 0 else vermelho
    )

    fonte_titulo = Font(color=branco, bold=True, size=16)
    fonte_subtitulo = Font(color=branco, bold=True, size=11)
    fonte_header = Font(color=branco, bold=True)
    fonte_negrito = Font(bold=True, color=azul)
    fonte_resultado = Font(color=branco, bold=True, size=13)

    borda = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = "FRIGODATTA — DRE Gerencial Industrial"
    ws["A1"].fill = fill_topo
    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Competência: {competencia}"
    ws["A2"].fill = fill_topo
    ws["A2"].font = fonte_subtitulo
    ws["A2"].alignment = Alignment(horizontal="center")

    linha = 4

    ws[f"A{linha}"] = "Indicador"
    ws[f"B{linha}"] = "Valor"
    ws[f"C{linha}"] = "% Receita"
    ws[f"D{linha}"] = "Observação"

    for col in range(1, 5):
        celula = ws.cell(row=linha, column=col)
        celula.fill = fill_laranja
        celula.font = fonte_header
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda

    def perc_excel(valor, base):
        return (valor / base * 100) if base else 0

    kpis = [
        ("Receita Bruta", dados["receita_bruta"], 100 if dados["receita_bruta"] > 0 else 0, "Central de Movimentacoes"),
        ("Deducoes da Receita", dados["deducoes_receita"], perc_excel(dados["deducoes_receita"], dados["receita_bruta"]), "Reducoes validas da receita"),
        ("Receita Operacional Liquida", dados["receita_operacional_liquida"], perc_excel(dados["receita_operacional_liquida"], dados["receita_bruta"]), "Receita Bruta - Deducoes"),
        ("CMV", dados["cmv_total"], dados["cmv_percentual"], "Custo das vendas"),
        ("Margem Bruta", dados["margem_bruta"], dados["margem_bruta_percentual"], "Receita - CMV"),
        ("Custos Operacionais", dados["custos_operacionais_total"], dados["custos_operacionais_percentual"], "Custos mensais"),
        ("Resultado Operacional", dados["resultado_operacional"], dados["margem_operacional_percentual"], "Margem Bruta - Custos"),
        ("Resultado Nao Operacional", dados["resultado_nao_operacional"], perc_excel(dados["resultado_nao_operacional"], dados["receita_bruta"]), "Financeiro e eventos nao operacionais"),
        ("Resultado Liquido Gerencial", dados["resultado_gerencial_periodo"], perc_excel(dados["resultado_gerencial_periodo"], dados["receita_bruta"]), "Resultado Operacional + Nao Operacional")
    ]

    for item in kpis:
        linha += 1
        ws[f"A{linha}"] = item[0]
        ws[f"B{linha}"] = item[1]
        ws[f"C{linha}"] = item[2] / 100
        ws[f"D{linha}"] = item[3]

        for col in range(1, 5):
            celula = ws.cell(row=linha, column=col)
            celula.border = borda
            celula.fill = fill_cinza
            if col == 1:
                celula.font = fonte_negrito

        ws[f"B{linha}"].number_format = 'R$ #,##0.00'
        ws[f"C{linha}"].number_format = '0.00%'

    linha += 3
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    ws.cell(row=linha, column=1).value = "DRE"
    ws.cell(row=linha, column=1).fill = fill_topo
    ws.cell(row=linha, column=1).font = fonte_header
    ws.cell(row=linha, column=1).alignment = Alignment(horizontal="center")

    linhas_dre = []

    linhas_dre.append(("Receita Bruta", dados["receita_bruta"], "receita"))
    linhas_dre.append(("(-) Deducoes da Receita", dados["deducoes_receita"], "normal"))
    linhas_dre.append(("= Receita Operacional Liquida", dados["receita_operacional_liquida"], "total"))

    linhas_dre.append(("(-) CMV", dados["cmv_total"], "normal"))

    for cmv in dados["cmv_por_sku"]:
        linhas_dre.append((f"  CMV — {cmv['sku']}", cmv["cmv"], "subitem"))

    linhas_dre.append(("= Margem Bruta", dados["margem_bruta"], "total"))
    linhas_dre.append(("Custos Operacionais", None, "grupo"))

    for custo in dados["linhas_custos"]:
        linhas_dre.append((f"(-) {custo['categoria']}", custo["valor"], "normal"))

    linhas_dre.append(("Total de Custos Operacionais", dados["custos_operacionais_total"], "total"))
    linhas_dre.append(("= Resultado Operacional", dados["resultado_operacional"], "resultado"))
    linhas_dre.append(("(+/-) Resultado Nao Operacional", dados["resultado_nao_operacional"], "normal"))
    linhas_dre.append(("= Resultado Liquido Gerencial", dados["resultado_gerencial_periodo"], "resultado"))

    for descricao, valor, tipo in linhas_dre:
        linha += 1

        ws[f"A{linha}"] = descricao

        if valor is not None:
            ws[f"B{linha}"] = valor
            ws[f"B{linha}"].number_format = 'R$ #,##0.00'

        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=4)

        for col in range(1, 5):
            celula = ws.cell(row=linha, column=col)
            celula.border = borda

        if tipo == "grupo":
            for col in range(1, 5):
                ws.cell(row=linha, column=col).fill = fill_topo
                ws.cell(row=linha, column=col).font = fonte_header
        elif tipo == "resultado":
            for col in range(1, 5):
                ws.cell(row=linha, column=col).fill = fill_resultado
                ws.cell(row=linha, column=col).font = fonte_resultado
        elif tipo == "total":
            for col in range(1, 5):
                ws.cell(row=linha, column=col).fill = fill_cinza
                ws.cell(row=linha, column=col).font = fonte_negrito
        elif tipo == "subitem":
            ws[f"A{linha}"].font = Font(color="64748B")
            ws[f"B{linha}"].font = Font(color="64748B")
        else:
            ws[f"A{linha}"].font = fonte_negrito

    linha += 3
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    ws.cell(row=linha, column=1).value = (
        "Leitura executiva: "
        + ("A operação apresentou resultado operacional positivo." if dados["resultado_operacional"] >= 0 else "A operação apresentou resultado operacional negativo.")
    )
    ws.cell(row=linha, column=1).alignment = Alignment(wrap_text=True)
    ws.cell(row=linha, column=1).fill = PatternFill("solid", fgColor="FFF7ED")

    larguras = {
        "A": 38,
        "B": 18,
        "C": 14,
        "D": 32
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="center",
                horizontal="right" if cell.column >= 2 else "left",
                wrap_text=True
            )

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = "A4"

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return arquivo






# ============================================================
# MÓDULO ALMOXARIFADO
# ============================================================

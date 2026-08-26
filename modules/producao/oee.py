"""Motor oficial e auditavel de Disponibilidade, Performance, Qualidade e OEE."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from database import DATABASE_URL, conectar, q
from .disponibilidade import FUSO_MANAUS, calcular_disponibilidade
from .performance import calcular_performance


SITUACOES = {"CALCULAVEL", "EM_ANDAMENTO", "NAO_CALCULAVEL", "INCONSISTENTE"}
STATUS_HISTORICOS = {"ESTORNADA", "ESTORNADO", "CANCELADA", "CANCELADO"}


def _pk():
    return "SERIAL PRIMARY KEY" if DATABASE_URL else "INTEGER PRIMARY KEY AUTOINCREMENT"


def criar_tabelas_oee():
    """Cria somente a configuracao fisica; nenhum valor historico e inferido."""
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""CREATE TABLE IF NOT EXISTS linha_abate_configuracoes_fisicas (
            id {_pk()}, vigencia_inicio TEXT NOT NULL, vigencia_fim TEXT,
            noria_1_ganchos_instalados INTEGER NOT NULL,
            noria_1_ganchos_operacionais INTEGER NOT NULL,
            noria_2_ganchos_instalados INTEGER NOT NULL,
            noria_2_ganchos_operacionais INTEGER NOT NULL,
            justificativa TEXT NOT NULL, registrado_por TEXT NOT NULL,
            registrado_por_id INTEGER, criado_em TEXT NOT NULL,
            ativo_logico INTEGER NOT NULL DEFAULT 1, versao INTEGER NOT NULL DEFAULT 1
        )""")
        cursor.execute("""CREATE INDEX IF NOT EXISTS idx_linha_config_fisica_vigencia
            ON linha_abate_configuracoes_fisicas(vigencia_inicio,vigencia_fim,ativo_logico)""")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _data(valor, nome):
    try:
        return date.fromisoformat(str(valor or ""))
    except ValueError as erro:
        raise ValueError(f"{nome} invalida.") from erro


def registrar_configuracao_fisica(
    vigencia_inicio, vigencia_fim,
    noria_1_ganchos_instalados, noria_1_ganchos_operacionais,
    noria_2_ganchos_instalados, noria_2_ganchos_operacionais,
    justificativa, *, usuario=None, usuario_id=None, perfil=None,
):
    if str(perfil or "").lower() != "admin":
        raise PermissionError("Somente Administrador pode registrar a configuracao fisica da linha.")
    inicio = _data(vigencia_inicio, "Vigencia inicial")
    fim = _data(vigencia_fim, "Vigencia final") if vigencia_fim else None
    if fim and fim < inicio:
        raise ValueError("A vigencia final deve ser posterior a inicial.")
    valores = []
    for nome, valor in (
        ("ganchos instalados da Noria 1", noria_1_ganchos_instalados),
        ("ganchos operacionais da Noria 1", noria_1_ganchos_operacionais),
        ("ganchos instalados da Noria 2", noria_2_ganchos_instalados),
        ("ganchos operacionais da Noria 2", noria_2_ganchos_operacionais),
    ):
        try:
            numero = int(valor)
        except (TypeError, ValueError) as erro:
            raise ValueError(f"Informe {nome}.") from erro
        if numero <= 0:
            raise ValueError(f"{nome.capitalize()} deve ser maior que zero.")
        valores.append(numero)
    n1i, n1o, n2i, n2o = valores
    if n1o > n1i or n2o > n2i:
        raise ValueError("Ganchos operacionais nao podem exceder os instalados.")
    justificativa = str(justificativa or "").strip()
    if not justificativa:
        raise ValueError("A justificativa tecnica e obrigatoria.")

    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute(q("""SELECT id FROM linha_abate_configuracoes_fisicas
            WHERE ativo_logico=1 AND vigencia_inicio<=?
              AND COALESCE(vigencia_fim,'9999-12-31')>=?"""),
            ((fim or date.max).isoformat(), inicio.isoformat()))
        if cursor.fetchone():
            raise ValueError("Existe configuracao fisica vigente no intervalo informado.")
        criado_em = datetime.now(FUSO_MANAUS).isoformat(timespec="seconds")
        parametros = (
            inicio.isoformat(), fim.isoformat() if fim else None,
            n1i, n1o, n2i, n2o, justificativa,
            usuario or "Sistema", usuario_id, criado_em,
        )
        sql = """INSERT INTO linha_abate_configuracoes_fisicas (
            vigencia_inicio,vigencia_fim,noria_1_ganchos_instalados,
            noria_1_ganchos_operacionais,noria_2_ganchos_instalados,
            noria_2_ganchos_operacionais,justificativa,registrado_por,
            registrado_por_id,criado_em
        ) VALUES (?,?,?,?,?,?,?,?,?,?)"""
        if DATABASE_URL:
            cursor.execute(q(sql + " RETURNING id"), parametros)
            registro_id = cursor.fetchone()["id"]
        else:
            cursor.execute(sql, parametros)
            registro_id = cursor.lastrowid
        conn.commit()
        return registro_id
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def listar_configuracoes_fisicas():
    conn = conectar()
    try:
        cursor = conn.cursor()
        cursor.execute("""SELECT * FROM linha_abate_configuracoes_fisicas
            WHERE ativo_logico=1 ORDER BY vigencia_inicio DESC,id DESC""")
        return cursor.fetchall()
    finally:
        conn.close()


def calcular_qualidade(op_id, *, conn=None):
    """Nao converte rendimento, PNC ou condenacao em Qualidade por conveniencia."""
    propria = conn is None
    conn = conn or conectar()
    try:
        cursor = conn.cursor()
        cursor.execute(q("SELECT id,status FROM ordens_producao WHERE id=?"), (op_id,))
        op = cursor.fetchone()
        base = {
            "situacao": "NAO_CALCULAVEL", "unidades_boas": None,
            "unidades_processadas": None, "qualidade": None,
            "motivos": [
                "Qualidade nao calculavel com os dados atualmente disponiveis: "
                "nao existe contagem oficial independente de unidades processadas e boas."
            ],
            "alertas": [], "inconsistencias": [],
        }
        if not op:
            base["motivos"] = ["OP nao encontrada."]
            return base
        status = str(op["status"] or "").upper()
        if status not in {"ENCERRADA"} and status not in STATUS_HISTORICOS:
            base["situacao"] = "EM_ANDAMENTO"
            base["motivos"] = ["OP em andamento; Qualidade final nao e calculada."]
        return base
    finally:
        if propria:
            conn.close()


def _situacao_oee(op_status, componentes):
    status = str(op_status or "").upper()
    if any(item.get("situacao") == "INCONSISTENTE" for item in componentes):
        return "INCONSISTENTE"
    if status not in {"ENCERRADA"} and status not in STATUS_HISTORICOS:
        return "EM_ANDAMENTO"
    if any(item.get("situacao") == "EM_ANDAMENTO" for item in componentes):
        return "EM_ANDAMENTO"
    if any(item.get("situacao") != "CALCULAVEL" for item in componentes):
        return "NAO_CALCULAVEL"
    return "CALCULAVEL"


def calcular_oee(op_id, *, conn=None, disponibilidade=None, performance=None, qualidade=None):
    propria = conn is None
    conn = conn or conectar()
    try:
        cursor = conn.cursor()
        cursor.execute(q("SELECT * FROM ordens_producao WHERE id=?"), (op_id,))
        op = cursor.fetchone()
        if not op:
            return {
                "op_id": op_id, "situacao": "NAO_CALCULAVEL", "oee": None,
                "motivos": ["OP nao encontrada."], "alertas": [], "inconsistencias": [],
            }
        disponibilidade = disponibilidade or calcular_disponibilidade(op_id, conn=conn)
        performance = performance or calcular_performance(
            op_id, conn=conn, disponibilidade=disponibilidade,
        )
        qualidade = qualidade or calcular_qualidade(op_id, conn=conn)
        situacao = _situacao_oee(
            op["status"], [disponibilidade, performance, qualidade],
        )
        oee = None
        if situacao == "CALCULAVEL":
            oee = (
                Decimal(str(disponibilidade["disponibilidade"]))
                * Decimal(str(performance["performance"]))
                * Decimal(str(qualidade["qualidade"]))
                / Decimal("10000")
            )
        alertas = list(disponibilidade.get("alertas") or []) + list(performance.get("alertas") or [])
        if oee is not None and oee > Decimal("100"):
            alertas.append("OEE acima de 100% decorre da Performance acima de 100%; revise as bases oficiais.")
        motivos = []
        for componente in (disponibilidade, performance, qualidade):
            motivos.extend(componente.get("motivos") or [])
        return {
            "op_id": op_id, "data": op["data"], "status": op["status"],
            "fornecedor": op["fornecedor"], "sku": op["sku"],
            "disponibilidade": disponibilidade, "performance": performance,
            "qualidade": qualidade, "oee": oee, "situacao": situacao,
            "motivos": motivos, "alertas": alertas,
            "inconsistencias": (
                list(disponibilidade.get("inconsistencias") or [])
                + list(performance.get("inconsistencias") or [])
                + list(qualidade.get("inconsistencias") or [])
            ),
        }
    finally:
        if propria:
            conn.close()


def normalizar_filtros_oee(args):
    hoje = date.today()
    status = args.get("status") or "Encerrada"
    if status not in {"Encerrada", "Aberta", "Historico", "Todas"}:
        status = "Encerrada"
    return {
        "data_inicio": args.get("data_inicio") or hoje.replace(day=1).isoformat(),
        "data_fim": args.get("data_fim") or hoje.isoformat(),
        "op_id": args.get("op_id") or "",
        "status": status,
        "fornecedor": args.get("fornecedor") or "Todos",
    }


def _buscar_ops(filtros, conn):
    condicoes = ["data BETWEEN ? AND ?"]
    parametros = [filtros["data_inicio"], filtros["data_fim"]]
    if filtros["op_id"]:
        try:
            condicoes.append("id=?")
            parametros.append(int(filtros["op_id"]))
        except ValueError:
            condicoes.append("1=0")
    if filtros["status"] == "Encerrada":
        condicoes.append("UPPER(COALESCE(status,''))='ENCERRADA'")
    elif filtros["status"] == "Aberta":
        condicoes.append("UPPER(COALESCE(status,'')) NOT IN ('ENCERRADA','ESTORNADA','ESTORNADO','CANCELADA','CANCELADO')")
    elif filtros["status"] == "Historico":
        condicoes.append("UPPER(COALESCE(status,'')) IN ('ESTORNADA','ESTORNADO','CANCELADA','CANCELADO')")
    if filtros["fornecedor"] != "Todos":
        condicoes.append("fornecedor=?")
        parametros.append(filtros["fornecedor"])
    cursor = conn.cursor()
    cursor.execute(q(f"SELECT id FROM ordens_producao WHERE {' AND '.join(condicoes)} ORDER BY data DESC,id DESC"), tuple(parametros))
    return [item["id"] for item in cursor.fetchall()]


def _estado_consolidado(linhas, chave):
    estados = [linha[chave]["situacao"] for linha in linhas]
    if not estados:
        return "NAO_CALCULAVEL"
    if "INCONSISTENTE" in estados:
        return "INCONSISTENTE"
    if "EM_ANDAMENTO" in estados:
        return "EM_ANDAMENTO"
    if any(estado != "CALCULAVEL" for estado in estados):
        return "NAO_CALCULAVEL"
    return "CALCULAVEL"


def consolidar_oee(filtros):
    conn = conectar()
    try:
        linhas = [calcular_oee(op_id, conn=conn) for op_id in _buscar_ops(filtros, conn)]
        estado_d = _estado_consolidado(linhas, "disponibilidade")
        estado_p = _estado_consolidado(linhas, "performance")
        estado_q = _estado_consolidado(linhas, "qualidade")
        planejado = sum((Decimal(str(l["disponibilidade"].get("tempo_planejado_liquido_minutos") or 0)) for l in linhas if l["disponibilidade"]["situacao"] == "CALCULAVEL"), Decimal("0"))
        operacional = sum((Decimal(str(l["disponibilidade"].get("tempo_operacional_minutos") or 0)) for l in linhas if l["disponibilidade"]["situacao"] == "CALCULAVEL"), Decimal("0"))
        aves = sum((Decimal(str(l["performance"].get("quantidade_total_considerada") or 0)) for l in linhas if l["performance"]["situacao"] == "CALCULAVEL"), Decimal("0"))
        teorica = sum((Decimal(str(l["performance"].get("producao_teorica") or 0)) for l in linhas if l["performance"]["situacao"] == "CALCULAVEL"), Decimal("0"))
        disponibilidade = operacional / planejado * Decimal("100") if estado_d == "CALCULAVEL" and planejado > 0 else None
        performance = aves / teorica * Decimal("100") if estado_p == "CALCULAVEL" and teorica > 0 else None
        boas = sum((Decimal(str(l["qualidade"].get("unidades_boas") or 0)) for l in linhas if l["qualidade"]["situacao"] == "CALCULAVEL"), Decimal("0"))
        processadas = sum((Decimal(str(l["qualidade"].get("unidades_processadas") or 0)) for l in linhas if l["qualidade"]["situacao"] == "CALCULAVEL"), Decimal("0"))
        qualidade = boas / processadas * Decimal("100") if estado_q == "CALCULAVEL" and processadas > 0 else None
        estado = _situacao_oee("Encerrada", [
            {"situacao": estado_d}, {"situacao": estado_p}, {"situacao": estado_q},
        ])
        oee = None
        if estado == "CALCULAVEL":
            oee = disponibilidade * performance * qualidade / Decimal("10000")
        return {
            "linhas": linhas,
            "totais": {
                "ops": len(linhas), "tempo_planejado_minutos": planejado,
                "tempo_operacional_minutos": operacional,
                "disponibilidade": disponibilidade, "situacao_disponibilidade": estado_d,
                "aves_consideradas": aves, "capacidade_teorica": teorica,
                "performance": performance, "situacao_performance": estado_p,
                "unidades_boas": boas if estado_q == "CALCULAVEL" else None,
                "unidades_processadas": processadas if estado_q == "CALCULAVEL" else None,
                "qualidade": qualidade, "situacao_qualidade": estado_q,
                "oee": oee, "situacao_oee": estado,
                "ops_disponibilidade_calculavel": sum(l["disponibilidade"]["situacao"] == "CALCULAVEL" for l in linhas),
                "ops_performance_calculavel": sum(l["performance"]["situacao"] == "CALCULAVEL" for l in linhas),
            },
        }
    finally:
        conn.close()


def montar_contexto_oee(args, slug="oee"):
    filtros = normalizar_filtros_oee(args)
    dados = consolidar_oee(filtros)
    conn = conectar()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT fornecedor FROM ordens_producao WHERE COALESCE(TRIM(fornecedor),'')<>'' ORDER BY fornecedor")
        fornecedores = [item["fornecedor"] for item in cursor.fetchall()]
    finally:
        conn.close()
    titulos = {
        "disponibilidade": "Disponibilidade da Linha",
        "performance": "Performance da Linha",
        "oee": "OEE",
    }
    return {
        "slug": slug, "titulo": titulos.get(slug, "OEE"),
        "filtros": filtros, "fornecedores": fornecedores,
        "linhas": dados["linhas"], "totais": dados["totais"],
        "limitacoes": [
            "Qualidade permanece nao calculavel: PNC, condenacoes e rendimento nao substituem unidades boas/processadas.",
            "Sem Qualidade oficial, OEE permanece N/A; nao e apresentado OEE parcial.",
            "Ganchos operacionais sao configuracao fisica auditavel e nao multiplicam a Performance.",
        ],
    }


def gerar_excel_oee(contexto):
    wb = Workbook()
    ws = wb.active
    ws.title = "OEE auditavel"
    cabecalhos = [
        "OP", "Data", "Status", "Estado", "Planejado min", "Operacional min",
        "Disponibilidade %", "Aves consideradas", "Velocidade aves/h",
        "Capacidade teorica", "Performance %", "Qualidade %", "OEE %", "Observacoes",
    ]
    ws.append(cabecalhos)
    for celula in ws[1]:
        celula.font = Font(bold=True)
    for linha in contexto["linhas"]:
        d, p, ql = linha["disponibilidade"], linha["performance"], linha["qualidade"]
        data_op = date.fromisoformat(str(linha["data"])[:10]) if linha.get("data") else None
        ws.append([
            linha["op_id"], data_op, linha["status"], linha["situacao"],
            float(d["tempo_planejado_liquido_minutos"]) if d.get("tempo_planejado_liquido_minutos") is not None else None,
            float(d["tempo_operacional_minutos"]) if d.get("tempo_operacional_minutos") is not None else None,
            float(d["disponibilidade"]) if d.get("disponibilidade") is not None else None,
            float(p["quantidade_total_considerada"]) if p.get("quantidade_total_considerada") is not None else None,
            float(p["velocidade_ideal_aves_hora"]) if p.get("velocidade_ideal_aves_hora") is not None else None,
            float(p["producao_teorica"]) if p.get("producao_teorica") is not None else None,
            float(p["performance"]) if p.get("performance") is not None else None,
            float(ql["qualidade"]) if ql.get("qualidade") is not None else None,
            float(linha["oee"]) if linha.get("oee") is not None else None,
            " | ".join(linha.get("motivos") or []),
        ])
    for celula in ws["B"][1:]:
        celula.number_format = "dd/mm/yyyy"
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo

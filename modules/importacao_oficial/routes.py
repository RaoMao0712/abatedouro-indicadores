"""Importacao oficial de Maio/2026."""

from datetime import datetime

from flask import flash, render_template, request
from openpyxl import load_workbook

from database import DATABASE_URL, conectar, q
from modules.auth.decorators import perfil_permitido

_criar_banco = None
_criar_tabela_tempos_setor = None


def configurar_integracoes(criar_banco=None, criar_tabela_tempos_setor=None):
    global _criar_banco, _criar_tabela_tempos_setor
    _criar_banco = criar_banco
    _criar_tabela_tempos_setor = criar_tabela_tempos_setor

MOTIVOS_DESCARTE_IMPORTACAO_MAIO = [
    "Hematomas",
    "Contaminação no processo",
    "Aspecto repugnante",
    "Doença",
    "Cozimento",
    "Caquexia",
    "Fratura",
    "Carcaça incompleta",
    "Outra",
]


def normalizar_cabecalho_importacao(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    mapa = str.maketrans("áàâãéêíóôõúç", "aaaaeeiooouc")
    return texto.translate(mapa).replace("_", " ")


def valor_excel_para_data_texto(valor):
    if not valor:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    for formato in ["%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"]:
        try:
            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except Exception:
            pass
    raise ValueError(f"Data inválida: {valor}")


def numero_importacao(valor, padrao=0):
    if valor is None or str(valor).strip() == "":
        return padrao
    if isinstance(valor, str):
        valor = valor.replace("R$", "").replace(".", "").replace(",", ".").strip()
    return float(valor)


def inteiro_importacao(valor, padrao=0):
    return int(round(numero_importacao(valor, padrao)))


def texto_importacao(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def mapa_colunas_por_cabecalho(ws):
    cabecalhos = {}
    for idx, celula in enumerate(ws[1], start=1):
        cabecalhos[normalizar_cabecalho_importacao(celula.value)] = idx
    return cabecalhos


def localizar_coluna(cabecalhos, nomes_possiveis):
    for nome in nomes_possiveis:
        chave = normalizar_cabecalho_importacao(nome)
        if chave in cabecalhos:
            return cabecalhos[chave]
    return None


def ler_planilha_importacao_maio(arquivo_excel):
    wb = load_workbook(arquivo_excel, data_only=True)
    erros = []
    avisos = []
    ops = []
    descartes = []

    if "OPs" not in wb.sheetnames:
        return {"erros": ["A planilha precisa ter uma aba chamada OPs."], "avisos": [], "ops": [], "descartes": []}

    ws_ops = wb["OPs"]
    cab = mapa_colunas_por_cabecalho(ws_ops)
    col = {
        "data": localizar_coluna(cab, ["data"]),
        "sku": localizar_coluna(cab, ["sku"]),
        "fornecedor": localizar_coluna(cab, ["fornecedor"]),
        "gta": localizar_coluna(cab, ["gta"]),
        "nota_fiscal": localizar_coluna(cab, ["nota fiscal", "nf", "nota_fiscal"]),
        "quantidade_aves": localizar_coluna(cab, ["quantidade de aves", "aves", "quantidade_aves"]),
        "mortes_antes_pendura": localizar_coluna(cab, ["mortes na gaiola", "mortes_antes_pendura"]),
        "peso_vivo": localizar_coluna(cab, ["peso vivo", "peso_vivo"]),
        "unidades_produzidas": localizar_coluna(cab, ["unidades produzidas", "unidades_produzidas"]),
        "kg_produzidos": localizar_coluna(cab, ["kg produzidos", "kg_produzidos"]),
        "observacoes": localizar_coluna(cab, ["observações", "observacoes", "obs"]),
    }

    for chave in ["data", "sku", "fornecedor", "quantidade_aves", "mortes_antes_pendura", "peso_vivo", "unidades_produzidas", "kg_produzidos"]:
        if not col.get(chave):
            erros.append(f"Aba OPs: coluna obrigatória ausente: {chave}.")

    if erros:
        return {"erros": erros, "avisos": avisos, "ops": ops, "descartes": descartes}

    for linha in range(2, ws_ops.max_row + 1):
        if not ws_ops.cell(linha, col["data"]).value:
            continue
        try:
            op = {
                "linha": linha,
                "data": valor_excel_para_data_texto(ws_ops.cell(linha, col["data"]).value),
                "sku": texto_importacao(ws_ops.cell(linha, col["sku"]).value) or "Galinha Cortada",
                "fornecedor": texto_importacao(ws_ops.cell(linha, col["fornecedor"]).value),
                "gta": texto_importacao(ws_ops.cell(linha, col["gta"]).value) if col.get("gta") else "",
                "nota_fiscal": texto_importacao(ws_ops.cell(linha, col["nota_fiscal"]).value) if col.get("nota_fiscal") else "",
                "quantidade_aves": inteiro_importacao(ws_ops.cell(linha, col["quantidade_aves"]).value),
                "mortes_antes_pendura": inteiro_importacao(ws_ops.cell(linha, col["mortes_antes_pendura"]).value),
                "peso_vivo": numero_importacao(ws_ops.cell(linha, col["peso_vivo"]).value),
                "unidades_produzidas": numero_importacao(ws_ops.cell(linha, col["unidades_produzidas"]).value),
                "kg_produzidos": numero_importacao(ws_ops.cell(linha, col["kg_produzidos"]).value),
                "observacoes": texto_importacao(ws_ops.cell(linha, col["observacoes"]).value) if col.get("observacoes") else "",
            }
            if not op["fornecedor"]:
                erros.append(f"Aba OPs linha {linha}: fornecedor vazio.")
            if op["quantidade_aves"] <= 0:
                erros.append(f"Aba OPs linha {linha}: quantidade de aves precisa ser maior que zero.")
            if op["peso_vivo"] <= 0:
                erros.append(f"Aba OPs linha {linha}: peso vivo precisa ser maior que zero.")
            if op["unidades_produzidas"] <= 0:
                erros.append(f"Aba OPs linha {linha}: unidades produzidas precisa ser maior que zero.")
            if op["kg_produzidos"] <= 0:
                erros.append(f"Aba OPs linha {linha}: kg produzidos está vazio ou zerado.")
            ops.append(op)
        except Exception as erro:
            erros.append(f"Aba OPs linha {linha}: {erro}")

    if "Descartes" in wb.sheetnames:
        ws_desc = wb["Descartes"]
        cab_desc = mapa_colunas_por_cabecalho(ws_desc)
        col_data = localizar_coluna(cab_desc, ["data"])
        if not col_data:
            erros.append("Aba Descartes: coluna Data ausente.")
        else:
            col_motivos = []
            for motivo in MOTIVOS_DESCARTE_IMPORTACAO_MAIO:
                c = localizar_coluna(cab_desc, [motivo])
                if c:
                    col_motivos.append((motivo, c))
            if not col_motivos:
                avisos.append("Aba Descartes existe, mas nenhum motivo conhecido foi encontrado.")
            for linha in range(2, ws_desc.max_row + 1):
                data_raw = ws_desc.cell(linha, col_data).value
                if not data_raw:
                    continue
                try:
                    data = valor_excel_para_data_texto(data_raw)
                    for motivo, c in col_motivos:
                        qtd = numero_importacao(ws_desc.cell(linha, c).value)
                        if qtd > 0:
                            descartes.append({
                                "linha": linha,
                                "data": data,
                                "setor": "Não informado",
                                "categoria": "Condenação / Descarte",
                                "motivo": motivo,
                                "quantidade": qtd,
                                "unidade": "aves",
                                "observacoes": "Importado da planilha oficial de maio/2026. Setor não informado na origem.",
                            })
                except Exception as erro:
                    erros.append(f"Aba Descartes linha {linha}: {erro}")
    else:
        avisos.append("A planilha não possui aba Descartes. Apenas OPs serão importadas.")

    if not ops:
        erros.append("Nenhuma OP válida encontrada para importação.")

    return {"erros": erros, "avisos": avisos, "ops": ops, "descartes": descartes}


def resumir_importacao_maio(dados):
    ops = dados.get("ops", [])
    descartes = dados.get("descartes", [])
    total_descartes = sum(item["quantidade"] for item in descartes)
    motivos = {}
    for item in descartes:
        motivos[item["motivo"]] = motivos.get(item["motivo"], 0) + item["quantidade"]
    return {
        "total_ops": len(ops),
        "total_aves": round(sum(op["quantidade_aves"] for op in ops), 2),
        "total_mortes": round(sum(op["mortes_antes_pendura"] for op in ops), 2),
        "total_peso_vivo": round(sum(op["peso_vivo"] for op in ops), 2),
        "total_unidades": round(sum(op["unidades_produzidas"] for op in ops), 2),
        "total_kg": round(sum(op["kg_produzidos"] for op in ops), 2),
        "total_descartes": round(total_descartes, 2),
        "motivos": [
            {"motivo": m, "quantidade": round(q, 2), "percentual": round((q / total_descartes * 100), 2) if total_descartes > 0 else 0}
            for m, q in sorted(motivos.items(), key=lambda par: par[1], reverse=True)
        ],
    }


def obter_id_inserido(cursor):
    if DATABASE_URL:
        cursor.execute("SELECT LASTVAL() as id")
        return cursor.fetchone()["id"]
    return cursor.lastrowid


def excluir_historico_maio_2026(cursor):
    cursor.execute(q("""
    SELECT id FROM ordens_producao
    WHERE data BETWEEN ? AND ?
      AND observacoes LIKE ?
    """), ("2026-05-01", "2026-05-31", "%Importação oficial Maio/2026%"))
    ids = [item["id"] for item in cursor.fetchall()]
    if not ids:
        return 0
    placeholders = ",".join(["?"] * len(ids))
    for tabela in ["apontamentos_setor", "apontamentos_producao", "apontamentos_mao_obra", "apontamentos_paradas", "apontamentos_descartes", "apontamentos_tempos_setor"]:
        cursor.execute(q(f"DELETE FROM {tabela} WHERE op_id IN ({placeholders})"), tuple(ids))
    cursor.execute(q(f"DELETE FROM ordens_producao WHERE id IN ({placeholders})"), tuple(ids))
    return len(ids)


def importar_dados_oficiais_maio(dados, substituir=True):
    _criar_banco()
    _criar_tabela_tempos_setor()
    if dados.get("erros"):
        raise ValueError("A planilha possui erros de validação e não pode ser importada.")
    conn = conectar()
    cursor = conn.cursor()
    try:
        removidas = excluir_historico_maio_2026(cursor) if substituir else 0
        op_por_data = {}
        ops_importadas = 0
        descartes_importados = 0
        for op in dados["ops"]:
            peso_medio = op["peso_vivo"] / op["quantidade_aves"] if op["quantidade_aves"] else 0
            obs = (op.get("observacoes") or "").strip()
            obs = (obs + " | " if obs else "") + "Importação oficial Maio/2026"
            cursor.execute(q("""
            INSERT INTO ordens_producao (
                data, sku, fornecedor, gta, nota_fiscal, quantidade_aves,
                mortes_antes_pendura, peso_vivo, peso_medio, observacoes, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """), (op["data"], op["sku"], op["fornecedor"], op["gta"], op["nota_fiscal"], op["quantidade_aves"], op["mortes_antes_pendura"], op["peso_vivo"], peso_medio, obs, "Encerrada"))
            op_id = obter_id_inserido(cursor)
            op_por_data[op["data"]] = op_id
            ops_importadas += 1
            cursor.execute(q("""
            INSERT INTO apontamentos_producao (op_id, data, setor, quantidade, unidade, observacoes)
            VALUES (?, ?, ?, ?, ?, ?)
            """), (op_id, op["data"], "Expedição", op["unidades_produzidas"], "unidades", "Produção final importada da planilha oficial de maio/2026."))
            cursor.execute(q("""
            INSERT INTO apontamentos_producao (op_id, data, setor, quantidade, unidade, observacoes)
            VALUES (?, ?, ?, ?, ?, ?)
            """), (op_id, op["data"], "Expedição", op["kg_produzidos"], "kg", "Kg final produzido importado da planilha oficial de maio/2026."))
        for descarte in dados["descartes"]:
            op_id = op_por_data.get(descarte["data"])
            if not op_id:
                continue
            cursor.execute(q("""
            INSERT INTO apontamentos_descartes (op_id, data, setor, categoria, motivo, quantidade, unidade, observacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """), (op_id, descarte["data"], descarte["setor"], descarte["categoria"], descarte["motivo"], descarte["quantidade"], descarte["unidade"], descarte["observacoes"]))
            descartes_importados += 1
        conn.commit()
        conn.close()
        return {"removidas": removidas, "ops_importadas": ops_importadas, "descartes_importados": descartes_importados}
    except Exception:
        conn.rollback()
        conn.close()
        raise




def register_importacao_routes(app, integracoes=None):
    integracoes = integracoes or {}
    configurar_integracoes(
        criar_banco=integracoes.get("criar_banco"),
        criar_tabela_tempos_setor=integracoes.get("criar_tabela_tempos_setor"),
    )
    @app.route("/importar-maio", methods=["GET", "POST"])
    @perfil_permitido("pcp")
    def importar_maio():
        resultado = None
        resumo = None
        erros = []
        avisos = []
        importado = False
        if request.method == "POST":
            arquivo = request.files.get("arquivo")
            acao = request.form.get("acao", "validar")
            substituir = request.form.get("substituir") == "sim"
            if not arquivo or not arquivo.filename:
                erros.append("Selecione a planilha de maio em Excel.")
            else:
                try:
                    dados = ler_planilha_importacao_maio(arquivo)
                    erros = dados.get("erros", [])
                    avisos = dados.get("avisos", [])
                    resumo = resumir_importacao_maio(dados)
                    if acao == "importar" and not erros:
                        if not substituir:
                            erros.append("Marque a confirmação de substituição segura dos dados oficiais de Maio/2026 importados anteriormente.")
                        else:
                            resultado = importar_dados_oficiais_maio(dados, substituir=True)
                            importado = True
                            flash("Importação oficial de Maio/2026 concluída com sucesso.")
                except Exception as erro:
                    erros.append(str(erro))
        return render_template("importar_maio.html", resumo=resumo, erros=erros, avisos=avisos, resultado=resultado, importado=importado)

"""Servicos de Financeiro e Movimentacoes."""

import calendar
import hashlib
import json
import re
import time
import unicodedata
import uuid
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.datavalidation import DataValidation

from database import DATABASE_URL, conectar, q
from database.migrations import executar_alteracao_segura
from modules.financeiro.services import (
    campos_derivados_conta,
    categorias_entradas_financeiras,
    categorias_saidas_financeiras,
    criar_tabela_plano_contas_mestre,
    listar_plano_contas,
    opcoes_reclassificacao_plano,
    resolver_conta_plano,
)
from modules.movimentacoes.origens import (
    IMPORTACAO_FINANCEIRA,
    LEGADO_IMPORTADO,
    MANUAL_CONTROLADO,
    MODOS_ORIGEM,
    STATUS_LINHA_CONFLITANTE,
    STATUS_LINHA_IDENTICA,
    STATUS_LINHA_IMPORTADA,
    STATUS_LINHA_REJEITADA,
    aplicar_origem_leitura,
    criar_lote_importacao_cursor,
    criar_tabelas_governanca_financeira,
    falhar_lote_importacao_cursor,
    finalizar_lote_importacao_cursor,
    hash_normalizado,
    hash_sha256,
    registrar_linha_importacao_cursor,
    registrar_origem_principal_cursor,
    sanitizar_nome_arquivo,
    vincular_auditoria_origem_cursor,
)


def tentar_alter_table(cursor, conn, comando):
    executar_alteracao_segura(cursor, conn, comando)


CATEGORIAS_FINANCEIRAS_ENTRADA = categorias_entradas_financeiras()
CATEGORIAS_FINANCEIRAS_SAIDA = categorias_saidas_financeiras()
CATEGORIA_NAO_CLASSIFICADO = "Não Classificado"
CATEGORIA_RECEITA_BRUTA = "Receita Bruta"
ORIGEM_IMPORTACAO_DESPESAS = "excel_movimentacoes"
ORIGEM_IMPORTACAO_VENDAS = "IMPORTACAO VENDAS"
_PLANO_CONTAS_SYNC_EXECUTADO = False
_SCHEMA_MOVIMENTACOES_FINANCEIRAS_INICIALIZADO = False
TAMANHO_LOTE_GRAVACAO_IMPORTACAO = 500
HOTFIX_APORTES_NATUREZA_CHAVE = "20260714_aportes_entrada_fluxo"
TABELA_BACKUP_APORTES_NATUREZA = "movimentacoes_financeiras_backup_aportes_natureza_20260714"
ACOES_AUDITORIA_FINANCEIRA = {
    "CRIACAO_MANUAL",
    "CRIACAO_IMPORTACAO",
    "EDICAO",
    "CANCELAMENTO",
    "REABERTURA",
    "RECLASSIFICACAO",
    "CONFLITO_IMPORTACAO",
    "ALTERACAO_ORIGEM",
    "HOTFIX_APORTES",
}

CAMPOS_PLANILHA_OFICIAL = [
    "Data do Documento",
    "Data de Vencimento",
    "Data de Pagamento",
    "Numero do Documento",
    "CNPJ/CPF",
    "Favorecido/Cliente",
    "Descricao",
    "Historico",
    "Tipo/Natureza",
    "Categoria",
    "Subcategoria",
    "Centro de Analise",
    "Valor do Documento",
    "Valor Pago",
    "Valor Liquido",
    "Origem",
    "Observacoes",
]


def derivar_plano_movimentacao(categoria, plano_conta_id=None):
    conta = resolver_conta_plano(categoria=categoria, plano_conta_id=plano_conta_id)
    campos = campos_derivados_conta(conta)
    if not campos["categoria_movimentacao"]:
        campos["categoria_movimentacao"] = categoria or CATEGORIA_NAO_CLASSIFICADO
    return campos


def somar_tempo_importacao(resultado, etapa, segundos):
    resultado["tempos_etapas"][etapa] = resultado["tempos_etapas"].get(etapa, 0) + float(segundos or 0)


def finalizar_tempos_importacao(resultado):
    resultado["tempo_processamento_segundos"] = round(resultado["tempo_processamento_segundos"], 2)
    resultado["tempos_etapas"] = {
        etapa: round(segundos, 4)
        for etapa, segundos in resultado.get("tempos_etapas", {}).items()
    }
    return resultado


def normalizar_texto_plano(valor):
    return normalizar_cabecalho_importacao(valor)


def conta_eh_aportes(conta=None, categoria=None):
    if conta:
        campos = [
            conta.get("nome"),
            conta.get("categoria"),
            conta.get("categoria_plano"),
        ]
        if any(normalizar_texto_plano(campo) == "aportes" for campo in campos):
            return True
    return normalizar_texto_plano(categoria) == "aportes"


def resolver_tipo_caixa_importacao(tipo_informado, conta_plano):
    if conta_eh_aportes(conta_plano):
        return "Entrada"
    return tipo_informado


def buscar_conta_plano_importacao(categoria, subcategoria=""):
    categoria_norm = normalizar_texto_plano(categoria)
    subcategoria_norm = normalizar_texto_plano(subcategoria)
    if not categoria_norm:
        return None

    candidatos = []
    for conta in listar_plano_contas():
        if normalizar_texto_plano(conta["categoria"]) == categoria_norm:
            candidatos.append(conta)
        elif normalizar_texto_plano(conta["nome"]) == categoria_norm and not subcategoria_norm:
            return conta

    if subcategoria_norm:
        for conta in candidatos:
            if normalizar_texto_plano(conta["subcategoria"]) == subcategoria_norm:
                return conta
        return None

    sem_subcategoria = [conta for conta in candidatos if not conta["subcategoria"]]
    if len(sem_subcategoria) == 1:
        return sem_subcategoria[0]
    if len(candidatos) == 1:
        return candidatos[0]
    return resolver_conta_plano(categoria=categoria)


FORMAS_PAGAMENTO_FINANCEIRO = [
    "Pix",
    "Dinheiro",
    "Boleto",
    "Cartão",
    "Transferência",
    "Cheque",
    "Outro"
]

# Status que o usuário escolhe no lançamento.
# "Em atraso" não deve ser escolhido manualmente; o sistema calcula pela data de vencimento.
STATUS_FINANCEIRO = [
    "Pendente",
    "Realizado",
    "Cancelado"
]

# Opções usadas nos filtros e na leitura gerencial da tela.
STATUS_FINANCEIRO_FILTRO = [
    "Todos",
    "A vencer",
    "Em atraso",
    "Realizado",
    "Cancelado"
]


def calcular_status_financeiro_visual(item, data_referencia=None):
    """
    Calcula o status visual/gerencial sem alterar o status gravado no banco.

    Regra:
    - Cancelado permanece Cancelado;
    - Realizado vira Liquidado;
    - Pendente com vencimento anterior à data de referência vira Em atraso;
    - Pendente com vencimento igual ou posterior vira A vencer.

    Observação de segurança:
    Se existir algum registro antigo com status "Atrasado", ele será exibido como "Em atraso"
    para manter compatibilidade com dados já lançados.
    """
    if data_referencia is None:
        data_referencia = datetime.now().date()

    status = (item.get("status") if hasattr(item, "get") else item["status"]) or "Pendente"
    data_vencimento = (item.get("data_vencimento") if hasattr(item, "get") else item["data_vencimento"]) or ""

    if status == "Cancelado":
        return "Cancelado"

    if status == "Realizado":
        return "Liquidado"

    if status == "Atrasado":
        return "Em atraso"

    try:
        vencimento = datetime.strptime(data_vencimento, "%Y-%m-%d").date()
    except Exception:
        return "A vencer"

    if vencimento < data_referencia:
        return "Em atraso"

    return "A vencer"


def preparar_movimentacoes_financeiras_para_tela(movimentacoes, status_filtro="Todos"):
    """
    Converte as linhas retornadas do banco em dicionários e adiciona campos calculados.
    Isso evita alterar a estrutura do banco e protege a lógica já existente.
    """
    hoje_data = datetime.now().date()
    resultado = []

    for item in movimentacoes:
        item_dict = aplicar_origem_leitura(dict(item))
        status_visual = calcular_status_financeiro_visual(item_dict, hoje_data)
        item_dict["status_original"] = item_dict.get("status", "Pendente")
        item_dict["status_visual"] = status_visual

        # Classe CSS simples para futura estilização do badge, sem obrigar mudança no template.
        item_dict["status_classe"] = (
            status_visual.lower()
            .replace(" ", "-")
            .replace("ç", "c")
            .replace("ã", "a")
        )

        if status_filtro == "Todos" or status_visual == status_filtro:
            resultado.append(item_dict)

    return resultado


def criar_tabela_movimentacoes_financeiras():
    global _SCHEMA_MOVIMENTACOES_FINANCEIRAS_INICIALIZADO

    if _SCHEMA_MOVIMENTACOES_FINANCEIRAS_INICIALIZADO:
        return

    criar_tabela_plano_contas_mestre()
    conn = conectar()
    cursor = conn.cursor()

    try:
        if DATABASE_URL:
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS movimentacoes_financeiras (
                id SERIAL PRIMARY KEY,
                data_vencimento TEXT NOT NULL,
                data_realizacao TEXT,
                tipo TEXT NOT NULL,
                categoria TEXT NOT NULL,
                descricao TEXT NOT NULL,
                valor REAL NOT NULL,
                forma_pagamento TEXT,
                status TEXT DEFAULT 'Pendente',
                parcelas INTEGER DEFAULT 1,
                parcela_atual INTEGER DEFAULT 1,
                intervalo_dias INTEGER DEFAULT 30,
                documento_id TEXT,
                data_documento TEXT,
                valor_documento REAL DEFAULT 0,
                prazo_medio_dias REAL DEFAULT 0,
                observacoes TEXT,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
        else:
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS movimentacoes_financeiras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_vencimento TEXT NOT NULL,
                data_realizacao TEXT,
                tipo TEXT NOT NULL,
                categoria TEXT NOT NULL,
                descricao TEXT NOT NULL,
                valor REAL NOT NULL,
                forma_pagamento TEXT,
                status TEXT DEFAULT 'Pendente',
                parcelas INTEGER DEFAULT 1,
                parcela_atual INTEGER DEFAULT 1,
                intervalo_dias INTEGER DEFAULT 30,
                documento_id TEXT,
                data_documento TEXT,
                valor_documento REAL DEFAULT 0,
                prazo_medio_dias REAL DEFAULT 0,
                observacoes TEXT,
                criado_em TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """)

        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN intervalo_dias INTEGER DEFAULT 30")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN documento_id TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN data_documento TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN valor_documento REAL DEFAULT 0")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN prazo_medio_dias REAL DEFAULT 0")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN import_key TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN cnpj_cpf TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN numero_documento TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN favorecido TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN parceiro TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN historico TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN valor_pago REAL DEFAULT 0")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN valor_liquido REAL DEFAULT 0")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN origem_importacao TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN plano_conta_id INTEGER")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN grupo_gerencial TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN categoria_plano TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN subcategoria TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN centro_analise TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN linha_dre TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN tipo_conta TEXT")
        tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN impacta_fluxo_caixa INTEGER DEFAULT 1")
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_movimentacoes_financeiras_import_key "
            "ON movimentacoes_financeiras (import_key)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_vencimento_tipo_status "
            "ON movimentacoes_financeiras (data_vencimento, tipo, status)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_realizacao_tipo_status "
            "ON movimentacoes_financeiras (data_realizacao, tipo, status)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_tipo_documento_status_categoria "
            "ON movimentacoes_financeiras (tipo, data_documento, status, categoria)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_categoria "
            "ON movimentacoes_financeiras (categoria)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_origem_importacao "
            "ON movimentacoes_financeiras (origem_importacao)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_plano_conta "
            "ON movimentacoes_financeiras (plano_conta_id)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_linha_dre "
            "ON movimentacoes_financeiras (linha_dre, data_documento, status)"
        )
        tentar_alter_table(
            cursor,
            conn,
            "CREATE INDEX IF NOT EXISTS idx_mov_fin_impacta_fluxo "
            "ON movimentacoes_financeiras (impacta_fluxo_caixa, data_vencimento, data_realizacao)"
        )
        criar_tabela_auditoria_movimentacoes_financeiras(cursor)

        conn.commit()
        _SCHEMA_MOVIMENTACOES_FINANCEIRAS_INICIALIZADO = True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def criar_tabela_auditoria_movimentacoes_financeiras(cursor=None):
    """Cria somente a estrutura aditiva da trilha; nunca altera movimentos."""
    conexao_propria = cursor is None
    conn = conectar() if conexao_propria else None
    cursor = cursor or conn.cursor()
    id_pk = "SERIAL PRIMARY KEY" if DATABASE_URL else "INTEGER PRIMARY KEY AUTOINCREMENT"
    timestamp_type = "TIMESTAMP" if DATABASE_URL else "TEXT"
    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS movimentacoes_financeiras_auditoria (
        id {id_pk},
        movimentacao_id INTEGER NOT NULL,
        acao TEXT NOT NULL,
        estado_anterior TEXT,
        estado_posterior TEXT,
        usuario_id INTEGER,
        usuario_nome TEXT NOT NULL,
        perfil TEXT NOT NULL,
        data_hora {timestamp_type} NOT NULL DEFAULT CURRENT_TIMESTAMP,
        justificativa TEXT,
        origem_acao TEXT NOT NULL,
        referencia_importacao TEXT
    )
    """)
    cursor.execute("""
    CREATE INDEX IF NOT EXISTS idx_mov_fin_auditoria_movimentacao_data
    ON movimentacoes_financeiras_auditoria (movimentacao_id, data_hora, id)
    """)
    if conexao_propria:
        conn.commit()
        conn.close()


def _estado_json(estado):
    if estado is None:
        return None
    if not isinstance(estado, dict):
        estado = dict(estado)
    return json.dumps(estado, ensure_ascii=False, sort_keys=True, default=str)


def _registrar_auditoria_cursor(
    cursor,
    movimentacao_id,
    acao,
    estado_anterior,
    estado_posterior,
    usuario_id=None,
    usuario_nome="Sistema",
    perfil="sistema",
    justificativa="",
    origem_acao="sistema",
    referencia_importacao=None,
):
    if acao not in ACOES_AUDITORIA_FINANCEIRA:
        raise ValueError("Acao de auditoria financeira invalida.")
    cursor.execute(q("""
    INSERT INTO movimentacoes_financeiras_auditoria (
        movimentacao_id, acao, estado_anterior, estado_posterior,
        usuario_id, usuario_nome, perfil, justificativa,
        origem_acao, referencia_importacao
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """), (
        movimentacao_id,
        acao,
        _estado_json(estado_anterior),
        _estado_json(estado_posterior),
        usuario_id,
        usuario_nome or "Sistema",
        perfil or "sistema",
        (justificativa or "").strip(),
        origem_acao or "sistema",
        referencia_importacao,
    ))
    if DATABASE_URL:
        cursor.execute("SELECT LASTVAL() AS id")
        return cursor.fetchone()["id"]
    return cursor.lastrowid


def buscar_historico_movimentacao_financeira(movimentacao_id):
    criar_tabela_movimentacoes_financeiras()
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q("""
    SELECT *
    FROM movimentacoes_financeiras_auditoria
    WHERE movimentacao_id = ?
    ORDER BY data_hora ASC, id ASC
    """), (movimentacao_id,))
    eventos = [dict(item) for item in cursor.fetchall()]
    conn.close()
    for evento in eventos:
        anterior = json.loads(evento["estado_anterior"]) if evento.get("estado_anterior") else {}
        posterior = json.loads(evento["estado_posterior"]) if evento.get("estado_posterior") else {}
        campos = sorted({
            campo
            for campo in set(anterior) | set(posterior)
            if anterior.get(campo) != posterior.get(campo)
        })
        evento["campos_alterados"] = campos
        evento["resumo_alteracoes"] = ", ".join(campos) if campos else "Sem alteração de campos"
    return eventos


def valor_realizado_movimentacao(movimentacao):
    """Valor de caixa efetivo, com fallback integral restrito ao legado quitado."""
    item = movimentacao if hasattr(movimentacao, "get") else dict(movimentacao)
    if (item.get("status") or "Pendente") == "Cancelado":
        return 0.0
    if not str(item.get("data_realizacao") or "").strip():
        return 0.0
    if (item.get("status") or "") not in ("Pago", "Recebido", "Realizado"):
        return 0.0
    try:
        valor_pago = float(item.get("valor_pago") or 0)
    except (TypeError, ValueError):
        valor_pago = 0.0
    if valor_pago > 0:
        return round(valor_pago, 2)
    try:
        valor_legado = float(item.get("valor") or item.get("valor_documento") or 0)
    except (TypeError, ValueError):
        valor_legado = 0.0
    return round(max(valor_legado, 0), 2)


def sincronizar_movimentacoes_plano_contas(
    justificativa="",
    usuario_nome="CLI administrativo",
    perfil="admin",
):
    from .reset_financeiro import require_financial_writes_enabled

    require_financial_writes_enabled("Sincronizacao historica do plano de contas")
    global _PLANO_CONTAS_SYNC_EXECUTADO
    justificativa = (justificativa or "").strip()
    if not justificativa:
        raise ValueError("A sincronizacao exige justificativa explicita.")
    if _PLANO_CONTAS_SYNC_EXECUTADO:
        return {"executado": False, "motivo": "sincronizacao_ja_executada_no_processo"}

    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM movimentacoes_financeiras")
        anteriores = {item["id"]: dict(item) for item in cursor.fetchall()}
        for conta in listar_plano_contas():
            campos = campos_derivados_conta(conta)
            chaves = [conta["nome"], conta["categoria"]]
            chaves.extend(conta.get("aliases", []))
            for categoria in {chave for chave in chaves if chave}:
                cursor.execute(q("""
                UPDATE movimentacoes_financeiras
                SET plano_conta_id = ?,
                    grupo_gerencial = ?,
                    categoria_plano = ?,
                    subcategoria = ?,
                    centro_analise = ?,
                    linha_dre = ?,
                    tipo_conta = ?,
                    impacta_fluxo_caixa = ?
                WHERE (
                    plano_conta_id IS NULL
                    OR plano_conta_id = 0
                    OR COALESCE(linha_dre, '') = ''
                    OR COALESCE(impacta_fluxo_caixa, -1) <> ?
                    OR linha_dre <> ?
                    OR tipo_conta <> ?
                )
                  AND categoria = ?
                """), (
                    campos["plano_conta_id"],
                    campos["grupo_gerencial"],
                    campos["categoria_plano"],
                    campos["subcategoria"],
                    campos["centro_analise"],
                    campos["linha_dre"],
                    campos["tipo_conta"],
                    int(campos["impacta_fluxo_caixa"]),
                    int(campos["impacta_fluxo_caixa"]),
                    campos["linha_dre"],
                    campos["tipo_conta"],
                    categoria,
                ))
        atualizadas = 0
        for movimentacao_id, anterior in anteriores.items():
            cursor.execute(
                q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"),
                (movimentacao_id,),
            )
            posterior = cursor.fetchone()
            if posterior and _estado_json(anterior) != _estado_json(posterior):
                atualizadas += 1
                _registrar_auditoria_cursor(
                    cursor,
                    movimentacao_id,
                    "RECLASSIFICACAO",
                    anterior,
                    posterior,
                    usuario_nome=usuario_nome,
                    perfil=perfil,
                    justificativa=justificativa,
                    origem_acao="cli_sincronizacao_plano",
                    referencia_importacao=anterior.get("import_key"),
                )
        conn.commit()
        _PLANO_CONTAS_SYNC_EXECUTADO = True
        return {"executado": True, "quantidade_atualizada": atualizadas}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def criar_tabelas_controle_hotfix_aportes(cursor):
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS hotfix_aportes_natureza_fluxo (
        chave TEXT PRIMARY KEY,
        executado_em TEXT NOT NULL,
        quantidade_corrigida INTEGER NOT NULL,
        valor_corrigido REAL NOT NULL
    )
    """)
    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS {TABELA_BACKUP_APORTES_NATUREZA} AS
    SELECT *
    FROM movimentacoes_financeiras
    WHERE 1 = 0
    """)


def hotfix_aportes_ja_executado(cursor):
    cursor.execute(q("""
    SELECT chave
    FROM hotfix_aportes_natureza_fluxo
    WHERE chave = ?
    """), (HOTFIX_APORTES_NATUREZA_CHAVE,))
    return cursor.fetchone() is not None


def corrigir_natureza_aportes_fluxo_caixa(
    justificativa="",
    usuario_nome="CLI administrativo",
    perfil="admin",
):
    from .reset_financeiro import require_financial_writes_enabled

    require_financial_writes_enabled("Hotfix historico de aportes")
    justificativa = (justificativa or "").strip()
    if not justificativa:
        raise ValueError("O hotfix exige justificativa explicita.")
    conn = conectar()
    cursor = conn.cursor()
    try:
        criar_tabelas_controle_hotfix_aportes(cursor)
        if hotfix_aportes_ja_executado(cursor):
            conn.commit()
            return {"executado": False, "motivo": "hotfix_ja_aplicado"}

        cursor.execute(q("""
        SELECT COUNT(*) as quantidade,
               COALESCE(SUM(valor), 0) as valor_total
        FROM movimentacoes_financeiras
        WHERE categoria = ?
          AND COALESCE(tipo, '') <> ?
        """), ("Aportes", "Entrada"))
        alvo = cursor.fetchone()
        quantidade = int(alvo["quantidade"] or 0)
        valor_total = round(float(alvo["valor_total"] or 0), 2)
        cursor.execute(q("""
        SELECT *
        FROM movimentacoes_financeiras
        WHERE categoria = ?
          AND COALESCE(tipo, '') <> ?
        """), ("Aportes", "Entrada"))
        anteriores = [dict(item) for item in cursor.fetchall()]

        cursor.execute(q(f"""
        INSERT INTO {TABELA_BACKUP_APORTES_NATUREZA}
        SELECT m.*
        FROM movimentacoes_financeiras m
        WHERE m.categoria = ?
          AND COALESCE(m.tipo, '') <> ?
          AND NOT EXISTS (
              SELECT 1
              FROM {TABELA_BACKUP_APORTES_NATUREZA} b
              WHERE b.id = m.id
          )
        """), ("Aportes", "Entrada"))

        cursor.execute(q("""
        UPDATE movimentacoes_financeiras
        SET tipo = ?,
            tipo_conta = ?,
            linha_dre = ?,
            impacta_fluxo_caixa = ?
        WHERE categoria = ?
          AND COALESCE(tipo, '') <> ?
        """), ("Entrada", "Entrada", "Neutro", 1, "Aportes", "Entrada"))

        cursor.execute(q("""
        INSERT INTO hotfix_aportes_natureza_fluxo (
            chave, executado_em, quantidade_corrigida, valor_corrigido
        ) VALUES (?, ?, ?, ?)
        """), (
            HOTFIX_APORTES_NATUREZA_CHAVE,
            datetime.now().isoformat(timespec="seconds"),
            quantidade,
            valor_total,
        ))
        for anterior in anteriores:
            cursor.execute(
                q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"),
                (anterior["id"],),
            )
            posterior = cursor.fetchone()
            _registrar_auditoria_cursor(
                cursor,
                anterior["id"],
                "HOTFIX_APORTES",
                anterior,
                posterior,
                usuario_nome=usuario_nome,
                perfil=perfil,
                justificativa=justificativa,
                origem_acao="cli_hotfix_aportes",
                referencia_importacao=anterior.get("import_key"),
            )
        conn.commit()
        return {
            "executado": True,
            "quantidade_corrigida": quantidade,
            "valor_corrigido": valor_total,
            "backup": TABELA_BACKUP_APORTES_NATUREZA,
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def adicionar_meses(data_base, meses):
    ano = data_base.year
    mes = data_base.month + meses

    while mes > 12:
        mes -= 12
        ano += 1

    while mes < 1:
        mes += 12
        ano -= 1

    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dia = min(data_base.day, ultimo_dia)

    return data_base.replace(year=ano, month=mes, day=dia)


def salvar_movimentacao_financeira(
    form,
    usuario_id=None,
    usuario_nome="Sistema",
    perfil="sistema",
    origem_acao="interface_manual",
):
    from .reset_financeiro import require_financial_writes_enabled

    require_financial_writes_enabled("Criacao manual de movimentacao financeira")
    criar_tabela_movimentacoes_financeiras()
    criar_tabelas_governanca_financeira()

    tipo = form.get("tipo", "").strip()
    categoria = form.get("categoria", "").strip()
    descricao = form.get("descricao", "").strip()
    data_documento = form.get("data_documento") or datetime.now().strftime("%Y-%m-%d")
    data_realizacao = form.get("data_realizacao", "")
    forma_pagamento = form.get("forma_pagamento", "")
    status = form.get("status", "Pendente")
    observacoes = form.get("observacoes", "")
    valor_documento = float(form.get("valor") or 0)
    justificativa_manual = (form.get("justificativa") or "").strip()
    referencia_evidencia = (form.get("referencia_evidencia") or "").strip()

    vencimentos = form.getlist("parcela_vencimento[]")
    valores = form.getlist("parcela_valor[]")

    if not usuario_id or not str(usuario_nome or "").strip():
        raise ValueError("Usuário autenticado é obrigatório para lançamento manual.")
    if len(justificativa_manual) < 5:
        raise ValueError("Informe uma justificativa objetiva para o lançamento manual.")
    if len(referencia_evidencia) < 3:
        raise ValueError("Informe a referência documental ou evidência do lançamento.")

    if tipo not in ["Entrada", "Saída"]:
        raise ValueError("Tipo de movimentação inválido.")

    if not descricao:
        raise ValueError("Informe uma descrição para a movimentação.")

    if valor_documento <= 0:
        raise ValueError("O valor total do documento deve ser maior que zero.")

    parcelas_validas = []

    for vencimento, valor in zip(vencimentos, valores):
        vencimento = (vencimento or "").strip()
        valor = float(valor or 0)

        if vencimento and valor > 0:
            parcelas_validas.append({
                "vencimento": vencimento,
                "valor": round(valor, 2)
            })

    if not parcelas_validas:
        parcelas_validas.append({
            "vencimento": data_documento,
            "valor": round(valor_documento, 2)
        })

    soma_parcelas = round(sum(item["valor"] for item in parcelas_validas), 2)

    if abs(soma_parcelas - round(valor_documento, 2)) > 0.02:
        raise ValueError(
            f"A soma das parcelas (R$ {soma_parcelas:.2f}) precisa bater com o valor total do documento (R$ {valor_documento:.2f})."
        )

    data_base = datetime.strptime(data_documento, "%Y-%m-%d")
    prazo_ponderado = 0

    for item in parcelas_validas:
        data_vencimento = datetime.strptime(item["vencimento"], "%Y-%m-%d")
        dias = (data_vencimento - data_base).days
        prazo_ponderado += item["valor"] * dias

    prazo_medio_dias = prazo_ponderado / valor_documento if valor_documento > 0 else 0
    documento_id = uuid.uuid4().hex
    total_parcelas = len(parcelas_validas)
    plano = derivar_plano_movimentacao(categoria, form.get("plano_conta_id"))
    if not plano["plano_conta_id"] or not plano["grupo_gerencial"] or not plano["linha_dre"]:
        raise ValueError("Selecione uma conta do Plano de Contas com grupo e linha DRE estruturados.")
    categoria = plano["categoria_movimentacao"]

    conn = conectar()
    cursor = conn.cursor()

    try:
        for indice, parcela in enumerate(parcelas_validas, start=1):
            descricao_parcela = descricao

            if total_parcelas > 1:
                descricao_parcela = f"{descricao} ({indice}/{total_parcelas})"

            cursor.execute(q("""
            INSERT INTO movimentacoes_financeiras (
                data_vencimento,
                data_realizacao,
                tipo,
                categoria,
                descricao,
                valor,
                forma_pagamento,
                status,
                parcelas,
                parcela_atual,
                intervalo_dias,
                documento_id,
                data_documento,
                valor_documento,
                prazo_medio_dias,
                observacoes,
                plano_conta_id,
                grupo_gerencial,
                categoria_plano,
                subcategoria,
                centro_analise,
                linha_dre,
                tipo_conta,
                impacta_fluxo_caixa
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """), (
                parcela["vencimento"],
                data_realizacao if status == "Realizado" else "",
                tipo,
                categoria,
                descricao_parcela,
                parcela["valor"],
                forma_pagamento,
                status,
                total_parcelas,
                indice,
                0,
                documento_id,
                data_documento,
                valor_documento,
                round(prazo_medio_dias, 2),
                observacoes,
                plano["plano_conta_id"],
                plano["grupo_gerencial"],
                plano["categoria_plano"],
                plano["subcategoria"],
                plano["centro_analise"],
                plano["linha_dre"],
                plano["tipo_conta"],
                int(plano["impacta_fluxo_caixa"]),
            ))
            if DATABASE_URL:
                cursor.execute("SELECT LASTVAL() AS id")
                movimentacao_id = cursor.fetchone()["id"]
            else:
                movimentacao_id = cursor.lastrowid
            origem_id = registrar_origem_principal_cursor(
                cursor,
                movimentacao_id,
                MANUAL_CONTROLADO,
                sistema_origem="FrigoDatta",
                modulo_origem="Central de Movimentações",
                tipo_evento="LANCAMENTO_MANUAL",
                evento_id_interno=documento_id,
                chave_externa=referencia_evidencia,
                chave_idempotente=f"manual:{documento_id}:{indice}",
                usuario_id=usuario_id,
                usuario_nome=usuario_nome,
                metadados={
                    "justificativa": justificativa_manual,
                    "referencia_evidencia": referencia_evidencia,
                    "parcela": indice,
                    "total_parcelas": total_parcelas,
                },
            )
            cursor.execute(q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"), (movimentacao_id,))
            posterior = cursor.fetchone()
            auditoria_id = _registrar_auditoria_cursor(
                cursor,
                movimentacao_id,
                "CRIACAO_MANUAL",
                None,
                posterior,
                usuario_id=usuario_id,
                usuario_nome=usuario_nome,
                perfil=perfil,
                justificativa=justificativa_manual,
                origem_acao=origem_acao,
            )
            vincular_auditoria_origem_cursor(cursor, origem_id, auditoria_id)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def buscar_movimentacao_financeira_por_id(movimentacao_id):
    criar_tabela_movimentacoes_financeiras()
    criar_tabelas_governanca_financeira()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT
        m.*,
        o.id AS origem_id,
        o.modo AS modo_origem,
        o.sistema_origem,
        o.modulo_origem,
        o.tipo_evento AS origem_tipo_evento,
        o.chave_externa AS origem_chave_externa,
        o.chave_idempotente AS origem_chave_idempotente,
        o.lote_importacao_id,
        o.linha_arquivo AS origem_linha_arquivo,
        o.metadados AS origem_metadados,
        l.arquivo_nome AS lote_arquivo_nome
    FROM movimentacoes_financeiras m
    LEFT JOIN movimentacoes_financeiras_origens o
      ON o.movimentacao_id = m.id
     AND o.papel = 'PRINCIPAL'
     AND o.status = 'ATIVA'
    LEFT JOIN movimentacoes_financeiras_importacao_lotes l
      ON l.id = o.lote_importacao_id
    WHERE m.id = ?
    """), (movimentacao_id,))

    movimentacao = cursor.fetchone()
    conn.close()

    return aplicar_origem_leitura(dict(movimentacao)) if movimentacao else None


def atualizar_movimentacao_financeira(
    movimentacao_id,
    form,
    usuario_id=None,
    usuario_nome="Sistema",
    perfil="sistema",
    origem_acao="interface_edicao",
):
    from .reset_financeiro import require_financial_writes_enabled

    require_financial_writes_enabled("Edicao, realizacao ou reabertura de movimentacao financeira")
    criar_tabela_movimentacoes_financeiras()
    justificativa = (form.get("justificativa") or "").strip()
    if not justificativa:
        raise ValueError("Informe a justificativa da alteração.")
    plano = derivar_plano_movimentacao(form.get("categoria", ""), form.get("plano_conta_id"))

    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute(q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"), (movimentacao_id,))
        anterior = cursor.fetchone()
        if not anterior:
            raise ValueError("Movimentação financeira não encontrada.")
        cursor.execute(q("""
        UPDATE movimentacoes_financeiras
        SET data_vencimento = ?,
            data_realizacao = ?,
            tipo = ?,
            categoria = ?,
            descricao = ?,
            valor = ?,
            forma_pagamento = ?,
            status = ?,
            intervalo_dias = ?,
            observacoes = ?,
            plano_conta_id = ?,
            grupo_gerencial = ?,
            categoria_plano = ?,
            subcategoria = ?,
            centro_analise = ?,
            linha_dre = ?,
            tipo_conta = ?,
            impacta_fluxo_caixa = ?
        WHERE id = ?
        """), (
            form.get("data_vencimento", ""),
            form.get("data_realizacao", ""),
            form.get("tipo", ""),
            plano["categoria_movimentacao"],
            form.get("descricao", ""),
            float(form.get("valor") or 0),
            form.get("forma_pagamento", ""),
            form.get("status", ""),
            int(form.get("intervalo_dias") or 30),
            form.get("observacoes", ""),
            plano["plano_conta_id"],
            plano["grupo_gerencial"],
            plano["categoria_plano"],
            plano["subcategoria"],
            plano["centro_analise"],
            plano["linha_dre"],
            plano["tipo_conta"],
            int(plano["impacta_fluxo_caixa"]),
            movimentacao_id
        ))
        cursor.execute(q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"), (movimentacao_id,))
        posterior = cursor.fetchone()
        _registrar_auditoria_cursor(
            cursor,
            movimentacao_id,
            (
                "REABERTURA"
                if anterior["status"] == "Cancelado" and posterior["status"] != "Cancelado"
                else "EDICAO"
            ),
            anterior,
            posterior,
            usuario_id=usuario_id,
            usuario_nome=usuario_nome,
            perfil=perfil,
            justificativa=justificativa,
            origem_acao=origem_acao,
            referencia_importacao=anterior["import_key"],
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def excluir_movimentacao_financeira(
    movimentacao_id,
    justificativa="",
    usuario_id=None,
    usuario_nome="Sistema",
    perfil="sistema",
    origem_acao="alias_exclusao",
):
    """Compatibilidade do alias antigo: cancela e audita, sem exclusão física."""
    from .reset_financeiro import require_financial_writes_enabled

    require_financial_writes_enabled("Cancelamento de movimentacao financeira")
    criar_tabela_movimentacoes_financeiras()
    justificativa = (justificativa or "").strip()
    if not justificativa:
        raise ValueError("Informe a justificativa do cancelamento.")

    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute(q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"), (movimentacao_id,))
        anterior = cursor.fetchone()
        if not anterior:
            raise ValueError("Movimentação financeira não encontrada.")
        if anterior["status"] == "Cancelado":
            raise ValueError("A movimentação já está cancelada.")
        cursor.execute(q("""
        UPDATE movimentacoes_financeiras
        SET status = ?
        WHERE id = ?
        """), ("Cancelado", movimentacao_id))
        cursor.execute(q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"), (movimentacao_id,))
        posterior = cursor.fetchone()
        _registrar_auditoria_cursor(
            cursor,
            movimentacao_id,
            "CANCELAMENTO",
            anterior,
            posterior,
            usuario_id=usuario_id,
            usuario_nome=usuario_nome,
            perfil=perfil,
            justificativa=justificativa,
            origem_acao=origem_acao,
            referencia_importacao=anterior["import_key"],
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def alterar_origem_principal_movimentacao(
    movimentacao_id,
    novo_modo,
    justificativa,
    usuario_id,
    usuario_nome,
    perfil="pcp",
):
    """Altera uma origem já persistida com justificativa e trilha da Onda 0."""
    from .reset_financeiro import require_financial_writes_enabled

    require_financial_writes_enabled("Alteracao de origem de movimentacao financeira")
    criar_tabela_movimentacoes_financeiras()
    criar_tabelas_governanca_financeira()
    justificativa = (justificativa or "").strip()
    if not justificativa:
        raise ValueError("A alteração de origem exige justificativa.")
    if novo_modo not in {MANUAL_CONTROLADO, IMPORTACAO_FINANCEIRA}:
        raise ValueError("Modo de origem não permitido para alteração nesta etapa.")
    if not usuario_id or not str(usuario_nome or "").strip():
        raise ValueError("Usuário autenticado é obrigatório.")

    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute(q("""
        SELECT *
        FROM movimentacoes_financeiras_origens
        WHERE movimentacao_id = ?
          AND papel = 'PRINCIPAL'
          AND status = 'ATIVA'
        """), (movimentacao_id,))
        anterior = cursor.fetchone()
        if not anterior:
            raise ValueError("Origem persistida não encontrada; legado não é materializado em edição.")
        if anterior["modo"] == novo_modo:
            raise ValueError("A origem informada já é a origem atual.")

        cursor.execute(q("""
        UPDATE movimentacoes_financeiras_origens
        SET modo = ?
        WHERE id = ?
        """), (novo_modo, anterior["id"]))
        cursor.execute(q("""
        SELECT *
        FROM movimentacoes_financeiras_origens
        WHERE id = ?
        """), (anterior["id"],))
        posterior = cursor.fetchone()
        auditoria_id = _registrar_auditoria_cursor(
            cursor,
            movimentacao_id,
            "ALTERACAO_ORIGEM",
            anterior,
            posterior,
            usuario_id=usuario_id,
            usuario_nome=usuario_nome,
            perfil=perfil,
            justificativa=justificativa,
            origem_acao="governanca_origem",
            referencia_importacao=posterior["chave_idempotente"],
        )
        vincular_auditoria_origem_cursor(cursor, anterior["id"], auditoria_id)
        conn.commit()
        return dict(posterior)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def buscar_movimentacoes_financeiras(
    data_inicio,
    data_fim,
    tipo_filtro,
    status_filtro,
    modo_origem_filtro="Todos",
):
    criar_tabelas_governanca_financeira()
    condicoes = ["m.data_vencimento BETWEEN ? AND ?"]
    parametros = [data_inicio, data_fim]

    if tipo_filtro in ["Entrada", "Saída"]:
        condicoes.append("m.tipo = ?")
        parametros.append(tipo_filtro)

    # O filtro por status visual é aplicado depois da consulta, porque "A vencer" e
    # "Em atraso" são calculados pela data de vencimento, não gravados no banco.
    if status_filtro == "Realizado":
        condicoes.append("m.status = ?")
        parametros.append("Realizado")
    elif status_filtro == "Cancelado":
        condicoes.append("m.status = ?")
        parametros.append("Cancelado")

    if modo_origem_filtro == LEGADO_IMPORTADO:
        condicoes.append("o.id IS NULL")
    elif modo_origem_filtro in MODOS_ORIGEM and modo_origem_filtro != LEGADO_IMPORTADO:
        condicoes.append("o.modo = ?")
        parametros.append(modo_origem_filtro)

    where_sql = " AND ".join(condicoes)

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q(f"""
    SELECT
        m.*,
        o.id AS origem_id,
        o.modo AS modo_origem,
        o.sistema_origem,
        o.modulo_origem,
        o.tipo_evento AS origem_tipo_evento,
        o.chave_externa AS origem_chave_externa,
        o.chave_idempotente AS origem_chave_idempotente,
        o.lote_importacao_id,
        o.linha_arquivo AS origem_linha_arquivo,
        l.arquivo_nome AS lote_arquivo_nome
    FROM movimentacoes_financeiras m
    LEFT JOIN movimentacoes_financeiras_origens o
      ON o.movimentacao_id = m.id
     AND o.papel = 'PRINCIPAL'
     AND o.status = 'ATIVA'
    LEFT JOIN movimentacoes_financeiras_importacao_lotes l
      ON l.id = o.lote_importacao_id
    WHERE {where_sql}
    ORDER BY m.data_vencimento ASC, m.id ASC
    """), tuple(parametros))

    movimentacoes = cursor.fetchall()
    conn.close()

    return preparar_movimentacoes_financeiras_para_tela(movimentacoes, status_filtro)


def calcular_resumo_financeiro(movimentacoes):
    entradas_previstas = 0
    saidas_previstas = 0
    entradas_realizadas = 0
    saidas_realizadas = 0

    for item in movimentacoes:
        valor = float(item["valor"] or 0)
        valor_realizado = valor_realizado_movimentacao(item)
        tipo = item["tipo"]
        status = item.get("status_original", item.get("status", "Pendente")) if hasattr(item, "get") else item["status"]

        # Cancelados não entram no previsto nem no realizado.
        if status == "Cancelado":
            continue

        if tipo == "Entrada":
            entradas_previstas += valor

            if status == "Realizado":
                entradas_realizadas += valor_realizado

        elif tipo == "Saída":
            saidas_previstas += valor

            if status == "Realizado":
                saidas_realizadas += valor_realizado

    saldo_previsto = entradas_previstas - saidas_previstas
    saldo_realizado = entradas_realizadas - saidas_realizadas

    return {
        "entradas_previstas": round(entradas_previstas, 2),
        "saidas_previstas": round(saidas_previstas, 2),
        "saldo_previsto": round(saldo_previsto, 2),
        "entradas_realizadas": round(entradas_realizadas, 2),
        "saidas_realizadas": round(saidas_realizadas, 2),
        "saldo_realizado": round(saldo_realizado, 2)
    }


def agrupar_fluxo_por_dia(movimentacoes):
    fluxo = {}

    for item in movimentacoes:
        data = item["data_vencimento"]
        valor = float(item["valor"] or 0)

        if data not in fluxo:
            fluxo[data] = {
                "data": data,
                "entradas": 0,
                "saidas": 0,
                "saldo": 0
            }

        status = item.get("status_original", item.get("status", "Pendente")) if hasattr(item, "get") else item["status"]

        # Fluxo diário ignora documentos cancelados.
        if status == "Cancelado":
            continue

        if item["tipo"] == "Entrada":
            fluxo[data]["entradas"] += valor
        else:
            fluxo[data]["saidas"] += valor

        fluxo[data]["saldo"] = fluxo[data]["entradas"] - fluxo[data]["saidas"]

    return [
        {
            "data": item["data"],
            "entradas": round(item["entradas"], 2),
            "saidas": round(item["saidas"], 2),
            "saldo": round(item["saldo"], 2)
        }
        for item in fluxo.values()
    ]


def normalizar_cabecalho_importacao(valor):
    texto = str(valor or "").strip().lower()
    substituicoes = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for origem, destino in substituicoes.items():
        texto = texto.replace(origem, destino)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", texto)


MAPEAMENTO_CABECALHOS_IMPORTACAO = {
    "grupo_gerencial": ["grupo", "grupogerencial", "grupooperacional", "linha", "linhagerencial"],
    "categoria": ["categoria", "classificacao", "classificacaofinanceira"],
    "subcategoria": ["subcategoria", "subcategoriafinanceira"],
    "centro_analise": ["centroanalise", "centrodeanalise", "centrodecusto", "centro"],
    "natureza": ["natureza", "tipo", "tiponatureza", "receitadespesa", "entradasaida", "entradaousaida"],
    "data_documento": ["datadocumento", "datadodocumento", "dataemissao", "emissao", "data", "datavenda", "datadavenda"],
    "cnpj_cpf": ["cnpjcpf", "cnpj", "cpf"],
    "valor_documento": ["valordocumento", "valordodocumento", "valor", "valorvenda", "valordavenda", "total", "totalvenda"],
    "numero_documento": ["numerodocumento", "numerododocumento", "documento", "numero", "nf", "notafiscal", "nota", "cupom", "pedido"],
    "data_vencimento": ["datavencimento", "datadevencimento", "vencimento", "datarecebimento", "previsaorecebimento"],
    "valor_pago": ["valorpago", "pago", "valorpagamento", "valorrecebido", "recebido"],
    "data_pagamento": ["datapagamento", "datadepagamento", "pagamento", "datarealizacao", "datarecebido", "datarecebimento"],
    "favorecido": ["favorecido", "favorecidocliente", "fornecedor", "cliente", "comprador"],
    "valor_liquido": ["valorliquido", "liquido", "valorliquidado", "valorliquidovenda"],
    "descricao": ["descricao", "descricaohistorico", "produto", "item"],
    "parceiro": ["parceiro", "razaosocial", "nome", "nomecliente"],
    "historico": ["historico", "histórico", "observacao", "observacoes", "historicodevenda"],
    "forma_pagamento": ["formapagamento", "formaderecebimento", "meio", "meiopagamento", "meiorecebimento"],
    "origem": ["origem", "origemimportacao", "origemdolancamento"],
    "observacoes": ["observacoes", "observacao", "comentarios", "comentario"],
}


MAPEAMENTO_CONTAS_IMPORTACAO = {
    ("receitabruta", "vendadeproducaopropria"): "Venda de Producao Propria",
    ("receitabruta", "vendadeproduoprpria"): "Venda de Producao Propria",
    ("deducoesdareceita", "devolucoesdevenda"): "Devolucoes",
    ("deduesdareceita", "devoluesdevenda"): "Devolucoes",
    ("cmv", "materiaprima"): "Materia Prima",
    ("cmv", "matriaprima"): "Materia Prima",
    ("cmv", "embalagens"): "Embalagens",
    ("financeiro", "despesasfinanceiras"): "Despesas Financeiras",
    ("maodeobra", "salariosclt"): "Mao de Obra - CLT",
    ("modeobra", "salriosclt"): "Mao de Obra - CLT",
    ("maodeobra", "clt"): "Mao de Obra - CLT",
    ("modeobra", "clt"): "Mao de Obra - CLT",
    ("maodeobra", "terceirizados"): "Mao de Obra - Terceiros",
    ("modeobra", "terceirizados"): "Mao de Obra - Terceiros",
    ("maodeobra", "terceiros"): "Mao de Obra - Terceiros",
    ("modeobra", "terceiros"): "Mao de Obra - Terceiros",
    ("servicos", "terceiros"): "Mao de Obra - Terceiros",
    ("servios", "terceiros"): "Mao de Obra - Terceiros",
    ("servicos", "consultoria"): "Outras Despesas Operacionais",
    ("servios", "consultoria"): "Outras Despesas Operacionais",
    ("servicos", "responsabilidadetecnica"): "Outras Despesas Operacionais",
    ("servios", "responsabilidadetcnica"): "Outras Despesas Operacionais",
    ("materiaisdeapoio", "produtosquimicos"): "Produtos Quimicos",
    ("materiaisdeapoio", "produtosqumicos"): "Produtos Quimicos",
    ("materiaisdeapoio", "usoeconsumo"): "Outras Despesas Operacionais",
    ("materiaisdeapoio", "epis"): "EPIs",
    ("materiaisdeapoio", "uniformes"): "Uniformes",
    ("limpeza", "limpeza"): "Limpeza",
    ("impostos", "encargos"): "Impostos",
    ("comercial", "marketing"): "Marketing",
    ("patrimonial", "aquisicaodeequipamentos"): "Equipamentos",
    ("patrimonial", "aquisiesdeequipamentos"): "Equipamentos",
    ("patrimonial", "obrasebenfeitorias"): "Obras e Benfeitorias",
    ("manutencao", "manutencaodeveiculos"): "Manutencao de Veiculos",
    ("manuteno", "manutenodeveculos"): "Manutencao de Veiculos",
    ("manutencao", "manutencaodeequipamentos"): "Manutencao de Equipamentos",
    ("manuteno", "manutenodeequipamentos"): "Manutencao de Equipamentos",
    ("manutencao", "manutencaopredial"): "Manutencao Predial",
    ("manuteno", "manutenopredial"): "Manutencao Predial",
    ("neutras", "aportes"): "Aportes",
    ("neutra", "aportes"): "Aportes",
}


def conta_padrao_mao_obra_importacao(subcategoria):
    subcategoria_norm = normalizar_texto_plano(subcategoria)
    if not subcategoria_norm:
        return None

    if any(termo in subcategoria_norm for termo in [
        "fgts", "inss", "irrf", "darf", "previdenciario", "previdencirio",
        "encargo", "encargos",
    ]):
        return "Mao de Obra - Encargos"
    if any(termo in subcategoria_norm for termo in [
        "rescisao", "resciso", "rescisorio", "rescisrio", "termode",
        "acordojudicial", "acordotrabalhista",
    ]):
        return "Mao de Obra - Rescisoes"
    if any(termo in subcategoria_norm for termo in ["ferias", "frias"]):
        return "Mao de Obra - Ferias"
    if any(termo in subcategoria_norm for termo in [
        "folha", "salario", "salrios", "pagamento", "horasextras",
    ]):
        return "Mao de Obra - CLT"
    return None


def localizar_conta_plano_por_nome(nome_conta):
    nome_norm = normalizar_texto_plano(nome_conta)
    for conta in listar_plano_contas():
        if normalizar_texto_plano(conta["nome"]) == nome_norm:
            return conta
    return resolver_conta_plano(categoria=nome_conta)


def combinacoes_classificacao_importacao(grupo_gerencial, categoria, subcategoria):
    partes = [
        normalizar_texto_plano(valor)
        for valor in [grupo_gerencial, categoria, subcategoria]
        if normalizar_texto_plano(valor)
    ]

    deduplicadas = []
    for parte in partes:
        if not deduplicadas or deduplicadas[-1] != parte:
            deduplicadas.append(parte)

    candidatos = []
    if len(deduplicadas) >= 2:
        candidatos.append((deduplicadas[0], deduplicadas[-1]))
        candidatos.append(tuple(deduplicadas[-2:]))
    if len(deduplicadas) >= 3:
        candidatos.append(tuple(deduplicadas))

    categoria_norm = normalizar_texto_plano(categoria)
    subcategoria_norm = normalizar_texto_plano(subcategoria)
    grupo_norm = normalizar_texto_plano(grupo_gerencial)
    if categoria_norm and subcategoria_norm:
        candidatos.append((categoria_norm, subcategoria_norm))
    if grupo_norm and categoria_norm:
        candidatos.append((grupo_norm, categoria_norm))

    unicos = []
    for candidato in candidatos:
        if candidato not in unicos:
            unicos.append(candidato)
    return unicos


def formatar_rejeicao_plano_importacao(grupo_gerencial, categoria, subcategoria, conta_esperada, motivo):
    return (
        "Categoria/Subcategoria fora do Plano de Contas. "
        f"Grupo recebido: {grupo_gerencial or '-'} | "
        f"Categoria recebida: {categoria or '-'} | "
        f"Subcategoria recebida: {subcategoria or '-'} | "
        f"Conta esperada: {conta_esperada or '-'} | "
        f"Motivo: {motivo}"
    )


def resolver_conta_plano_importacao(grupo_gerencial, categoria, subcategoria=""):
    if normalizar_texto_plano(categoria) in ["maodeobra", "modeobra"]:
        conta_esperada = conta_padrao_mao_obra_importacao(subcategoria)
        if conta_esperada:
            conta = localizar_conta_plano_por_nome(conta_esperada)
            if conta:
                return conta

    for chave in combinacoes_classificacao_importacao(grupo_gerencial, categoria, subcategoria):
        conta_esperada = MAPEAMENTO_CONTAS_IMPORTACAO.get(chave)
        if not conta_esperada:
            continue

        conta = localizar_conta_plano_por_nome(conta_esperada)
        if conta:
            return conta

        raise ValueError(formatar_rejeicao_plano_importacao(
            grupo_gerencial,
            categoria,
            subcategoria,
            conta_esperada,
            "Conta esperada nao existe no Plano de Contas.",
        ))

    conta = buscar_conta_plano_importacao(categoria, subcategoria)
    if conta:
        return conta

    conta_esperada = None
    for chave in combinacoes_classificacao_importacao(grupo_gerencial, categoria, subcategoria):
        if chave in MAPEAMENTO_CONTAS_IMPORTACAO:
            conta_esperada = MAPEAMENTO_CONTAS_IMPORTACAO[chave]
            break

    raise ValueError(formatar_rejeicao_plano_importacao(
        grupo_gerencial,
        categoria,
        subcategoria,
        conta_esperada,
        "Sem mapeamento para a classificacao operacional recebida.",
    ))


def mapear_cabecalhos_importacao(ws):
    cabecalhos = {}
    for indice, celula in enumerate(ws[1], start=1):
        chave = normalizar_cabecalho_importacao(celula.value)
        if chave:
            cabecalhos[chave] = indice

    colunas = {}
    for campo, aliases in MAPEAMENTO_CABECALHOS_IMPORTACAO.items():
        for alias in aliases:
            chave = normalizar_cabecalho_importacao(alias)
            if chave in cabecalhos:
                colunas[campo] = cabecalhos[chave]
                break

    obrigatorios = ["data_documento", "valor_documento", "natureza", "categoria"]
    ausentes = [campo for campo in obrigatorios if campo not in colunas]
    return colunas, ausentes


def texto_importacao(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def data_importacao(valor):
    if valor in [None, ""]:
        return ""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, (int, float)) and valor > 20000:
        try:
            return from_excel(valor).date().isoformat()
        except Exception:
            return ""

    texto = str(valor).strip()
    for formato in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"]:
        try:
            return datetime.strptime(texto[:10], formato).date().isoformat()
        except ValueError:
            pass
    return ""


def numero_importacao(valor):
    if valor in [None, ""]:
        return 0
    if isinstance(valor, (int, float)):
        return float(valor or 0)
    texto = str(valor).strip()
    if not texto:
        return 0
    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0


def valor_celula(ws, linha, colunas, campo):
    coluna = colunas.get(campo)
    if not coluna:
        return None
    return ws.cell(linha, coluna).value


def natureza_por_categoria(categoria):
    if categoria in CATEGORIAS_FINANCEIRAS_ENTRADA:
        return "Entrada"
    if categoria in CATEGORIAS_FINANCEIRAS_SAIDA:
        return "Saída"
    return ""


def normalizar_natureza_importacao(valor):
    texto = normalizar_cabecalho_importacao(valor)
    if texto in ["receita", "receitas", "entrada", "entradas", "receber", "credito", "creditos"]:
        return "Entrada"
    if texto in ["despesa", "despesas", "saida", "saidas", "pagar", "debito", "debitos"]:
        return "Saida"
    if texto in ["neutro", "neutra", "transferencia", "transferencias"]:
        return "Neutro"
    return ""


def resolver_tipo_importacao(ws, linha, colunas, categoria, natureza_padrao):
    natureza_planilha = normalizar_natureza_importacao(valor_celula(ws, linha, colunas, "natureza"))
    if natureza_planilha:
        return natureza_planilha

    natureza_padrao_normalizada = normalizar_cabecalho_importacao(natureza_padrao)
    if natureza_padrao_normalizada in ["despesa", "despesas", "saida", "saidas"]:
        return "Saida"
    if natureza_padrao_normalizada in ["receita", "receitas", "entrada", "entradas"]:
        return "Entrada"

    return natureza_por_categoria(categoria) or "Saida"


def montar_import_key(dados, incluir_origem=False):
    partes = [
        dados.get("numero_documento", ""),
        dados.get("data_documento", ""),
        dados.get("data_vencimento", ""),
        dados.get("favorecido", "") or dados.get("parceiro", ""),
        f"{float(dados.get('valor_documento') or 0):.2f}",
        dados.get("historico", "") or dados.get("descricao", ""),
    ]
    if incluir_origem:
        partes.extend([dados.get("tipo", ""), dados.get("origem_importacao", "")])
    ocorrencia = int(dados.get("ocorrencia_import_key") or 1)
    if ocorrencia > 1:
        partes.append(f"ocorrencia:{ocorrencia}")
    base = "|".join(str(parte).strip().lower() for parte in partes)
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def aplicar_ocorrencia_import_key(dados, ocorrencia, incluir_origem=False):
    dados["ocorrencia_import_key"] = ocorrencia
    dados["import_key"] = montar_import_key(dados, incluir_origem=incluir_origem)
    return dados["import_key"]


def linha_importacao_vazia(dados):
    return not any(str(valor or "").strip() for valor in dados.values())


def linha_planilha_importacao_vazia(ws, linha, colunas):
    return not any(
        str(ws.cell(linha, indice).value or "").strip()
        for indice in colunas.values()
    )


def preparar_linha_importacao(
    ws,
    linha,
    colunas,
    natureza_padrao="DESPESA",
    categoria_padrao=CATEGORIA_NAO_CLASSIFICADO,
    origem_importacao=ORIGEM_IMPORTACAO_DESPESAS,
    incluir_origem_import_key=False,
    resultado_metricas=None,
):
    valor_documento_original = numero_importacao(valor_celula(ws, linha, colunas, "valor_documento"))
    valor_liquido_original = numero_importacao(valor_celula(ws, linha, colunas, "valor_liquido"))
    valor_pago_original = numero_importacao(valor_celula(ws, linha, colunas, "valor_pago"))
    grupo_gerencial_informado = texto_importacao(valor_celula(ws, linha, colunas, "grupo_gerencial"))
    categoria_informada = texto_importacao(valor_celula(ws, linha, colunas, "categoria"))
    subcategoria_informada = texto_importacao(valor_celula(ws, linha, colunas, "subcategoria"))
    centro_analise_informado = texto_importacao(valor_celula(ws, linha, colunas, "centro_analise"))
    categoria = categoria_informada or categoria_padrao
    descricao = texto_importacao(valor_celula(ws, linha, colunas, "descricao"))
    historico = texto_importacao(valor_celula(ws, linha, colunas, "historico"))
    favorecido = texto_importacao(valor_celula(ws, linha, colunas, "favorecido"))
    parceiro = texto_importacao(valor_celula(ws, linha, colunas, "parceiro"))
    natureza_informada = texto_importacao(valor_celula(ws, linha, colunas, "natureza"))

    valor_referencia = valor_liquido_original or valor_documento_original or valor_pago_original
    if not natureza_informada:
        raise ValueError("tipo/natureza obrigatorio")
    tipo = resolver_tipo_importacao(ws, linha, colunas, categoria, natureza_padrao)
    if not tipo:
        raise ValueError("tipo/natureza obrigatorio")
    inicio_plano = time.perf_counter()
    try:
        conta_plano = resolver_conta_plano_importacao(
            grupo_gerencial_informado,
            categoria,
            subcategoria_informada,
        )
    finally:
        if resultado_metricas is not None:
            somar_tempo_importacao(
                resultado_metricas,
                "resolucao_plano_contas",
                time.perf_counter() - inicio_plano,
            )
    plano = derivar_plano_movimentacao(categoria, conta_plano["id"])
    categoria = plano["categoria_movimentacao"]
    tipo = resolver_tipo_caixa_importacao(tipo, conta_plano)

    valor_documento = abs(valor_documento_original or valor_referencia)
    valor_liquido = abs(valor_liquido_original)
    valor_pago = abs(valor_pago_original)

    data_documento = data_importacao(valor_celula(ws, linha, colunas, "data_documento"))
    data_vencimento = data_importacao(valor_celula(ws, linha, colunas, "data_vencimento")) or data_documento
    data_pagamento = data_importacao(valor_celula(ws, linha, colunas, "data_pagamento"))
    descricao_final = descricao or historico or favorecido or parceiro or "Movimentacao importada"
    status = "Realizado" if data_pagamento or valor_pago > 0 else "Pendente"

    dados = {
        "data_vencimento": data_vencimento,
        "data_realizacao": data_pagamento if status == "Realizado" else "",
        "tipo": tipo,
        "categoria": categoria,
        "descricao": descricao_final,
        "valor": abs(valor_referencia) if valor_referencia else valor_documento,
        "forma_pagamento": texto_importacao(valor_celula(ws, linha, colunas, "forma_pagamento")),
        "status": status,
        "parcelas": 1,
        "parcela_atual": 1,
        "intervalo_dias": 0,
        "documento_id": texto_importacao(valor_celula(ws, linha, colunas, "numero_documento")),
        "data_documento": data_documento,
        "valor_documento": valor_documento,
        "prazo_medio_dias": 0,
        "observacoes": texto_importacao(valor_celula(ws, linha, colunas, "observacoes")) or f"Importado do Excel. Historico: {historico}".strip(),
        "cnpj_cpf": texto_importacao(valor_celula(ws, linha, colunas, "cnpj_cpf")),
        "numero_documento": texto_importacao(valor_celula(ws, linha, colunas, "numero_documento")),
        "favorecido": favorecido,
        "parceiro": parceiro,
        "historico": historico,
        "valor_pago": valor_pago,
        "valor_liquido": valor_liquido,
        "origem_importacao": texto_importacao(valor_celula(ws, linha, colunas, "origem")) or origem_importacao,
        "plano_conta_id": plano["plano_conta_id"],
        "grupo_gerencial": plano["grupo_gerencial"],
        "categoria_plano": plano["categoria_plano"],
        "subcategoria": plano["subcategoria"],
        "centro_analise": centro_analise_informado or plano["centro_analise"],
        "linha_dre": plano["linha_dre"],
        "tipo_conta": plano["tipo_conta"],
        "impacta_fluxo_caixa": int(plano["impacta_fluxo_caixa"]),
    }
    aplicar_ocorrencia_import_key(dados, 1, incluir_origem=incluir_origem_import_key)
    return dados


def separar_em_lotes(itens, tamanho=900):
    for indice in range(0, len(itens), tamanho):
        yield itens[indice:indice + tamanho]


def buscar_import_keys_existentes(cursor, import_keys):
    existentes = {}
    for lote in separar_em_lotes(list(import_keys)):
        placeholders = ", ".join(["?"] * len(lote))
        cursor.execute(q(f"""
        SELECT *
        FROM movimentacoes_financeiras
        WHERE import_key IN ({placeholders})
        """), tuple(lote))
        for item in cursor.fetchall():
            existentes[item["import_key"]] = dict(item)
    return existentes


CAMPOS_COMPARACAO_IMPORTACAO = [
    "data_vencimento", "data_realizacao", "tipo", "categoria", "descricao", "valor",
    "forma_pagamento", "status", "parcelas", "parcela_atual", "intervalo_dias",
    "documento_id", "data_documento", "valor_documento", "prazo_medio_dias",
    "observacoes", "cnpj_cpf", "numero_documento", "favorecido", "parceiro",
    "historico", "valor_pago", "valor_liquido", "origem_importacao",
    "plano_conta_id", "grupo_gerencial", "categoria_plano", "subcategoria",
    "centro_analise", "linha_dre", "tipo_conta", "impacta_fluxo_caixa",
]
CAMPOS_NUMERICOS_IMPORTACAO = {
    "valor", "parcelas", "parcela_atual", "intervalo_dias", "valor_documento",
    "prazo_medio_dias", "valor_pago", "valor_liquido", "plano_conta_id",
    "impacta_fluxo_caixa",
}


def _valor_comparavel_importacao(campo, valor):
    if campo in CAMPOS_NUMERICOS_IMPORTACAO:
        try:
            return round(float(valor or 0), 4)
        except (TypeError, ValueError):
            return 0.0
    return str(valor or "").strip()


def importacao_identica(existente, dados):
    return all(
        _valor_comparavel_importacao(campo, existente.get(campo))
        == _valor_comparavel_importacao(campo, dados.get(campo))
        for campo in CAMPOS_COMPARACAO_IMPORTACAO
    )


def valores_insert_importacao(dados):
    return (
        dados["data_vencimento"], dados["data_realizacao"], dados["tipo"],
        dados["categoria"], dados["descricao"], dados["valor"], dados["forma_pagamento"],
        dados["status"], dados["parcelas"], dados["parcela_atual"], dados["intervalo_dias"],
        dados["documento_id"], dados["data_documento"], dados["valor_documento"],
        dados["prazo_medio_dias"], dados["observacoes"], dados["import_key"], dados["cnpj_cpf"],
        dados["numero_documento"], dados["favorecido"], dados["parceiro"], dados["historico"],
        dados["valor_pago"], dados["valor_liquido"], dados["origem_importacao"],
        dados["plano_conta_id"], dados["grupo_gerencial"], dados["categoria_plano"],
        dados["subcategoria"], dados["centro_analise"], dados["linha_dre"], dados["tipo_conta"],
        dados["impacta_fluxo_caixa"],
    )


SQL_INSERT_IMPORTACAO_SQLITE = """
INSERT INTO movimentacoes_financeiras (
    data_vencimento, data_realizacao, tipo, categoria, descricao, valor,
    forma_pagamento, status, parcelas, parcela_atual, intervalo_dias,
    documento_id, data_documento, valor_documento, prazo_medio_dias, observacoes,
    import_key, cnpj_cpf, numero_documento, favorecido, parceiro, historico,
    valor_pago, valor_liquido, origem_importacao, plano_conta_id, grupo_gerencial,
    categoria_plano, subcategoria, centro_analise, linha_dre, tipo_conta,
    impacta_fluxo_caixa
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""


SQL_INSERT_IMPORTACAO_POSTGRES = """
INSERT INTO movimentacoes_financeiras (
    data_vencimento, data_realizacao, tipo, categoria, descricao, valor,
    forma_pagamento, status, parcelas, parcela_atual, intervalo_dias,
    documento_id, data_documento, valor_documento, prazo_medio_dias, observacoes,
    import_key, cnpj_cpf, numero_documento, favorecido, parceiro, historico,
    valor_pago, valor_liquido, origem_importacao, plano_conta_id, grupo_gerencial,
    categoria_plano, subcategoria, centro_analise, linha_dre, tipo_conta,
    impacta_fluxo_caixa
) VALUES %s
"""


def executar_inserts_importacao(cursor, novos, tamanho_lote=TAMANHO_LOTE_GRAVACAO_IMPORTACAO):
    for lote in separar_em_lotes(novos, tamanho_lote):
        if DATABASE_URL:
            from psycopg2.extras import execute_values

            execute_values(cursor, SQL_INSERT_IMPORTACAO_POSTGRES, lote, page_size=len(lote))
        else:
            cursor.executemany(q(SQL_INSERT_IMPORTACAO_SQLITE), lote)


def importar_movimentacoes_financeiras_excel(
    arquivo_excel,
    natureza_padrao="DESPESA",
    categoria_padrao=CATEGORIA_NAO_CLASSIFICADO,
    origem_importacao=ORIGEM_IMPORTACAO_DESPESAS,
    incluir_origem_import_key=False,
    usuario_id=None,
    usuario_nome="",
    perfil="pcp",
):
    from .reset_financeiro import require_financial_imports_enabled

    require_financial_imports_enabled()
    inicio_processamento = time.perf_counter()
    criar_tabela_movimentacoes_financeiras()
    criar_tabelas_governanca_financeira()
    if not usuario_id or not str(usuario_nome or "").strip():
        raise ValueError("Usuário autenticado é obrigatório para importar.")

    inicio_leitura = time.perf_counter()
    if hasattr(arquivo_excel, "read"):
        nome_arquivo = sanitizar_nome_arquivo(getattr(arquivo_excel, "filename", None))
        if hasattr(arquivo_excel, "seek"):
            arquivo_excel.seek(0)
        conteudo_arquivo = arquivo_excel.read()
        if hasattr(arquivo_excel, "seek"):
            arquivo_excel.seek(0)
    else:
        caminho_arquivo = Path(arquivo_excel)
        nome_arquivo = sanitizar_nome_arquivo(caminho_arquivo.name)
        conteudo_arquivo = caminho_arquivo.read_bytes()
    if not conteudo_arquivo:
        raise ValueError("Arquivo de importação vazio.")
    hash_arquivo = hash_sha256(conteudo_arquivo)
    wb = load_workbook(BytesIO(conteudo_arquivo), data_only=True)
    ws = wb.active
    colunas, ausentes = mapear_cabecalhos_importacao(ws)
    if ausentes:
        raise ValueError("Cabecalhos obrigatorios ausentes: " + ", ".join(ausentes))

    resultado = {
        "linhas_lidas": 0,
        "importadas": 0,
        "atualizadas": 0,
        "conflitos": 0,
        "identicas": 0,
        "rejeitadas": 0,
        "ignoradas": 0,
        "erros": 0,
        "classificadas": 0,
        "nao_classificadas": 0,
        "detalhes_erros": [],
        "detalhes_conflitos": [],
        "tempo_processamento_segundos": 0,
        "tempos_etapas": {
            "leitura_excel": 0,
            "normalizacao_dados": 0,
            "resolucao_plano_contas": 0,
            "consulta_existentes": 0,
            "preparacao_novos": 0,
            "preparacao_atualizados": 0,
            "insert": 0,
            "update": 0,
            "commit": 0,
        },
        "lotes_insert": 0,
        "lotes_update": 0,
        "tamanho_lote_gravacao": TAMANHO_LOTE_GRAVACAO_IMPORTACAO,
        "transacao": "rollback_integral_antes_do_commit",
        "lote_id": None,
        "arquivo_nome": nome_arquivo,
        "arquivo_hash": hash_arquivo,
    }
    somar_tempo_importacao(resultado, "leitura_excel", time.perf_counter() - inicio_leitura)

    registros_por_import_key = {}
    ocorrencias_por_import_key_base = {}
    linhas_governanca = []

    for linha in range(2, ws.max_row + 1):
        resultado["linhas_lidas"] += 1
        campos_brutos = {
            campo: valor_celula(ws, linha, colunas, campo)
            for campo in sorted(colunas)
        }
        registro_linha = {
            "numero_linha": linha,
            "hash_normalizado": hash_normalizado(campos_brutos),
            "status": None,
            "movimentacao_id": None,
            "chave_encontrada": None,
            "mensagem": "",
            "campos_normalizados": campos_brutos,
            "auditoria_id": None,
        }
        try:
            if linha_planilha_importacao_vazia(ws, linha, colunas):
                resultado["ignoradas"] += 1
                resultado["rejeitadas"] += 1
                registro_linha["status"] = STATUS_LINHA_REJEITADA
                registro_linha["mensagem"] = "Linha vazia."
                linhas_governanca.append(registro_linha)
                continue

            inicio_normalizacao = time.perf_counter()
            tempo_plano_antes = resultado["tempos_etapas"]["resolucao_plano_contas"]
            dados = preparar_linha_importacao(
                ws,
                linha,
                colunas,
                natureza_padrao=natureza_padrao,
                categoria_padrao=categoria_padrao,
                origem_importacao=origem_importacao,
                incluir_origem_import_key=incluir_origem_import_key,
                resultado_metricas=resultado,
            )
            tempo_plano_linha = resultado["tempos_etapas"]["resolucao_plano_contas"] - tempo_plano_antes
            somar_tempo_importacao(
                resultado,
                "normalizacao_dados",
                max((time.perf_counter() - inicio_normalizacao) - tempo_plano_linha, 0),
            )
            if linha_importacao_vazia(dados):
                resultado["ignoradas"] += 1
                resultado["rejeitadas"] += 1
                registro_linha["status"] = STATUS_LINHA_REJEITADA
                registro_linha["mensagem"] = "Linha sem conteúdo financeiro."
                linhas_governanca.append(registro_linha)
                continue
            if not dados["data_documento"]:
                raise ValueError("linha sem data do documento")
            if float(dados["valor"] or 0) <= 0:
                raise ValueError("linha sem valor financeiro valido")

            import_key_base = dados["import_key"]
            ocorrencia_import_key = ocorrencias_por_import_key_base.get(import_key_base, 0) + 1
            ocorrencias_por_import_key_base[import_key_base] = ocorrencia_import_key
            aplicar_ocorrencia_import_key(
                dados,
                ocorrencia_import_key,
                incluir_origem=incluir_origem_import_key,
            )

            if dados["import_key"] in registros_por_import_key:
                resultado["ignoradas"] += 1
                resultado["identicas"] += 1
                registro_linha["status"] = STATUS_LINHA_IDENTICA
                registro_linha["chave_encontrada"] = dados["import_key"]
                registro_linha["mensagem"] = "Linha duplicada no mesmo arquivo; nenhum novo registro criado."
                registro_linha["campos_normalizados"] = dados
                linhas_governanca.append(registro_linha)
                continue

            if dados["categoria"] == CATEGORIA_NAO_CLASSIFICADO:
                resultado["nao_classificadas"] += 1
            else:
                resultado["classificadas"] += 1

            registros_por_import_key[dados["import_key"]] = dados
            registro_linha["chave_encontrada"] = dados["import_key"]
            registro_linha["campos_normalizados"] = dados
            linhas_governanca.append(registro_linha)
        except Exception as erro:
            resultado["erros"] += 1
            resultado["rejeitadas"] += 1
            resultado["detalhes_erros"].append(f"Linha {linha}: {erro}")
            registro_linha["status"] = STATUS_LINHA_REJEITADA
            registro_linha["mensagem"] = str(erro)
            linhas_governanca.append(registro_linha)

    wb.close()

    conn = conectar()
    cursor = conn.cursor()
    try:
        tipo_importador = "VENDAS" if incluir_origem_import_key else "MOVIMENTACOES"
        lote_id = criar_lote_importacao_cursor(
            cursor,
            nome_arquivo,
            hash_arquivo,
            tipo_importador,
            usuario_id,
            usuario_nome,
            metadados_tecnicos={
                "natureza_padrao": natureza_padrao,
                "categoria_padrao": categoria_padrao,
                "origem_importacao": origem_importacao,
            },
        )
        resultado["lote_id"] = lote_id
        linhas_principais = {
            item["chave_encontrada"]: item
            for item in linhas_governanca
            if item["chave_encontrada"] and item["status"] is None
        }
        inicio_consulta = time.perf_counter()
        existentes = buscar_import_keys_existentes(cursor, registros_por_import_key.keys())
        somar_tempo_importacao(resultado, "consulta_existentes", time.perf_counter() - inicio_consulta)
        novos_dados = []
        conflitos = []
        movimentacoes_por_chave = {}

        for import_key, dados in registros_por_import_key.items():
            linha_resultado = linhas_principais[import_key]
            existente = existentes.get(import_key)
            if existente:
                movimentacoes_por_chave[import_key] = existente["id"]
                if importacao_identica(existente, dados):
                    resultado["ignoradas"] += 1
                    resultado["identicas"] += 1
                    linha_resultado["status"] = STATUS_LINHA_IDENTICA
                    linha_resultado["movimentacao_id"] = existente["id"]
                    linha_resultado["mensagem"] = "Registro idêntico já existente; nenhum novo registro criado."
                else:
                    conflitos.append((existente, dados, linha_resultado))
            else:
                inicio_preparacao_insert = time.perf_counter()
                novos_dados.append(dados)
                somar_tempo_importacao(resultado, "preparacao_novos", time.perf_counter() - inicio_preparacao_insert)

        for existente, dados, linha_resultado in conflitos:
            auditoria_id = _registrar_auditoria_cursor(
                cursor,
                existente["id"],
                "CONFLITO_IMPORTACAO",
                existente,
                dados,
                usuario_id=usuario_id,
                usuario_nome=usuario_nome,
                perfil=perfil,
                justificativa="Mesma chave de importação recebida com conteúdo divergente; registro preservado.",
                origem_acao="importacao_excel",
                referencia_importacao=dados["import_key"],
            )
            linha_resultado["status"] = STATUS_LINHA_CONFLITANTE
            linha_resultado["movimentacao_id"] = existente["id"]
            linha_resultado["auditoria_id"] = auditoria_id
            linha_resultado["mensagem"] = "Conteúdo divergente; registro existente preservado."
            resultado["detalhes_conflitos"].append(
                f"Chave {dados['import_key']}: divergência registrada sem sobrescrita."
            )
        resultado["conflitos"] = len(conflitos)

        if novos_dados:
            inicio_insert = time.perf_counter()
            novos = [valores_insert_importacao(dados) for dados in novos_dados]
            executar_inserts_importacao(cursor, novos)
            somar_tempo_importacao(resultado, "insert", time.perf_counter() - inicio_insert)
            resultado["importadas"] = len(novos)
            resultado["lotes_insert"] = len(list(separar_em_lotes(novos, TAMANHO_LOTE_GRAVACAO_IMPORTACAO)))

            inseridos = buscar_import_keys_existentes(
                cursor,
                [dados["import_key"] for dados in novos_dados],
            )
            for dados in novos_dados:
                movimentacao = inseridos[dados["import_key"]]
                linha_resultado = linhas_principais[dados["import_key"]]
                linha_resultado["status"] = STATUS_LINHA_IMPORTADA
                linha_resultado["movimentacao_id"] = movimentacao["id"]
                linha_resultado["mensagem"] = "Movimentação importada."
                origem_id = registrar_origem_principal_cursor(
                    cursor,
                    movimentacao["id"],
                    IMPORTACAO_FINANCEIRA,
                    "FRIGODATTA",
                    "IMPORTACAO_EXCEL",
                    usuario_id,
                    usuario_nome,
                    modulo_origem="MOVIMENTACOES_FINANCEIRAS",
                    chave_externa=dados["import_key"],
                    chave_idempotente=dados["import_key"],
                    lote_importacao_id=lote_id,
                    linha_arquivo=linha_resultado["numero_linha"],
                    metadados={"arquivo_hash": hash_arquivo},
                )
                auditoria_id = _registrar_auditoria_cursor(
                    cursor,
                    movimentacao["id"],
                    "CRIACAO_IMPORTACAO",
                    None,
                    movimentacao,
                    usuario_id=usuario_id,
                    usuario_nome=usuario_nome,
                    perfil=perfil,
                    justificativa="Criação por importação financeira rastreada.",
                    origem_acao="importacao_excel",
                    referencia_importacao=dados["import_key"],
                )
                vincular_auditoria_origem_cursor(cursor, origem_id, auditoria_id)
                linha_resultado["auditoria_id"] = auditoria_id

        for linha_resultado in linhas_governanca:
            registrar_linha_importacao_cursor(
                cursor,
                lote_id,
                linha_resultado["numero_linha"],
                linha_resultado["hash_normalizado"],
                linha_resultado["status"] or STATUS_LINHA_REJEITADA,
                movimentacao_id=linha_resultado["movimentacao_id"],
                chave_encontrada=linha_resultado["chave_encontrada"],
                mensagem=linha_resultado["mensagem"],
                campos_normalizados=linha_resultado["campos_normalizados"],
                auditoria_id=linha_resultado["auditoria_id"],
            )

        resultado["rejeitadas"] = sum(
            item["status"] == STATUS_LINHA_REJEITADA for item in linhas_governanca
        )
        finalizar_lote_importacao_cursor(
            cursor,
            lote_id,
            len(linhas_governanca),
            resultado["importadas"],
            resultado["identicas"],
            resultado["conflitos"],
            resultado["rejeitadas"],
            "Importação concluída com rastreabilidade por lote e linha.",
        )

        inicio_commit = time.perf_counter()
        conn.commit()
        somar_tempo_importacao(resultado, "commit", time.perf_counter() - inicio_commit)
        resultado["transacao"] = "commit_unico_concluido"
    except Exception as erro:
        conn.rollback()
        resultado["transacao"] = "rollback_integral_executado"
        try:
            lote_falho_id = criar_lote_importacao_cursor(
                cursor,
                nome_arquivo,
                hash_arquivo,
                tipo_importador,
                usuario_id,
                usuario_nome,
                metadados_tecnicos={"falha_integral": True},
            )
            falhar_lote_importacao_cursor(
                cursor,
                lote_falho_id,
                len(linhas_governanca),
                f"Falha integral: {type(erro).__name__}",
            )
            conn.commit()
        except Exception:
            conn.rollback()
        raise
    finally:
        conn.close()

    resultado["tempo_processamento_segundos"] = time.perf_counter() - inicio_processamento
    return finalizar_tempos_importacao(resultado)


def gerar_planilha_modelo_importacao_financeira():
    wb = Workbook()
    ws = wb.active
    ws.title = "Importacao Financeira"
    ws.append(CAMPOS_PLANILHA_OFICIAL)

    exemplo = [
        "2026-01-01",
        "2026-01-10",
        "",
        "NF-0001",
        "",
        "Cliente ou fornecedor",
        "Descricao do lancamento",
        "Historico complementar",
        "Receita",
        "Venda de Producao Propria",
        "",
        "",
        0,
        0,
        0,
        "NOVA BASE OFICIAL",
        "",
    ]
    ws.append(exemplo)
    ws.freeze_panes = "A2"

    for indice, titulo in enumerate(CAMPOS_PLANILHA_OFICIAL, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = max(16, len(titulo) + 2)

    plano_ws = wb.create_sheet("Plano de Contas")
    plano_ws.append([
        "ID",
        "Categoria",
        "Subcategoria",
        "Conta",
        "Grupo Gerencial",
        "Linha DRE",
        "Tipo Conta",
    ])
    contas = listar_plano_contas()
    for conta in contas:
        plano_ws.append([
            conta["id"],
            conta["categoria"],
            conta["subcategoria"],
            conta["nome"],
            conta["grupo_gerencial"],
            conta["linha_dre"],
            conta["tipo_conta"],
        ])

    categorias_unicas = sorted({conta["categoria"] for conta in contas if conta["categoria"]})
    subcategorias_unicas = sorted({conta["subcategoria"] for conta in contas if conta["subcategoria"]})
    plano_ws["I1"] = "Categorias"
    for linha, categoria in enumerate(categorias_unicas, start=2):
        plano_ws.cell(linha, 9).value = categoria
    plano_ws["J1"] = "Subcategorias"
    for linha, subcategoria in enumerate(subcategorias_unicas, start=2):
        plano_ws.cell(linha, 10).value = subcategoria

    for coluna in range(1, 11):
        plano_ws.column_dimensions[get_column_letter(coluna)].width = 24

    max_linhas = 1000
    validacoes = [
        DataValidation(type="list", formula1='"Receita,Despesa,Neutro"', allow_blank=False),
        DataValidation(type="list", formula1=f"'Plano de Contas'!$I$2:$I${len(categorias_unicas) + 1}", allow_blank=False),
        DataValidation(type="list", formula1=f"'Plano de Contas'!$J$2:$J${len(subcategorias_unicas) + 1}", allow_blank=True),
    ]
    for validacao in validacoes:
        ws.add_data_validation(validacao)

    validacoes[0].add(f"I2:I{max_linhas}")
    validacoes[1].add(f"J2:J{max_linhas}")
    if subcategorias_unicas:
        validacoes[2].add(f"K2:K{max_linhas}")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def buscar_pendencias_classificacao():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q("""
    SELECT *
    FROM movimentacoes_financeiras
    WHERE categoria = ?
    ORDER BY data_documento ASC, data_vencimento ASC, id ASC
    """), (CATEGORIA_NAO_CLASSIFICADO,))
    pendencias = cursor.fetchall()
    conn.close()

    total = sum(float(item["valor"] or 0) for item in pendencias)
    return {
        "quantidade": len(pendencias),
        "valor_total": round(total, 2),
        "movimentacoes": pendencias,
    }


def categorias_reclassificacao_financeira():
    return [
        item["nome"]
        for item in listar_plano_contas()
    ]


def contas_reclassificacao_financeira():
    return opcoes_reclassificacao_plano()


def categorias_filtro_auditoria():
    categorias = set(categorias_reclassificacao_financeira())
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q("""
    SELECT DISTINCT categoria
    FROM movimentacoes_financeiras
    WHERE categoria IS NOT NULL
      AND TRIM(categoria) <> ''
    ORDER BY categoria ASC
    """))
    for item in cursor.fetchall():
        categorias.add(item["categoria"])
    conn.close()
    return [CATEGORIA_NAO_CLASSIFICADO] + sorted(
        categoria for categoria in categorias
        if categoria != CATEGORIA_NAO_CLASSIFICADO
    )


def normalizar_filtros_auditoria(args):
    try:
        pagina = int(args.get("pagina", 1) or 1)
    except (TypeError, ValueError):
        pagina = 1

    try:
        por_pagina = int(args.get("por_pagina", 50) or 50)
    except (TypeError, ValueError):
        por_pagina = 50

    if pagina < 1:
        pagina = 1

    if por_pagina not in [50, 100, 200]:
        por_pagina = 50

    return {
        "data_inicio": args.get("data_inicio", "").strip(),
        "data_fim": args.get("data_fim", "").strip(),
        "categoria": args.get("categoria", "Todas").strip() or "Todas",
        "favorecido": args.get("favorecido", "").strip(),
        "descricao": args.get("descricao", "").strip(),
        "historico": args.get("historico", "").strip(),
        "mes": args.get("mes", "").strip(),
        "somente_nao_classificados": args.get("somente_nao_classificados") == "1",
        "pagina": pagina,
        "por_pagina": por_pagina,
    }


def montar_where_auditoria(filtros, somente_pendencias=False):
    condicoes = ["(import_key IS NOT NULL OR origem_importacao = ?)"]
    parametros = ["excel_movimentacoes"]

    if filtros["data_inicio"]:
        condicoes.append("COALESCE(data_documento, data_vencimento) >= ?")
        parametros.append(filtros["data_inicio"])

    if filtros["data_fim"]:
        condicoes.append("COALESCE(data_documento, data_vencimento) <= ?")
        parametros.append(filtros["data_fim"])

    if filtros["categoria"] != "Todas":
        condicoes.append("categoria = ?")
        parametros.append(filtros["categoria"])

    if filtros["favorecido"]:
        condicoes.append("(COALESCE(favorecido, '') LIKE ? OR COALESCE(parceiro, '') LIKE ?)")
        termo = f"%{filtros['favorecido']}%"
        parametros.extend([termo, termo])

    if filtros["descricao"]:
        condicoes.append("COALESCE(descricao, '') LIKE ?")
        parametros.append(f"%{filtros['descricao']}%")

    if filtros["historico"]:
        condicoes.append("(COALESCE(historico, '') LIKE ? OR COALESCE(observacoes, '') LIKE ?)")
        termo = f"%{filtros['historico']}%"
        parametros.extend([termo, termo])

    if filtros["mes"]:
        condicoes.append("substr(COALESCE(data_documento, data_vencimento), 1, 7) = ?")
        parametros.append(filtros["mes"])

    if filtros["somente_nao_classificados"]:
        condicoes.append("categoria = ?")
        parametros.append(CATEGORIA_NAO_CLASSIFICADO)

    if somente_pendencias and not filtros["somente_nao_classificados"]:
        condicoes.append("categoria = ?")
        parametros.append(CATEGORIA_NAO_CLASSIFICADO)

    return " AND ".join(condicoes), parametros


def buscar_movimentacoes_auditoria(filtros, somente_pendencias=False, limite=None, offset=0):
    where_sql, parametros = montar_where_auditoria(filtros, somente_pendencias)

    conn = conectar()
    cursor = conn.cursor()

    limite_sql = ""
    if limite is not None:
        limite_sql = " LIMIT ? OFFSET ?"
        parametros.extend([limite, offset])

    cursor.execute(q(f"""
    SELECT *
    FROM movimentacoes_financeiras
    WHERE {where_sql}
    ORDER BY COALESCE(data_documento, data_vencimento) ASC, id ASC
    {limite_sql}
    """), tuple(parametros))
    movimentacoes = [dict(item) for item in cursor.fetchall()]
    conn.close()
    return movimentacoes


def buscar_movimentacoes_exportacao_auditoria(filtros):
    """Retorna a população completa da Auditoria com governança, sem paginação."""
    where_sql, parametros = montar_where_auditoria(filtros)

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(f"""
    WITH auditoria_ordenada AS (
        SELECT
            a.*,
            ROW_NUMBER() OVER (
                PARTITION BY a.movimentacao_id
                ORDER BY a.data_hora ASC, a.id ASC
            ) AS ordem_primeira,
            ROW_NUMBER() OVER (
                PARTITION BY a.movimentacao_id
                ORDER BY a.data_hora DESC, a.id DESC
            ) AS ordem_ultima
        FROM movimentacoes_financeiras_auditoria a
    ),
    cancelamentos_ordenados AS (
        SELECT
            a.*,
            ROW_NUMBER() OVER (
                PARTITION BY a.movimentacao_id
                ORDER BY a.data_hora DESC, a.id DESC
            ) AS ordem_cancelamento
        FROM movimentacoes_financeiras_auditoria a
        WHERE a.acao = 'CANCELAMENTO'
    )
    SELECT
        m.*,
        p.grupo_gerencial AS plano_grupo_gerencial,
        p.categoria AS plano_categoria,
        p.subcategoria AS plano_subcategoria,
        o.id AS origem_id,
        o.modo AS modo_origem,
        o.sistema_origem,
        o.modulo_origem,
        o.tipo_evento AS origem_tipo_evento,
        o.chave_externa AS origem_chave_externa,
        o.chave_idempotente AS origem_chave_idempotente,
        o.lote_importacao_id,
        o.linha_arquivo AS origem_linha_arquivo,
        o.usuario_id AS origem_usuario_id,
        o.usuario_nome AS origem_usuario_nome,
        o.criado_em AS origem_criado_em,
        l.arquivo_nome AS lote_arquivo_nome,
        l.arquivo_hash AS lote_arquivo_hash,
        primeira.usuario_id AS primeiro_usuario_id,
        primeira.usuario_nome AS primeiro_usuario_nome,
        ultima.usuario_id AS ultimo_usuario_id,
        ultima.usuario_nome AS ultimo_usuario_nome,
        ultima.data_hora AS ultima_alteracao_em,
        cancelamento.data_hora AS cancelado_em,
        cancelamento.justificativa AS cancelamento_justificativa
    FROM (
        SELECT
            id, data_vencimento, data_realizacao, tipo, categoria,
            descricao, valor, status, parcelas, parcela_atual,
            documento_id, data_documento, valor_documento, observacoes,
            criado_em, import_key, cnpj_cpf, numero_documento, favorecido,
            parceiro, historico, valor_pago, valor_liquido,
            origem_importacao, plano_conta_id, grupo_gerencial,
            categoria_plano, subcategoria, centro_analise, linha_dre
        FROM movimentacoes_financeiras
        WHERE {where_sql}
    ) m
    LEFT JOIN plano_contas_mestre p
      ON p.id = m.plano_conta_id
    LEFT JOIN movimentacoes_financeiras_origens o
      ON o.movimentacao_id = m.id
     AND o.papel = 'PRINCIPAL'
     AND o.status = 'ATIVA'
    LEFT JOIN movimentacoes_financeiras_importacao_lotes l
      ON l.id = o.lote_importacao_id
    LEFT JOIN auditoria_ordenada primeira
      ON primeira.movimentacao_id = m.id
     AND primeira.ordem_primeira = 1
    LEFT JOIN auditoria_ordenada ultima
      ON ultima.movimentacao_id = m.id
     AND ultima.ordem_ultima = 1
    LEFT JOIN cancelamentos_ordenados cancelamento
      ON cancelamento.movimentacao_id = m.id
     AND cancelamento.ordem_cancelamento = 1
    ORDER BY COALESCE(m.data_documento, m.data_vencimento) ASC, m.id ASC
    """), tuple(parametros))
    movimentacoes = [aplicar_origem_leitura(dict(item)) for item in cursor.fetchall()]
    conn.close()
    return movimentacoes


def buscar_indicadores_auditoria(filtros):
    where_sql, parametros = montar_where_auditoria(filtros)

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(f"""
    SELECT
        COUNT(*) as total_movimentacoes,
        COALESCE(SUM(CASE WHEN tipo = ? THEN valor ELSE 0 END), 0) as total_receitas,
        COALESCE(SUM(CASE WHEN tipo = ? THEN 0 ELSE valor END), 0) as total_despesas,
        COALESCE(SUM(CASE WHEN categoria = ? THEN 1 ELSE 0 END), 0) as qtd_nao_classificado,
        COALESCE(SUM(CASE WHEN categoria = ? THEN valor ELSE 0 END), 0) as valor_nao_classificado
    FROM movimentacoes_financeiras
    WHERE {where_sql}
    """), tuple(["Entrada", "Entrada", CATEGORIA_NAO_CLASSIFICADO, CATEGORIA_NAO_CLASSIFICADO] + parametros))
    item = cursor.fetchone()
    conn.close()

    total_receitas = float(item["total_receitas"] or 0)
    total_despesas = float(item["total_despesas"] or 0)
    return {
        "total_movimentacoes": int(item["total_movimentacoes"] or 0),
        "total_receitas": round(total_receitas, 2),
        "total_despesas": round(total_despesas, 2),
        "saldo_liquido": round(total_receitas - total_despesas, 2),
        "qtd_nao_classificado": int(item["qtd_nao_classificado"] or 0),
        "valor_nao_classificado": round(float(item["valor_nao_classificado"] or 0), 2),
    }


def buscar_totais_categoria_auditoria(filtros):
    where_sql, parametros = montar_where_auditoria(filtros)
    indicadores = buscar_indicadores_auditoria(filtros)
    total_geral = indicadores["total_receitas"] + indicadores["total_despesas"]

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(f"""
    SELECT
        COALESCE(categoria, ?) as categoria,
        COUNT(*) as quantidade,
        COALESCE(SUM(valor), 0) as valor_total
    FROM movimentacoes_financeiras
    WHERE {where_sql}
    GROUP BY COALESCE(categoria, ?)
    ORDER BY valor_total DESC
    """), tuple([CATEGORIA_NAO_CLASSIFICADO] + parametros + [CATEGORIA_NAO_CLASSIFICADO]))
    linhas = cursor.fetchall()
    conn.close()

    resultado = []
    for item in linhas:
        valor_total = float(item["valor_total"] or 0)
        percentual = (valor_total / total_geral * 100) if total_geral else 0
        resultado.append({
            "categoria": item["categoria"],
            "quantidade": int(item["quantidade"] or 0),
            "valor_total": round(valor_total, 2),
            "percentual": round(percentual, 2),
        })
    return resultado


def buscar_totais_mes_auditoria(filtros):
    where_sql, parametros = montar_where_auditoria(filtros)

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(f"""
    SELECT
        COALESCE(substr(COALESCE(data_documento, data_vencimento), 1, 7), 'Sem data') as mes,
        COALESCE(SUM(CASE WHEN tipo = ? THEN valor ELSE 0 END), 0) as total_receitas,
        COALESCE(SUM(CASE WHEN tipo = ? THEN 0 ELSE valor END), 0) as total_despesas,
        COUNT(*) as quantidade
    FROM movimentacoes_financeiras
    WHERE {where_sql}
    GROUP BY COALESCE(substr(COALESCE(data_documento, data_vencimento), 1, 7), 'Sem data')
    ORDER BY mes ASC
    """), tuple(["Entrada", "Entrada"] + parametros))
    linhas = cursor.fetchall()
    conn.close()

    resultado = []
    for item in linhas:
        total_receitas = float(item["total_receitas"] or 0)
        total_despesas = float(item["total_despesas"] or 0)
        resultado.append({
            "mes": item["mes"],
            "total_receitas": round(total_receitas, 2),
            "total_despesas": round(total_despesas, 2),
            "saldo": round(total_receitas - total_despesas, 2),
            "quantidade": int(item["quantidade"] or 0),
        })
    return resultado


def contar_pendencias_auditoria(filtros):
    where_sql, parametros = montar_where_auditoria(filtros, somente_pendencias=True)
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(f"""
    SELECT COUNT(*) as total
    FROM movimentacoes_financeiras
    WHERE {where_sql}
    """), tuple(parametros))
    total = int(cursor.fetchone()["total"] or 0)
    conn.close()
    return total


def montar_query_paginacao(filtros, pagina):
    parametros = {
        "data_inicio": filtros["data_inicio"],
        "data_fim": filtros["data_fim"],
        "categoria": filtros["categoria"],
        "favorecido": filtros["favorecido"],
        "descricao": filtros["descricao"],
        "historico": filtros["historico"],
        "mes": filtros["mes"],
        "por_pagina": filtros["por_pagina"],
        "pagina": pagina,
    }
    if filtros["somente_nao_classificados"]:
        parametros["somente_nao_classificados"] = "1"
    return urlencode({chave: valor for chave, valor in parametros.items() if valor not in ["", None]})


def montar_contexto_auditoria_financeira(args, exportar=False):
    filtros = normalizar_filtros_auditoria(args)
    total_pendencias = contar_pendencias_auditoria(filtros)
    total_paginas = max(1, (total_pendencias + filtros["por_pagina"] - 1) // filtros["por_pagina"])

    if filtros["pagina"] > total_paginas:
        filtros["pagina"] = total_paginas

    offset = (filtros["pagina"] - 1) * filtros["por_pagina"]
    limite = None if exportar else filtros["por_pagina"]
    pendencias = buscar_movimentacoes_auditoria(
        filtros,
        somente_pendencias=True,
        limite=limite,
        offset=offset,
    )
    pagina_anterior = max(1, filtros["pagina"] - 1)
    proxima_pagina = min(total_paginas, filtros["pagina"] + 1)

    return {
        "filtros": filtros,
        "indicadores": buscar_indicadores_auditoria(filtros),
        "totais_categoria": buscar_totais_categoria_auditoria(filtros),
        "totais_mes": buscar_totais_mes_auditoria(filtros),
        "pendencias": pendencias,
        "movimentacoes": (
            buscar_movimentacoes_exportacao_auditoria(filtros)
            if exportar
            else []
        ),
        "paginacao": {
            "pagina": filtros["pagina"],
            "por_pagina": filtros["por_pagina"],
            "total_paginas": total_paginas,
            "total_registros": total_pendencias,
            "tem_anterior": filtros["pagina"] > 1,
            "tem_proxima": filtros["pagina"] < total_paginas,
            "query_anterior": montar_query_paginacao(filtros, pagina_anterior),
            "query_proxima": montar_query_paginacao(filtros, proxima_pagina),
        },
        "categorias_reclassificacao": categorias_reclassificacao_financeira(),
        "contas_reclassificacao": contas_reclassificacao_financeira(),
        "categorias_filtro": categorias_filtro_auditoria(),
    }


STATUS_LIQUIDACAO_OPCOES = [
    "Todos",
    "Liquidado",
    "Parcialmente liquidado",
    "Em aberto",
    "Vencido",
    "Inconsistente",
]


FAIXAS_ATRASO_LIQUIDACAO = {
    "1_30": (1, 30),
    "31_60": (31, 60),
    "61_90": (61, 90),
    "acima_90": (91, None),
}


def data_iso_para_date(valor):
    texto = str(valor or "").strip()
    if not texto:
        return None
    try:
        return datetime.strptime(texto[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def valor_float(valor):
    try:
        return round(float(valor or 0), 2)
    except (TypeError, ValueError):
        return 0


def calcular_status_liquidacao(movimentacao, data_referencia=None):
    data_ref = data_referencia or date.today()
    data_vencimento = data_iso_para_date(movimentacao.get("data_vencimento"))
    data_baixa = data_iso_para_date(movimentacao.get("data_realizacao"))
    valor_documento = valor_float(movimentacao.get("valor_documento") or movimentacao.get("valor"))
    valor_baixado = valor_float(movimentacao.get("valor_pago"))
    baixa_tem_data = data_baixa is not None
    baixa_tem_valor = valor_baixado > 0
    dias_em_atraso = 0
    status = "Em aberto"
    inconsistente = False
    motivo = ""

    if data_vencimento and data_vencimento < data_ref:
        dias_em_atraso = (data_ref - data_vencimento).days

    if baixa_tem_data and not baixa_tem_valor:
        status = "Inconsistente"
        inconsistente = True
        motivo = "Data da baixa preenchida sem valor da baixa."
    elif baixa_tem_valor and not baixa_tem_data:
        status = "Inconsistente"
        inconsistente = True
        motivo = "Valor da baixa preenchido sem data da baixa."
    elif valor_baixado > valor_documento and valor_documento > 0:
        status = "Inconsistente"
        inconsistente = True
        motivo = "Valor da baixa superior ao valor do documento."
    elif baixa_tem_data and baixa_tem_valor and valor_baixado >= valor_documento:
        status = "Liquidado"
    elif baixa_tem_valor and valor_baixado < valor_documento:
        status = "Parcialmente liquidado"
    elif data_vencimento and data_vencimento < data_ref:
        status = "Vencido"
    else:
        status = "Em aberto"

    saldo_em_aberto = max(valor_documento - valor_baixado, 0)
    if status == "Inconsistente" and valor_baixado > valor_documento:
        saldo_em_aberto = 0
    if status == "Liquidado":
        saldo_em_aberto = 0

    return {
        "status_liquidacao": status,
        "valor_documento_liquidacao": valor_documento,
        "valor_baixado": valor_baixado,
        "saldo_em_aberto": round(saldo_em_aberto, 2),
        "dias_em_atraso": dias_em_atraso if status in ["Vencido", "Parcialmente liquidado", "Inconsistente"] else 0,
        "inconsistente": inconsistente,
        "motivo_inconsistencia": motivo,
    }


def movimentacao_compativel_contas_pagar(movimentacao):
    if movimentacao.get("status") == "Cancelado":
        return False
    if normalizar_texto_plano(movimentacao.get("tipo")) != "saida":
        return False
    if movimentacao.get("tipo_conta") == "Neutro":
        return False
    if movimentacao.get("linha_dre") == "Neutro":
        return False
    return True


def normalizar_filtros_liquidacao(args):
    try:
        pagina = int(args.get("pagina", 1) or 1)
    except (TypeError, ValueError):
        pagina = 1
    try:
        por_pagina = int(args.get("por_pagina", 50) or 50)
    except (TypeError, ValueError):
        por_pagina = 50
    if pagina < 1:
        pagina = 1
    if por_pagina not in [50, 100, 200]:
        por_pagina = 50
    return {
        "data_documento_inicio": args.get("data_documento_inicio", "").strip(),
        "data_documento_fim": args.get("data_documento_fim", "").strip(),
        "data_vencimento_inicio": args.get("data_vencimento_inicio", "").strip(),
        "data_vencimento_fim": args.get("data_vencimento_fim", "").strip(),
        "status": args.get("status", "Todos").strip() or "Todos",
        "favorecido": args.get("favorecido", "").strip(),
        "categoria": args.get("categoria", "Todas").strip() or "Todas",
        "subcategoria": args.get("subcategoria", "Todas").strip() or "Todas",
        "natureza": args.get("natureza", "Saida").strip() or "Saida",
        "somente_vencidos": args.get("somente_vencidos") == "1",
        "somente_abertos": args.get("somente_abertos") == "1",
        "somente_parciais": args.get("somente_parciais") == "1",
        "somente_inconsistentes": args.get("somente_inconsistentes") == "1",
        "faixa_atraso": args.get("faixa_atraso", "Todas").strip() or "Todas",
        "relatorio": args.get("relatorio", "").strip(),
        "pagina": pagina,
        "por_pagina": por_pagina,
    }


def montar_where_liquidacao(filtros):
    condicoes = ["(import_key IS NOT NULL OR origem_importacao = ?)"]
    parametros = ["excel_movimentacoes"]
    if filtros["data_documento_inicio"]:
        condicoes.append("COALESCE(data_documento, data_vencimento) >= ?")
        parametros.append(filtros["data_documento_inicio"])
    if filtros["data_documento_fim"]:
        condicoes.append("COALESCE(data_documento, data_vencimento) <= ?")
        parametros.append(filtros["data_documento_fim"])
    if filtros["data_vencimento_inicio"]:
        condicoes.append("data_vencimento >= ?")
        parametros.append(filtros["data_vencimento_inicio"])
    if filtros["data_vencimento_fim"]:
        condicoes.append("data_vencimento <= ?")
        parametros.append(filtros["data_vencimento_fim"])
    if filtros["favorecido"]:
        condicoes.append("(COALESCE(favorecido, '') LIKE ? OR COALESCE(parceiro, '') LIKE ?)")
        termo = f"%{filtros['favorecido']}%"
        parametros.extend([termo, termo])
    if filtros["categoria"] != "Todas":
        condicoes.append("categoria = ?")
        parametros.append(filtros["categoria"])
    if filtros["subcategoria"] != "Todas":
        condicoes.append("COALESCE(subcategoria, '') = ?")
        parametros.append(filtros["subcategoria"])
    if filtros["natureza"] != "Todas":
        if normalizar_texto_plano(filtros["natureza"]) == "saida":
            condicoes.append("tipo IN (?, ?)")
            parametros.extend(["Saida", "Saída"])
            return " AND ".join(condicoes), parametros
        condicoes.append("tipo = ?")
        parametros.append(filtros["natureza"])
    return " AND ".join(condicoes), parametros


def buscar_movimentacoes_base_liquidacao(filtros):
    where_sql, parametros = montar_where_liquidacao(filtros)
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(f"""
    SELECT *
    FROM movimentacoes_financeiras
    WHERE {where_sql}
    ORDER BY data_vencimento ASC, COALESCE(valor_documento, valor) DESC, id ASC
    """), tuple(parametros))
    itens = [dict(item) for item in cursor.fetchall()]
    conn.close()
    return itens


def aplicar_filtros_liquidacao(movimentacoes, filtros, data_referencia=None):
    resultado = []
    for item in movimentacoes:
        if not movimentacao_compativel_contas_pagar(item):
            continue
        liquidacao = calcular_status_liquidacao(item, data_referencia)
        item = {**item, **liquidacao}

        if filtros["relatorio"] == "vencidos_sem_baixa":
            if item["status_liquidacao"] not in ["Vencido", "Parcialmente liquidado"]:
                continue
            if item["dias_em_atraso"] <= 0 or item["saldo_em_aberto"] <= 0:
                continue
        if filtros["status"] != "Todos" and item["status_liquidacao"] != filtros["status"]:
            continue
        if filtros["somente_vencidos"] and item["status_liquidacao"] != "Vencido":
            continue
        if filtros["somente_abertos"] and item["status_liquidacao"] not in ["Em aberto", "Vencido", "Parcialmente liquidado"]:
            continue
        if filtros["somente_parciais"] and item["status_liquidacao"] != "Parcialmente liquidado":
            continue
        if filtros["somente_inconsistentes"] and item["status_liquidacao"] != "Inconsistente":
            continue
        if filtros["faixa_atraso"] in FAIXAS_ATRASO_LIQUIDACAO:
            inicio, fim = FAIXAS_ATRASO_LIQUIDACAO[filtros["faixa_atraso"]]
            dias = item["dias_em_atraso"]
            if dias < inicio or (fim is not None and dias > fim):
                continue
        resultado.append(item)

    resultado.sort(key=lambda item: (-int(item["dias_em_atraso"] or 0), -float(item["saldo_em_aberto"] or 0), item.get("data_vencimento") or ""))
    return resultado


def resumir_liquidacao(movimentacoes):
    resumo = {
        "quantidade_total": len(movimentacoes),
        "total_titulos": 0,
        "total_liquidado": 0,
        "saldo_em_aberto": 0,
        "saldo_vencido": 0,
        "saldo_a_vencer": 0,
        "qtd_liquidado": 0,
        "qtd_em_aberto": 0,
        "qtd_vencido": 0,
        "qtd_a_vencer": 0,
        "qtd_parcial": 0,
        "valor_parcial": 0,
        "qtd_inconsistente": 0,
        "valor_inconsistente": 0,
    }
    for item in movimentacoes:
        resumo["total_titulos"] += float(item["valor_documento_liquidacao"] or 0)
        resumo["total_liquidado"] += float(item["valor_baixado"] or 0)
        resumo["saldo_em_aberto"] += float(item["saldo_em_aberto"] or 0)
        status = item["status_liquidacao"]
        if status == "Liquidado":
            resumo["qtd_liquidado"] += 1
        elif status == "Vencido":
            resumo["qtd_vencido"] += 1
            resumo["saldo_vencido"] += float(item["saldo_em_aberto"] or 0)
        elif status == "Em aberto":
            resumo["qtd_em_aberto"] += 1
            resumo["qtd_a_vencer"] += 1
            resumo["saldo_a_vencer"] += float(item["saldo_em_aberto"] or 0)
        elif status == "Parcialmente liquidado":
            resumo["qtd_parcial"] += 1
            resumo["valor_parcial"] += float(item["saldo_em_aberto"] or 0)
            if int(item["dias_em_atraso"] or 0) > 0:
                resumo["qtd_vencido"] += 1
                resumo["saldo_vencido"] += float(item["saldo_em_aberto"] or 0)
            else:
                resumo["qtd_a_vencer"] += 1
                resumo["saldo_a_vencer"] += float(item["saldo_em_aberto"] or 0)
        elif status == "Inconsistente":
            resumo["qtd_inconsistente"] += 1
            resumo["valor_inconsistente"] += float(item["valor_documento_liquidacao"] or 0)

    for chave, valor in list(resumo.items()):
        if chave.startswith("qtd") or chave == "quantidade_total":
            resumo[chave] = int(valor)
        else:
            resumo[chave] = round(float(valor or 0), 2)
    return resumo


def filtros_query_liquidacao(filtros, pagina=None):
    dados = {chave: valor for chave, valor in filtros.items() if valor not in ["", None, False, "Todas"]}
    dados.pop("pagina", None)
    if pagina is not None:
        dados["pagina"] = pagina
    return urlencode(dados)


def montar_contexto_liquidacao_financeira(args, exportar=False):
    filtros = normalizar_filtros_liquidacao(args)
    base = buscar_movimentacoes_base_liquidacao(filtros)
    movimentacoes = aplicar_filtros_liquidacao(base, filtros)
    resumo = resumir_liquidacao(movimentacoes)
    total_paginas = max(1, (len(movimentacoes) + filtros["por_pagina"] - 1) // filtros["por_pagina"])
    if filtros["pagina"] > total_paginas:
        filtros["pagina"] = total_paginas
    inicio = (filtros["pagina"] - 1) * filtros["por_pagina"]
    fim = inicio + filtros["por_pagina"]
    itens_pagina = movimentacoes if exportar else movimentacoes[inicio:fim]
    return {
        "filtros": filtros,
        "resumo": resumo,
        "movimentacoes": itens_pagina,
        "total_filtrado": len(movimentacoes),
        "status_opcoes": STATUS_LIQUIDACAO_OPCOES,
        "categorias_filtro": categorias_filtro_auditoria(),
        "subcategorias_filtro": sorted({item.get("subcategoria") for item in base if item.get("subcategoria")}),
        "naturezas_filtro": ["Todas", "Saida", "Entrada", "Neutro"],
        "faixas_atraso": FAIXAS_ATRASO_LIQUIDACAO,
        "paginacao": {
            "pagina": filtros["pagina"],
            "total_paginas": total_paginas,
            "tem_anterior": filtros["pagina"] > 1,
            "tem_proxima": filtros["pagina"] < total_paginas,
            "query_anterior": filtros_query_liquidacao(filtros, filtros["pagina"] - 1),
            "query_proxima": filtros_query_liquidacao(filtros, filtros["pagina"] + 1),
        },
    }


def gerar_excel_liquidacao_financeira(contexto):
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidacao"
    ws.append(["Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    ws.append(["Filtros"])
    for chave, valor in contexto["filtros"].items():
        ws.append([chave, valor])
    ws.append([])
    ws.append(["Indicador", "Valor"])
    for chave, valor in contexto["resumo"].items():
        ws.append([chave, valor])
    ws.append([])
    ws.append([
        "Data documento", "Data vencimento", "Data baixa", "Numero documento", "Favorecido",
        "Categoria", "Subcategoria", "Historico", "Valor documento", "Valor baixa",
        "Saldo em aberto", "Status", "Dias em atraso", "Inconsistencia"
    ])
    for item in contexto["movimentacoes"]:
        ws.append([
            item.get("data_documento") or "",
            item.get("data_vencimento") or "",
            item.get("data_realizacao") or "",
            item.get("numero_documento") or item.get("documento_id") or "",
            item.get("favorecido") or item.get("parceiro") or "",
            item.get("categoria") or "",
            item.get("subcategoria") or "",
            item.get("historico") or item.get("observacoes") or "",
            item.get("valor_documento_liquidacao") or 0,
            item.get("valor_baixado") or 0,
            item.get("saldo_em_aberto") or 0,
            item.get("status_liquidacao") or "",
            item.get("dias_em_atraso") or 0,
            item.get("motivo_inconsistencia") or "",
        ])
    ws.append([])
    ws.append(["Totais consolidados"])
    for chave, valor in contexto["resumo"].items():
        ws.append([chave, valor])
    for coluna in range(1, 15):
        ws.column_dimensions[get_column_letter(coluna)].width = 22
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def reclassificar_movimentacoes(
    ids,
    plano_conta_id,
    justificativa="",
    usuario_id=None,
    usuario_nome="Sistema",
    perfil="sistema",
):
    from .reset_financeiro import require_financial_writes_enabled

    require_financial_writes_enabled("Reclassificacao financeira em lote")
    criar_tabela_movimentacoes_financeiras()
    justificativa = (justificativa or "").strip()
    if not justificativa:
        raise ValueError("Informe a justificativa da reclassificação.")

    ids_validos = []
    for item in ids:
        try:
            ids_validos.append(int(item))
        except (TypeError, ValueError):
            pass

    if not ids_validos:
        raise ValueError("Selecione ao menos uma movimentacao.")

    plano = derivar_plano_movimentacao("", plano_conta_id)
    if not plano["plano_conta_id"]:
        raise ValueError("Categoria invalida para reclassificacao.")

    conn = conectar()
    cursor = conn.cursor()
    atualizadas = 0
    try:
        for movimentacao_id in ids_validos:
            cursor.execute(q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"), (movimentacao_id,))
            anterior = cursor.fetchone()
            if not anterior:
                continue
            cursor.execute(q("""
            UPDATE movimentacoes_financeiras
            SET categoria = ?,
                plano_conta_id = ?,
                grupo_gerencial = ?,
                categoria_plano = ?,
                subcategoria = ?,
                centro_analise = ?,
                linha_dre = ?,
                tipo_conta = ?,
                impacta_fluxo_caixa = ?
            WHERE id = ?
            """), (
                plano["categoria_movimentacao"],
                plano["plano_conta_id"],
                plano["grupo_gerencial"],
                plano["categoria_plano"],
                plano["subcategoria"],
                plano["centro_analise"],
                plano["linha_dre"],
                plano["tipo_conta"],
                int(plano["impacta_fluxo_caixa"]),
                movimentacao_id,
            ))
            atualizadas += cursor.rowcount
            cursor.execute(q("SELECT * FROM movimentacoes_financeiras WHERE id = ?"), (movimentacao_id,))
            posterior = cursor.fetchone()
            _registrar_auditoria_cursor(
                cursor,
                movimentacao_id,
                "RECLASSIFICACAO",
                anterior,
                posterior,
                usuario_id=usuario_id,
                usuario_nome=usuario_nome,
                perfil=perfil,
                justificativa=justificativa,
                origem_acao="auditoria_financeira",
                referencia_importacao=anterior["import_key"],
            )
        conn.commit()
        return atualizadas
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def gerar_excel_auditoria_financeira(contexto):
    wb = Workbook()
    ws = wb.active
    ws.title = "Visao geral"
    indicadores = contexto["indicadores"]
    ws.append(["Indicador", "Valor"])
    ws.append(["Total de movimentacoes importadas", indicadores["total_movimentacoes"]])
    ws.append(["Total de receitas", indicadores["total_receitas"]])
    ws.append(["Total de despesas", indicadores["total_despesas"]])
    ws.append(["Saldo liquido", indicadores["saldo_liquido"]])
    ws.append(["Qtd. Nao Classificado", indicadores["qtd_nao_classificado"]])
    ws.append(["Valor Nao Classificado", indicadores["valor_nao_classificado"]])

    ws_cat = wb.create_sheet("Por categoria")
    ws_cat.append(["Categoria", "Quantidade", "Valor total", "Percentual"])
    for item in contexto["totais_categoria"]:
        ws_cat.append([
            texto_excel_seguro(item["categoria"]),
            item["quantidade"],
            item["valor_total"],
            item["percentual"],
        ])

    ws_mes = wb.create_sheet("Por mes")
    ws_mes.append(["Mes", "Receitas", "Despesas", "Saldo", "Quantidade"])
    for item in contexto["totais_mes"]:
        ws_mes.append([
            texto_excel_seguro(item["mes"]),
            item["total_receitas"],
            item["total_despesas"],
            item["saldo"],
            item["quantidade"],
        ])

    ws_pend = wb.create_sheet("Pendencias")
    ws_pend.append([
        "Data documento", "Data vencimento", "Favorecido", "Descricao", "Historico",
        "Numero documento", "Valor documento", "Valor pago", "Categoria atual"
    ])
    for item in contexto["pendencias"]:
        ws_pend.append([
            data_excel(item.get("data_documento")),
            data_excel(item.get("data_vencimento")),
            texto_excel_seguro(item.get("favorecido") or item.get("parceiro") or ""),
            texto_excel_seguro(item.get("descricao", "")),
            texto_excel_seguro(item.get("historico", "")),
            texto_excel_seguro(item.get("numero_documento") or item.get("documento_id") or ""),
            float(item.get("valor_documento") or item.get("valor") or 0),
            float(item.get("valor_pago") or 0),
            texto_excel_seguro(item.get("categoria", "")),
        ])

    ws_mov = wb.create_sheet("Movimentações")
    cabecalhos = [
        "ID movimentacao", "Natureza", "Situacao", "Cancelada",
        "Data cancelamento", "Justificativa cancelamento",
        "Data documento", "Data vencimento", "Data pagamento/realizacao",
        "Numero documento", "Parcela", "Total parcelas", "CNPJ/CPF",
        "Favorecido/Cliente", "Descricao", "Historico", "Observacoes",
        "Grupo gerencial", "Categoria", "Subcategoria", "Centro de analise",
        "ID conta Plano de Contas", "Conta Plano de Contas", "Classificacao DRE",
        "Valor documental", "Valor pago", "Valor liquido",
        "Valor regra financeira", "Saldo aberto",
        "Import key", "Origem de importacao", "Modo de origem resolvido",
        "Origem persistida", "Resolucao da origem", "Sistema de origem",
        "Modulo de origem", "Tipo de evento", "Chave externa",
        "Chave idempotente", "ID lote", "Arquivo do lote", "SHA-256 do lote",
        "Linha do arquivo", "Usuario de criacao", "Data/hora de criacao",
        "Usuario da ultima alteracao", "Data/hora da ultima alteracao",
    ]
    ws_mov.append(cabecalhos)

    for item in contexto.get("movimentacoes", []):
        cancelada = (item.get("status") or "") == "Cancelado"
        conta_partes = [
            item.get("plano_grupo_gerencial") or item.get("grupo_gerencial"),
            item.get("plano_categoria") or item.get("categoria_plano"),
            item.get("plano_subcategoria") or item.get("subcategoria"),
        ]
        conta_rotulo = " / ".join(str(parte) for parte in conta_partes if parte)
        liquidacao = calcular_status_liquidacao(item)
        saldo_aberto = "" if cancelada else liquidacao["saldo_em_aberto"]
        usuario_criacao = (
            item.get("origem_usuario_nome")
            or item.get("primeiro_usuario_nome")
            or ""
        )
        criado_em = item.get("criado_em") or item.get("origem_criado_em")
        ws_mov.append([
            int(item["id"]),
            texto_excel_seguro(item.get("tipo")),
            texto_excel_seguro(item.get("status")),
            "Sim" if cancelada else "Nao",
            data_excel(item.get("cancelado_em"), incluir_hora=True) if cancelada else "",
            texto_excel_seguro(item.get("cancelamento_justificativa")) if cancelada else "",
            data_excel(item.get("data_documento")),
            data_excel(item.get("data_vencimento")),
            data_excel(item.get("data_realizacao")),
            texto_excel_seguro(item.get("numero_documento") or item.get("documento_id")),
            int(item.get("parcela_atual") or 1),
            int(item.get("parcelas") or 1),
            texto_excel_seguro(item.get("cnpj_cpf")),
            texto_excel_seguro(item.get("favorecido") or item.get("parceiro")),
            texto_excel_seguro(item.get("descricao")),
            texto_excel_seguro(item.get("historico")),
            texto_excel_seguro(item.get("observacoes")),
            texto_excel_seguro(item.get("grupo_gerencial")),
            texto_excel_seguro(item.get("categoria_plano") or item.get("categoria")),
            texto_excel_seguro(item.get("subcategoria")),
            texto_excel_seguro(item.get("centro_analise")),
            item.get("plano_conta_id") or "",
            texto_excel_seguro(conta_rotulo),
            texto_excel_seguro(item.get("linha_dre")),
            float(item.get("valor_documento") or item.get("valor") or 0),
            float(item.get("valor_pago") or 0),
            float(item.get("valor_liquido") or 0),
            float(item.get("valor") or 0),
            saldo_aberto,
            texto_excel_seguro(item.get("import_key")),
            texto_excel_seguro(item.get("origem_importacao")),
            texto_excel_seguro(item.get("modo_origem")),
            "Sim" if item.get("origem_persistida") else "Nao",
            "Persistida" if item.get("origem_persistida") else "Fallback legado",
            texto_excel_seguro(
                item.get("sistema_origem")
                or ("FrigoDatta legado" if not item.get("origem_persistida") else "")
            ),
            texto_excel_seguro(item.get("modulo_origem")),
            texto_excel_seguro(item.get("origem_tipo_evento")),
            texto_excel_seguro(item.get("origem_chave_externa")),
            texto_excel_seguro(item.get("origem_chave_idempotente")),
            item.get("lote_importacao_id") or "",
            texto_excel_seguro(item.get("lote_arquivo_nome")),
            texto_excel_seguro(item.get("lote_arquivo_hash")),
            item.get("origem_linha_arquivo") or "",
            texto_excel_seguro(usuario_criacao),
            data_excel(criado_em, incluir_hora=True),
            texto_excel_seguro(item.get("ultimo_usuario_nome")),
            data_excel(item.get("ultima_alteracao_em"), incluir_hora=True),
        ])

    formatar_aba_movimentacoes_auditoria(ws_mov)

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo


def texto_excel_seguro(valor):
    """Neutraliza formula injection somente em campos textuais."""
    if valor is None:
        return ""
    texto = str(valor)
    if texto.startswith(("=", "+", "-", "@")):
        return f"'{texto}"
    return texto


def data_excel(valor, incluir_hora=False):
    """Converte datas ISO/SQL em tipos nativos do Excel quando reconhecíveis."""
    if valor in (None, ""):
        return ""
    if isinstance(valor, datetime):
        return valor if incluir_hora else valor.date()
    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()
    formatos = (
        ("%Y-%m-%d %H:%M:%S.%f", True),
        ("%Y-%m-%d %H:%M:%S", True),
        ("%Y-%m-%dT%H:%M:%S", True),
        ("%Y-%m-%d", False),
    )
    for formato, tem_hora in formatos:
        try:
            convertido = datetime.strptime(texto[:26], formato)
            return convertido if incluir_hora and tem_hora else convertido.date()
        except ValueError:
            continue
    return texto_excel_seguro(texto)


def formatar_aba_movimentacoes_auditoria(ws):
    """Aplica formatação limitada, tipada e adequada a grandes populações."""
    ultima_coluna = get_column_letter(ws.max_column)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ultima_coluna}{max(ws.max_row, 1)}"
    ws.sheet_view.showGridLines = False

    preenchimento = PatternFill("solid", fgColor="1F4E78")
    for celula in ws[1]:
        celula.fill = preenchimento
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    colunas_data = {5, 7, 8, 9}
    colunas_data_hora = {45, 47}
    colunas_monetarias = {25, 26, 27, 28, 29}
    colunas_texto_forcado = {10, 13, 30, 38, 39, 42}

    if ws.max_row >= 2:
        for coluna in colunas_data:
            for celula in ws.iter_cols(min_col=coluna, max_col=coluna, min_row=2):
                celula[0].number_format = "dd/mm/yyyy"
        for coluna in colunas_data_hora:
            for celula in ws.iter_cols(min_col=coluna, max_col=coluna, min_row=2):
                celula[0].number_format = "dd/mm/yyyy hh:mm:ss"
        for coluna in colunas_monetarias:
            for celula in ws.iter_cols(min_col=coluna, max_col=coluna, min_row=2):
                celula[0].number_format = '"R$" #,##0.00'
        for coluna in colunas_texto_forcado:
            for celula in ws.iter_cols(min_col=coluna, max_col=coluna, min_row=2):
                celula[0].number_format = "@"

    larguras = {
        1: 14, 2: 12, 3: 18, 4: 11, 5: 19, 6: 34,
        7: 15, 8: 15, 9: 22, 10: 20, 11: 10, 12: 12,
        13: 20, 14: 30, 15: 34, 16: 34, 17: 34, 18: 24,
        19: 24, 20: 24, 21: 22, 22: 20, 23: 38, 24: 28,
        25: 18, 26: 18, 27: 18, 28: 20, 29: 18, 30: 44,
        31: 22, 32: 25, 33: 18, 34: 20, 35: 24, 36: 22,
        37: 22, 38: 30, 39: 36, 40: 14, 41: 28, 42: 66,
        43: 16, 44: 24, 45: 22, 46: 28, 47: 24,
    }
    for indice, largura in larguras.items():
        ws.column_dimensions[get_column_letter(indice)].width = min(largura, 66)

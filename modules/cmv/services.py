"""CMV FIFO independente do estoque fisico.

O modulo e um subledger gerencial: registra somente custos explicitamente
conhecidos e nunca altera as tabelas operacionais de estoque. Quantidade sem
custo permanece identificada como lacuna, em vez de receber custo zero.
"""

from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from database import DATABASE_URL, conectar, q, transaction


ESTADOS_CALCULO = ("CALCULAVEL", "PARCIAL", "NAO_CALCULAVEL", "INCONSISTENTE")
TIPOS_EVENTO = ("VENDA", "ESTORNO_VENDA", "DESCARTE")
_SCHEMA_INICIALIZADO = False
CENTAVO = Decimal("0.01")
SEIS_CASAS = Decimal("0.000001")


def _decimal(valor, campo="valor", permite_zero=True):
    try:
        numero = Decimal(str(valor))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{campo} invalido.")
    if numero < 0 or (not permite_zero and numero == 0):
        raise ValueError(f"{campo} deve ser maior que zero." if not permite_zero else f"{campo} nao pode ser negativo.")
    return numero


def _data_iso(valor):
    texto = str(valor or "").strip()
    try:
        return date.fromisoformat(texto[:10]).isoformat()
    except ValueError:
        raise ValueError("Data deve estar no formato AAAA-MM-DD.")


def criar_tabelas_cmv():
    global _SCHEMA_INICIALIZADO
    if _SCHEMA_INICIALIZADO:
        return
    id_pk = "SERIAL PRIMARY KEY" if DATABASE_URL else "INTEGER PRIMARY KEY AUTOINCREMENT"
    timestamp_type = "TIMESTAMP" if DATABASE_URL else "TEXT"
    with transaction() as conn:
        cursor = conn.cursor()
        cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS cmv_camadas (
            id {id_pk}, produto TEXT NOT NULL, unidade TEXT NOT NULL,
            data_entrada TEXT NOT NULL, quantidade_inicial REAL NOT NULL,
            quantidade_disponivel REAL NOT NULL, custo_unitario REAL,
            custo_conhecido INTEGER NOT NULL DEFAULT 0, origem_tipo TEXT NOT NULL,
            origem_id TEXT, documento TEXT, op_id INTEGER, lote TEXT,
            idempotency_key TEXT NOT NULL UNIQUE, status TEXT NOT NULL DEFAULT 'ATIVA',
            criado_por TEXT NOT NULL, criado_em {timestamp_type} NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """)
        cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS cmv_eventos (
            id {id_pk}, tipo TEXT NOT NULL, data_evento TEXT NOT NULL,
            documento TEXT NOT NULL, produto TEXT NOT NULL, unidade TEXT NOT NULL,
            quantidade REAL NOT NULL, quantidade_com_custo REAL NOT NULL DEFAULT 0,
            quantidade_sem_custo REAL NOT NULL DEFAULT 0, custo_total REAL,
            estado_calculo TEXT NOT NULL, evento_original_id INTEGER,
            origem_tipo TEXT NOT NULL, origem_id TEXT, idempotency_key TEXT NOT NULL UNIQUE,
            justificativa TEXT, criado_por TEXT NOT NULL,
            criado_em {timestamp_type} NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """)
        cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS cmv_consumos (
            id {id_pk}, evento_id INTEGER NOT NULL, camada_id INTEGER,
            quantidade REAL NOT NULL, custo_unitario REAL, custo_total REAL,
            custo_conhecido INTEGER NOT NULL DEFAULT 0, ordem_fifo INTEGER NOT NULL,
            restaurado INTEGER NOT NULL DEFAULT 0,
            criado_em {timestamp_type} NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """)
        cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS cmv_auditoria (
            id {id_pk}, entidade TEXT NOT NULL, entidade_id INTEGER NOT NULL,
            acao TEXT NOT NULL, dados TEXT, usuario TEXT NOT NULL,
            criado_em {timestamp_type} NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cmv_camadas_fifo ON cmv_camadas(produto,unidade,status,data_entrada,id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cmv_eventos_periodo ON cmv_eventos(data_evento,tipo,estado_calculo)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cmv_consumos_evento ON cmv_consumos(evento_id,ordem_fifo)")
    _SCHEMA_INICIALIZADO = True


def _ultimo_id(cursor):
    if DATABASE_URL:
        cursor.execute("SELECT LASTVAL() AS id")
        return cursor.fetchone()["id"]
    return cursor.lastrowid


def _auditar(cursor, entidade, entidade_id, acao, dados, usuario):
    cursor.execute(q("INSERT INTO cmv_auditoria(entidade,entidade_id,acao,dados,usuario) VALUES (?,?,?,?,?)"),
                   (entidade, entidade_id, acao, str(dados), usuario or "Sistema"))


def registrar_camada(*, produto, unidade, data_entrada, quantidade, custo_unitario=None,
                     custo_conhecido=True, origem_tipo, origem_id=None, documento=None,
                     op_id=None, lote=None, idempotency_key, usuario="Sistema"):
    """Registra uma entrada valorizada ou explicitamente desconhecida, uma unica vez."""
    criar_tabelas_cmv()
    produto = str(produto or "").strip()
    unidade = str(unidade or "").strip().upper()
    quantidade = _decimal(quantidade, "Quantidade", permite_zero=False)
    if not produto or not unidade or not str(idempotency_key or "").strip():
        raise ValueError("Produto, unidade e chave de idempotencia sao obrigatorios.")
    custo = _decimal(custo_unitario, "Custo unitario") if custo_conhecido else None
    if custo_conhecido and custo_unitario is None:
        raise ValueError("Custo unitario e obrigatorio quando o custo e conhecido.")
    with transaction() as conn:
        cursor = conn.cursor()
        cursor.execute(q("SELECT * FROM cmv_camadas WHERE idempotency_key=?"), (idempotency_key,))
        existente = cursor.fetchone()
        if existente:
            return dict(existente), False
        cursor.execute(q("""
        INSERT INTO cmv_camadas(produto,unidade,data_entrada,quantidade_inicial,
          quantidade_disponivel,custo_unitario,custo_conhecido,origem_tipo,origem_id,
          documento,op_id,lote,idempotency_key,criado_por)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """), (produto, unidade, _data_iso(data_entrada), float(quantidade), float(quantidade),
                float(custo) if custo is not None else None, 1 if custo_conhecido else 0,
                origem_tipo, origem_id, documento, op_id, lote, idempotency_key, usuario))
        camada_id = _ultimo_id(cursor)
        _auditar(cursor, "CAMADA", camada_id, "CRIACAO", {"quantidade": str(quantidade), "custo": str(custo)}, usuario)
        cursor.execute(q("SELECT * FROM cmv_camadas WHERE id=?"), (camada_id,))
        return dict(cursor.fetchone()), True


def calcular_custo_op(op_id):
    """Reconcilia somente consumo oficial valorizado e PA efetivamente pesado.

    A funcao e deliberadamente conservadora: ausencia de consumo, lote ou valor
    torna o custo nao calculavel. Nao usa parametros atuais nem peso estimado.
    """
    conn = conectar()
    try:
        cursor = conn.cursor()
        cursor.execute(q("SELECT id,data,sku,status FROM ordens_producao WHERE id=?"), (op_id,))
        op = cursor.fetchone()
        if not op:
            raise ValueError("OP nao encontrada.")
        cursor.execute(q("""SELECT id,lote_id,quantidade,valor_unitario,valor_total
          FROM almoxarifado_movimentacoes WHERE op_id=? AND tipo='SAIDA_OP' ORDER BY id"""), (op_id,))
        consumos = cursor.fetchall()
        custo_conhecido = bool(consumos) and all(item["lote_id"] is not None and item["valor_unitario"] is not None and item["valor_total"] is not None for item in consumos)
        custo_total = sum(Decimal(str(item["valor_total"] or 0)) for item in consumos) if custo_conhecido else None
        # Caixas compostas sao rateadas pela quantidade oficial da composicao.
        cursor.execute(q("""SELECT cx.id,cx.sku,cx.peso_liquido,cx.quantidade_bandejas,
          c.quantidade_bandejas quantidade_op,
          (SELECT SUM(c2.quantidade_bandejas) FROM pa_caixa_composicao c2 WHERE c2.caixa_id=cx.id) quantidade_caixa
          FROM pa_caixas cx JOIN pa_caixa_composicao c ON c.caixa_id=cx.id
          WHERE c.op_id=? AND UPPER(COALESCE(cx.status,'')) NOT IN ('ESTORNADA','ESTORNADO','CANCELADA','CANCELADO')"""), (op_id,))
        caixas = cursor.fetchall()
        sku = op["sku"] or (caixas[0]["sku"] if caixas else "Produto nao identificado")
        if sku == "Galinha Inteira":
            unidade = "UN"
            quantidade = sum(Decimal(str(item["quantidade_op"] or 0)) for item in caixas)
        else:
            unidade = "KG"
            quantidade = Decimal("0")
            for item in caixas:
                total_caixa = Decimal(str(item["quantidade_caixa"] or 0))
                proporcao = Decimal(str(item["quantidade_op"] or 0)) / total_caixa if total_caixa > 0 else Decimal("0")
                quantidade += Decimal(str(item["peso_liquido"] or 0)) * proporcao
        if quantidade <= 0:
            custo_conhecido = False
        custo_unitario = custo_total / quantidade if custo_conhecido and custo_total is not None else None
        return {"op_id": int(op_id), "data": op["data"], "produto": sku, "unidade": unidade,
                "quantidade_pa": float(quantidade), "componentes_oficiais": len(consumos),
                "custo_total": float(custo_total) if custo_total is not None else None,
                "custo_unitario": float(custo_unitario) if custo_unitario is not None else None,
                "estado_calculo": "CALCULAVEL" if custo_conhecido else "NAO_CALCULAVEL"}
    except Exception as erro:
        # Bancos legados sem uma das fontes oficiais sao uma lacuna de dados,
        # nao motivo para estimar custo.
        if isinstance(erro, ValueError):
            raise
        return {"op_id": int(op_id), "quantidade_pa": 0, "componentes_oficiais": 0,
                "custo_total": None, "custo_unitario": None,
                "estado_calculo": "NAO_CALCULAVEL", "limitacao": str(erro)}
    finally:
        conn.close()


def registrar_camada_op(op_id, *, usuario="Sistema"):
    reconciliacao = calcular_custo_op(op_id)
    if reconciliacao["quantidade_pa"] <= 0:
        raise ValueError("OP sem quantidade oficial de produto acabado para valorizar.")
    return registrar_camada(
        produto=reconciliacao["produto"], unidade=reconciliacao["unidade"],
        data_entrada=reconciliacao["data"], quantidade=reconciliacao["quantidade_pa"],
        custo_unitario=reconciliacao["custo_unitario"],
        custo_conhecido=reconciliacao["estado_calculo"] == "CALCULAVEL",
        origem_tipo="OP", origem_id=str(op_id), op_id=op_id,
        documento=f"OP-{op_id}", idempotency_key=f"CMV:OP:{op_id}", usuario=usuario,
    )


def _estado(quantidade, com_custo, sem_custo, inconsistente=False):
    if inconsistente:
        return "INCONSISTENTE"
    if com_custo == quantidade and sem_custo == 0:
        return "CALCULAVEL"
    if com_custo > 0:
        return "PARCIAL"
    return "NAO_CALCULAVEL"


def registrar_saida(*, data_evento, documento, produto, unidade, quantidade,
                    origem_tipo, origem_id=None, idempotency_key, usuario="Sistema",
                    tipo_evento="VENDA"):
    """Consome FIFO com lock. Falta de camada vira consumo sem custo, nunca zero."""
    criar_tabelas_cmv()
    quantidade = _decimal(quantidade, "Quantidade", permite_zero=False)
    produto = str(produto or "").strip()
    unidade = str(unidade or "").strip().upper()
    if tipo_evento not in ("VENDA", "DESCARTE"):
        raise ValueError("Tipo de evento de saida invalido.")
    with transaction() as conn:
        cursor = conn.cursor()
        cursor.execute(q("SELECT * FROM cmv_eventos WHERE idempotency_key=?"), (idempotency_key,))
        existente = cursor.fetchone()
        if existente:
            return dict(existente), False
        lock = " FOR UPDATE" if DATABASE_URL else ""
        cursor.execute(q("""SELECT * FROM cmv_camadas
          WHERE produto=? AND unidade=? AND status='ATIVA' AND quantidade_disponivel>0
          ORDER BY data_entrada,id""" + lock), (produto, unidade))
        camadas = cursor.fetchall()
        restante = quantidade
        alocacoes = []
        custo_total = Decimal("0")
        com_custo = Decimal("0")
        ordem = 0
        for camada in camadas:
            if restante <= 0:
                break
            disponivel = Decimal(str(camada["quantidade_disponivel"] or 0))
            consumir = min(restante, disponivel)
            if consumir <= 0:
                continue
            ordem += 1
            conhecido = bool(camada["custo_conhecido"])
            unitario = Decimal(str(camada["custo_unitario"] or 0)) if conhecido else None
            total = consumir * unitario if conhecido else None
            if conhecido:
                com_custo += consumir
                custo_total += total
            alocacoes.append((camada["id"], consumir, unitario, total, conhecido, ordem))
            novo_saldo = disponivel - consumir
            cursor.execute(q("UPDATE cmv_camadas SET quantidade_disponivel=?,status=? WHERE id=?"),
                           (float(novo_saldo), "ESGOTADA" if novo_saldo == 0 else "ATIVA", camada["id"]))
            restante -= consumir
        sem_custo = quantidade - com_custo
        estado = _estado(quantidade, com_custo, sem_custo)
        custo_gravado = float(custo_total.quantize(SEIS_CASAS)) if com_custo > 0 else None
        cursor.execute(q("""INSERT INTO cmv_eventos(tipo,data_evento,documento,produto,unidade,
          quantidade,quantidade_com_custo,quantidade_sem_custo,custo_total,estado_calculo,
          origem_tipo,origem_id,idempotency_key,criado_por)
          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""),
          (tipo_evento, _data_iso(data_evento), documento, produto, unidade, float(quantidade), float(com_custo),
           float(sem_custo), custo_gravado, estado, origem_tipo, origem_id, idempotency_key, usuario))
        evento_id = _ultimo_id(cursor)
        for camada_id, qtd, unitario, total, conhecido, ordem in alocacoes:
            cursor.execute(q("""INSERT INTO cmv_consumos(evento_id,camada_id,quantidade,custo_unitario,
              custo_total,custo_conhecido,ordem_fifo) VALUES (?,?,?,?,?,?,?)"""),
              (evento_id, camada_id, float(qtd), float(unitario) if unitario is not None else None,
               float(total) if total is not None else None, 1 if conhecido else 0, ordem))
        if restante > 0:
            cursor.execute(q("""INSERT INTO cmv_consumos(evento_id,camada_id,quantidade,custo_unitario,
              custo_total,custo_conhecido,ordem_fifo) VALUES (?,?,?,?,?,?,?)"""),
              (evento_id, None, float(restante), None, None, 0, ordem + 1))
        _auditar(cursor, "EVENTO", evento_id, "CONSUMO_FIFO", {"estado": estado, "cobertura": str(com_custo)}, usuario)
        cursor.execute(q("SELECT * FROM cmv_eventos WHERE id=?"), (evento_id,))
        return dict(cursor.fetchone()), True


def estornar_saida(evento_id, *, data_evento, idempotency_key, justificativa, usuario="Sistema"):
    """Restaura exatamente as camadas consumidas pelo evento original."""
    criar_tabelas_cmv()
    if not str(justificativa or "").strip():
        raise ValueError("Justificativa e obrigatoria.")
    with transaction() as conn:
        cursor = conn.cursor()
        cursor.execute(q("SELECT * FROM cmv_eventos WHERE idempotency_key=?"), (idempotency_key,))
        existente = cursor.fetchone()
        if existente:
            return dict(existente), False
        lock = " FOR UPDATE" if DATABASE_URL else ""
        cursor.execute(q("SELECT * FROM cmv_eventos WHERE id=?" + lock), (evento_id,))
        original = cursor.fetchone()
        if not original or original["tipo"] != "VENDA":
            raise ValueError("Evento de venda original nao encontrado.")
        cursor.execute(q("SELECT id FROM cmv_eventos WHERE tipo='ESTORNO_VENDA' AND evento_original_id=?" + lock), (evento_id,))
        if cursor.fetchone():
            raise ValueError("Venda ja estornada.")
        cursor.execute(q("SELECT * FROM cmv_consumos WHERE evento_id=? ORDER BY ordem_fifo" + lock), (evento_id,))
        consumos = cursor.fetchall()
        for consumo in consumos:
            if consumo["camada_id"]:
                cursor.execute(q("""UPDATE cmv_camadas SET quantidade_disponivel=quantidade_disponivel+?,
                  status='ATIVA' WHERE id=?"""), (consumo["quantidade"], consumo["camada_id"]))
            cursor.execute(q("UPDATE cmv_consumos SET restaurado=1 WHERE id=?"), (consumo["id"],))
        custo = -float(original["custo_total"]) if original["custo_total"] is not None else None
        cursor.execute(q("""INSERT INTO cmv_eventos(tipo,data_evento,documento,produto,unidade,
          quantidade,quantidade_com_custo,quantidade_sem_custo,custo_total,estado_calculo,
          evento_original_id,origem_tipo,origem_id,idempotency_key,justificativa,criado_por)
          VALUES ('ESTORNO_VENDA',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""),
          (_data_iso(data_evento), original["documento"], original["produto"], original["unidade"],
           -float(original["quantidade"]), -float(original["quantidade_com_custo"]),
           -float(original["quantidade_sem_custo"]), custo, original["estado_calculo"], evento_id,
           original["origem_tipo"], original["origem_id"], idempotency_key, justificativa, usuario))
        estorno_id = _ultimo_id(cursor)
        _auditar(cursor, "EVENTO", estorno_id, "ESTORNO", {"evento_original": evento_id}, usuario)
        cursor.execute(q("SELECT * FROM cmv_eventos WHERE id=?"), (estorno_id,))
        return dict(cursor.fetchone()), True


def registrar_descarte(**kwargs):
    """Consome camadas, mas classifica o evento como perda, fora do CMV de venda."""
    return registrar_saida(**kwargs, tipo_evento="DESCARTE")


def resumo_periodo(data_inicio, data_fim):
    criar_tabelas_cmv()
    inicio, fim = _data_iso(data_inicio), _data_iso(data_fim)
    conn = conectar()
    try:
        cursor = conn.cursor()
        cursor.execute(q("""SELECT produto,unidade,
          COALESCE(SUM(quantidade),0) quantidade,
          COALESCE(SUM(quantidade_com_custo),0) quantidade_com_custo,
          COALESCE(SUM(quantidade_sem_custo),0) quantidade_sem_custo,
          SUM(custo_total) custo_total
          FROM cmv_eventos WHERE data_evento BETWEEN ? AND ? AND tipo IN ('VENDA','ESTORNO_VENDA')
          GROUP BY produto,unidade ORDER BY produto,unidade"""), (inicio, fim))
        linhas = [dict(item) for item in cursor.fetchall()]
        quantidade = sum(Decimal(str(item["quantidade"] or 0)) for item in linhas)
        com_custo = sum(Decimal(str(item["quantidade_com_custo"] or 0)) for item in linhas)
        sem_custo = sum(Decimal(str(item["quantidade_sem_custo"] or 0)) for item in linhas)
        custos = [Decimal(str(item["custo_total"])) for item in linhas if item["custo_total"] is not None]
        estado = _estado(quantidade, com_custo, sem_custo) if quantidade > 0 else ("CALCULAVEL" if linhas else "NAO_CALCULAVEL")
        cobertura = (com_custo / quantidade * 100) if quantidade > 0 else Decimal("0")
        return {"data_inicio": inicio, "data_fim": fim, "estado_calculo": estado,
                "quantidade_vendida": float(quantidade), "quantidade_com_custo": float(com_custo),
                "quantidade_sem_custo": float(sem_custo), "cobertura_percentual": float(cobertura.quantize(CENTAVO)),
                "cmv_total": float(sum(custos).quantize(CENTAVO)) if com_custo > 0 or (linhas and quantidade == 0) else None,
                "linhas": linhas}
    finally:
        conn.close()


def estoque_valorizado():
    criar_tabelas_cmv()
    conn = conectar()
    try:
        cursor = conn.cursor()
        cursor.execute("""SELECT produto,unidade,COUNT(*) camadas,
          SUM(quantidade_disponivel) quantidade,
          SUM(CASE WHEN custo_conhecido=1 THEN quantidade_disponivel ELSE 0 END) quantidade_com_custo,
          SUM(CASE WHEN custo_conhecido=0 THEN quantidade_disponivel ELSE 0 END) quantidade_sem_custo,
          SUM(CASE WHEN custo_conhecido=1 THEN quantidade_disponivel*custo_unitario ELSE 0 END) valor_conhecido
          FROM cmv_camadas WHERE quantidade_disponivel>0 GROUP BY produto,unidade ORDER BY produto,unidade""")
        linhas = [dict(item) for item in cursor.fetchall()]
        for item in linhas:
            qtd = Decimal(str(item["quantidade"] or 0))
            conhecida = Decimal(str(item["quantidade_com_custo"] or 0))
            item["cobertura_percentual"] = float((conhecida / qtd * 100).quantize(CENTAVO)) if qtd > 0 else 0
        return linhas
    finally:
        conn.close()


def detalhamento_periodo(data_inicio, data_fim):
    criar_tabelas_cmv()
    conn = conectar()
    try:
        cursor = conn.cursor()
        cursor.execute(q("""SELECT e.id evento_id,e.tipo,e.data_evento,e.documento,e.produto,e.unidade,
          e.estado_calculo,c.ordem_fifo,c.quantidade,c.custo_unitario,c.custo_total,c.custo_conhecido,
          l.id camada_id,l.data_entrada,l.origem_tipo origem_custo,l.origem_id origem_custo_id,l.lote
          FROM cmv_eventos e JOIN cmv_consumos c ON c.evento_id=e.id
          LEFT JOIN cmv_camadas l ON l.id=c.camada_id
          WHERE e.data_evento BETWEEN ? AND ? ORDER BY e.data_evento,e.id,c.ordem_fifo"""),
          (_data_iso(data_inicio), _data_iso(data_fim)))
        return [dict(item) for item in cursor.fetchall()]
    finally:
        conn.close()


def gerar_excel(data_inicio, data_fim):
    resumo = resumo_periodo(data_inicio, data_fim)
    detalhe = detalhamento_periodo(data_inicio, data_fim)
    estoque = estoque_valorizado()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo CMV"
    ws.append(["CMV FIFO auditavel", resumo["data_inicio"], resumo["data_fim"]])
    ws.append(["Estado", resumo["estado_calculo"]])
    ws.append(["Cobertura (%)", resumo["cobertura_percentual"]])
    ws.append(["CMV conhecido", resumo["cmv_total"] if resumo["cmv_total"] is not None else "N/A"])
    for cell in ws[1]: cell.font = Font(bold=True)
    wd = wb.create_sheet("Consumos FIFO")
    colunas = ["evento_id","tipo","data_evento","documento","produto","unidade","estado_calculo","ordem_fifo","quantidade","camada_id","data_entrada","custo_unitario","custo_total","origem_custo","origem_custo_id","lote"]
    wd.append(colunas)
    for item in detalhe: wd.append([item.get(c) for c in colunas])
    we = wb.create_sheet("Estoque valorizado")
    col_estoque = ["produto","unidade","camadas","quantidade","quantidade_com_custo","quantidade_sem_custo","cobertura_percentual","valor_conhecido"]
    we.append(col_estoque)
    for item in estoque: we.append([item.get(c) for c in col_estoque])
    for planilha in (wd, we):
        for cell in planilha[1]: cell.font = Font(bold=True)
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = planilha.dimensions
    arquivo = BytesIO(); wb.save(arquivo); arquivo.seek(0)
    return arquivo


def gerar_pdf(data_inicio, data_fim):
    resumo = resumo_periodo(data_inicio, data_fim)
    estoque = estoque_valorizado()
    arquivo = BytesIO()
    doc = SimpleDocTemplate(arquivo, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    estilos = getSampleStyleSheet()
    elementos = [Paragraph("CMV e estoque valorizado", estilos["Title"]),
                 Paragraph(f"Periodo: {resumo['data_inicio']} a {resumo['data_fim']} | Estado: {resumo['estado_calculo']} | Cobertura: {resumo['cobertura_percentual']:.2f}%", estilos["Normal"]), Spacer(1, 6*mm)]
    cmv = "N/A" if resumo["cmv_total"] is None else f"R$ {resumo['cmv_total']:,.2f}"
    dados = [["CMV conhecido", "Qtd. vendida", "Qtd. com custo", "Qtd. sem custo"], [cmv, resumo["quantidade_vendida"], resumo["quantidade_com_custo"], resumo["quantidade_sem_custo"]]]
    tabela = Table(dados, repeatRows=1); tabela.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F3B4D")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.25,colors.grey),("ALIGN",(1,1),(-1,-1),"RIGHT")]))
    elementos.extend([tabela, Spacer(1, 6*mm), Paragraph("Estoque valorizado", estilos["Heading2"])])
    linhas = [["Produto","Unidade","Quantidade","Com custo","Sem custo","Cobertura","Valor conhecido"]]
    for item in estoque:
        linhas.append([item["produto"],item["unidade"],item["quantidade"],item["quantidade_com_custo"],item["quantidade_sem_custo"],f"{item['cobertura_percentual']:.2f}%",f"R$ {float(item['valor_conhecido'] or 0):,.2f}"])
    t2 = Table(linhas, repeatRows=1); t2.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F3B4D")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.25,colors.grey)]))
    elementos.append(t2); doc.build(elementos); arquivo.seek(0)
    return arquivo

"""Servicos compartilhados para relatorios financeiros oficiais."""

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database import conectar, q


LINHA_RECEITA_BRUTA = "Receita Bruta"
LINHA_DEDUCOES_RECEITA = "Deducoes da Receita"
LINHA_DESPESAS_OPERACIONAIS = "Despesas Operacionais"
LINHA_RESULTADO_NAO_OPERACIONAL = "Resultado Nao Operacional"
LINHA_NEUTRA = "Neutro"
NATUREZAS_FILTRO = ["Todas", "Entrada", "Saida"]
STATUS_REALIZADOS = ["Pago", "Recebido", "Realizado"]
COMPETENCIA_SEM_DATA_DOCUMENTO = "Sem data do documento"


RELATORIOS_FINANCEIROS = {
    "entradas-caixa": {
        "catalogo_id": "financeiro-entradas-caixa",
        "titulo": "Entradas de Caixa",
        "objetivo": "Identificar fontes que geraram ou gerarao caixa e quando isso acontece.",
        "familia": "caixa",
        "tipo": "Entrada",
        "data_padrao": "vencimento",
        "excluir_aportes": False,
        "somente_aportes": False,
    },
    "saidas-caixa": {
        "catalogo_id": "financeiro-saidas-caixa",
        "titulo": "Saidas de Caixa",
        "objetivo": "Evidenciar para onde o caixa foi ou sera direcionado.",
        "familia": "caixa",
        "tipo": "Saida",
        "data_padrao": "vencimento",
        "excluir_aportes": True,
        "somente_aportes": False,
    },
    "contas-pagar": {
        "catalogo_id": "financeiro-contas-pagar",
        "titulo": "Contas a Pagar",
        "objetivo": "Listar obrigacoes a pagar, vencidas, a vencer e liquidadas.",
        "familia": "contas",
        "tipo": "Saida",
        "data_padrao": "vencimento",
        "excluir_aportes": True,
        "somente_aportes": False,
    },
    "contas-receber": {
        "catalogo_id": "financeiro-contas-receber",
        "titulo": "Contas a Receber",
        "objetivo": "Listar valores a receber, vencidos, a vencer e recebidos.",
        "familia": "contas",
        "tipo": "Entrada",
        "data_padrao": "vencimento",
        "excluir_aportes": True,
        "somente_aportes": False,
    },
    "despesas-categoria": {
        "catalogo_id": "financeiro-despesas-categoria",
        "titulo": "Despesas por Categoria",
        "objetivo": "Mostrar categorias que mais pressionam o resultado gerencial.",
        "familia": "dre_agrupado",
        "tipo": "Saida",
        "data_padrao": "documento",
        "agrupamento": "categoria",
    },
    "despesas-subcategoria": {
        "catalogo_id": "financeiro-despesas-subcategoria",
        "titulo": "Despesas por Subcategoria",
        "objetivo": "Explicar as despesas pela hierarquia categoria, subcategoria e favorecido.",
        "familia": "dre_agrupado",
        "tipo": "Saida",
        "data_padrao": "documento",
        "agrupamento": "subcategoria",
    },
    "receitas": {
        "catalogo_id": "financeiro-receitas",
        "titulo": "Receitas",
        "objetivo": "Consolidar receitas oficiais sem misturar aportes ou transferencias.",
        "familia": "receitas",
        "tipo": "Entrada",
        "data_padrao": "documento",
    },
    "aportes": {
        "catalogo_id": "financeiro-aportes",
        "titulo": "Aportes",
        "objetivo": "Acompanhar aportes que entram no caixa e permanecem fora da DRE.",
        "familia": "caixa",
        "tipo": "Entrada",
        "data_padrao": "vencimento",
        "excluir_aportes": False,
        "somente_aportes": True,
    },
    "evolucao-financeira": {
        "catalogo_id": "financeiro-evolucao-financeira",
        "titulo": "Evolucao Financeira",
        "objetivo": "Comparar a evolucao mensal de DRE, caixa e aportes destacados.",
        "familia": "evolucao",
        "tipo": "Todos",
        "data_padrao": "documento",
    },
    "competencia-realizacao": {
        "catalogo_id": "financeiro-competencia-realizacao",
        "titulo": "Competencia x Realizacao",
        "objetivo": "Comparar quando o evento nasceu, venceu e virou caixa sem duplicar eventos.",
        "familia": "competencia_realizacao",
        "tipo": "Todos",
        "data_padrao": "documento",
    },
}


REFERENCIAS_DATA = {
    "documento": "COALESCE(NULLIF(data_documento, ''), data_vencimento)",
    "vencimento": "data_vencimento",
    "realizacao": "data_realizacao",
}


def hoje_iso():
    return date.today().isoformat()


def primeiro_dia_mes():
    agora = date.today()
    return agora.replace(day=1).isoformat()


def valor_float(valor):
    try:
        return round(float(valor or 0), 2)
    except (TypeError, ValueError):
        return 0


def normalizar_tipo_sql(tipo):
    if tipo == "Entrada":
        return "tipo = ?"
    if tipo == "Saida":
        return "COALESCE(tipo, '') <> ?"
    return ""


def eh_aporte_sql():
    return "(COALESCE(categoria_plano, categoria, '') = ? OR COALESCE(categoria, '') = ?)"


def normalizar_filtros(args, config):
    data_inicio = args.get("data_inicio") or primeiro_dia_mes()
    data_fim = args.get("data_fim") or hoje_iso()
    referencia_data = args.get("referencia_data") or config.get("data_padrao", "documento")
    if referencia_data not in REFERENCIAS_DATA:
        referencia_data = config.get("data_padrao", "documento")
    natureza = args.get("natureza") or "Todas"
    if natureza not in NATUREZAS_FILTRO:
        natureza = "Todas"

    return {
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "referencia_data": referencia_data,
        "natureza": natureza,
        "categoria": args.get("categoria") or "Todas",
        "subcategoria": args.get("subcategoria") or "Todas",
        "favorecido": args.get("favorecido") or "",
        "origem": args.get("origem") or "Todas",
        "situacao": args.get("situacao") or "Todas",
        "termo": args.get("termo") or "",
        "por_pagina": 100,
    }


def montar_condicoes_base(config, filtros, aplicar_periodo=True):
    condicoes = ["COALESCE(status, 'Pendente') <> ?"]
    parametros = ["Cancelado"]
    campo_data = REFERENCIAS_DATA[filtros["referencia_data"]]

    if aplicar_periodo:
        condicoes.append(f"{campo_data} BETWEEN ? AND ?")
        parametros.extend([filtros["data_inicio"], filtros["data_fim"]])

    familia = config["familia"]
    tipo = config.get("tipo")
    cond_tipo = normalizar_tipo_sql(tipo)
    if cond_tipo:
        condicoes.append(cond_tipo)
        parametros.append("Entrada")
    elif filtros.get("natureza") and filtros["natureza"] != "Todas":
        condicoes.append(normalizar_tipo_sql(filtros["natureza"]))
        parametros.append("Entrada")

    if familia in ["caixa", "contas"]:
        condicoes.append("COALESCE(impacta_fluxo_caixa, 1) = 1")

    if familia in ["dre_agrupado", "receitas"]:
        condicoes.append("COALESCE(linha_dre, '') <> ?")
        condicoes.append("COALESCE(linha_dre, '') <> ?")
        parametros.extend(["", LINHA_NEUTRA])

    if familia == "dre_agrupado":
        condicoes.append("COALESCE(linha_dre, '') IN (?, ?)")
        parametros.extend([LINHA_DESPESAS_OPERACIONAIS, LINHA_RESULTADO_NAO_OPERACIONAL])

    if familia == "receitas":
        condicoes.append("COALESCE(linha_dre, '') IN (?, ?, ?)")
        parametros.extend([LINHA_RECEITA_BRUTA, LINHA_DEDUCOES_RECEITA, LINHA_RESULTADO_NAO_OPERACIONAL])

    if config.get("somente_aportes"):
        condicoes.append(eh_aporte_sql())
        parametros.extend(["Aportes", "Aportes"])
    elif config.get("excluir_aportes"):
        condicoes.append(f"NOT {eh_aporte_sql()}")
        parametros.extend(["Aportes", "Aportes"])

    if filtros["categoria"] != "Todas":
        condicoes.append("COALESCE(NULLIF(categoria_plano, ''), categoria, '') = ?")
        parametros.append(filtros["categoria"])
    if filtros["subcategoria"] != "Todas":
        condicoes.append("COALESCE(subcategoria, '') = ?")
        parametros.append(filtros["subcategoria"])
    if filtros["origem"] != "Todas":
        condicoes.append("COALESCE(origem_importacao, '') = ?")
        parametros.append(filtros["origem"])
    if filtros["favorecido"]:
        condicoes.append("(COALESCE(favorecido, '') LIKE ? OR COALESCE(parceiro, '') LIKE ?)")
        termo = f"%{filtros['favorecido']}%"
        parametros.extend([termo, termo])
    if filtros["termo"]:
        condicoes.append("""(
            COALESCE(descricao, '') LIKE ?
            OR COALESCE(historico, '') LIKE ?
            OR COALESCE(numero_documento, '') LIKE ?
            OR COALESCE(documento_id, '') LIKE ?
        )""")
        termo = f"%{filtros['termo']}%"
        parametros.extend([termo, termo, termo, termo])

    return condicoes, parametros, campo_data


def executar_lista(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), tuple(parametros))
    registros = [dict(item) for item in cursor.fetchall()]
    conn.close()
    return registros


def executar_um(sql, parametros=()):
    registros = executar_lista(sql, parametros)
    return registros[0] if registros else {}


def sinal_item(item):
    return 1 if item.get("tipo") == "Entrada" else -1


def valor_evento(item):
    return valor_float(item.get("valor") or item.get("valor_documento"))


def valor_baixado(item):
    pago = valor_float(item.get("valor_pago"))
    if pago > 0:
        return pago
    if item.get("data_realizacao") and item.get("status") in ["Pago", "Recebido", "Realizado"]:
        return valor_evento(item)
    return 0


def status_realizado(item):
    return bool(item.get("data_realizacao")) and item.get("status") in STATUS_REALIZADOS


def competencia_origem(data_documento):
    data = (data_documento or "")[:10]
    if len(data) >= 7:
        return data[:7]
    return COMPETENCIA_SEM_DATA_DOCUMENTO


def documento_ou_historico(item):
    return (
        item.get("numero_documento")
        or item.get("documento_id")
        or item.get("historico")
        or item.get("descricao")
        or "Nao informado"
    )


def analise_pagamentos_ativa(filtros):
    return filtros.get("referencia_data") == "realizacao" and filtros.get("natureza") == "Saida"


def percentual(valor, total):
    return round((float(valor or 0) / total * 100) if total else 0, 2)


def dias_entre(data_a, data_b):
    try:
        a = datetime.strptime((data_a or "")[:10], "%Y-%m-%d").date()
        b = datetime.strptime((data_b or "")[:10], "%Y-%m-%d").date()
        return (b - a).days
    except ValueError:
        return None


def status_conta(item, data_referencia=None):
    ref = datetime.strptime(data_referencia or hoje_iso(), "%Y-%m-%d").date()
    data_venc = item.get("data_vencimento") or ""
    valor = valor_evento(item)
    baixado = valor_baixado(item)
    saldo = max(valor - baixado, 0)
    if saldo <= 0 and status_realizado(item):
        return "Realizado", 0, 0
    if not data_venc:
        return "Sem vencimento", saldo, 0
    venc = datetime.strptime(data_venc[:10], "%Y-%m-%d").date()
    if venc < ref:
        return "Vencido", saldo, (ref - venc).days
    return "A vencer", saldo, 0


def faixa_aging(dias, situacao):
    if situacao == "A vencer":
        return "A vencer"
    if situacao == "Sem vencimento":
        return "Sem vencimento"
    if dias <= 7:
        return "Vencido 1 a 7 dias"
    if dias <= 15:
        return "Vencido 8 a 15 dias"
    if dias <= 30:
        return "Vencido 16 a 30 dias"
    if dias <= 60:
        return "Vencido 31 a 60 dias"
    return "Vencido acima de 60 dias"


def buscar_opcoes_filtro():
    categorias = executar_lista("""
        SELECT DISTINCT COALESCE(NULLIF(categoria_plano, ''), categoria, '') as valor
        FROM movimentacoes_financeiras
        WHERE COALESCE(NULLIF(categoria_plano, ''), categoria, '') <> ''
        ORDER BY valor
    """)
    subcategorias = executar_lista("""
        SELECT DISTINCT COALESCE(subcategoria, '') as valor
        FROM movimentacoes_financeiras
        WHERE COALESCE(subcategoria, '') <> ''
        ORDER BY valor
    """)
    origens = executar_lista("""
        SELECT DISTINCT COALESCE(origem_importacao, '') as valor
        FROM movimentacoes_financeiras
        WHERE COALESCE(origem_importacao, '') <> ''
        ORDER BY valor
    """)
    return {
        "categorias": [item["valor"] for item in categorias],
        "subcategorias": [item["valor"] for item in subcategorias],
        "origens": [item["valor"] for item in origens],
    }


def buscar_detalhes(config, filtros, limite=100):
    condicoes, parametros, campo_data = montar_condicoes_base(config, filtros)
    sql = f"""
        SELECT
            id, data_documento, data_vencimento, data_realizacao, tipo,
            COALESCE(NULLIF(categoria_plano, ''), categoria, '') as categoria,
            COALESCE(subcategoria, '') as subcategoria,
            COALESCE(favorecido, parceiro, '') as favorecido,
            COALESCE(origem_importacao, '') as origem_importacao,
            COALESCE(status, '') as status,
            COALESCE(linha_dre, '') as linha_dre,
            COALESCE(tipo_conta, '') as tipo_conta,
            descricao, historico, numero_documento,
            valor, valor_documento, valor_pago, impacta_fluxo_caixa,
            {campo_data} as data_referencia
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(condicoes)}
        ORDER BY data_referencia ASC, id ASC
        LIMIT ?
    """
    return executar_lista(sql, parametros + [limite])


def aplicar_situacao_contas(itens, filtros):
    resultado = []
    for item in itens:
        situacao, saldo, dias = status_conta(item, filtros["data_fim"])
        item["situacao"] = situacao
        item["saldo_em_aberto"] = round(saldo, 2)
        item["valor_baixado"] = valor_baixado(item)
        item["dias_atraso"] = dias
        item["aging"] = faixa_aging(dias, situacao)
        if filtros["situacao"] != "Todas" and situacao != filtros["situacao"]:
            continue
        resultado.append(item)
    return resultado


def montar_resumo_caixa(config, filtros):
    condicoes, parametros, _ = montar_condicoes_base(config, filtros, aplicar_periodo=False)
    cond_prev = condicoes + ["data_vencimento BETWEEN ? AND ?"]
    par_prev = parametros + [filtros["data_inicio"], filtros["data_fim"]]
    cond_real = condicoes + [
        "data_realizacao BETWEEN ? AND ?",
        "COALESCE(data_realizacao, '') <> ?",
        "COALESCE(status, '') IN (?, ?, ?)",
    ]
    par_real = parametros + [filtros["data_inicio"], filtros["data_fim"], "", "Pago", "Recebido", "Realizado"]

    previsto = executar_um(f"""
        SELECT COUNT(*) as quantidade, COALESCE(SUM(valor), 0) as total
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(cond_prev)}
    """, par_prev)
    realizado = executar_um(f"""
        SELECT COUNT(*) as quantidade, COALESCE(SUM(valor), 0) as total
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(cond_real)}
    """, par_real)
    detalhes = buscar_detalhes(config, filtros)
    aberto = 0
    vencido = 0
    for item in detalhes:
        if status_realizado(item):
            continue
        valor = valor_evento(item)
        aberto += valor
        situacao, _, _ = status_conta(item, filtros["data_fim"])
        if situacao == "Vencido":
            vencido += valor

    return [
        {"rotulo": "Total previsto", "valor": valor_float(previsto.get("total")), "tipo": "moeda"},
        {"rotulo": "Total realizado", "valor": valor_float(realizado.get("total")), "tipo": "moeda"},
        {"rotulo": "Total em aberto", "valor": round(aberto, 2), "tipo": "moeda"},
        {"rotulo": "Total vencido", "valor": round(vencido, 2), "tipo": "moeda"},
        {"rotulo": "Eventos", "valor": int(previsto.get("quantidade") or 0), "tipo": "numero"},
    ]


def montar_resumo_gerencial_financeiro(slug, args):
    config = RELATORIOS_FINANCEIROS[slug]
    filtros = normalizar_filtros(args, config)
    if config["familia"] != "caixa":
        return montar_contexto_relatorio_financeiro(slug, args)

    condicoes, parametros, _ = montar_condicoes_base(config, filtros, aplicar_periodo=False)
    condicoes.append("data_vencimento BETWEEN ? AND ?")
    parametros.extend([filtros["data_inicio"], filtros["data_fim"]])
    linha = executar_um(f"""
        SELECT COUNT(*) AS quantidade, COALESCE(SUM(valor), 0) AS total
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(condicoes)}
    """, parametros)
    resumo = [
        {"rotulo": "Total previsto", "valor": valor_float(linha.get("total")), "tipo": "moeda"},
        {"rotulo": "Eventos", "valor": int(linha.get("quantidade") or 0), "tipo": "numero"},
    ]
    return {"resumo": resumo, "tem_dados": bool(linha.get("quantidade"))}


def montar_resumo_contas(config, filtros):
    itens = aplicar_situacao_contas(buscar_detalhes(config, filtros, limite=5000), filtros)
    totais = defaultdict(float)
    qtd = defaultdict(int)
    for item in itens:
        situacao = item["situacao"]
        saldo = float(item["saldo_em_aberto"] or 0)
        totais[situacao] += saldo
        qtd[situacao] += 1
    return [
        {"rotulo": "Total em aberto", "valor": round(sum(totais.values()), 2), "tipo": "moeda"},
        {"rotulo": "Vencido", "valor": round(totais["Vencido"], 2), "tipo": "moeda"},
        {"rotulo": "A vencer", "valor": round(totais["A vencer"], 2), "tipo": "moeda"},
        {"rotulo": "Titulos", "valor": len(itens), "tipo": "numero"},
        {"rotulo": "Titulos vencidos", "valor": qtd["Vencido"], "tipo": "numero"},
    ]


def montar_agrupamento(config, filtros):
    condicoes, parametros, campo_data = montar_condicoes_base(config, filtros)
    if config["familia"] == "dre_agrupado" and config.get("agrupamento") == "subcategoria":
        chave = "COALESCE(NULLIF(categoria_plano, ''), categoria, '')"
        chave2 = "COALESCE(NULLIF(subcategoria, ''), 'Sem subcategoria')"
        group = f"{chave}, {chave2}"
        select = f"{chave} as categoria, {chave2} as subcategoria"
    else:
        chave = "COALESCE(NULLIF(categoria_plano, ''), categoria, '')"
        group = chave
        select = f"{chave} as categoria, '' as subcategoria"
    linhas = executar_lista(f"""
        SELECT {select}, COUNT(*) as quantidade, COALESCE(SUM(valor), 0) as total
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(condicoes)}
        GROUP BY {group}
        ORDER BY total DESC
        LIMIT 80
    """, parametros)
    total = sum(float(item["total"] or 0) for item in linhas)
    for item in linhas:
        item["percentual"] = round((float(item["total"] or 0) / total * 100) if total else 0, 2)
    evolucao = executar_lista(f"""
        SELECT SUBSTR({campo_data}, 1, 7) as mes, COALESCE(SUM(valor), 0) as total, COUNT(*) as quantidade
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(condicoes)}
        GROUP BY SUBSTR({campo_data}, 1, 7)
        ORDER BY mes
    """, parametros)
    return linhas, evolucao


def montar_receitas(config, filtros):
    linhas, evolucao = montar_agrupamento(config, filtros)
    receita_bruta = sum(float(item["total"] or 0) for item in linhas if item["categoria"])
    deducoes = sum(float(item["total"] or 0) for item in linhas if "dedu" in item["categoria"].lower())
    return linhas, evolucao, [
        {"rotulo": "Receitas totais", "valor": round(receita_bruta, 2), "tipo": "moeda"},
        {"rotulo": "Deducoes", "valor": round(deducoes, 2), "tipo": "moeda"},
        {"rotulo": "Receita liquida", "valor": round(receita_bruta - deducoes, 2), "tipo": "moeda"},
        {"rotulo": "Grupos", "valor": len(linhas), "tipo": "numero"},
    ]


def montar_evolucao(config, filtros):
    condicoes_fluxo, parametros_fluxo, campo_data = montar_condicoes_base(
        {"familia": "caixa", "tipo": "Todos"},
        filtros,
    )
    linhas = executar_lista(f"""
        SELECT
            SUBSTR({campo_data}, 1, 7) as mes,
            COALESCE(SUM(CASE WHEN tipo = 'Entrada' THEN valor ELSE 0 END), 0) as entradas,
            COALESCE(SUM(CASE WHEN tipo <> 'Entrada' THEN valor ELSE 0 END), 0) as saidas,
            COALESCE(SUM(CASE WHEN {eh_aporte_sql()} THEN valor ELSE 0 END), 0) as aportes
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(condicoes_fluxo)}
        GROUP BY SUBSTR({campo_data}, 1, 7)
        ORDER BY mes
    """, ["Aportes", "Aportes"] + parametros_fluxo)
    for item in linhas:
        item["saldo"] = round(float(item["entradas"] or 0) - float(item["saidas"] or 0), 2)
    resumo = [
        {"rotulo": "Entradas", "valor": round(sum(float(i["entradas"] or 0) for i in linhas), 2), "tipo": "moeda"},
        {"rotulo": "Saidas", "valor": round(sum(float(i["saidas"] or 0) for i in linhas), 2), "tipo": "moeda"},
        {"rotulo": "Saldo do periodo", "valor": round(sum(float(i["saldo"] or 0) for i in linhas), 2), "tipo": "moeda"},
        {"rotulo": "Aportes destacados", "valor": round(sum(float(i["aportes"] or 0) for i in linhas), 2), "tipo": "moeda"},
    ]
    return linhas, resumo


def montar_analise_pagamentos_origem(itens, filtros):
    ativo = analise_pagamentos_ativa(filtros)
    analise = {
        "ativa": ativo,
        "orientacao": (
            "A analise de origem dos pagamentos exige Referencia = Realizacao / baixa "
            "e Natureza = Saida."
        ),
        "resumo": {
            "total_pago": 0,
            "quantidade_pagamentos": 0,
            "ticket_medio": 0,
            "maior_pagamento": 0,
            "principal_competencia": "",
            "principal_percentual": 0,
            "eventos_sem_data_documento": 0,
            "valor_top5": 0,
            "percentual_top5": 0,
            "reconciliacao_ok": True,
        },
        "top5": [],
        "origem_competencia": [],
        "dados_detalhados": [],
    }
    if not ativo:
        return analise

    pagamentos = []
    for item in itens:
        if not status_realizado(item):
            continue
        if item.get("tipo") == "Entrada":
            continue
        if int(item.get("impacta_fluxo_caixa") or 0) != 1:
            continue
        if (item.get("categoria") or "") == "Aportes":
            continue
        valor_realizado = valor_baixado(item)
        if valor_realizado <= 0:
            continue
        item["valor_realizado"] = valor_realizado
        item["competencia_origem"] = competencia_origem(item.get("data_documento"))
        item["documento_ou_historico"] = documento_ou_historico(item)
        pagamentos.append(item)

    pagamentos.sort(key=lambda item: (-valor_float(item.get("valor_realizado")), item.get("data_realizacao") or "", item.get("id") or 0))
    total_pago = round(sum(valor_float(item.get("valor_realizado")) for item in pagamentos), 2)
    quantidade = len(pagamentos)
    top5 = pagamentos[:5]
    valor_top5 = round(sum(valor_float(item.get("valor_realizado")) for item in top5), 2)

    por_competencia = defaultdict(lambda: {
        "competencia": "",
        "quantidade": 0,
        "valor": 0,
        "percentual": 0,
        "valor_medio": 0,
        "maior_pagamento": 0,
        "principal": False,
    })
    for item in pagamentos:
        chave = item["competencia_origem"]
        linha = por_competencia[chave]
        linha["competencia"] = chave
        linha["quantidade"] += 1
        linha["valor"] = round(linha["valor"] + valor_float(item.get("valor_realizado")), 2)
        linha["maior_pagamento"] = max(linha["maior_pagamento"], valor_float(item.get("valor_realizado")))

    origem_competencia = list(por_competencia.values())
    origem_competencia.sort(key=lambda item: (item["competencia"] == COMPETENCIA_SEM_DATA_DOCUMENTO, item["competencia"]))
    principal = None
    for linha in origem_competencia:
        linha["percentual"] = percentual(linha["valor"], total_pago)
        linha["valor_medio"] = round((linha["valor"] / linha["quantidade"]) if linha["quantidade"] else 0, 2)
        if principal is None or linha["valor"] > principal["valor"]:
            principal = linha
    if principal:
        principal["principal"] = True

    soma_competencias = round(sum(linha["valor"] for linha in origem_competencia), 2)
    analise["resumo"] = {
        "total_pago": total_pago,
        "quantidade_pagamentos": quantidade,
        "ticket_medio": round((total_pago / quantidade) if quantidade else 0, 2),
        "maior_pagamento": valor_float(top5[0].get("valor_realizado")) if top5 else 0,
        "principal_competencia": principal["competencia"] if principal else "",
        "principal_percentual": principal["percentual"] if principal else 0,
        "eventos_sem_data_documento": sum(1 for item in pagamentos if item["competencia_origem"] == COMPETENCIA_SEM_DATA_DOCUMENTO),
        "valor_top5": valor_top5,
        "percentual_top5": percentual(valor_top5, total_pago),
        "reconciliacao_ok": abs(total_pago - soma_competencias) < 0.01,
    }
    analise["top5"] = top5
    analise["origem_competencia"] = origem_competencia
    analise["dados_detalhados"] = pagamentos
    return analise


def montar_competencia_realizacao(config, filtros):
    condicoes, parametros, campo_data = montar_condicoes_base({"familia": "caixa", "tipo": "Todos"}, filtros)
    itens = executar_lista(f"""
        SELECT id, documento_id, data_documento, data_vencimento, data_realizacao, tipo,
               COALESCE(NULLIF(categoria_plano, ''), categoria, '') as categoria,
               COALESCE(subcategoria, '') as subcategoria,
               COALESCE(favorecido, parceiro, '') as favorecido,
               COALESCE(origem_importacao, '') as origem_importacao,
               valor, valor_documento, valor_pago, valor_liquido,
               status, linha_dre, tipo_conta, COALESCE(impacta_fluxo_caixa, 1) as impacta_fluxo_caixa,
               descricao, historico, numero_documento,
               {campo_data} as data_referencia
        FROM movimentacoes_financeiras
        WHERE {" AND ".join(condicoes)}
        ORDER BY data_referencia ASC, id ASC
    """, parametros)
    total_competencia = 0
    total_realizado = 0
    prazos_doc_baixa = []
    prazos_venc_baixa = []
    sem_baixa = 0
    for item in itens:
        valor = valor_evento(item) * sinal_item(item)
        realizado = valor_baixado(item)
        total_competencia += valor
        if status_realizado(item):
            total_realizado += realizado * sinal_item(item)
            prazo = dias_entre(item.get("data_documento") or item.get("data_vencimento"), item.get("data_realizacao"))
            prazo_venc = dias_entre(item.get("data_vencimento"), item.get("data_realizacao"))
            if prazo is not None:
                prazos_doc_baixa.append(prazo)
            if prazo_venc is not None:
                prazos_venc_baixa.append(prazo_venc)
        else:
            sem_baixa += 1
        item["valor_assinado"] = round(valor, 2)
        item["valor_realizado"] = realizado
        item["valor_realizado_assinado"] = round(realizado * sinal_item(item), 2)
        item["competencia_origem"] = competencia_origem(item.get("data_documento"))
        item["documento_ou_historico"] = documento_ou_historico(item)
        item["prazo_documento_baixa"] = dias_entre(item.get("data_documento") or item.get("data_vencimento"), item.get("data_realizacao"))
        item["prazo_vencimento_baixa"] = dias_entre(item.get("data_vencimento"), item.get("data_realizacao"))
    media_doc = sum(prazos_doc_baixa) / len(prazos_doc_baixa) if prazos_doc_baixa else 0
    media_venc = sum(prazos_venc_baixa) / len(prazos_venc_baixa) if prazos_venc_baixa else 0
    resumo = [
        {"rotulo": "Valor por competencia", "valor": round(total_competencia, 2), "tipo": "moeda"},
        {"rotulo": "Valor realizado", "valor": round(total_realizado, 2), "tipo": "moeda"},
        {"rotulo": "Diferenca em aberto", "valor": round(total_competencia - total_realizado, 2), "tipo": "moeda"},
        {"rotulo": "Eventos sem baixa", "valor": sem_baixa, "tipo": "numero"},
        {"rotulo": "Prazo medio doc. baixa", "valor": round(media_doc, 1), "tipo": "numero"},
        {"rotulo": "Prazo medio venc. baixa", "valor": round(media_venc, 1), "tipo": "numero"},
    ]
    return itens, resumo, montar_analise_pagamentos_origem(itens, filtros)


def montar_contexto_relatorio_financeiro(slug, args):
    config = RELATORIOS_FINANCEIROS[slug]
    filtros = normalizar_filtros(args, config)
    detalhes = buscar_detalhes(config, filtros)
    agrupamentos = []
    evolucao = []
    resumo = []
    analise_pagamentos = {
        "ativa": False,
        "orientacao": "",
        "resumo": {},
        "top5": [],
        "origem_competencia": [],
        "dados_detalhados": [],
    }

    if config["familia"] == "caixa":
        resumo = montar_resumo_caixa(config, filtros)
        if slug == "aportes":
            detalhes = [item for item in detalhes if (item.get("categoria") == "Aportes")]
        agrupamentos, evolucao = montar_agrupamento(config, filtros)
    elif config["familia"] == "contas":
        detalhes = aplicar_situacao_contas(buscar_detalhes(config, filtros, limite=500), filtros)
        resumo = montar_resumo_contas(config, filtros)
        por_aging = defaultdict(lambda: {"categoria": "", "subcategoria": "", "quantidade": 0, "total": 0, "percentual": 0})
        for item in detalhes:
            agg = por_aging[item["aging"]]
            agg["categoria"] = item["aging"]
            agg["quantidade"] += 1
            agg["total"] += float(item["saldo_em_aberto"] or 0)
        total_aging = sum(item["total"] for item in por_aging.values())
        agrupamentos = []
        for item in por_aging.values():
            item["total"] = round(item["total"], 2)
            item["percentual"] = round((item["total"] / total_aging * 100) if total_aging else 0, 2)
            agrupamentos.append(item)
    elif config["familia"] == "dre_agrupado":
        agrupamentos, evolucao = montar_agrupamento(config, filtros)
        resumo = [
            {"rotulo": "Total", "valor": round(sum(float(i["total"] or 0) for i in agrupamentos), 2), "tipo": "moeda"},
            {"rotulo": "Grupos", "valor": len(agrupamentos), "tipo": "numero"},
            {"rotulo": "Eventos detalhados", "valor": len(detalhes), "tipo": "numero"},
        ]
    elif config["familia"] == "receitas":
        agrupamentos, evolucao, resumo = montar_receitas(config, filtros)
    elif config["familia"] == "evolucao":
        evolucao, resumo = montar_evolucao(config, filtros)
        detalhes = []
    elif config["familia"] == "competencia_realizacao":
        detalhes, resumo, analise_pagamentos = montar_competencia_realizacao(config, filtros)

    return {
        "slug": slug,
        "config": config,
        "filtros": filtros,
        "opcoes": buscar_opcoes_filtro(),
        "resumo": resumo,
        "agrupamentos": agrupamentos,
        "evolucao": evolucao,
        "detalhes": detalhes,
        "analise_pagamentos": analise_pagamentos,
        "query_string": urlencode({k: v for k, v in filtros.items() if v not in ["", "Todas"]}),
        "referencias_data": [
            ("documento", "Documento / competencia"),
            ("vencimento", "Vencimento"),
            ("realizacao", "Realizacao / baixa"),
        ],
        "naturezas": [
            ("Todas", "Todas"),
            ("Entrada", "Entrada"),
            ("Saida", "Saida"),
        ],
        "situacoes": ["Todas", "A vencer", "Vencido", "Realizado", "Sem vencimento"],
    }


def ajustar_planilha(ws):
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for coluna in range(1, ws.max_column + 1):
        largura = 14
        for linha in range(1, min(ws.max_row, 60) + 1):
            valor = ws.cell(linha, coluna).value
            if valor is not None:
                largura = max(largura, min(len(str(valor)) + 2, 42))
        ws.column_dimensions[get_column_letter(coluna)].width = largura
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3B4D")


def formatar_abas_analise_pagamentos(wb):
    for ws in wb.worksheets:
        ajustar_planilha(ws)
        for row in ws.iter_rows():
            for cell in row:
                cabecalho = str(ws.cell(1, cell.column).value or "").lower()
                if "valor" in cabecalho or "ticket" in cabecalho or "pagamento" in cabecalho or "total" in cabecalho:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                if "percentual" in cabecalho or "participacao" in cabecalho:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'


def gerar_excel_competencia_pagamentos(contexto):
    wb = Workbook()
    analise = contexto["analise_pagamentos"]
    resumo = analise["resumo"]

    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Indicador", "Valor"])
    linhas_resumo = [
        ("Periodo", f"{contexto['filtros']['data_inicio']} a {contexto['filtros']['data_fim']}"),
        ("Referencia", contexto["filtros"]["referencia_data"]),
        ("Natureza", contexto["filtros"]["natureza"]),
        ("Total pago", resumo.get("total_pago", 0)),
        ("Quantidade", resumo.get("quantidade_pagamentos", 0)),
        ("Ticket medio", resumo.get("ticket_medio", 0)),
        ("Maior pagamento", resumo.get("maior_pagamento", 0)),
        ("Principal competencia", resumo.get("principal_competencia") or "Nao informado"),
        ("Participacao principal competencia", (resumo.get("principal_percentual", 0) or 0) / 100),
        ("Valor Top 5", resumo.get("valor_top5", 0)),
        ("Participacao Top 5", (resumo.get("percentual_top5", 0) or 0) / 100),
        ("Eventos sem Data do Documento", resumo.get("eventos_sem_data_documento", 0)),
        ("Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Base oficial", "Central de Movimentacoes"),
        ("Reconciliacao", "OK" if resumo.get("reconciliacao_ok") else "Divergente"),
    ]
    for linha in linhas_resumo:
        ws.append(list(linha))
    for row in range(2, ws.max_row + 1):
        indicador = str(ws.cell(row, 1).value or "").lower()
        if "participacao" in indicador:
            ws.cell(row, 2).number_format = "0.00%"

    ws = wb.create_sheet("Top 5 Pagamentos")
    ws.append([
        "Posicao", "Data Realizacao", "Data Documento", "Competencia Origem",
        "Favorecido", "Categoria", "Subcategoria", "Documento ou Historico",
        "Origem", "Valor Realizado",
    ])
    for posicao, item in enumerate(analise["top5"], start=1):
        ws.append([
            posicao,
            item.get("data_realizacao") or "Nao informado",
            item.get("data_documento") or "Nao informado",
            item.get("competencia_origem") or "Nao informado",
            item.get("favorecido") or "Nao informado",
            item.get("categoria") or "Nao informado",
            item.get("subcategoria") or "Nao informado",
            item.get("documento_ou_historico") or "Nao informado",
            item.get("origem_importacao") or "Nao informado",
            valor_float(item.get("valor_realizado")),
        ])

    ws = wb.create_sheet("Origem por Competencia")
    ws.append(["Competencia", "Quantidade", "Valor", "Percentual", "Valor medio", "Maior pagamento"])
    for item in analise["origem_competencia"]:
        ws.append([
            item.get("competencia") or "Nao informado",
            item.get("quantidade", 0),
            item.get("valor", 0),
            (item.get("percentual", 0) or 0) / 100,
            item.get("valor_medio", 0),
            item.get("maior_pagamento", 0),
        ])

    ws = wb.create_sheet("Dados Detalhados")
    ws.append([
        "Data Realizacao", "Data Documento", "Data Vencimento", "Competencia Origem",
        "Favorecido", "Categoria", "Subcategoria", "Documento ou Historico",
        "Origem", "Status", "Valor Documento", "Valor Realizado",
        "Impacta Fluxo Caixa", "Linha DRE",
    ])
    for item in analise["dados_detalhados"]:
        ws.append([
            item.get("data_realizacao") or "Nao informado",
            item.get("data_documento") or "Nao informado",
            item.get("data_vencimento") or "Nao informado",
            item.get("competencia_origem") or "Nao informado",
            item.get("favorecido") or "Nao informado",
            item.get("categoria") or "Nao informado",
            item.get("subcategoria") or "Nao informado",
            item.get("documento_ou_historico") or "Nao informado",
            item.get("origem_importacao") or "Nao informado",
            item.get("status") or "Nao informado",
            valor_evento(item),
            valor_float(item.get("valor_realizado")),
            item.get("impacta_fluxo_caixa"),
            item.get("linha_dre") or "",
        ])

    ws = wb.create_sheet("Parametros")
    ws.append(["Campo", "Valor"])
    for chave, valor in contexto["filtros"].items():
        ws.append([chave, valor])
    ws.append(["Definicao", "Pagamento realizado = Saida com data_realizacao no periodo, status realizado e impacta_fluxo_caixa = 1."])
    ws.append(["Valor", "Usa valor_pago quando preenchido; senao usa valor do evento somente para registros realizados."])
    ws.append(["Competencia de origem", "Mes da data_documento; ausente permanece como Sem data do documento."])
    ws.append(["Transferencias", "Eventos neutros do fluxo ficam fora por impacta_fluxo_caixa = 0."])
    ws.append(["Top 5", "Cinco eventos individuais por maior valor realizado, sem agrupamento."])
    ws.append(["Reconciliacao", "Total pago = soma por competencia = soma dos dados detalhados."])

    formatar_abas_analise_pagamentos(wb)
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo


def gerar_excel_relatorio_financeiro(contexto):
    if contexto.get("slug") == "competencia-realizacao" and contexto.get("analise_pagamentos", {}).get("ativa"):
        return gerar_excel_competencia_pagamentos(contexto)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    ws.append([contexto["config"]["titulo"]])
    ws.append([contexto["config"]["objetivo"]])
    ws.append([])
    ws.append(["Filtros"])
    for chave, valor in contexto["filtros"].items():
        ws.append([chave, valor])
    ws.append([])
    ws.append(["Indicador", "Valor"])
    for item in contexto["resumo"]:
        ws.append([item["rotulo"], item["valor"]])
    ws.append([])
    if contexto["agrupamentos"]:
        ws.append(["Agrupamentos"])
        ws.append(["Categoria", "Subcategoria", "Quantidade", "Total", "Percentual"])
        for item in contexto["agrupamentos"]:
            ws.append([
                item.get("categoria", ""),
                item.get("subcategoria", ""),
                item.get("quantidade", 0),
                item.get("total", 0),
                item.get("percentual", 0),
            ])
        ws.append([])
    if contexto["evolucao"]:
        ws.append(["Evolucao"])
        chaves = list(contexto["evolucao"][0].keys())
        ws.append(chaves)
        for item in contexto["evolucao"]:
            ws.append([item.get(chave, "") for chave in chaves])
        ws.append([])
    ws.append(["Detalhes"])
    colunas = [
        "id", "data_documento", "data_vencimento", "data_realizacao", "tipo",
        "categoria", "subcategoria", "favorecido", "valor", "valor_pago",
        "status", "situacao", "saldo_em_aberto", "linha_dre", "impacta_fluxo_caixa",
        "descricao", "historico",
    ]
    ws.append(colunas)
    for item in contexto["detalhes"]:
        ws.append([item.get(coluna, "") for coluna in colunas])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3B4D")
    for coluna in range(1, 18):
        ws.column_dimensions[get_column_letter(coluna)].width = 20
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo

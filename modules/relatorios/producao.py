"""Servicos compartilhados para relatorios oficiais de producao."""

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database import DATABASE_URL, conectar, q


META_RENDIMENTO = 63.0


RELATORIOS_PRODUCAO = {
    "producao-por-op": {
        "catalogo_id": "producao-op",
        "titulo": "Producao por OP",
        "objetivo": "Avaliar desempenho, perdas e rendimento por ordem de producao.",
        "familia": "op",
        "agrupamento": "op",
    },
    "producao-por-sku": {
        "catalogo_id": "producao-sku",
        "titulo": "Producao por SKU",
        "objetivo": "Comparar volume produzido, rendimento e perdas por SKU oficial.",
        "familia": "producao",
        "agrupamento": "sku",
    },
    "producao-por-fornecedor": {
        "catalogo_id": "producao-fornecedor",
        "titulo": "Producao por Fornecedor",
        "objetivo": "Medir resultado produtivo e qualidade da materia-prima por fornecedor.",
        "familia": "producao",
        "agrupamento": "fornecedor",
    },
    "producao-por-periodo": {
        "catalogo_id": "producao-periodo",
        "titulo": "Producao por Periodo",
        "objetivo": "Acompanhar a evolucao da producao por dia, semana ou mes.",
        "familia": "periodo",
        "agrupamento": "periodo",
    },
    "rendimento": {
        "catalogo_id": "producao-rendimento",
        "titulo": "Rendimento",
        "objetivo": "Preservar o calculo homologado de kg produzido sobre peso vivo.",
        "familia": "rendimento",
        "agrupamento": "periodo_fornecedor",
        "somente_encerradas": True,
    },
    "condenacoes": {
        "catalogo_id": "producao-condenacoes",
        "titulo": "Condenacoes",
        "objetivo": "Analisar causas, fornecedores e OPs com condenacoes oficiais.",
        "familia": "perdas",
        "agrupamento": "causa",
        "tipo_perda": "condenacao",
    },
    "perdas": {
        "catalogo_id": "producao-perdas",
        "titulo": "Perdas",
        "objetivo": "Consolidar perdas oficiais sem sobrepor condenacoes, descartes e mortes.",
        "familia": "perdas",
        "agrupamento": "tipo",
        "tipo_perda": "todas",
    },
    "eficiencia": {
        "catalogo_id": "producao-eficiencia",
        "titulo": "Eficiencia da Producao",
        "objetivo": "Acompanhar produtividade horaria oficial a partir de producao, tempos e paradas registrados.",
        "familia": "eficiencia",
        "agrupamento": "periodo",
    },
}


def hoje_iso():
    return date.today().isoformat()


def primeiro_dia_mes():
    hoje = date.today()
    return hoje.replace(day=1).isoformat()


def valor_float(valor):
    try:
        return round(float(valor or 0), 4)
    except (TypeError, ValueError):
        return 0.0


def valor_int(valor):
    try:
        return int(float(valor or 0))
    except (TypeError, ValueError):
        return 0


def percentual(numerador, denominador):
    denominador = float(denominador or 0)
    if denominador <= 0:
        return 0.0
    return round(float(numerador or 0) / denominador * 100, 2)


def dividir(numerador, denominador):
    denominador = float(denominador or 0)
    if denominador <= 0:
        return 0.0
    return round(float(numerador or 0) / denominador, 4)


def horas_entre(hora_inicio, hora_fim):
    try:
        inicio = datetime.strptime(str(hora_inicio or ""), "%H:%M")
        fim = datetime.strptime(str(hora_fim or ""), "%H:%M")
    except ValueError:
        return 0.0

    diferenca = (fim - inicio).total_seconds() / 3600
    if diferenca < 0:
        diferenca += 24
    return round(diferenca, 4)


def normalizar_filtros(args):
    granularidade = args.get("granularidade") or "mes"
    if granularidade not in ["dia", "semana", "mes"]:
        granularidade = "mes"

    return {
        "data_inicio": args.get("data_inicio") or primeiro_dia_mes(),
        "data_fim": args.get("data_fim") or hoje_iso(),
        "op_id": args.get("op_id") or "",
        "status": args.get("status") or "Todos",
        "sku": args.get("sku") or "Todos",
        "fornecedor": args.get("fornecedor") or "Todos",
        "lote": args.get("lote") or "",
        "causa": args.get("causa") or "Todos",
        "setor": args.get("setor") or "Todos",
        "situacao": args.get("situacao") or "Todas",
        "granularidade": granularidade,
        "por_pagina": 200,
    }


def expressao_periodo(campo, granularidade):
    if granularidade == "dia":
        return campo
    if granularidade == "semana":
        if DATABASE_URL:
            return f"TO_CHAR(TO_DATE({campo}, 'YYYY-MM-DD'), 'IYYY-\"W\"IW')"
        return f"strftime('%Y-W%W', {campo})"
    return f"SUBSTR({campo}, 1, 7)"


def montar_condicoes_ops(filtros, alias="o"):
    condicoes = [f"{alias}.data BETWEEN ? AND ?"]
    parametros = [filtros["data_inicio"], filtros["data_fim"]]

    if filtros["op_id"]:
        try:
            condicoes.append(f"{alias}.id = ?")
            parametros.append(int(filtros["op_id"]))
        except ValueError:
            condicoes.append("1 = 0")

    if filtros["status"] != "Todos":
        condicoes.append(f"COALESCE({alias}.status, 'Aberta') = ?")
        parametros.append(filtros["status"])

    if filtros["sku"] != "Todos":
        condicoes.append(f"COALESCE({alias}.sku, 'Galinha Cortada') = ?")
        parametros.append(filtros["sku"])

    if filtros["fornecedor"] != "Todos":
        condicoes.append(f"{alias}.fornecedor = ?")
        parametros.append(filtros["fornecedor"])

    if filtros["lote"]:
        condicoes.append(f"{alias}.id = ?")
        try:
            parametros.append(int(str(filtros["lote"]).upper().replace("OP-", "").lstrip("0") or 0))
        except ValueError:
            parametros.append(-1)

    if filtros["causa"] == "Mortes antes da pendura":
        condicoes.append(f"COALESCE({alias}.mortes_antes_pendura, 0) > 0")
    elif filtros["causa"] != "Todos":
        condicoes.append(f"""
            EXISTS (
                SELECT 1 FROM apontamentos_descartes df
                WHERE df.op_id = {alias}.id
                  AND COALESCE(df.motivo, '') = ?
            )
        """)
        parametros.append(filtros["causa"])

    if filtros["setor"] != "Todos":
        condicoes.append(f"""
            EXISTS (
                SELECT 1 FROM apontamentos_descartes ds
                WHERE ds.op_id = {alias}.id
                  AND COALESCE(ds.setor, '') = ?
            )
        """)
        parametros.append(filtros["setor"])

    return condicoes, parametros


def cte_ops_agregadas(filtros, somente_encerradas=False):
    condicoes, parametros = montar_condicoes_ops(filtros)
    if somente_encerradas:
        condicoes.append("COALESCE(o.status, 'Aberta') = ?")
        parametros.append("Encerrada")
    where_ops = " AND ".join(condicoes)

    sql = f"""
    WITH op_base AS (
        SELECT
            o.id AS op_id,
            o.data AS data_op,
            o.fornecedor,
            COALESCE(o.sku, 'Galinha Cortada') AS sku,
            COALESCE(o.status, 'Aberta') AS status,
            COALESCE(o.quantidade_aves, 0) AS aves_recebidas,
            COALESCE(o.mortes_antes_pendura, 0) AS mortes_antes_pendura,
            COALESCE(o.peso_vivo, 0) AS peso_vivo,
            COALESCE(o.peso_medio, 0) AS peso_medio
        FROM ordens_producao o
        WHERE {where_ops}
    ),
    prod AS (
        SELECT
            op_id,
            COALESCE(SUM(CASE WHEN LOWER(COALESCE(unidade, '')) = 'kg' THEN quantidade ELSE 0 END), 0) AS kg_produzidos,
            COALESCE(SUM(CASE
                WHEN COALESCE(setor, '') = 'Expedição'
                 AND LOWER(COALESCE(unidade, '')) IN ('unidades', 'unidade', 'aves', 'ave', 'bandejas', 'caixas')
                THEN quantidade ELSE 0 END), 0) AS unidades_finais
        FROM apontamentos_producao
        GROUP BY op_id
    ),
    primaria AS (
        SELECT
            op_id,
            COALESCE(SUM(quantidade_bandejas), 0) AS producao_primaria
        FROM embalagem_primaria_apontamentos
        GROUP BY op_id
    ),
    pa AS (
        SELECT
            comp.op_id,
            COUNT(DISTINCT cx.id) AS caixas,
            COALESCE(SUM(comp.quantidade_bandejas), 0) AS producao_secundaria,
            COALESCE(SUM(cx.peso_liquido), 0) AS peso_liquido_pa,
            COALESCE(SUM(cx.peso_bruto), 0) AS peso_bruto_pa
        FROM pa_caixa_composicao comp
        INNER JOIN pa_caixas cx ON cx.id = comp.caixa_id
        WHERE COALESCE(cx.status, '') <> 'Cancelada'
        GROUP BY comp.op_id
    ),
    perdas AS (
        SELECT
            op_id,
            COALESCE(SUM(CASE
                WHEN LOWER(COALESCE(unidade, '')) IN ('aves', 'ave', 'unidade', 'unidades')
                 AND (LOWER(COALESCE(categoria, '')) LIKE '%%conden%%'
                   OR LOWER(COALESCE(motivo, '')) LIKE '%%conden%%')
                THEN quantidade ELSE 0 END), 0) AS condenacoes_aves,
            COALESCE(SUM(CASE
                WHEN LOWER(COALESCE(unidade, '')) IN ('aves', 'ave', 'unidade', 'unidades')
                 AND LOWER(TRIM(COALESCE(motivo, ''))) = 'morte na gaiola'
                THEN quantidade ELSE 0 END), 0) AS mortes_na_gaiola,
            COALESCE(SUM(CASE
                WHEN LOWER(COALESCE(unidade, '')) IN ('aves', 'ave', 'unidade', 'unidades')
                 AND LOWER(TRIM(COALESCE(motivo, ''))) <> 'morte na gaiola'
                 AND LOWER(COALESCE(categoria, '')) NOT LIKE '%%conden%%'
                 AND LOWER(COALESCE(motivo, '')) NOT LIKE '%%conden%%'
                THEN quantidade ELSE 0 END), 0) AS descartes_aves,
            COALESCE(SUM(CASE
                WHEN LOWER(COALESCE(unidade, '')) = 'kg'
                THEN quantidade ELSE 0 END), 0) AS perdas_kg
        FROM apontamentos_descartes
        GROUP BY op_id
    )
    """
    return sql, parametros


def executar_lista(sql, parametros=None):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), tuple(parametros or []))
    linhas = cursor.fetchall()
    conn.close()
    return [dict(item) for item in linhas]


def buscar_ops_agregadas(filtros, somente_encerradas=False, limite=500):
    cte, parametros = cte_ops_agregadas(filtros, somente_encerradas)
    sql = cte + f"""
    SELECT
        op.op_id, op.data_op, op.fornecedor, op.sku, op.status,
        op.aves_recebidas, op.peso_vivo, op.peso_medio,
        op.mortes_antes_pendura,
        COALESCE(perdas.mortes_na_gaiola, 0) AS mortes_na_gaiola,
        COALESCE(perdas.descartes_aves, 0) AS descartes_aves,
        COALESCE(perdas.condenacoes_aves, 0) AS condenacoes_aves,
        COALESCE(perdas.perdas_kg, 0) AS perdas_kg,
        COALESCE(primaria.producao_primaria, 0) AS producao_primaria,
        COALESCE(pa.producao_secundaria, 0) AS producao_secundaria,
        COALESCE(pa.caixas, 0) AS caixas,
        COALESCE(pa.peso_liquido_pa, 0) AS peso_liquido_pa,
        COALESCE(pa.peso_bruto_pa, 0) AS peso_bruto_pa,
        COALESCE(prod.kg_produzidos, 0) AS kg_produzidos,
        COALESCE(prod.unidades_finais, 0) AS unidades_finais
    FROM op_base op
    LEFT JOIN prod ON prod.op_id = op.op_id
    LEFT JOIN primaria ON primaria.op_id = op.op_id
    LEFT JOIN pa ON pa.op_id = op.op_id
    LEFT JOIN perdas ON perdas.op_id = op.op_id
    ORDER BY op.data_op DESC, op.op_id DESC
    LIMIT {int(limite)}
    """
    linhas = executar_lista(sql, parametros)
    for linha in linhas:
        enriquecer_linha_op(linha)
    return linhas


def enriquecer_linha_op(linha):
    mortes_total = valor_float(linha.get("mortes_antes_pendura")) + valor_float(linha.get("mortes_na_gaiola"))
    perdas_aves = mortes_total + valor_float(linha.get("descartes_aves")) + valor_float(linha.get("condenacoes_aves"))
    aves_recebidas = valor_float(linha.get("aves_recebidas"))
    kg_base = valor_float(linha.get("kg_produzidos"))
    if kg_base <= 0:
        kg_base = valor_float(linha.get("peso_liquido_pa"))
    linha["lote"] = f"OP-{valor_int(linha.get('op_id')):05d}"
    linha["mortes_total"] = round(mortes_total, 2)
    linha["perdas_aves"] = round(perdas_aves, 2)
    linha["aves_viaveis"] = round(max(0, aves_recebidas - perdas_aves), 2)
    linha["peso_produzido"] = round(kg_base, 2)
    linha["rendimento"] = percentual(kg_base, linha.get("peso_vivo"))
    linha["taxa_condenacao"] = percentual(linha.get("condenacoes_aves"), aves_recebidas)
    linha["taxa_perda"] = percentual(perdas_aves, aves_recebidas)
    linha["desvio_meta"] = round(linha["rendimento"] - META_RENDIMENTO, 2)
    return linha


def somar_ops(linhas):
    totais = defaultdict(float)
    for linha in linhas:
        for chave in [
            "aves_recebidas", "peso_vivo", "producao_primaria", "producao_secundaria",
            "caixas", "peso_liquido_pa", "peso_bruto_pa", "kg_produzidos",
            "unidades_finais", "descartes_aves", "condenacoes_aves", "mortes_total",
            "perdas_aves", "perdas_kg", "peso_produzido",
        ]:
            totais[chave] += valor_float(linha.get(chave))
    totais["ops"] = len(linhas)
    totais["rendimento"] = percentual(totais["peso_produzido"], totais["peso_vivo"])
    totais["taxa_condenacao"] = percentual(totais["condenacoes_aves"], totais["aves_recebidas"])
    totais["taxa_perda"] = percentual(totais["perdas_aves"], totais["aves_recebidas"])
    return dict(totais)


def resumo_kpis(totais):
    return [
        {"rotulo": "OPs", "valor": valor_int(totais.get("ops")), "tipo": "numero", "unidade": "OPs"},
        {"rotulo": "Aves recebidas", "valor": round(totais.get("aves_recebidas", 0), 2), "tipo": "numero", "unidade": "aves"},
        {"rotulo": "Peso entrada", "valor": round(totais.get("peso_vivo", 0), 2), "tipo": "decimal", "unidade": "kg"},
        {"rotulo": "PI produzido", "valor": round(totais.get("producao_primaria", 0), 2), "tipo": "numero", "unidade": "unid."},
        {"rotulo": "Caixas PA", "valor": round(totais.get("caixas", 0), 2), "tipo": "numero", "unidade": "caixas"},
        {"rotulo": "Peso produzido", "valor": round(totais.get("peso_produzido", 0), 2), "tipo": "decimal", "unidade": "kg"},
        {"rotulo": "Perdas", "valor": round(totais.get("perdas_aves", 0), 2), "tipo": "numero", "unidade": "aves"},
        {"rotulo": "Rendimento", "valor": round(totais.get("rendimento", 0), 2), "tipo": "percentual", "unidade": "%"},
    ]


def agrupar_linhas(linhas, chave):
    grupos = {}
    for linha in linhas:
        valor = linha.get(chave) or "Nao informado"
        if valor not in grupos:
            grupos[valor] = {"chave": valor, "detalhes": []}
        grupos[valor]["detalhes"].append(linha)

    agrupados = []
    total_peso = sum(valor_float(l.get("peso_produzido")) for l in linhas)
    for grupo in grupos.values():
        totais = somar_ops(grupo["detalhes"])
        agrupados.append({
            "grupo": grupo["chave"],
            "ops": valor_int(totais["ops"]),
            "aves_recebidas": round(totais["aves_recebidas"], 2),
            "peso_vivo": round(totais["peso_vivo"], 2),
            "producao_primaria": round(totais["producao_primaria"], 2),
            "producao_secundaria": round(totais["producao_secundaria"], 2),
            "caixas": round(totais["caixas"], 2),
            "peso_produzido": round(totais["peso_produzido"], 2),
            "rendimento": round(totais["rendimento"], 2),
            "condenacoes_aves": round(totais["condenacoes_aves"], 2),
            "taxa_condenacao": round(totais["taxa_condenacao"], 2),
            "perdas_aves": round(totais["perdas_aves"], 2),
            "taxa_perda": round(totais["taxa_perda"], 2),
            "participacao": percentual(totais["peso_produzido"], total_peso),
        })
    return sorted(agrupados, key=lambda item: item["peso_produzido"], reverse=True)


def agrupar_periodo(linhas, granularidade):
    grupos = {}
    for linha in linhas:
        data_op = linha.get("data_op") or ""
        if granularidade == "dia":
            periodo = data_op
        elif granularidade == "semana":
            periodo = periodo_semana_python(data_op)
        else:
            periodo = data_op[:7]
        grupos.setdefault(periodo or "Sem data", []).append(linha)

    ordenados = []
    anterior = None
    for periodo in sorted(grupos.keys()):
        totais = somar_ops(grupos[periodo])
        peso = round(totais["peso_produzido"], 2)
        variacao_abs = 0 if anterior is None else round(peso - anterior, 2)
        variacao_pct = 0 if not anterior else percentual(peso - anterior, anterior)
        ordenados.append({
            "grupo": periodo,
            "ops": valor_int(totais["ops"]),
            "aves_recebidas": round(totais["aves_recebidas"], 2),
            "peso_vivo": round(totais["peso_vivo"], 2),
            "producao_primaria": round(totais["producao_primaria"], 2),
            "producao_secundaria": round(totais["producao_secundaria"], 2),
            "caixas": round(totais["caixas"], 2),
            "peso_produzido": peso,
            "media_diaria": round(peso / max(1, len({l.get("data_op") for l in grupos[periodo]})), 2),
            "rendimento": round(totais["rendimento"], 2),
            "condenacoes_aves": round(totais["condenacoes_aves"], 2),
            "perdas_aves": round(totais["perdas_aves"], 2),
            "variacao_abs": variacao_abs,
            "variacao_pct": variacao_pct,
        })
        anterior = peso
    return ordenados


def periodo_semana_python(data_texto):
    try:
        ano, semana, _ = date.fromisoformat(data_texto).isocalendar()
        return f"{ano}-W{semana:02d}"
    except Exception:
        return data_texto[:7] if data_texto else "Sem data"


def buscar_perdas_detalhadas(filtros, tipo="todas", limite=500):
    condicoes, parametros = montar_condicoes_ops(filtros, "o")
    condicoes.append("LOWER(COALESCE(d.unidade, '')) IN ('aves', 'ave', 'unidade', 'unidades', 'kg')")
    if tipo == "condenacao":
        condicoes.append("(LOWER(COALESCE(d.categoria, '')) LIKE ? OR LOWER(COALESCE(d.motivo, '')) LIKE ?)")
        parametros.extend(["%conden%", "%conden%"])
    where_sql = " AND ".join(condicoes)
    linhas = executar_lista(f"""
    SELECT
        d.id, d.op_id, o.data AS data_op, o.fornecedor, COALESCE(o.sku, 'Galinha Cortada') AS sku,
        d.data, COALESCE(d.setor, 'Sem setor') AS setor,
        COALESCE(d.categoria, '') AS categoria,
        COALESCE(d.motivo, 'Sem motivo') AS motivo,
        COALESCE(d.quantidade, 0) AS quantidade,
        COALESCE(d.unidade, '') AS unidade,
        COALESCE(o.quantidade_aves, 0) AS aves_recebidas,
        CASE
            WHEN LOWER(TRIM(COALESCE(d.motivo, ''))) = 'morte na gaiola' THEN 'Morte na gaiola'
            WHEN LOWER(COALESCE(d.categoria, '')) LIKE '%%conden%%'
              OR LOWER(COALESCE(d.motivo, '')) LIKE '%%conden%%' THEN 'Condenacao'
            WHEN LOWER(COALESCE(d.unidade, '')) = 'kg' THEN 'Perda em kg'
            ELSE 'Descarte operacional'
        END AS tipo_perda
    FROM apontamentos_descartes d
    INNER JOIN ordens_producao o ON o.id = d.op_id
    WHERE {where_sql}
    ORDER BY d.data DESC, d.id DESC
    LIMIT {int(limite)}
    """, parametros)
    for linha in linhas:
        linha["lote"] = f"OP-{valor_int(linha.get('op_id')):05d}"
        linha["taxa"] = percentual(linha["quantidade"], linha["aves_recebidas"]) if str(linha.get("unidade", "")).lower() != "kg" else 0

    if tipo == "todas" and filtros["setor"] == "Todos" and filtros["causa"] in ["Todos", "Mortes antes da pendura"]:
        condicoes_op, parametros_op = montar_condicoes_ops(filtros, "o")
        condicoes_op.append("COALESCE(o.mortes_antes_pendura, 0) > 0")
        legado = executar_lista(f"""
        SELECT
            NULL AS id,
            o.id AS op_id,
            o.data AS data_op,
            o.fornecedor,
            COALESCE(o.sku, 'Galinha Cortada') AS sku,
            o.data AS data,
            'Recebimento' AS setor,
            'Morte' AS categoria,
            'Mortes antes da pendura' AS motivo,
            COALESCE(o.mortes_antes_pendura, 0) AS quantidade,
            'aves' AS unidade,
            COALESCE(o.quantidade_aves, 0) AS aves_recebidas,
            'Mortes antes da pendura' AS tipo_perda
        FROM ordens_producao o
        WHERE {" AND ".join(condicoes_op)}
        ORDER BY o.data DESC, o.id DESC
        """, parametros_op)
        for linha in legado:
            linha["lote"] = f"OP-{valor_int(linha.get('op_id')):05d}"
            linha["taxa"] = percentual(linha["quantidade"], linha["aves_recebidas"])
        linhas.extend(legado)

    return linhas


def agrupar_perdas(linhas, tipo="todas"):
    chave = "motivo" if tipo == "condenacao" else "tipo_perda"
    grupos = defaultdict(lambda: {"grupo": "", "quantidade_aves": 0.0, "quantidade_kg": 0.0, "ocorrencias": 0, "ops": set()})
    for linha in linhas:
        grupo = linha.get(chave) or "Nao informado"
        agg = grupos[grupo]
        agg["grupo"] = grupo
        agg["ocorrencias"] += 1
        agg["ops"].add(linha.get("op_id"))
        if str(linha.get("unidade", "")).lower() == "kg":
            agg["quantidade_kg"] += valor_float(linha.get("quantidade"))
        else:
            agg["quantidade_aves"] += valor_float(linha.get("quantidade"))
    total_aves = sum(item["quantidade_aves"] for item in grupos.values())
    saida = []
    for item in grupos.values():
        saida.append({
            "grupo": item["grupo"],
            "ops": len(item["ops"]),
            "ocorrencias": item["ocorrencias"],
            "quantidade_aves": round(item["quantidade_aves"], 2),
            "quantidade_kg": round(item["quantidade_kg"], 2),
            "participacao": percentual(item["quantidade_aves"], total_aves),
        })
    return sorted(saida, key=lambda item: (item["quantidade_aves"], item["quantidade_kg"]), reverse=True)


def buscar_opcoes_filtro():
    return {
        "status": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT COALESCE(status, 'Aberta') AS valor
            FROM ordens_producao
            WHERE COALESCE(status, '') <> ''
            ORDER BY valor
        """)],
        "skus": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT COALESCE(sku, 'Galinha Cortada') AS valor
            FROM ordens_producao
            WHERE COALESCE(sku, '') <> ''
            ORDER BY valor
        """)],
        "fornecedores": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT fornecedor AS valor
            FROM ordens_producao
            WHERE COALESCE(fornecedor, '') <> ''
            ORDER BY valor
        """)],
        "causas": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT COALESCE(motivo, '') AS valor
            FROM apontamentos_descartes
            WHERE COALESCE(motivo, '') <> ''
            ORDER BY valor
        """)] + ["Mortes antes da pendura"],
        "setores": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT COALESCE(setor, '') AS valor
            FROM apontamentos_descartes
            WHERE COALESCE(setor, '') <> ''
            ORDER BY valor
        """)],
    }


def buscar_eficiencia_dados(filtros):
    condicoes, parametros = montar_condicoes_ops(filtros, "o")
    where_ops = " AND ".join(condicoes)

    linhas = executar_lista(f"""
    WITH op_base AS (
        SELECT
            o.id AS op_id,
            o.data AS data_op,
            o.fornecedor,
            COALESCE(o.sku, 'Galinha Cortada') AS sku,
            COALESCE(o.status, 'Aberta') AS status,
            COALESCE(o.quantidade_aves, 0) AS aves_recebidas,
            COALESCE(o.peso_vivo, 0) AS peso_vivo
        FROM ordens_producao o
        WHERE {where_ops}
    ),
    prod AS (
        SELECT
            op_id,
            COALESCE(SUM(CASE WHEN LOWER(COALESCE(unidade, '')) = 'kg' THEN quantidade ELSE 0 END), 0) AS kg_produzidos,
            COALESCE(SUM(CASE
                WHEN LOWER(COALESCE(unidade, '')) IN ('unidades', 'unidade', 'aves', 'ave', 'bandejas')
                THEN quantidade ELSE 0 END), 0) AS unidades_produzidas
        FROM apontamentos_producao
        GROUP BY op_id
    ),
    pa AS (
        SELECT
            comp.op_id,
            COUNT(DISTINCT cx.id) AS caixas,
            COALESCE(SUM(cx.peso_liquido), 0) AS peso_liquido_pa
        FROM pa_caixa_composicao comp
        INNER JOIN pa_caixas cx ON cx.id = comp.caixa_id
        WHERE COALESCE(cx.status, '') <> 'Cancelada'
        GROUP BY comp.op_id
    ),
    tempos AS (
        SELECT
            op_id,
            data,
            setor,
            hora_inicio,
            hora_fim
        FROM apontamentos_tempos_setor
    ),
    paradas AS (
        SELECT
            op_id,
            data,
            setor,
            COALESCE(SUM(horas_paradas), 0) AS horas_paradas
        FROM apontamentos_paradas
        GROUP BY op_id, data, setor
    )
    SELECT
        op.op_id,
        op.data_op,
        op.fornecedor,
        op.sku,
        op.status,
        op.aves_recebidas,
        op.peso_vivo,
        COALESCE(prod.kg_produzidos, 0) AS kg_produzidos,
        COALESCE(prod.unidades_produzidas, 0) AS unidades_produzidas,
        COALESCE(pa.caixas, 0) AS caixas,
        COALESCE(pa.peso_liquido_pa, 0) AS peso_liquido_pa,
        tempos.data AS data_tempo,
        tempos.setor,
        tempos.hora_inicio,
        tempos.hora_fim,
        COALESCE(paradas.horas_paradas, 0) AS horas_paradas
    FROM op_base op
    LEFT JOIN prod ON prod.op_id = op.op_id
    LEFT JOIN pa ON pa.op_id = op.op_id
    LEFT JOIN tempos ON tempos.op_id = op.op_id
    LEFT JOIN paradas
      ON paradas.op_id = tempos.op_id
     AND paradas.data = tempos.data
     AND paradas.setor = tempos.setor
    ORDER BY op.data_op DESC, op.op_id DESC, tempos.setor ASC
    """, parametros)

    ops = {}
    for linha in linhas:
        op_id = linha.get("op_id")
        op = ops.setdefault(op_id, {
            "op_id": op_id,
            "lote": f"OP-{valor_int(op_id):05d}",
            "data_op": linha.get("data_op"),
            "fornecedor": linha.get("fornecedor"),
            "sku": linha.get("sku"),
            "status": linha.get("status"),
            "aves_recebidas": valor_float(linha.get("aves_recebidas")),
            "peso_vivo": valor_float(linha.get("peso_vivo")),
            "kg_produzidos": valor_float(linha.get("kg_produzidos")),
            "unidades_produzidas": valor_float(linha.get("unidades_produzidas")),
            "caixas": valor_float(linha.get("caixas")),
            "peso_liquido_pa": valor_float(linha.get("peso_liquido_pa")),
            "horas_programadas": 0.0,
            "horas_paradas": 0.0,
            "horas_produtivas": 0.0,
            "setores_registrados": set(),
        })

        setor = linha.get("setor")
        if setor:
            horas_programadas = horas_entre(linha.get("hora_inicio"), linha.get("hora_fim"))
            horas_paradas = min(valor_float(linha.get("horas_paradas")), horas_programadas)
            op["horas_programadas"] += horas_programadas
            op["horas_paradas"] += horas_paradas
            op["setores_registrados"].add(setor)

    detalhes = []
    for op in ops.values():
        op["horas_programadas"] = round(op["horas_programadas"], 4)
        op["horas_paradas"] = round(op["horas_paradas"], 4)
        op["horas_produtivas"] = round(max(op["horas_programadas"] - op["horas_paradas"], 0), 4)
        op["setores"] = ", ".join(sorted(op["setores_registrados"]))
        op["qtd_setores"] = len(op["setores_registrados"])
        op["peso_produzido"] = op["kg_produzidos"] or op["peso_liquido_pa"]
        op["kg_por_hora_setor"] = dividir(op["peso_produzido"], op["horas_produtivas"])
        op["caixas_por_hora_setor"] = dividir(op["caixas"], op["horas_produtivas"])
        op["aves_por_hora_setor"] = dividir(op["aves_recebidas"], op["horas_produtivas"])
        op["situacao_eficiencia"] = "Com tempo oficial" if op["horas_produtivas"] > 0 else "Sem tempo oficial"
        op.pop("setores_registrados", None)
        detalhes.append(op)

    return detalhes


def somar_eficiencia(linhas):
    totais = defaultdict(float)
    for linha in linhas:
        for chave in [
            "aves_recebidas", "peso_vivo", "peso_produzido", "kg_produzidos",
            "unidades_produzidas", "caixas", "horas_programadas",
            "horas_paradas", "horas_produtivas",
        ]:
            totais[chave] += valor_float(linha.get(chave))
        if linha.get("horas_produtivas", 0) > 0:
            totais["ops_com_tempo"] += 1

    totais["ops"] = len(linhas)
    totais["kg_por_hora_setor"] = dividir(totais["peso_produzido"], totais["horas_produtivas"])
    totais["caixas_por_hora_setor"] = dividir(totais["caixas"], totais["horas_produtivas"])
    totais["aves_por_hora_setor"] = dividir(totais["aves_recebidas"], totais["horas_produtivas"])
    return dict(totais)


def resumo_kpis_eficiencia(totais):
    return [
        {"rotulo": "OPs", "valor": valor_int(totais.get("ops")), "tipo": "numero", "unidade": "OPs"},
        {"rotulo": "OPs com tempo", "valor": valor_int(totais.get("ops_com_tempo")), "tipo": "numero", "unidade": "OPs"},
        {"rotulo": "Peso produzido", "valor": round(totais.get("peso_produzido", 0), 2), "tipo": "decimal", "unidade": "kg"},
        {"rotulo": "Caixas PA", "valor": round(totais.get("caixas", 0), 2), "tipo": "numero", "unidade": "caixas"},
        {"rotulo": "Horas registradas", "valor": round(totais.get("horas_programadas", 0), 2), "tipo": "decimal", "unidade": "hora-setor"},
        {"rotulo": "Horas paradas", "valor": round(totais.get("horas_paradas", 0), 2), "tipo": "decimal", "unidade": "hora-setor"},
        {"rotulo": "Horas produtivas", "valor": round(totais.get("horas_produtivas", 0), 2), "tipo": "decimal", "unidade": "hora-setor"},
        {"rotulo": "Kg por hora", "valor": round(totais.get("kg_por_hora_setor", 0), 2), "tipo": "decimal", "unidade": "kg/hora-setor"},
    ]


def agrupar_eficiencia_periodo(linhas, granularidade):
    grupos = defaultdict(list)
    for linha in linhas:
        data_op = linha.get("data_op") or ""
        if granularidade == "dia":
            periodo = data_op
        elif granularidade == "semana":
            periodo = periodo_semana_python(data_op)
        else:
            periodo = data_op[:7]
        grupos[periodo or "Sem data"].append(linha)

    saida = []
    for periodo in sorted(grupos.keys()):
        totais = somar_eficiencia(grupos[periodo])
        saida.append({
            "grupo": periodo,
            "ops": valor_int(totais.get("ops")),
            "ops_com_tempo": valor_int(totais.get("ops_com_tempo")),
            "peso_produzido": round(totais.get("peso_produzido", 0), 2),
            "caixas": round(totais.get("caixas", 0), 2),
            "horas_programadas": round(totais.get("horas_programadas", 0), 2),
            "horas_paradas": round(totais.get("horas_paradas", 0), 2),
            "horas_produtivas": round(totais.get("horas_produtivas", 0), 2),
            "kg_por_hora_setor": round(totais.get("kg_por_hora_setor", 0), 2),
            "caixas_por_hora_setor": round(totais.get("caixas_por_hora_setor", 0), 2),
        })
    return saida


def montar_contexto_eficiencia(config, filtros):
    detalhes = buscar_eficiencia_dados(filtros)
    totais = somar_eficiencia(detalhes)
    limitacoes = [
        "Base parcial: o PRUMO possui producao, tempos e paradas oficiais, mas nao possui meta/capacidade oficial por OP.",
        "Este relatorio mede produtividade horaria por hora-setor registrada; nao calcula percentual de eficiencia, aderencia a meta ou OEE.",
        "Rendimento industrial permanece no relatorio Rendimento e nao foi reutilizado como eficiencia.",
    ]

    if not detalhes:
        limitacoes.append("Nenhuma OP encontrada para os filtros selecionados.")

    return {
        "slug": "eficiencia",
        "config": config,
        "filtros": filtros,
        "opcoes": buscar_opcoes_filtro(),
        "resumo": resumo_kpis_eficiencia(totais),
        "totais": totais,
        "agrupamentos": agrupar_eficiencia_periodo(detalhes, filtros["granularidade"]),
        "detalhes": detalhes,
        "limitacoes": limitacoes,
        "query_string": urlencode({k: v for k, v in filtros.items() if v not in ["", "Todos", "Todas"]}),
        "granularidades": [("dia", "Dia"), ("semana", "Semana"), ("mes", "Mes")],
    }


def montar_contexto_relatorio_producao(slug, args):
    config = RELATORIOS_PRODUCAO[slug]
    filtros = normalizar_filtros(args)
    if config["familia"] == "eficiencia":
        return montar_contexto_eficiencia(config, filtros)

    somente_encerradas = bool(config.get("somente_encerradas"))

    detalhes = buscar_ops_agregadas(filtros, somente_encerradas=somente_encerradas)
    agrupamentos = []
    perdas_detalhes = []
    limitacoes = []

    if config["familia"] == "perdas":
        perdas_detalhes = buscar_perdas_detalhadas(filtros, config.get("tipo_perda", "todas"))
        agrupamentos = agrupar_perdas(perdas_detalhes, config.get("tipo_perda", "todas"))
        totais = somar_ops(detalhes)
        totais["perdas_aves"] = sum(valor_float(i.get("quantidade")) for i in perdas_detalhes if str(i.get("unidade", "")).lower() != "kg")
        totais["perdas_kg"] = sum(valor_float(i.get("quantidade")) for i in perdas_detalhes if str(i.get("unidade", "")).lower() == "kg")
        if config.get("tipo_perda") == "condenacao":
            totais["condenacoes_aves"] = totais["perdas_aves"]
            limitacoes.append("Condenacoes em kg sao exibidas separadamente e nao sao convertidas em aves.")
        detalhes = perdas_detalhes
    elif config["agrupamento"] == "sku":
        agrupamentos = agrupar_linhas(detalhes, "sku")
        totais = somar_ops(detalhes)
    elif config["agrupamento"] == "fornecedor":
        agrupamentos = agrupar_linhas(detalhes, "fornecedor")
        totais = somar_ops(detalhes)
    elif config["agrupamento"] == "periodo":
        agrupamentos = agrupar_periodo(detalhes, filtros["granularidade"])
        totais = somar_ops(detalhes)
    elif config["agrupamento"] == "periodo_fornecedor":
        agrupamentos = agrupar_linhas(detalhes, "fornecedor")
        totais = somar_ops(detalhes)
        limitacoes.append("Formula preservada: kg produzido em apontamentos_producao / peso vivo das OPs encerradas.")
        limitacoes.append(f"Meta homologada preservada: {META_RENDIMENTO:.1f}%.")
    else:
        totais = somar_ops(detalhes)

    if not detalhes:
        limitacoes.append("Nenhum evento produtivo encontrado para os filtros selecionados.")
    if slug in ["producao-por-op", "producao-por-periodo"]:
        limitacoes.append("Data oficial utilizada: data da OP. Nao existe campo dedicado de data de encerramento.")
    if slug == "perdas":
        limitacoes.append("Perdas em kg nao sao convertidas para aves e nao entram na taxa de perda em aves.")

    return {
        "slug": slug,
        "config": config,
        "filtros": filtros,
        "opcoes": buscar_opcoes_filtro(),
        "resumo": resumo_kpis(totais),
        "totais": totais,
        "agrupamentos": agrupamentos,
        "detalhes": detalhes,
        "limitacoes": limitacoes,
        "query_string": urlencode({k: v for k, v in filtros.items() if v not in ["", "Todos", "Todas"]}),
        "granularidades": [("dia", "Dia"), ("semana", "Semana"), ("mes", "Mes")],
    }


def gerar_excel_relatorio_producao(contexto):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    ws.append([contexto["config"]["titulo"]])
    ws.append([contexto["config"]["objetivo"]])
    ws.append([])
    ws.append(["Filtros"])
    for chave, valor in contexto["filtros"].items():
        ws.append([chave, valor])
    ws.append([])
    ws.append(["Indicador", "Valor", "Unidade"])
    for item in contexto["resumo"]:
        ws.append([item["rotulo"], item["valor"], item["unidade"]])
    ws.append([])
    if contexto["agrupamentos"]:
        ws.append(["Agrupamentos"])
        chaves = list(contexto["agrupamentos"][0].keys())
        ws.append(chaves)
        for item in contexto["agrupamentos"]:
            linha = []
            for chave in chaves:
                valor = item.get(chave, "")
                if isinstance(valor, set):
                    valor = len(valor)
                linha.append(valor)
            ws.append(linha)
        ws.append([])
    ws.append(["Detalhes"])
    if contexto["config"].get("familia") == "eficiencia":
        colunas = [
            "op_id", "lote", "data_op", "fornecedor", "sku", "status",
            "aves_recebidas", "peso_vivo", "peso_produzido", "kg_produzidos",
            "unidades_produzidas", "caixas", "setores", "qtd_setores",
            "horas_programadas", "horas_paradas", "horas_produtivas",
            "kg_por_hora_setor", "caixas_por_hora_setor",
            "aves_por_hora_setor", "situacao_eficiencia",
        ]
    else:
        colunas = [
            "op_id", "lote", "data_op", "fornecedor", "sku", "status",
            "aves_recebidas", "peso_vivo", "producao_primaria", "producao_secundaria",
            "caixas", "peso_produzido", "descartes_aves", "condenacoes_aves",
            "mortes_total", "perdas_aves", "perdas_kg", "rendimento",
            "taxa_condenacao", "taxa_perda", "tipo_perda", "motivo", "setor",
            "quantidade", "unidade",
        ]
    ws.append(colunas)
    for item in contexto["detalhes"]:
        ws.append([item.get(coluna, "") for coluna in colunas])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3B4D")
    for coluna in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(coluna)].width = 18
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo

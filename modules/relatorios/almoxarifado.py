"""Servicos compartilhados para relatorios oficiais de almoxarifado."""

from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database import conectar, q


RELATORIOS_ALMOXARIFADO = {
    "entradas": {
        "catalogo_id": "almoxarifado-entradas",
        "titulo": "Entradas",
        "objetivo": "Consultar entradas oficiais de insumos por periodo, documento, fornecedor e lote.",
        "familia": "movimentacoes",
        "tipos": ["ENTRADA"],
    },
    "consumo": {
        "catalogo_id": "almoxarifado-consumo",
        "titulo": "Consumo",
        "objetivo": "Consultar saidas e consumos de insumos sem alterar a regra operacional de baixa.",
        "familia": "movimentacoes",
        "tipos": ["SAIDA", "SAIDA_OP"],
    },
    "estoque-atual": {
        "catalogo_id": "almoxarifado-estoque-atual",
        "titulo": "Estoque Atual",
        "objetivo": "Exibir o saldo oficial atual dos insumos calculado pelos lotes.",
        "familia": "saldo",
    },
    "estoque-por-produto": {
        "catalogo_id": "almoxarifado-estoque-produto",
        "titulo": "Estoque por Produto",
        "objetivo": "Detalhar a distribuicao do saldo por produto, categoria e lotes.",
        "familia": "saldo",
    },
}

TIPOS_CONSUMO = ("SAIDA", "SAIDA_OP")
TIPOS_ENTRADA = ("ENTRADA", "ESTORNO_OP")
GIRO_BAIXO_LIMITE = 0.5
GIRO_ALTO_LIMITE = 2.0


def hoje_iso():
    return date.today().isoformat()


def primeiro_dia_mes():
    hoje = date.today()
    return hoje.replace(day=1).isoformat()


def valor_float(valor, casas=4):
    try:
        return round(float(valor or 0), casas)
    except (TypeError, ValueError):
        return 0.0


def parse_data(data_texto):
    try:
        return date.fromisoformat(str(data_texto or "")[:10])
    except ValueError:
        return None


def dias_periodo(data_inicio, data_fim):
    inicio = parse_data(data_inicio)
    fim = parse_data(data_fim)
    if not inicio or not fim or fim < inicio:
        return []
    dias = []
    atual = inicio
    while atual <= fim:
        dias.append(atual)
        atual += timedelta(days=1)
    return dias


def dividir(numerador, denominador, casas=4):
    denominador = float(denominador or 0)
    if denominador <= 0:
        return 0.0
    return round(float(numerador or 0) / denominador, casas)


def sinal_movimentacao(tipo):
    if tipo in TIPOS_ENTRADA:
        return 1
    if tipo in TIPOS_CONSUMO:
        return -1
    return 0


def executar_lista(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), tuple(parametros))
    linhas = cursor.fetchall()
    conn.close()
    return [dict(linha) for linha in linhas]


def normalizar_filtros(args, config):
    tipo = args.get("tipo") or "Todos"
    tipos_validos = ["Todos", "ENTRADA", "SAIDA", "SAIDA_OP", "ESTORNO_OP", "AJUSTE"]
    if tipo not in tipos_validos:
        tipo = "Todos"

    return {
        "data_inicio": args.get("data_inicio") or primeiro_dia_mes(),
        "data_fim": args.get("data_fim") or hoje_iso(),
        "categoria": args.get("categoria") or "Todas",
        "insumo": args.get("insumo") or "",
        "tipo": tipo,
        "fornecedor": args.get("fornecedor") or "",
        "numero_nf": args.get("numero_nf") or "",
        "lote": args.get("lote") or "",
        "op_id": args.get("op_id") or "",
        "status_lote": args.get("status_lote") or "Todos",
        "situacao": args.get("situacao") or "Todas",
        "status_fifo": args.get("status_fifo") or "Todos",
        "somente_com_saldo": args.get("somente_com_saldo") or ("Sim" if config["familia"] == "saldo" else "Nao"),
        "por_pagina": 300,
    }


def montar_condicoes_movimentacoes(config, filtros):
    condicoes = ["m.data_movimentacao BETWEEN ? AND ?"]
    parametros = [filtros["data_inicio"], filtros["data_fim"]]

    tipos_config = config.get("tipos")
    if filtros["tipo"] != "Todos":
        condicoes.append("m.tipo = ?")
        parametros.append(filtros["tipo"])
    elif tipos_config:
        placeholders = ", ".join(["?"] * len(tipos_config))
        condicoes.append(f"m.tipo IN ({placeholders})")
        parametros.extend(tipos_config)

    if filtros["categoria"] != "Todas":
        condicoes.append("i.categoria = ?")
        parametros.append(filtros["categoria"])
    if filtros["insumo"]:
        condicoes.append("LOWER(i.descricao) LIKE ?")
        parametros.append(f"%{filtros['insumo'].lower()}%")
    if filtros["fornecedor"]:
        condicoes.append("LOWER(COALESCE(m.fornecedor, '')) LIKE ?")
        parametros.append(f"%{filtros['fornecedor'].lower()}%")
    if filtros["numero_nf"]:
        condicoes.append("LOWER(COALESCE(m.numero_nf, '')) LIKE ?")
        parametros.append(f"%{filtros['numero_nf'].lower()}%")
    if filtros["lote"]:
        condicoes.append("LOWER(COALESCE(m.lote, '')) LIKE ?")
        parametros.append(f"%{filtros['lote'].lower()}%")
    if filtros["op_id"]:
        try:
            condicoes.append("m.op_id = ?")
            parametros.append(int(filtros["op_id"]))
        except ValueError:
            condicoes.append("1 = 0")

    return " AND ".join(condicoes), parametros


def montar_condicoes_saldo(filtros):
    condicoes = ["1 = 1"]
    parametros = []

    if filtros["categoria"] != "Todas":
        condicoes.append("i.categoria = ?")
        parametros.append(filtros["categoria"])
    if filtros["insumo"]:
        condicoes.append("LOWER(i.descricao) LIKE ?")
        parametros.append(f"%{filtros['insumo'].lower()}%")
    if filtros["fornecedor"]:
        condicoes.append("LOWER(COALESCE(l.fornecedor, '')) LIKE ?")
        parametros.append(f"%{filtros['fornecedor'].lower()}%")
    if filtros["numero_nf"]:
        condicoes.append("LOWER(COALESCE(l.numero_nf, '')) LIKE ?")
        parametros.append(f"%{filtros['numero_nf'].lower()}%")
    if filtros["lote"]:
        condicoes.append("LOWER(COALESCE(l.lote, '')) LIKE ?")
        parametros.append(f"%{filtros['lote'].lower()}%")
    if filtros["status_lote"] != "Todos":
        condicoes.append("COALESCE(l.status, 'Aberto') = ?")
        parametros.append(filtros["status_lote"])
    if filtros["somente_com_saldo"] == "Sim":
        condicoes.append("COALESCE(l.quantidade_atual, 0) > 0")

    return " AND ".join(condicoes), parametros


def buscar_movimentacoes(config, filtros):
    where_sql, parametros = montar_condicoes_movimentacoes(config, filtros)
    return executar_lista(f"""
        SELECT
            m.id,
            m.data_movimentacao,
            m.tipo,
            i.descricao AS insumo,
            i.categoria,
            i.unidade,
            m.quantidade,
            m.valor_unitario,
            m.valor_total,
            m.fornecedor,
            m.numero_nf,
            m.lote,
            m.origem,
            m.op_id,
            m.observacoes,
            m.criado_em
        FROM almoxarifado_movimentacoes m
        JOIN almoxarifado_insumos i ON i.id = m.insumo_id
        WHERE {where_sql}
        ORDER BY m.data_movimentacao DESC, m.id DESC
        LIMIT ?
    """, parametros + [filtros["por_pagina"]])


def buscar_saldos(filtros):
    where_sql, parametros = montar_condicoes_saldo(filtros)
    return executar_lista(f"""
        SELECT
            i.id AS insumo_id,
            i.descricao AS insumo,
            i.categoria,
            i.unidade,
            i.ativo,
            COUNT(l.id) AS lotes,
            COALESCE(SUM(l.quantidade_atual), 0) AS saldo_atual,
            COALESCE(SUM(l.quantidade_atual * l.valor_unitario), 0) AS valor_estoque,
            MIN(l.data_entrada) AS primeira_entrada,
            MAX(l.data_entrada) AS ultima_entrada
        FROM almoxarifado_insumos i
        LEFT JOIN almoxarifado_lotes l ON l.insumo_id = i.id
        WHERE {where_sql}
        GROUP BY i.id, i.descricao, i.categoria, i.unidade, i.ativo
        ORDER BY i.categoria ASC, i.descricao ASC
        LIMIT ?
    """, parametros + [filtros["por_pagina"]])


def buscar_lotes(filtros):
    where_sql, parametros = montar_condicoes_saldo(filtros)
    return executar_lista(f"""
        SELECT
            l.id,
            i.descricao AS insumo,
            i.categoria,
            i.unidade,
            l.data_entrada,
            l.lote,
            l.fornecedor,
            l.numero_nf,
            l.quantidade_inicial,
            l.quantidade_atual,
            l.valor_unitario,
            l.valor_total,
            l.status,
            l.criado_em
        FROM almoxarifado_lotes l
        JOIN almoxarifado_insumos i ON i.id = l.insumo_id
        WHERE {where_sql}
        ORDER BY l.data_entrada ASC, l.id ASC
        LIMIT ?
    """, parametros + [filtros["por_pagina"]])


def montar_condicoes_insumos(filtros, alias="i"):
    condicoes = ["1 = 1"]
    parametros = []
    if filtros["categoria"] != "Todas":
        condicoes.append(f"{alias}.categoria = ?")
        parametros.append(filtros["categoria"])
    if filtros["insumo"]:
        condicoes.append(f"LOWER({alias}.descricao) LIKE ?")
        parametros.append(f"%{filtros['insumo'].lower()}%")
    return " AND ".join(condicoes), parametros


def buscar_movimentacoes_periodo_com_abertura(filtros):
    where_insumo, parametros_insumo = montar_condicoes_insumos(filtros, "i")
    return executar_lista(f"""
        SELECT
            m.id,
            m.data_movimentacao,
            m.tipo,
            m.insumo_id,
            i.descricao AS insumo,
            i.categoria,
            i.unidade,
            m.lote_id,
            m.lote,
            m.quantidade,
            m.valor_unitario,
            m.valor_total,
            m.op_id,
            m.origem,
            m.criado_em
        FROM almoxarifado_movimentacoes m
        JOIN almoxarifado_insumos i ON i.id = m.insumo_id
        WHERE {where_insumo}
          AND m.data_movimentacao <= ?
        ORDER BY m.data_movimentacao ASC, m.id ASC
    """, parametros_insumo + [filtros["data_fim"]])


def buscar_saldo_atual_por_insumo(filtros):
    where_insumo, parametros = montar_condicoes_insumos(filtros, "i")
    linhas = executar_lista(f"""
        SELECT
            i.id AS insumo_id,
            COALESCE(SUM(l.quantidade_atual), 0) AS saldo_atual,
            MAX(l.data_entrada) AS ultima_entrada,
            MIN(CASE WHEN COALESCE(l.quantidade_atual, 0) > 0 THEN l.data_entrada ELSE NULL END) AS primeira_entrada_com_saldo
        FROM almoxarifado_insumos i
        LEFT JOIN almoxarifado_lotes l ON l.insumo_id = i.id
        WHERE {where_insumo}
        GROUP BY i.id
    """, parametros)
    return {item["insumo_id"]: item for item in linhas}


def buscar_insumos_filtrados(filtros):
    where_insumo, parametros = montar_condicoes_insumos(filtros, "i")
    return executar_lista(f"""
        SELECT
            i.id AS insumo_id,
            i.descricao AS insumo,
            i.categoria,
            i.unidade
        FROM almoxarifado_insumos i
        WHERE {where_insumo}
        ORDER BY i.categoria, i.descricao
    """, parametros)


def classificar_giro(item):
    saldo_final = valor_float(item.get("saldo_final"))
    consumo = valor_float(item.get("consumo"))
    giro = valor_float(item.get("giro"))
    reconciliado = item.get("saldo_reconciliado")

    if saldo_final < 0:
        return "Saldo negativo"
    if reconciliado is False:
        return "Historico insuficiente"
    if consumo <= 0:
        return "Sem consumo"
    if saldo_final == 0:
        return "Saldo zerado"
    if giro < GIRO_BAIXO_LIMITE:
        return "Giro baixo"
    if giro >= GIRO_ALTO_LIMITE:
        return "Giro alto"
    return "Giro intermediario"


def buscar_giro_estoque(filtros):
    dias = dias_periodo(filtros["data_inicio"], filtros["data_fim"])
    movimentos = buscar_movimentacoes_periodo_com_abertura(filtros)
    saldos_atuais = buscar_saldo_atual_por_insumo(filtros)
    inicio = parse_data(filtros["data_inicio"])
    fim = parse_data(filtros["data_fim"])

    grupos = {}
    for mov in movimentos:
        chave = (mov["insumo_id"], mov["unidade"])
        if chave not in grupos:
            grupos[chave] = {
                "insumo_id": mov["insumo_id"],
                "insumo": mov["insumo"],
                "categoria": mov["categoria"],
                "unidade": mov["unidade"],
                "saldo_inicial": 0.0,
                "entradas": 0.0,
                "consumo": 0.0,
                "estornos": 0.0,
                "ajustes": 0.0,
                "outras_saidas": 0.0,
                "movimentos_por_dia": defaultdict(float),
                "ultima_movimentacao": "",
                "ultima_entrada": "",
                "ultimo_consumo": "",
            }

        grupo = grupos[chave]
        data_mov = parse_data(mov["data_movimentacao"])
        tipo = mov["tipo"]
        quantidade = valor_float(mov["quantidade"])
        sinal = sinal_movimentacao(tipo)
        delta = quantidade * sinal

        if data_mov and inicio and data_mov < inicio:
            grupo["saldo_inicial"] += delta
        elif data_mov and fim and data_mov <= fim:
            if tipo == "ENTRADA":
                grupo["entradas"] += quantidade
            elif tipo in TIPOS_CONSUMO:
                grupo["consumo"] += quantidade
            elif tipo == "ESTORNO_OP":
                grupo["estornos"] += quantidade
            elif tipo == "AJUSTE":
                grupo["ajustes"] += quantidade
            elif sinal < 0:
                grupo["outras_saidas"] += quantidade
            grupo["movimentos_por_dia"][data_mov] += delta

        if not grupo["ultima_movimentacao"] or str(mov["data_movimentacao"]) > grupo["ultima_movimentacao"]:
            grupo["ultima_movimentacao"] = mov["data_movimentacao"]
        if tipo == "ENTRADA" and (not grupo["ultima_entrada"] or str(mov["data_movimentacao"]) > grupo["ultima_entrada"]):
            grupo["ultima_entrada"] = mov["data_movimentacao"]
        if tipo in TIPOS_CONSUMO and (not grupo["ultimo_consumo"] or str(mov["data_movimentacao"]) > grupo["ultimo_consumo"]):
            grupo["ultimo_consumo"] = mov["data_movimentacao"]

    insumos_por_id = {item["insumo_id"]: item for item in buscar_insumos_filtrados(filtros)}
    for insumo_id, saldo in saldos_atuais.items():
        if filtros["somente_com_saldo"] == "Sim" and valor_float(saldo.get("saldo_atual")) <= 0:
            continue
        if (insumo_id, None) not in grupos and not any(chave[0] == insumo_id for chave in grupos):
            item = insumos_por_id.get(insumo_id)
            if item:
                grupos[(insumo_id, item["unidade"])] = {
                    "insumo_id": insumo_id,
                    "insumo": item["insumo"],
                    "categoria": item["categoria"],
                    "unidade": item["unidade"],
                    "saldo_inicial": 0.0,
                    "entradas": 0.0,
                    "consumo": 0.0,
                    "estornos": 0.0,
                    "ajustes": 0.0,
                    "outras_saidas": 0.0,
                    "movimentos_por_dia": defaultdict(float),
                    "ultima_movimentacao": "",
                    "ultima_entrada": saldo.get("ultima_entrada") or "",
                    "ultimo_consumo": "",
                }

    detalhes = []
    data_ref = fim or date.today()
    for grupo in grupos.values():
        saldo = valor_float(grupo["saldo_inicial"])
        fechamentos = []
        for dia in dias:
            saldo += valor_float(grupo["movimentos_por_dia"].get(dia))
            fechamentos.append(saldo)

        estoque_medio = dividir(sum(fechamentos), len(fechamentos)) if fechamentos else 0.0
        saldo_final = round(saldo, 4)
        saldo_atual = valor_float((saldos_atuais.get(grupo["insumo_id"]) or {}).get("saldo_atual"))
        compara_saldo_atual = bool(fim and fim >= date.today())
        consumo_medio_diario = dividir(grupo["consumo"], len(dias))
        cobertura = None
        if saldo_final >= 0 and consumo_medio_diario > 0:
            cobertura = dividir(saldo_final, consumo_medio_diario, 2)
        giro = dividir(grupo["consumo"], estoque_medio, 4) if estoque_medio > 0 else 0.0
        ultimo_consumo_data = parse_data(grupo["ultimo_consumo"])
        ultima_mov_data = parse_data(grupo["ultima_movimentacao"])
        grupo_saida = {
            "insumo_id": grupo["insumo_id"],
            "insumo": grupo["insumo"],
            "categoria": grupo["categoria"],
            "unidade": grupo["unidade"],
            "saldo_inicial": round(grupo["saldo_inicial"], 4),
            "entradas": round(grupo["entradas"], 4),
            "consumo": round(grupo["consumo"], 4),
            "outras_saidas": round(grupo["outras_saidas"], 4),
            "ajustes": round(grupo["ajustes"], 4),
            "estornos": round(grupo["estornos"], 4),
            "saldo_final": saldo_final,
            "saldo_atual": saldo_atual,
            "saldo_reconciliado": abs(saldo_final - saldo_atual) < 0.0001 if compara_saldo_atual else None,
            "estoque_medio": round(estoque_medio, 4),
            "giro": giro,
            "consumo_medio_diario": round(consumo_medio_diario, 4),
            "cobertura_dias": cobertura,
            "ultima_entrada": grupo["ultima_entrada"],
            "ultimo_consumo": grupo["ultimo_consumo"],
            "ultima_movimentacao": grupo["ultima_movimentacao"],
            "dias_sem_consumo": (data_ref - ultimo_consumo_data).days if ultimo_consumo_data else None,
            "dias_sem_movimentacao": (data_ref - ultima_mov_data).days if ultima_mov_data else None,
        }
        grupo_saida["situacao"] = classificar_giro(grupo_saida)
        if filtros["situacao"] != "Todas" and grupo_saida["situacao"] != filtros["situacao"]:
            continue
        detalhes.append(grupo_saida)

    detalhes.sort(key=lambda item: (item["situacao"], item["categoria"], item["insumo"]))
    return detalhes


def buscar_lotes_para_fifo(filtros, aplicar_filtro_lote=False):
    where_insumo, parametros = montar_condicoes_insumos(filtros, "i")
    if aplicar_filtro_lote and filtros["lote"]:
        where_insumo += " AND LOWER(COALESCE(l.lote, '')) LIKE ?"
        parametros.append(f"%{filtros['lote'].lower()}%")
    return executar_lista(f"""
        SELECT
            l.id,
            l.insumo_id,
            i.descricao AS insumo,
            i.categoria,
            i.unidade,
            l.data_entrada,
            l.lote,
            l.quantidade_inicial,
            l.quantidade_atual,
            l.status,
            l.fornecedor,
            l.numero_nf,
            l.criado_em
        FROM almoxarifado_lotes l
        JOIN almoxarifado_insumos i ON i.id = l.insumo_id
        WHERE {where_insumo}
        ORDER BY l.insumo_id, l.data_entrada ASC, l.id ASC
    """, parametros)


def buscar_movimentacoes_fifo(filtros):
    where_insumo, parametros = montar_condicoes_insumos(filtros, "i")
    condicoes = [where_insumo, "m.data_movimentacao <= ?"]
    parametros.append(filtros["data_fim"])
    if filtros["lote"]:
        condicoes.append("LOWER(COALESCE(m.lote, '')) LIKE ?")
        parametros.append(f"%{filtros['lote'].lower()}%")
    if filtros["op_id"]:
        try:
            condicoes.append("m.op_id = ?")
            parametros.append(int(filtros["op_id"]))
        except ValueError:
            condicoes.append("1 = 0")
    return executar_lista(f"""
        SELECT
            m.id,
            m.data_movimentacao,
            m.tipo,
            m.insumo_id,
            i.descricao AS insumo,
            i.categoria,
            i.unidade,
            m.lote_id,
            m.lote,
            m.quantidade,
            m.valor_unitario,
            m.valor_total,
            m.op_id,
            m.origem,
            m.observacoes,
            m.criado_em
        FROM almoxarifado_movimentacoes m
        JOIN almoxarifado_insumos i ON i.id = m.insumo_id
        WHERE {" AND ".join(condicoes)}
          AND m.tipo IN ('ENTRADA', 'SAIDA', 'SAIDA_OP', 'ESTORNO_OP')
        ORDER BY m.data_movimentacao ASC, m.id ASC
    """, parametros)


def lote_mais_antigo_elegivel(lotes, saldos_lote, insumo_id, data_movimentacao):
    elegiveis = []
    data_consumo = parse_data(data_movimentacao)
    for lote in lotes:
        if lote["insumo_id"] != insumo_id:
            continue
        if valor_float(saldos_lote.get(lote["id"])) <= 0:
            continue
        data_entrada = parse_data(lote["data_entrada"])
        if data_consumo and data_entrada and data_entrada > data_consumo:
            continue
        elegiveis.append(lote)
    if not elegiveis:
        return None
    return sorted(elegiveis, key=lambda item: ((item.get("data_entrada") or ""), item["id"]))[0]


def classificar_fifo(mov, lote_consumido, elegivel, historico_incompleto):
    if not mov.get("lote_id"):
        return "Sem lote", "Consumo oficial sem lote_id; nao e possivel auditar a ordem FIFO."
    if historico_incompleto:
        return "Nao verificavel", "Ha consumo anterior sem lote para o mesmo insumo, impedindo saldo historico confiavel."
    if not lote_consumido:
        return "Nao verificavel", "O lote consumido nao foi localizado na base oficial de lotes."
    if not elegivel:
        return "Nao verificavel", "Nao foi identificado lote elegivel com saldo positivo antes do consumo."
    if int(lote_consumido["id"]) == int(elegivel["id"]):
        return "Conforme", "O lote consumido era o lote elegivel mais antigo pela data de entrada e ID."
    return "Possivel violacao", "Havia lote anterior elegivel com saldo positivo no momento do consumo."


def buscar_fifo_analitico(filtros):
    lotes = buscar_lotes_para_fifo(filtros, aplicar_filtro_lote=False)
    lotes_por_id = {item["id"]: item for item in lotes}
    movimentos = buscar_movimentacoes_fifo(filtros)
    inicio = parse_data(filtros["data_inicio"])
    fim = parse_data(filtros["data_fim"])
    saldos_lote = defaultdict(float)
    historico_incompleto_insumo = defaultdict(bool)
    detalhes = []

    for mov in movimentos:
        data_mov = parse_data(mov["data_movimentacao"])
        tipo = mov["tipo"]
        lote_id = mov.get("lote_id")
        quantidade = valor_float(mov.get("quantidade"))

        if tipo == "ENTRADA" and lote_id:
            saldos_lote[lote_id] += quantidade
            continue

        if tipo == "ESTORNO_OP" and lote_id:
            saldos_lote[lote_id] += quantidade
            continue

        if tipo not in TIPOS_CONSUMO:
            continue

        dentro_periodo = bool(data_mov and inicio and fim and inicio <= data_mov <= fim)
        lote_consumido = lotes_por_id.get(lote_id)
        elegivel = lote_mais_antigo_elegivel(lotes, saldos_lote, mov["insumo_id"], mov["data_movimentacao"])

        if dentro_periodo:
            situacao, justificativa = classificar_fifo(
                mov,
                lote_consumido,
                elegivel,
                historico_incompleto_insumo[mov["insumo_id"]],
            )
            detalhe = {
                "id": mov["id"],
                "data_movimentacao": mov["data_movimentacao"],
                "tipo": mov["tipo"],
                "insumo": mov["insumo"],
                "categoria": mov["categoria"],
                "unidade": mov["unidade"],
                "op_id": mov["op_id"],
                "quantidade": quantidade,
                "lote_consumido_id": lote_id,
                "lote_consumido": mov.get("lote") or (lote_consumido or {}).get("lote"),
                "entrada_lote_consumido": (lote_consumido or {}).get("data_entrada"),
                "lote_elegivel_id": (elegivel or {}).get("id"),
                "lote_elegivel": (elegivel or {}).get("lote"),
                "entrada_lote_elegivel": (elegivel or {}).get("data_entrada"),
                "saldo_lote_elegivel_antes": valor_float(saldos_lote.get((elegivel or {}).get("id"))),
                "situacao": situacao,
                "justificativa": justificativa,
                "evento_original": mov["id"],
            }
            if filtros["status_fifo"] == "Todos" or detalhe["situacao"] == filtros["status_fifo"]:
                detalhes.append(detalhe)

        if lote_id:
            saldos_lote[lote_id] -= quantidade
        else:
            historico_incompleto_insumo[mov["insumo_id"]] = True

    lotes_antigos = []
    data_ref = fim or date.today()
    for lote in lotes:
        saldo = valor_float(lote.get("quantidade_atual"))
        if saldo <= 0:
            continue
        data_entrada = parse_data(lote.get("data_entrada"))
        lotes_antigos.append({
            "insumo": lote["insumo"],
            "categoria": lote["categoria"],
            "unidade": lote["unidade"],
            "lote": lote["lote"],
            "data_entrada": lote["data_entrada"],
            "idade_dias": (data_ref - data_entrada).days if data_entrada else None,
            "saldo": saldo,
            "status": lote["status"],
            "fornecedor": lote["fornecedor"],
            "numero_nf": lote["numero_nf"],
        })
    lotes_antigos.sort(key=lambda item: (item["data_entrada"] or "", item["insumo"]))
    return detalhes, lotes_antigos[:100]


def agrupar_por_tipo(movimentacoes):
    grupos = defaultdict(lambda: {"grupo": "", "eventos": 0, "quantidade": 0.0, "valor_total": 0.0})
    for item in movimentacoes:
        chave = item["tipo"] or "Sem tipo"
        grupos[chave]["grupo"] = chave
        grupos[chave]["eventos"] += 1
        grupos[chave]["quantidade"] += valor_float(item["quantidade"])
        grupos[chave]["valor_total"] += valor_float(item["valor_total"], 2)
    return sorted(grupos.values(), key=lambda item: item["valor_total"], reverse=True)


def agrupar_por_categoria(linhas):
    grupos = defaultdict(lambda: {"grupo": "", "itens": 0, "lotes": 0, "saldo_atual": 0.0, "valor_estoque": 0.0})
    for item in linhas:
        chave = item.get("categoria") or "Sem categoria"
        grupos[chave]["grupo"] = chave
        grupos[chave]["itens"] += 1
        grupos[chave]["lotes"] += int(item.get("lotes") or 0)
        grupos[chave]["saldo_atual"] += valor_float(item.get("saldo_atual"))
        grupos[chave]["valor_estoque"] += valor_float(item.get("valor_estoque"), 2)
    return sorted(grupos.values(), key=lambda item: item["valor_estoque"], reverse=True)


def resumo_movimentacoes(linhas):
    return [
        {"rotulo": "Eventos", "valor": len(linhas), "tipo": "inteiro", "unidade": "movimentacoes"},
        {"rotulo": "Quantidade", "valor": round(sum(valor_float(i.get("quantidade")) for i in linhas), 4), "tipo": "decimal", "unidade": "unidades conforme cadastro"},
        {"rotulo": "Valor total", "valor": round(sum(valor_float(i.get("valor_total"), 2) for i in linhas), 2), "tipo": "moeda", "unidade": "R$"},
        {"rotulo": "Com OP", "valor": sum(1 for i in linhas if i.get("op_id")), "tipo": "inteiro", "unidade": "eventos"},
    ]


def resumo_saldos(saldos):
    return [
        {"rotulo": "Itens com saldo", "valor": sum(1 for i in saldos if valor_float(i.get("saldo_atual")) > 0), "tipo": "inteiro", "unidade": "insumos"},
        {"rotulo": "Lotes", "valor": sum(int(i.get("lotes") or 0) for i in saldos), "tipo": "inteiro", "unidade": "lotes"},
        {"rotulo": "Saldo total", "valor": round(sum(valor_float(i.get("saldo_atual")) for i in saldos), 4), "tipo": "decimal", "unidade": "unidades conforme cadastro"},
        {"rotulo": "Valor em estoque", "valor": round(sum(valor_float(i.get("valor_estoque"), 2) for i in saldos), 2), "tipo": "moeda", "unidade": "R$"},
    ]


def montar_resumo_gerencial_almoxarifado(slug, args):
    config = RELATORIOS_ALMOXARIFADO[slug]
    filtros = normalizar_filtros(args, config)
    if config["familia"] == "movimentacoes":
        where_sql, parametros = montar_condicoes_movimentacoes(config, filtros)
        linha = executar_lista(f"""
            SELECT
                COUNT(*) AS eventos,
                COALESCE(SUM(quantidade), 0) AS quantidade,
                COALESCE(SUM(valor_total), 0) AS valor_total,
                COALESCE(SUM(CASE WHEN op_id IS NOT NULL THEN 1 ELSE 0 END), 0) AS com_op
            FROM almoxarifado_movimentacoes m
            JOIN almoxarifado_insumos i ON i.id = m.insumo_id
            WHERE {where_sql}
        """, parametros)[0]
        resumo = [
            {"rotulo": "Eventos", "valor": int(linha.get("eventos") or 0), "tipo": "inteiro", "unidade": "movimentacoes"},
            {"rotulo": "Quantidade", "valor": valor_float(linha.get("quantidade")), "tipo": "decimal", "unidade": "unidades conforme cadastro"},
            {"rotulo": "Valor total", "valor": valor_float(linha.get("valor_total"), 2), "tipo": "moeda", "unidade": "R$"},
            {"rotulo": "Com OP", "valor": int(linha.get("com_op") or 0), "tipo": "inteiro", "unidade": "eventos"},
        ]
        return {"resumo": resumo, "tem_dados": bool(linha.get("eventos"))}

    where_sql, parametros = montar_condicoes_saldo(filtros)
    linha = executar_lista(f"""
        SELECT
            COALESCE(SUM(CASE WHEN saldo_atual > 0 THEN 1 ELSE 0 END), 0) AS itens_com_saldo,
            COALESCE(SUM(lotes), 0) AS lotes,
            COALESCE(SUM(saldo_atual), 0) AS saldo_total,
            COALESCE(SUM(valor_estoque), 0) AS valor_estoque
        FROM (
            SELECT
                i.id AS insumo_id,
                COUNT(l.id) AS lotes,
                COALESCE(SUM(l.quantidade_atual), 0) AS saldo_atual,
                COALESCE(SUM(l.quantidade_atual * l.valor_unitario), 0) AS valor_estoque
            FROM almoxarifado_insumos i
            LEFT JOIN almoxarifado_lotes l ON l.insumo_id = i.id
            WHERE {where_sql}
            GROUP BY i.id, i.descricao, i.categoria, i.unidade, i.ativo
        ) saldos
    """, parametros)[0]
    resumo = [
        {"rotulo": "Itens com saldo", "valor": int(linha.get("itens_com_saldo") or 0), "tipo": "inteiro", "unidade": "insumos"},
        {"rotulo": "Lotes", "valor": int(linha.get("lotes") or 0), "tipo": "inteiro", "unidade": "lotes"},
        {"rotulo": "Saldo total", "valor": valor_float(linha.get("saldo_total")), "tipo": "decimal", "unidade": "unidades conforme cadastro"},
        {"rotulo": "Valor em estoque", "valor": valor_float(linha.get("valor_estoque"), 2), "tipo": "moeda", "unidade": "R$"},
    ]
    return {"resumo": resumo, "tem_dados": bool(linha.get("lotes") or linha.get("itens_com_saldo"))}


def resumo_giro(detalhes):
    return [
        {"rotulo": "Produtos analisados", "valor": len(detalhes), "tipo": "inteiro", "unidade": "produto/unidade"},
        {"rotulo": "Consumo periodo", "valor": round(sum(valor_float(i.get("consumo")) for i in detalhes), 4), "tipo": "decimal", "unidade": "unidades compativeis por linha"},
        {"rotulo": "Sem consumo", "valor": sum(1 for i in detalhes if i.get("situacao") == "Sem consumo"), "tipo": "inteiro", "unidade": "itens"},
        {"rotulo": "Saldo negativo", "valor": sum(1 for i in detalhes if i.get("situacao") == "Saldo negativo"), "tipo": "inteiro", "unidade": "itens"},
        {"rotulo": "Historico insuf.", "valor": sum(1 for i in detalhes if i.get("situacao") == "Historico insuficiente"), "tipo": "inteiro", "unidade": "itens"},
        {"rotulo": "Giro medio", "valor": round(dividir(sum(valor_float(i.get("giro")) for i in detalhes), max(1, len(detalhes))), 4), "tipo": "decimal", "unidade": "media simples"},
    ]


def agrupar_giro(detalhes):
    grupos = defaultdict(lambda: {"grupo": "", "itens": 0, "consumo": 0.0, "saldo_final": 0.0, "estoque_medio": 0.0})
    for item in detalhes:
        grupo = item.get("situacao") or "Sem situacao"
        grupos[grupo]["grupo"] = grupo
        grupos[grupo]["itens"] += 1
        grupos[grupo]["consumo"] += valor_float(item.get("consumo"))
        grupos[grupo]["saldo_final"] += valor_float(item.get("saldo_final"))
        grupos[grupo]["estoque_medio"] += valor_float(item.get("estoque_medio"))
    return sorted(grupos.values(), key=lambda item: item["itens"], reverse=True)


def resumo_fifo(detalhes, lotes_antigos):
    total = len(detalhes)
    verificaveis = sum(1 for i in detalhes if i.get("situacao") in ["Conforme", "Possivel violacao"])
    conformes = sum(1 for i in detalhes if i.get("situacao") == "Conforme")
    violacoes = sum(1 for i in detalhes if i.get("situacao") == "Possivel violacao")
    return [
        {"rotulo": "Consumos analisados", "valor": total, "tipo": "inteiro", "unidade": "eventos"},
        {"rotulo": "Verificaveis", "valor": verificaveis, "tipo": "inteiro", "unidade": "eventos"},
        {"rotulo": "Conformes", "valor": conformes, "tipo": "inteiro", "unidade": "eventos"},
        {"rotulo": "Possiveis violacoes", "valor": violacoes, "tipo": "inteiro", "unidade": "eventos"},
        {"rotulo": "Sem lote", "valor": sum(1 for i in detalhes if i.get("situacao") == "Sem lote"), "tipo": "inteiro", "unidade": "eventos"},
        {"rotulo": "Cobertura verificavel", "valor": round(dividir(verificaveis, total, 4) * 100, 2), "tipo": "percentual", "unidade": "%"},
        {"rotulo": "Conformidade FIFO", "valor": round(dividir(conformes, verificaveis, 4) * 100, 2), "tipo": "percentual", "unidade": "% verificavel"},
        {"rotulo": "Lotes antigos em saldo", "valor": len(lotes_antigos), "tipo": "inteiro", "unidade": "lotes"},
    ]


def agrupar_fifo(detalhes):
    grupos = defaultdict(lambda: {"grupo": "", "eventos": 0, "quantidade": 0.0})
    for item in detalhes:
        grupo = item.get("situacao") or "Sem situacao"
        grupos[grupo]["grupo"] = grupo
        grupos[grupo]["eventos"] += 1
        grupos[grupo]["quantidade"] += valor_float(item.get("quantidade"))
    return sorted(grupos.values(), key=lambda item: item["eventos"], reverse=True)


def buscar_opcoes_filtro():
    categorias = [item["valor"] for item in executar_lista("""
        SELECT DISTINCT categoria AS valor
        FROM almoxarifado_insumos
        WHERE COALESCE(categoria, '') <> ''
        ORDER BY valor
    """)]
    status_lotes = [item["valor"] for item in executar_lista("""
        SELECT DISTINCT COALESCE(status, 'Aberto') AS valor
        FROM almoxarifado_lotes
        WHERE COALESCE(status, '') <> ''
        ORDER BY valor
    """)]
    return {
        "categorias": categorias,
        "tipos": ["ENTRADA", "SAIDA", "SAIDA_OP", "ESTORNO_OP", "AJUSTE"],
        "status_lotes": status_lotes,
        "situacoes_giro": [
            "Sem consumo",
            "Giro baixo",
            "Giro intermediario",
            "Giro alto",
            "Saldo zerado",
            "Saldo negativo",
            "Historico insuficiente",
        ],
        "status_fifo": [
            "Conforme",
            "Possivel violacao",
            "Nao verificavel",
            "Sem lote",
        ],
    }


def montar_contexto_relatorio_almoxarifado(slug, args):
    config = RELATORIOS_ALMOXARIFADO[slug]
    filtros = normalizar_filtros(args, config)
    limitacoes = []
    agrupamentos = []
    detalhes = []
    lotes = []
    lotes_antigos = []

    if config["familia"] == "movimentacoes":
        detalhes = buscar_movimentacoes(config, filtros)
        agrupamentos = agrupar_por_tipo(detalhes)
        resumo = resumo_movimentacoes(detalhes)
        if slug == "consumo":
            limitacoes.append("Consumo considera SAIDA e SAIDA_OP; ESTORNO_OP nao entra como consumo.")
        if not detalhes:
            limitacoes.append("Nenhuma movimentacao encontrada para os filtros selecionados.")
    else:
        if config["familia"] == "giro":
            detalhes = buscar_giro_estoque(filtros)
            resumo = resumo_giro(detalhes)
            agrupamentos = agrupar_giro(detalhes)
            limitacoes.append("Cenario B: giro usa somente eventos oficiais registrados em almoxarifado_movimentacoes.")
            limitacoes.append("Estoque medio = media dos saldos diarios reconstruidos por produto/unidade; unidades diferentes nao sao somadas.")
            limitacoes.append("Consumo considera exatamente SAIDA e SAIDA_OP, alinhado ao relatorio oficial de Consumo.")
            limitacoes.append("Cobertura nao e previsao de demanda nem prazo garantido de estoque.")
            if not detalhes:
                limitacoes.append("Nenhum produto encontrado para os filtros selecionados.")
            return {
                "slug": slug,
                "config": config,
                "filtros": filtros,
                "opcoes": buscar_opcoes_filtro(),
                "resumo": resumo,
                "agrupamentos": agrupamentos,
                "detalhes": detalhes,
                "lotes": [],
                "lotes_antigos": [],
                "limitacoes": limitacoes,
                "query_string": urlencode({k: v for k, v in filtros.items() if v not in ["", "Todos", "Todas", "Nao"]}),
            }

        if config["familia"] == "fifo":
            detalhes, lotes_antigos = buscar_fifo_analitico(filtros)
            resumo = resumo_fifo(detalhes, lotes_antigos)
            agrupamentos = agrupar_fifo(detalhes)
            limitacoes.append("Cenario B: FIFO e verificavel somente para consumos com lote_id e historico suficiente.")
            limitacoes.append("Regra auditada: lote mais antigo por data_entrada; desempate por ID do lote.")
            limitacoes.append("Eventos nao verificaveis nao entram como violacao nem como conformidade.")
            limitacoes.append("Nao ha validade oficial no Almoxarifado; o relatorio nao executa FEFO.")
            if not detalhes:
                limitacoes.append("Nenhum consumo encontrado para os filtros selecionados.")
            return {
                "slug": slug,
                "config": config,
                "filtros": filtros,
                "opcoes": buscar_opcoes_filtro(),
                "resumo": resumo,
                "agrupamentos": agrupamentos,
                "detalhes": detalhes,
                "lotes": [],
                "lotes_antigos": lotes_antigos,
                "limitacoes": limitacoes,
                "query_string": urlencode({k: v for k, v in filtros.items() if v not in ["", "Todos", "Todas", "Nao"]}),
            }

        detalhes = buscar_saldos(filtros)
        lotes = buscar_lotes(filtros)
        resumo = resumo_saldos(detalhes)
        agrupamentos = agrupar_por_categoria(detalhes)
        if not detalhes:
            limitacoes.append("Nenhum saldo encontrado para os filtros selecionados.")
        limitacoes.append("Saldo oficial: soma de almoxarifado_lotes.quantidade_atual.")

    if slug in ["estoque-atual", "estoque-por-produto"]:
        limitacoes.append("Valor em estoque usa quantidade atual do lote x valor unitario historico; nao e CMV nem custo medio.")

    return {
        "slug": slug,
        "config": config,
        "filtros": filtros,
        "opcoes": buscar_opcoes_filtro(),
        "resumo": resumo,
        "agrupamentos": agrupamentos,
        "detalhes": detalhes,
        "lotes": lotes,
        "limitacoes": limitacoes,
        "query_string": urlencode({k: v for k, v in filtros.items() if v not in ["", "Todos", "Todas", "Nao"]}),
    }


def gerar_excel_relatorio_almoxarifado(contexto):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    ws.append([contexto["config"]["titulo"]])
    ws.append([contexto["config"]["objetivo"]])
    ws.append([])
    ws.append(["Filtros"])
    for chave, valor in contexto["filtros"].items():
        ws.append([chave, valor])
    ws.append([])
    ws.append(["Indicador", "Valor", "Unidade"])
    for item in contexto["resumo"]:
        ws.append([item["rotulo"], item["valor"], item["unidade"]])
    ws.append([])

    if contexto["agrupamentos"]:
        ws.append(["Agrupamentos"])
        chaves = list(contexto["agrupamentos"][0].keys())
        ws.append(chaves)
        for item in contexto["agrupamentos"]:
            ws.append([item.get(chave, "") for chave in chaves])
        ws.append([])

    ws.append(["Detalhes"])
    familia = contexto["config"]["familia"]
    if familia == "movimentacoes":
        colunas = ["id", "data_movimentacao", "tipo", "insumo", "categoria", "unidade", "quantidade", "valor_unitario", "valor_total", "fornecedor", "numero_nf", "lote", "origem", "op_id", "observacoes", "criado_em"]
        ws.append(colunas)
        for item in contexto["detalhes"]:
            ws.append([item.get(coluna, "") for coluna in colunas])
    elif familia == "giro":
        colunas = [
            "insumo_id", "insumo", "categoria", "unidade", "saldo_inicial",
            "entradas", "consumo", "outras_saidas", "ajustes", "estornos",
            "saldo_final", "saldo_atual", "saldo_reconciliado", "estoque_medio",
            "giro", "consumo_medio_diario", "cobertura_dias",
            "ultima_entrada", "ultimo_consumo", "dias_sem_consumo",
            "ultima_movimentacao", "dias_sem_movimentacao", "situacao",
        ]
        ws.append(colunas)
        for item in contexto["detalhes"]:
            ws.append([item.get(coluna, "") for coluna in colunas])
    elif familia == "fifo":
        colunas = [
            "id", "data_movimentacao", "tipo", "insumo", "categoria",
            "unidade", "op_id", "quantidade", "lote_consumido_id",
            "lote_consumido", "entrada_lote_consumido", "lote_elegivel_id",
            "lote_elegivel", "entrada_lote_elegivel",
            "saldo_lote_elegivel_antes", "situacao", "justificativa",
            "evento_original",
        ]
        ws.append(colunas)
        for item in contexto["detalhes"]:
            ws.append([item.get(coluna, "") for coluna in colunas])
        ws.append([])
        ws.append(["Lotes antigos em estoque"])
        colunas_lote = ["insumo", "categoria", "unidade", "lote", "data_entrada", "idade_dias", "saldo", "status", "fornecedor", "numero_nf"]
        ws.append(colunas_lote)
        for item in contexto.get("lotes_antigos", []):
            ws.append([item.get(coluna, "") for coluna in colunas_lote])
    else:
        colunas = ["insumo_id", "insumo", "categoria", "unidade", "ativo", "lotes", "saldo_atual", "valor_estoque", "primeira_entrada", "ultima_entrada"]
        ws.append(colunas)
        for item in contexto["detalhes"]:
            ws.append([item.get(coluna, "") for coluna in colunas])
        ws.append([])
        ws.append(["Lotes"])
        colunas_lote = ["id", "insumo", "categoria", "unidade", "data_entrada", "lote", "fornecedor", "numero_nf", "quantidade_inicial", "quantidade_atual", "valor_unitario", "valor_total", "status", "criado_em"]
        ws.append(colunas_lote)
        for item in contexto["lotes"]:
            ws.append([item.get(coluna, "") for coluna in colunas_lote])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3B4D")
    for coluna in range(1, 18):
        ws.column_dimensions[get_column_letter(coluna)].width = 18
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo

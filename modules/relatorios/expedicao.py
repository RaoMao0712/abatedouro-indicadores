"""Servicos compartilhados para relatorios oficiais de expedicao."""

from collections import defaultdict
from datetime import date
from io import BytesIO
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database import conectar, q
from modules.expedicao.consolidado_estoque import consolidar_estoque_camara


LOCAL_ABATEDOURO = "Abatedouro"
LOCAL_CAMARA_FRIA = "Camara Fria LSM"


RELATORIOS_EXPEDICAO = {
    "transferencias": {
        "catalogo_id": "expedicao-transferencias",
        "titulo": "Transferencias",
        "objetivo": "Consultar transferencias fisicas realizadas entre Abatedouro e Camara Fria LSM.",
        "familia": "transferencias",
        "excel": True,
    },
    "estoque-camara-fria": {
        "catalogo_id": "expedicao-estoque-camara-fria",
        "titulo": "Estoque Camara Fria",
        "objetivo": "Exibir caixas cujo local atual oficial e a Camara Fria LSM.",
        "familia": "estoque",
        "excel": True,
    },
    "historico-por-caixa": {
        "catalogo_id": "expedicao-historico-caixa",
        "titulo": "Historico por Caixa",
        "objetivo": "Consultar a linha do tempo fisica de uma caixa de Produto Acabado.",
        "familia": "historico",
        "excel": False,
    },
}


def hoje_iso():
    return date.today().isoformat()


def primeiro_dia_mes():
    hoje = date.today()
    return hoje.replace(day=1).isoformat()


def valor_float(valor, casas=3):
    try:
        return round(float(valor or 0), casas)
    except (TypeError, ValueError):
        return 0.0


def valor_int(valor):
    try:
        return int(float(valor or 0))
    except (TypeError, ValueError):
        return 0


def executar_lista(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), tuple(parametros))
    linhas = cursor.fetchall()
    conn.close()
    return [dict(linha) for linha in linhas]


def normalizar_filtros(args, config):
    return {
        "data_inicio": args.get("data_inicio") or primeiro_dia_mes(),
        "data_fim": args.get("data_fim") or hoje_iso(),
        "caixa": args.get("caixa") or "",
        "op_id": args.get("op_id") or "",
        "sku": args.get("sku") or "Todos",
        "apresentacao": args.get("apresentacao") or "",
        "lote": args.get("lote") or "",
        "origem": args.get("origem") or "Todos",
        "destino": args.get("destino") or "Todos",
        "local_atual": args.get("local_atual") or "Todos",
        "validade_inicio": args.get("validade_inicio") or "",
        "validade_fim": args.get("validade_fim") or "",
        "status": args.get("status") or "Todos",
        "usuario": args.get("usuario") or "",
        "romaneio": args.get("romaneio") or "",
        "por_pagina": 300,
    }


def linhas_estoque_oficial(fotografia, filtros):
    """Projeta a fotografia oficial sem recalcular qualquer regra de estoque."""
    linhas = []
    sku_filtro = str(filtros.get("sku") or "Todos").strip().casefold()
    apresentacao_filtro = str(filtros.get("apresentacao") or "").strip().casefold()
    status_filtro = str(filtros.get("status") or "Todos").strip().casefold()
    origem_filtro = str(filtros.get("origem") or "Todos").strip().casefold()
    for grupo in fotografia.get("grupos", []):
        identidades = {
            str(grupo.get("sku_codigo") or "").strip().casefold(),
            str(grupo.get("produto") or "").strip().casefold(),
        }
        if sku_filtro != "todos" and sku_filtro not in identidades:
            continue
        if apresentacao_filtro and apresentacao_filtro not in str(grupo.get("apresentacao") or "").casefold():
            continue
        for situacao, dados in grupo.get("situacoes", {}).items():
            if status_filtro != "todos" and status_filtro != str(situacao).casefold():
                continue
            origens = list(dados.get("origens") or [])
            if origem_filtro != "todos" and origem_filtro not in {str(item).casefold() for item in origens}:
                continue
            quantidades = dados.get("quantidades") or {}
            if not any(quantidades.get(unidade, 0) for unidade in grupo.get("unidades", [])):
                continue
            linhas.append({
                "chave": grupo.get("chave"),
                "sku": grupo.get("sku_codigo") or "Não classificado",
                "produto": grupo.get("produto") or "Produto não classificado",
                "apresentacao": grupo.get("apresentacao") or "Não identificada",
                "classificado": bool(grupo.get("classificado")),
                "situacao": situacao,
                "situacao_rotulo": dados.get("rotulo") or situacao,
                "origens": ", ".join(origens) or "Não identificada",
                "caixas": quantidades.get("caixas", 0),
                "bandejas": quantidades.get("bandejas", 0),
                "peso_kg": quantidades.get("peso_kg", 0),
                "galinhas": quantidades.get("galinhas", 0),
                "pacotes": quantidades.get("pacotes", 0),
            })
    return linhas


def resumo_estoque_oficial(linhas):
    campos = (
        ("caixas", "Caixas", "caixas", "inteiro"),
        ("bandejas", "Bandejas", "bandejas", "inteiro"),
        ("peso_kg", "Peso físico", "kg", "decimal"),
        ("galinhas", "Galinhas", "aves", "inteiro"),
        ("pacotes", "Pacotes", "pacotes", "inteiro"),
    )
    return [
        {
            "rotulo": rotulo,
            "valor": sum((item.get(campo) or 0) for item in linhas),
            "tipo": tipo,
            "unidade": unidade,
        }
        for campo, rotulo, unidade, tipo in campos
    ]


def agrupar_estoque_oficial(linhas):
    grupos = {}
    for linha in linhas:
        chave = (linha["sku"], linha["produto"], linha["apresentacao"])
        grupo = grupos.setdefault(chave, {
            "sku": chave[0], "produto": chave[1], "apresentacao": chave[2],
            "caixas": 0, "bandejas": 0, "peso_kg": 0,
            "galinhas": 0, "pacotes": 0,
        })
        for campo in ("caixas", "bandejas", "peso_kg", "galinhas", "pacotes"):
            grupo[campo] += linha.get(campo) or 0
    return sorted(grupos.values(), key=lambda item: (item["produto"], item["apresentacao"]))


def reconciliar_estoque_oficial(fotografia, linhas, filtros):
    """Diagnóstico puro: serviço, tela e exportador devem compartilhar estes totais."""
    esperadas = linhas_estoque_oficial(fotografia, filtros)
    campos = ("caixas", "bandejas", "peso_kg", "galinhas", "pacotes")
    esperado = {campo: sum((item.get(campo) or 0) for item in esperadas) for campo in campos}
    exibido = {campo: sum((item.get(campo) or 0) for item in linhas) for campo in campos}
    return {"ok": esperado == exibido, "esperado": esperado, "exibido": exibido}


def montar_cte_caixas():
    return """
    WITH comp AS (
        SELECT
            caixa_id,
            MIN(op_id) AS op_id,
            COUNT(DISTINCT op_id) AS ops_composicao,
            COALESCE(SUM(quantidade_bandejas), 0) AS bandejas_composicao
        FROM pa_caixa_composicao
        GROUP BY caixa_id
    )
    """


def aplicar_filtros_caixa(filtros, alias="cx", comp_alias="comp"):
    condicoes = []
    parametros = []

    if filtros["caixa"]:
        condicoes.append(f"LOWER({alias}.codigo_caixa) LIKE ?")
        parametros.append(f"%{filtros['caixa'].lower()}%")
    if filtros["op_id"]:
        try:
            condicoes.append(f"{comp_alias}.op_id = ?")
            parametros.append(int(filtros["op_id"]))
        except ValueError:
            condicoes.append("1 = 0")
    if filtros["sku"] != "Todos":
        condicoes.append(f"{alias}.sku = ?")
        parametros.append(filtros["sku"])
    if filtros["lote"]:
        condicoes.append(f"""
            (
                LOWER({alias}.codigo_caixa) LIKE ?
                OR LOWER(CAST(COALESCE({comp_alias}.op_id, 0) AS TEXT)) LIKE ?
                OR CAST(COALESCE({comp_alias}.op_id, 0) AS TEXT) LIKE ?
            )
        """)
        termo = f"%{filtros['lote'].lower()}%"
        parametros.extend([termo, termo, termo])
    if filtros["validade_inicio"]:
        condicoes.append(f"COALESCE({alias}.data_validade, '') >= ?")
        parametros.append(filtros["validade_inicio"])
    if filtros["validade_fim"]:
        condicoes.append(f"COALESCE({alias}.data_validade, '') <= ?")
        parametros.append(filtros["validade_fim"])
    if filtros["status"] != "Todos":
        condicoes.append(f"COALESCE({alias}.status, '') = ?")
        parametros.append(filtros["status"])

    return condicoes, parametros


def buscar_transferencias(filtros):
    condicoes = ["mov.tipo = ?", "substr(CAST(mov.criado_em AS TEXT), 1, 10) BETWEEN ? AND ?"]
    parametros = ["TRANSFERENCIA", filtros["data_inicio"], filtros["data_fim"]]

    caixa_cond, caixa_params = aplicar_filtros_caixa(filtros)
    condicoes.extend(caixa_cond)
    parametros.extend(caixa_params)

    if filtros["origem"] != "Todos":
        condicoes.append("COALESCE(origem.nome, '') = ?")
        parametros.append(filtros["origem"])
    if filtros["destino"] != "Todos":
        condicoes.append("COALESCE(destino.nome, '') = ?")
        parametros.append(filtros["destino"])
    if filtros["usuario"]:
        condicoes.append("LOWER(COALESCE(mov.usuario, '')) LIKE ?")
        parametros.append(f"%{filtros['usuario'].lower()}%")
    if filtros["romaneio"]:
        condicoes.append("LOWER(COALESCE(e.numero_romaneio, '')) LIKE ?")
        parametros.append(f"%{filtros['romaneio'].lower()}%")

    where_sql = " AND ".join(condicoes)

    return executar_lista(montar_cte_caixas() + f"""
        SELECT
            mov.id AS evento_id,
            mov.criado_em AS data_evento,
            mov.tipo,
            mov.usuario,
            e.id AS expedicao_id,
            e.numero_romaneio,
            e.data AS data_romaneio,
            e.status AS status_romaneio,
            origem.nome AS origem_nome,
            destino.nome AS destino_nome,
            cx.id AS caixa_id,
            cx.codigo_caixa,
            cx.sku,
            comp.op_id,
            comp.ops_composicao,
            cx.data_fabricacao,
            cx.data_validade,
            cx.peso_bruto,
            cx.peso_liquido,
            cx.quantidade_bandejas,
            cx.status AS status_caixa,
            COALESCE(local_atual.nome, ?) AS local_atual
        FROM pa_movimentacoes mov
        INNER JOIN pa_caixas cx ON cx.id = mov.caixa_id
        LEFT JOIN comp ON comp.caixa_id = cx.id
        LEFT JOIN locais_estoque origem ON origem.id = mov.local_origem_id
        LEFT JOIN locais_estoque destino ON destino.id = mov.local_destino_id
        LEFT JOIN locais_estoque local_atual ON local_atual.id = cx.local_estoque_id
        LEFT JOIN expedicoes e ON e.id = mov.expedicao_id
        WHERE {where_sql}
        ORDER BY mov.criado_em DESC, mov.id DESC
        LIMIT ?
    """, [LOCAL_ABATEDOURO] + parametros + [filtros["por_pagina"]])


def buscar_caixas_para_selecao(filtros):
    condicoes, parametros = aplicar_filtros_caixa(filtros)
    if not condicoes:
        return []
    where_sql = " AND ".join(condicoes)
    return executar_lista(montar_cte_caixas() + f"""
        SELECT
            cx.id AS caixa_id,
            cx.codigo_caixa,
            cx.sku,
            comp.op_id,
            cx.data_fabricacao,
            cx.data_validade,
            cx.peso_liquido,
            cx.status AS status_caixa,
            COALESCE(local_atual.nome, ?) AS local_atual
        FROM pa_caixas cx
        LEFT JOIN comp ON comp.caixa_id = cx.id
        LEFT JOIN locais_estoque local_atual ON local_atual.id = cx.local_estoque_id
        WHERE {where_sql}
        ORDER BY cx.id DESC
        LIMIT ?
    """, [LOCAL_ABATEDOURO] + parametros + [50])


def buscar_caixa_por_id(caixa_id):
    linhas = executar_lista(montar_cte_caixas() + """
        SELECT
            cx.id AS caixa_id,
            cx.codigo_caixa,
            cx.sku,
            comp.op_id,
            comp.ops_composicao,
            comp.bandejas_composicao,
            cx.data_fabricacao,
            cx.data_validade,
            cx.peso_bruto,
            cx.peso_liquido,
            cx.quantidade_bandejas,
            cx.status AS status_caixa,
            cx.origem,
            cx.observacoes,
            cx.criado_em,
            COALESCE(local_atual.nome, ?) AS local_atual
        FROM pa_caixas cx
        LEFT JOIN comp ON comp.caixa_id = cx.id
        LEFT JOIN locais_estoque local_atual ON local_atual.id = cx.local_estoque_id
        WHERE cx.id = ?
    """, (LOCAL_ABATEDOURO, caixa_id))
    return linhas[0] if linhas else None


def buscar_historico_caixa(caixa_id):
    return executar_lista("""
        SELECT
            mov.id AS evento_id,
            mov.criado_em AS data_evento,
            mov.tipo,
            mov.usuario,
            origem.nome AS origem_nome,
            destino.nome AS destino_nome,
            e.id AS expedicao_id,
            e.numero_romaneio,
            e.status AS status_romaneio
        FROM pa_movimentacoes mov
        LEFT JOIN locais_estoque origem ON origem.id = mov.local_origem_id
        LEFT JOIN locais_estoque destino ON destino.id = mov.local_destino_id
        LEFT JOIN expedicoes e ON e.id = mov.expedicao_id
        WHERE mov.caixa_id = ?
        ORDER BY mov.criado_em ASC, mov.id ASC
    """, (caixa_id,))


def adicionar_lote(linha):
    op_id = linha.get("op_id")
    linha["lote"] = f"OP-{int(op_id):05d}" if op_id else ""
    return linha


def resumo_transferencias(linhas):
    caixas = {item["caixa_id"] for item in linhas}
    return [
        {"rotulo": "Transferencias", "valor": len(linhas), "tipo": "inteiro", "unidade": "eventos"},
        {"rotulo": "Caixas", "valor": len(caixas), "tipo": "inteiro", "unidade": "caixas unicas"},
        {"rotulo": "Peso liquido", "valor": sum(valor_float(i.get("peso_liquido")) for i in linhas), "tipo": "decimal", "unidade": "kg"},
        {"rotulo": "Peso bruto", "valor": sum(valor_float(i.get("peso_bruto")) for i in linhas), "tipo": "decimal", "unidade": "kg"},
    ]


def montar_resumo_gerencial_expedicao(slug, args):
    config = RELATORIOS_EXPEDICAO[slug]
    filtros = normalizar_filtros(args, config)
    if slug == "transferencias":
        condicoes = ["mov.tipo = ?", "substr(CAST(mov.criado_em AS TEXT), 1, 10) BETWEEN ? AND ?"]
        parametros = ["TRANSFERENCIA", filtros["data_inicio"], filtros["data_fim"]]
        caixa_cond, caixa_params = aplicar_filtros_caixa(filtros)
        condicoes.extend(caixa_cond)
        parametros.extend(caixa_params)
        if filtros["origem"] != "Todos":
            condicoes.append("COALESCE(origem.nome, '') = ?")
            parametros.append(filtros["origem"])
        if filtros["destino"] != "Todos":
            condicoes.append("COALESCE(destino.nome, '') = ?")
            parametros.append(filtros["destino"])
        if filtros["usuario"]:
            condicoes.append("LOWER(COALESCE(mov.usuario, '')) LIKE ?")
            parametros.append(f"%{filtros['usuario'].lower()}%")
        if filtros["romaneio"]:
            condicoes.append("LOWER(COALESCE(e.numero_romaneio, '')) LIKE ?")
            parametros.append(f"%{filtros['romaneio'].lower()}%")
        linha = executar_lista(montar_cte_caixas() + f"""
            SELECT
                COUNT(*) AS transferencias,
                COUNT(DISTINCT cx.id) AS caixas,
                COALESCE(SUM(cx.peso_liquido), 0) AS peso_liquido,
                COALESCE(SUM(cx.peso_bruto), 0) AS peso_bruto
            FROM pa_movimentacoes mov
            INNER JOIN pa_caixas cx ON cx.id = mov.caixa_id
            LEFT JOIN comp ON comp.caixa_id = cx.id
            LEFT JOIN locais_estoque origem ON origem.id = mov.local_origem_id
            LEFT JOIN locais_estoque destino ON destino.id = mov.local_destino_id
            LEFT JOIN locais_estoque local_atual ON local_atual.id = cx.local_estoque_id
            LEFT JOIN expedicoes e ON e.id = mov.expedicao_id
            WHERE {" AND ".join(condicoes)}
        """, parametros)[0]
        resumo = [
            {"rotulo": "Transferencias", "valor": int(linha.get("transferencias") or 0), "tipo": "inteiro", "unidade": "eventos"},
            {"rotulo": "Caixas", "valor": int(linha.get("caixas") or 0), "tipo": "inteiro", "unidade": "caixas unicas"},
            {"rotulo": "Peso liquido", "valor": valor_float(linha.get("peso_liquido")), "tipo": "decimal", "unidade": "kg"},
            {"rotulo": "Peso bruto", "valor": valor_float(linha.get("peso_bruto")), "tipo": "decimal", "unidade": "kg"},
        ]
        return {"resumo": resumo, "tem_dados": bool(linha.get("transferencias"))}

    fotografia = consolidar_estoque_camara(incluir_nao_conforme=True)
    linhas = linhas_estoque_oficial(fotografia, filtros)
    resumo = resumo_estoque_oficial(linhas)
    return {
        "resumo": resumo,
        "tem_dados": bool(linhas),
        "reconciliacao": reconciliar_estoque_oficial(fotografia, linhas, filtros),
        "fonte_oficial": "modules.expedicao.consolidado_estoque.consolidar_estoque_camara",
    }


def agrupar(linhas, chave, campo_valor="peso_liquido"):
    grupos = defaultdict(lambda: {"grupo": "", "caixas": set(), "eventos": 0, "peso_liquido": 0.0, "peso_bruto": 0.0})
    for item in linhas:
        valor = item.get(chave) or "Nao informado"
        grupos[valor]["grupo"] = valor
        grupos[valor]["caixas"].add(item.get("caixa_id"))
        grupos[valor]["eventos"] += 1
        grupos[valor]["peso_liquido"] += valor_float(item.get("peso_liquido"))
        grupos[valor]["peso_bruto"] += valor_float(item.get("peso_bruto"))
    saida = []
    for item in grupos.values():
        saida.append({
            "grupo": item["grupo"],
            "caixas": len(item["caixas"]),
            "eventos": item["eventos"],
            "peso_liquido": round(item["peso_liquido"], 3),
            "peso_bruto": round(item["peso_bruto"], 3),
        })
    return sorted(saida, key=lambda item: (item["peso_liquido"], item["caixas"]), reverse=True)


def buscar_opcoes_filtro():
    return {
        "skus": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT sku AS valor FROM pa_caixas
            WHERE COALESCE(sku, '') <> ''
            ORDER BY valor
        """)],
        "locais": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT nome AS valor FROM locais_estoque
            WHERE COALESCE(nome, '') <> ''
            ORDER BY valor
        """)],
        "status": [item["valor"] for item in executar_lista("""
            SELECT DISTINCT status AS valor FROM pa_caixas
            WHERE COALESCE(status, '') <> ''
            ORDER BY valor
        """)],
    }


def montar_contexto_relatorio_expedicao(slug, args):
    config = RELATORIOS_EXPEDICAO[slug]
    filtros = normalizar_filtros(args, config)
    limitacoes = []
    detalhes = []
    agrupamentos = []
    caixa = None
    historico = []
    sugestoes = []
    opcoes = {} if slug == "estoque-camara-fria" else buscar_opcoes_filtro()

    if slug == "transferencias":
        detalhes = [adicionar_lote(item) for item in buscar_transferencias(filtros)]
        resumo = resumo_transferencias(detalhes)
        agrupamentos = agrupar(detalhes, "sku")
        if not detalhes:
            limitacoes.append("Nenhuma transferencia encontrada para os filtros selecionados.")
        limitacoes.append("Transferencias sao eventos logisticos: nao geram financeiro, DRE ou CMV.")
    elif slug == "estoque-camara-fria":
        fotografia = consolidar_estoque_camara(incluir_nao_conforme=True)
        detalhes = linhas_estoque_oficial(fotografia, filtros)
        resumo = resumo_estoque_oficial(detalhes)
        agrupamentos = agrupar_estoque_oficial(detalhes)
        reconciliacao = reconciliar_estoque_oficial(fotografia, detalhes, filtros)
        opcoes = {
            **opcoes,
            "skus": sorted({grupo.get("sku_codigo") for grupo in fotografia["grupos"] if grupo.get("sku_codigo")}),
            "status": [
                {"valor": "disponivel", "rotulo": "Disponível para expedição"},
                {"valor": "reservado", "rotulo": "Reservado"},
                {"valor": "nao_conforme_bloqueado", "rotulo": "Não conforme bloqueado"},
                {"valor": "reprocessamento", "rotulo": "Em reprocessamento"},
                {"valor": "aguardando_liberacao", "rotulo": "Aguardando liberação"},
            ],
            "origens": sorted({
                origem
                for grupo in fotografia["grupos"]
                for dados in grupo["situacoes"].values()
                for origem in dados.get("origens", [])
            }),
        }
        if not detalhes:
            limitacoes.append("Nenhuma posição física encontrada para os filtros selecionados.")
        limitacoes.append(
            "Posição atual: mesma fotografia oficial da Câmara, com legado, pós-marco-zero, reservas e PNC remanescente."
        )
        limitacoes.append(
            "Físico = disponível + reservado + bloqueado; documentos históricos não são somados como estoque."
        )
    else:
        sugestoes = [adicionar_lote(item) for item in buscar_caixas_para_selecao(filtros)]
        caixa_id = args.get("caixa_id") or ""
        if caixa_id:
            try:
                caixa = buscar_caixa_por_id(int(caixa_id))
            except ValueError:
                caixa = None
            if caixa:
                caixa = adicionar_lote(caixa)
                historico = buscar_historico_caixa(caixa["caixa_id"])
        elif len(sugestoes) == 1:
            caixa = buscar_caixa_por_id(sugestoes[0]["caixa_id"])
            if caixa:
                caixa = adicionar_lote(caixa)
                historico = buscar_historico_caixa(caixa["caixa_id"])

        detalhes = sugestoes
        resumo = [
            {"rotulo": "Caixas encontradas", "valor": len(sugestoes), "tipo": "inteiro", "unidade": "caixas"},
            {"rotulo": "Eventos historicos", "valor": len(historico), "tipo": "inteiro", "unidade": "eventos"},
            {"rotulo": "Peso liquido", "valor": valor_float(caixa.get("peso_liquido") if caixa else 0), "tipo": "decimal", "unidade": "kg"},
            {"rotulo": "OP", "valor": valor_int(caixa.get("op_id") if caixa else 0), "tipo": "inteiro", "unidade": "origem"},
        ]
        limitacoes.append("Historico por Caixa nao inclui NF, cliente ou venda; rastreabilidade completa permanece futura.")
        limitacoes.append("Formato disponivel: tela e impressao do navegador. PDF dedicado nao foi declarado.")

    contexto = {
        "slug": slug,
        "config": config,
        "filtros": filtros,
        "opcoes": opcoes,
        "resumo": resumo,
        "agrupamentos": agrupamentos,
        "detalhes": detalhes,
        "caixa": caixa,
        "historico": historico,
        "limitacoes": limitacoes,
        "query_string": urlencode({
            k: v for k, v in filtros.items()
            if k != "por_pagina" and v not in ["", "Todos", "Todas"]
        }),
    }
    if slug == "estoque-camara-fria":
        contexto.update({
            "fotografia_estoque": fotografia,
            "reconciliacao": reconciliacao,
            "fonte_oficial": "modules.expedicao.consolidado_estoque.consolidar_estoque_camara",
        })
    return contexto


def gerar_excel_relatorio_expedicao(contexto):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    ws.append([contexto["config"]["titulo"]])
    ws.append([contexto["config"]["objetivo"]])
    ws.append([])
    ws.append(["Filtros"])
    for chave, valor in contexto["filtros"].items():
        ws.append([chave, valor])
    ws.append([])
    ws.append(["Indicador", "Valor", "Unidade"])
    for item in contexto["resumo"]:
        ws.append([item["rotulo"], item["valor"], item["unidade"]])
    ws.append([])

    if contexto["agrupamentos"]:
        ws.append(["Agrupamentos"])
        chaves = list(contexto["agrupamentos"][0].keys())
        ws.append(chaves)
        for item in contexto["agrupamentos"]:
            ws.append([item.get(chave, "") for chave in chaves])
        ws.append([])

    if contexto["slug"] == "transferencias":
        colunas = [
            "evento_id", "data_evento", "numero_romaneio", "status_romaneio",
            "origem_nome", "destino_nome", "codigo_caixa", "op_id", "sku", "lote",
            "peso_liquido", "peso_bruto", "data_validade", "usuario", "local_atual",
        ]
    elif contexto["slug"] == "estoque-camara-fria":
        colunas = [
            "sku", "produto", "apresentacao", "situacao_rotulo", "origens",
            "caixas", "bandejas", "peso_kg", "galinhas", "pacotes",
        ]
    else:
        colunas = [
            "caixa_id", "codigo_caixa", "op_id", "sku", "lote", "data_fabricacao",
            "data_validade", "idade_dias", "peso_liquido", "peso_bruto",
            "quantidade_bandejas", "status_caixa", "local_atual", "numero_romaneio",
            "data_transferencia", "usuario_transferencia",
        ]

    ws.append(["Detalhes"])
    ws.append(colunas)
    for item in contexto["detalhes"]:
        ws.append([item.get(coluna, "") for coluna in colunas])

    if contexto["slug"] == "estoque-camara-fria":
        cabecalho = ws.max_row - len(contexto["detalhes"])
        for indice, coluna in enumerate(colunas, start=1):
            for celula in ws.iter_rows(min_row=cabecalho + 1, max_row=ws.max_row,
                                       min_col=indice, max_col=indice):
                celula[0].number_format = "0.000" if coluna == "peso_kg" else "0"

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3B4D")
    for coluna in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(coluna)].width = 18
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo

"""Camada gerencial oficial: indicadores, comparativos e tendencias."""

from calendar import monthrange
from datetime import date, timedelta
from io import BytesIO
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.dre.services import buscar_resumo_dre_gerencial
from modules.fluxo_caixa.services import montar_resumo_gerencial_fluxo_caixa
from modules.relatorios.almoxarifado import montar_resumo_gerencial_almoxarifado
from modules.relatorios.expedicao import montar_resumo_gerencial_expedicao
from modules.relatorios.financeiro import montar_resumo_gerencial_financeiro
from modules.relatorios.producao import montar_resumo_gerencial_producao, montar_resumos_gerenciais_producao_periodos


STATUS_DISPONIVEL = "Disponivel"
STATUS_EVOLUCAO = "Disponivel - requer evolucao"
STATUS_SEM_DADOS = "Sem dados no periodo"
STATUS_ESTRUTURACAO = "Em estruturacao"
STATUS_CONGELADO = "Congelado"
STATUS_FUTURO = "Futuro"

TOLERANCIA_ESTAVEL = 0.02

RELATORIOS_GERENCIAIS = {
    "indicadores": {"titulo": "Indicadores"},
    "comparativos": {"titulo": "Comparativos"},
    "tendencias": {"titulo": "Tendencias"},
}


class ArgsDict(dict):
    def get(self, chave, padrao=None):
        return super().get(chave, padrao)


def hoje_iso():
    return date.today().isoformat()


def primeiro_dia_mes():
    return date.today().replace(day=1).isoformat()


def parse_data(texto):
    try:
        return date.fromisoformat(str(texto or "")[:10])
    except ValueError:
        return None


def periodo_meses(data_inicio, data_fim):
    inicio = parse_data(data_inicio)
    fim = parse_data(data_fim)
    if not inicio or not fim or fim < inicio:
        return []
    atual = inicio.replace(day=1)
    limite = fim.replace(day=1)
    meses = []
    while atual <= limite:
        meses.append(atual.strftime("%Y-%m"))
        if atual.month == 12:
            atual = atual.replace(year=atual.year + 1, month=1)
        else:
            atual = atual.replace(month=atual.month + 1)
    return meses


def ultimo_dia_mes(data):
    return monthrange(data.year, data.month)[1]


def adicionar_meses(data, meses):
    mes_base = data.month - 1 + meses
    ano = data.year + mes_base // 12
    mes = mes_base % 12 + 1
    dia = min(data.day, ultimo_dia_mes(date(ano, mes, 1)))
    return date(ano, mes, dia)


def periodo_mensal_completo(inicio, fim):
    return inicio.day == 1 and fim.day == ultimo_dia_mes(fim)


def deslocar_periodo_anterior(data_inicio, data_fim):
    inicio = parse_data(data_inicio)
    fim = parse_data(data_fim)
    if not inicio or not fim or fim < inicio:
        return data_inicio, data_fim
    if periodo_mensal_completo(inicio, fim):
        quantidade_meses = (fim.year - inicio.year) * 12 + fim.month - inicio.month + 1
        inicio_anterior = adicionar_meses(inicio, -quantidade_meses)
        fim_anterior_base = adicionar_meses(inicio, -1)
        fim_anterior = fim_anterior_base.replace(day=ultimo_dia_mes(fim_anterior_base))
        return inicio_anterior.isoformat(), fim_anterior.isoformat()
    dias = (fim - inicio).days + 1
    fim_anterior = inicio - timedelta(days=1)
    inicio_anterior = fim_anterior - timedelta(days=dias - 1)
    return inicio_anterior.isoformat(), fim_anterior.isoformat()


def periodos_serie(data_inicio, data_fim, granularidade):
    inicio = parse_data(data_inicio)
    fim = parse_data(data_fim)
    if not inicio or not fim or fim < inicio:
        return []
    periodos = []
    if granularidade == "dia":
        atual = inicio
        while atual <= fim:
            periodos.append((atual.isoformat(), atual.isoformat(), atual.isoformat()))
            atual += timedelta(days=1)
        return periodos
    if granularidade == "semana":
        atual = inicio
        while atual <= fim:
            fim_semana = min(fim, atual + timedelta(days=6))
            ano, semana, _ = atual.isocalendar()
            periodos.append((f"{ano}-W{semana:02d}", atual.isoformat(), fim_semana.isoformat()))
            atual = fim_semana + timedelta(days=1)
        return periodos

    atual = inicio.replace(day=1)
    while atual <= fim:
        prox = atual.replace(year=atual.year + 1, month=1) if atual.month == 12 else atual.replace(month=atual.month + 1)
        fim_mes = min(fim, prox - timedelta(days=1))
        inicio_mes = max(inicio, atual)
        periodos.append((atual.strftime("%Y-%m"), inicio_mes.isoformat(), fim_mes.isoformat()))
        atual = prox
    return periodos


def valor_numero(valor):
    try:
        return round(float(valor or 0), 4)
    except (TypeError, ValueError):
        return None


def arg_periodo(data_inicio, data_fim, extra=None):
    args = ArgsDict({"data_inicio": data_inicio, "data_fim": data_fim})
    if extra:
        args.update(extra)
    return args


def valor_resumo(contexto, rotulo):
    for item in contexto.get("resumo", []):
        if item.get("rotulo") == rotulo:
            return valor_numero(item.get("valor"))
    return None


def contexto_tem_linhas(contexto):
    if contexto.get("tem_dados"):
        return True
    for chave in ["detalhes", "agrupamentos", "evolucao", "linha_tempo", "movimentacoes"]:
        if contexto.get(chave):
            return True
    return False


def obter_cache(cache, chave, factory):
    if cache is None:
        return factory()
    if chave not in cache:
        cache[chave] = factory()
    return cache[chave]


def status_por_valor(valor, status_base=STATUS_DISPONIVEL):
    if valor is None:
        return STATUS_SEM_DADOS
    return status_base


def indicador_resultado(definicao, data_inicio, data_fim, valor, status=None, limitacao=""):
    return {
        **definicao,
        "valor": valor,
        "status_dados": status or status_por_valor(valor, definicao.get("status", STATUS_DISPONIVEL)),
        "periodo": f"{data_inicio} a {data_fim}",
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "limitacao_resultado": limitacao or definicao.get("limitacoes", ""),
    }


def resolver_dre(definicao, data_inicio, data_fim, cache=None):
    campo = definicao["campo_origem"]
    total = 0.0
    tem_base = False
    meses = periodo_meses(data_inicio, data_fim)
    if not meses:
        return indicador_resultado(definicao, data_inicio, data_fim, None)
    for competencia in meses:
        dados = obter_cache(cache, ("dre", competencia), lambda c=competencia: buscar_resumo_dre_gerencial(c))
        total += float(dados.get(campo) or 0)
        tem_base = tem_base or any(
            float(dados.get(chave) or 0) != 0
            for chave in [
                "receita_bruta",
                "deducoes_receita",
                "custos_operacionais_total",
                "resultado_nao_operacional",
                "resultado_gerencial_periodo",
            ]
        ) or bool(dados.get("linhas_custos"))
    valor = round(total, 2) if tem_base or total != 0 else None
    return indicador_resultado(definicao, data_inicio, data_fim, valor, limitacao="DRE consolidada por competencia mensal.")


def resolver_fluxo(definicao, data_inicio, data_fim, cache=None):
    contexto = obter_cache(
        cache,
        ("fluxo", data_inicio, data_fim),
        lambda: montar_resumo_gerencial_fluxo_caixa(arg_periodo(data_inicio, data_fim)),
    )
    valor = valor_numero(contexto.get("resumo", {}).get(definicao["campo_origem"]))
    if valor == 0 and not contexto_tem_linhas(contexto):
        valor = None
    return indicador_resultado(definicao, data_inicio, data_fim, valor)


def resolver_financeiro(definicao, data_inicio, data_fim, cache=None):
    contexto = obter_cache(
        cache,
        ("financeiro", definicao["slug_origem"], data_inicio, data_fim),
        lambda: montar_resumo_gerencial_financeiro(definicao["slug_origem"], arg_periodo(data_inicio, data_fim)),
    )
    valor = valor_resumo(contexto, definicao["campo_origem"])
    if valor == 0 and not contexto_tem_linhas(contexto):
        valor = None
    return indicador_resultado(definicao, data_inicio, data_fim, valor)


def resolver_producao(definicao, data_inicio, data_fim, cache=None):
    contexto = obter_cache(
        cache,
        ("producao", definicao["slug_origem"], data_inicio, data_fim),
        lambda: montar_resumo_gerencial_producao(definicao["slug_origem"], arg_periodo(data_inicio, data_fim)),
    )
    valor = valor_numero(contexto.get("totais", {}).get(definicao["campo_origem"]))
    if valor == 0 and not contexto_tem_linhas(contexto):
        valor = None
    status = definicao.get("status", STATUS_DISPONIVEL)
    return indicador_resultado(definicao, data_inicio, data_fim, valor, status=status)


def resolver_producao_contexto(definicao, data_inicio, data_fim, contexto):
    valor = valor_numero((contexto or {}).get("totais", {}).get(definicao["campo_origem"]))
    if valor == 0 and not contexto_tem_linhas(contexto or {}):
        valor = None
    return indicador_resultado(definicao, data_inicio, data_fim, valor, status=definicao.get("status", STATUS_DISPONIVEL))


def resolver_producao_resumo(definicao, data_inicio, data_fim, cache=None):
    contexto = obter_cache(
        cache,
        ("producao", definicao["slug_origem"], data_inicio, data_fim),
        lambda: montar_resumo_gerencial_producao(definicao["slug_origem"], arg_periodo(data_inicio, data_fim)),
    )
    valor = valor_resumo(contexto, definicao["campo_origem"])
    if valor == 0 and not contexto_tem_linhas(contexto):
        valor = None
    return indicador_resultado(definicao, data_inicio, data_fim, valor, status=definicao.get("status"))


def resolver_almoxarifado(definicao, data_inicio, data_fim, cache=None):
    contexto = obter_cache(
        cache,
        ("almoxarifado", definicao["slug_origem"], data_inicio, data_fim),
        lambda: montar_resumo_gerencial_almoxarifado(definicao["slug_origem"], arg_periodo(data_inicio, data_fim)),
    )
    valor = valor_resumo(contexto, definicao["campo_origem"])
    if valor == 0 and not contexto_tem_linhas(contexto):
        valor = None
    return indicador_resultado(definicao, data_inicio, data_fim, valor)


def resolver_expedicao(definicao, data_inicio, data_fim, cache=None):
    contexto = obter_cache(
        cache,
        ("expedicao", definicao["slug_origem"], data_inicio, data_fim),
        lambda: montar_resumo_gerencial_expedicao(definicao["slug_origem"], arg_periodo(data_inicio, data_fim)),
    )
    valor = valor_resumo(contexto, definicao["campo_origem"])
    if valor == 0 and not contexto_tem_linhas(contexto):
        valor = None
    return indicador_resultado(definicao, data_inicio, data_fim, valor)


RESOLVERS = {
    "dre": resolver_dre,
    "fluxo": resolver_fluxo,
    "financeiro": resolver_financeiro,
    "producao": resolver_producao,
    "producao_resumo": resolver_producao_resumo,
    "almoxarifado": resolver_almoxarifado,
    "expedicao": resolver_expedicao,
}


REGISTRO_INDICADORES = [
    {"id": "fin_receita_bruta", "nome": "Receita Bruta", "dominio": "Financeiro", "descricao": "Receita bruta pela DRE gerencial.", "origem": "DRE Gerencial", "tipo_origem": "dre", "campo_origem": "receita_bruta", "unidade": "R$", "referencia_temporal": "Competencia", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie_mensal", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "fin_deducoes", "nome": "Deducoes", "dominio": "Financeiro", "descricao": "Deducoes oficiais da receita.", "origem": "DRE Gerencial", "tipo_origem": "dre", "campo_origem": "deducoes_receita", "unidade": "R$", "referencia_temporal": "Competencia", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie_mensal", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "fin_receita_liquida", "nome": "Receita Operacional Liquida", "dominio": "Financeiro", "descricao": "Receita bruta menos deducoes.", "origem": "DRE Gerencial", "tipo_origem": "dre", "campo_origem": "receita_operacional_liquida", "unidade": "R$", "referencia_temporal": "Competencia", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie_mensal", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "fin_despesas_operacionais", "nome": "Despesas Operacionais", "dominio": "Financeiro", "descricao": "Custos/despesas operacionais da DRE.", "origem": "DRE Gerencial", "tipo_origem": "dre", "campo_origem": "custos_operacionais_total", "unidade": "R$", "referencia_temporal": "Competencia", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie_mensal", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "fin_resultado_operacional", "nome": "Resultado Operacional", "dominio": "Financeiro", "descricao": "Resultado operacional oficial da DRE.", "origem": "DRE Gerencial", "tipo_origem": "dre", "campo_origem": "resultado_operacional", "unidade": "R$", "referencia_temporal": "Competencia", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie_mensal", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "fin_resultado_nao_operacional", "nome": "Resultado Nao Operacional", "dominio": "Financeiro", "descricao": "Resultado nao operacional oficial.", "origem": "DRE Gerencial", "tipo_origem": "dre", "campo_origem": "resultado_nao_operacional", "unidade": "R$", "referencia_temporal": "Competencia", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie_mensal", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "fin_resultado_liquido", "nome": "Resultado Liquido Gerencial", "dominio": "Financeiro", "descricao": "Resultado operacional mais nao operacional.", "origem": "DRE Gerencial", "tipo_origem": "dre", "campo_origem": "resultado_gerencial_periodo", "unidade": "R$", "referencia_temporal": "Competencia", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie_mensal", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "fin_entradas_caixa", "nome": "Entradas de Caixa", "dominio": "Financeiro", "descricao": "Entradas realizadas no fluxo de caixa.", "origem": "Fluxo de Caixa", "tipo_origem": "fluxo", "campo_origem": "entradas_realizadas", "unidade": "R$", "referencia_temporal": "Realizacao", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "fin_saidas_caixa", "nome": "Saidas de Caixa", "dominio": "Financeiro", "descricao": "Saidas realizadas no fluxo de caixa.", "origem": "Fluxo de Caixa", "tipo_origem": "fluxo", "campo_origem": "saidas_realizadas", "unidade": "R$", "referencia_temporal": "Realizacao", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "fin_saldo_caixa", "nome": "Saldo de Caixa", "dominio": "Financeiro", "descricao": "Saldo realizado do fluxo de caixa no periodo.", "origem": "Fluxo de Caixa", "tipo_origem": "fluxo", "campo_origem": "saldo_realizado", "unidade": "R$", "referencia_temporal": "Realizacao", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "fin_aportes", "nome": "Aportes", "dominio": "Financeiro", "descricao": "Aportes que impactam caixa e nao compoem DRE.", "origem": "Aportes", "tipo_origem": "financeiro", "slug_origem": "aportes", "campo_origem": "Total previsto", "unidade": "R$", "referencia_temporal": "Vencimento", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "prod_ops", "nome": "Quantidade de OPs", "dominio": "Producao", "descricao": "OPs no periodo.", "origem": "Producao por OP", "tipo_origem": "producao", "slug_origem": "producao-por-op", "campo_origem": "ops", "unidade": "OPs", "referencia_temporal": "Data da OP", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "prod_aves", "nome": "Aves Processadas", "dominio": "Producao", "descricao": "Aves recebidas/processadas nas OPs.", "origem": "Producao por OP", "tipo_origem": "producao", "slug_origem": "producao-por-op", "campo_origem": "aves_recebidas", "unidade": "aves", "referencia_temporal": "Data da OP", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "prod_peso", "nome": "Peso Produzido", "dominio": "Producao", "descricao": "Peso produzido oficial.", "origem": "Producao por OP", "tipo_origem": "producao", "slug_origem": "producao-por-op", "campo_origem": "peso_produzido", "unidade": "kg", "referencia_temporal": "Data da OP", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "prod_caixas", "nome": "Caixas Produzidas", "dominio": "Producao", "descricao": "Caixas PA vinculadas as OPs.", "origem": "Producao por OP", "tipo_origem": "producao", "slug_origem": "producao-por-op", "campo_origem": "caixas", "unidade": "caixas", "referencia_temporal": "Data da OP", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "prod_rendimento", "nome": "Rendimento", "dominio": "Producao", "descricao": "Kg produzido sobre peso vivo em OPs encerradas.", "origem": "Rendimento", "tipo_origem": "producao", "slug_origem": "rendimento", "campo_origem": "rendimento", "unidade": "%", "referencia_temporal": "Data da OP", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"], "limitacoes": "Rendimento nao e eficiencia."},
    {"id": "prod_condenacoes", "nome": "Condenacoes", "dominio": "Producao", "descricao": "Condenacoes oficiais em aves.", "origem": "Condenacoes", "tipo_origem": "producao", "slug_origem": "condenacoes", "campo_origem": "condenacoes_aves", "unidade": "aves", "referencia_temporal": "Data da OP", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "prod_perdas", "nome": "Perdas", "dominio": "Producao", "descricao": "Perdas oficiais em aves.", "origem": "Perdas", "tipo_origem": "producao", "slug_origem": "perdas", "campo_origem": "perdas_aves", "unidade": "aves", "referencia_temporal": "Data da OP", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "prod_produtividade_hora_setor", "nome": "Produtividade por Hora-Setor", "dominio": "Producao", "descricao": "Produtividade horaria oficial da Eficiencia.", "origem": "Eficiencia", "tipo_origem": "producao", "slug_origem": "eficiencia", "campo_origem": "kg_por_hora_setor", "unidade": "kg/hora-setor", "referencia_temporal": "Data da OP", "status": STATUS_EVOLUCAO, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"], "limitacoes": "Nao e percentual de eficiencia, meta ou OEE."},
    {"id": "alm_entradas", "nome": "Entradas de Almoxarifado", "dominio": "Almoxarifado", "descricao": "Entradas oficiais de insumos.", "origem": "Entradas", "tipo_origem": "almoxarifado", "slug_origem": "entradas", "campo_origem": "Quantidade", "unidade": "unidades conforme cadastro", "referencia_temporal": "Data de movimentacao", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "alm_consumo", "nome": "Consumo de Almoxarifado", "dominio": "Almoxarifado", "descricao": "Consumo oficial de insumos.", "origem": "Consumo", "tipo_origem": "almoxarifado", "slug_origem": "consumo", "campo_origem": "Quantidade", "unidade": "unidades conforme cadastro", "referencia_temporal": "Data de movimentacao", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "alm_itens_saldo", "nome": "Produtos com Saldo", "dominio": "Almoxarifado", "descricao": "Itens com saldo atual positivo.", "origem": "Estoque Atual", "tipo_origem": "almoxarifado", "slug_origem": "estoque-atual", "campo_origem": "Itens com saldo", "unidade": "insumos", "referencia_temporal": "Saldo atual", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "exp_transferencias", "nome": "Transferencias", "dominio": "Expedicao", "descricao": "Eventos logisticos de transferencia.", "origem": "Transferencias", "tipo_origem": "expedicao", "slug_origem": "transferencias", "campo_origem": "Transferencias", "unidade": "eventos", "referencia_temporal": "Data do evento", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "exp_caixas_transferidas", "nome": "Caixas Transferidas", "dominio": "Expedicao", "descricao": "Caixas transferidas fisicamente.", "origem": "Transferencias", "tipo_origem": "expedicao", "slug_origem": "transferencias", "campo_origem": "Caixas", "unidade": "caixas", "referencia_temporal": "Data do evento", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "exp_peso_transferido", "nome": "Peso Transferido", "dominio": "Expedicao", "descricao": "Peso liquido transferido.", "origem": "Transferencias", "tipo_origem": "expedicao", "slug_origem": "transferencias", "campo_origem": "Peso liquido", "unidade": "kg", "referencia_temporal": "Data do evento", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["dia", "semana", "mes"]},
    {"id": "exp_caixas_camara", "nome": "Caixas na Camara Fria", "dominio": "Expedicao", "descricao": "Caixas atualmente na Camara Fria LSM.", "origem": "Estoque Camara Fria", "tipo_origem": "expedicao", "slug_origem": "estoque-camara-fria", "campo_origem": "Caixas", "unidade": "caixas", "referencia_temporal": "Posicao atual", "status": STATUS_DISPONIVEL, "comparacao": "periodo_equivalente", "tendencia": "serie", "direcao": "neutra", "permissao": "pcp", "granularidades": ["mes"]},
    {"id": "cmv", "nome": "CMV", "dominio": "Almoxarifado", "descricao": "CMV congelado por decisao arquitetural.", "origem": "CMV", "tipo_origem": "bloqueado", "unidade": "R$", "referencia_temporal": "Nao aplicavel", "status": STATUS_CONGELADO, "comparacao": "bloqueado", "tendencia": "bloqueado", "direcao": "neutra", "permissao": "pcp", "granularidades": []},
    {"id": "oee", "nome": "OEE", "dominio": "Producao", "descricao": "Indicador futuro, motor nao implantado.", "origem": "OEE", "tipo_origem": "bloqueado", "unidade": "%", "referencia_temporal": "Nao aplicavel", "status": STATUS_FUTURO, "comparacao": "bloqueado", "tendencia": "bloqueado", "direcao": "neutra", "permissao": "pcp", "granularidades": []},
]


def normalizar_filtros(args):
    granularidade = args.get("granularidade") or "mes"
    if granularidade not in ["dia", "semana", "mes"]:
        granularidade = "mes"
    return {
        "data_inicio": args.get("data_inicio") or primeiro_dia_mes(),
        "data_fim": args.get("data_fim") or hoje_iso(),
        "dominio": args.get("dominio") or "Todos",
        "indicador": args.get("indicador") or "Todos",
        "status": args.get("status") or "Todos",
        "unidade": args.get("unidade") or "Todas",
        "granularidade": granularidade,
        "carregar_todos": str(args.get("carregar_todos") or "") == "1",
    }


def filtrar_registro(filtros):
    indicadores = []
    for indicador in REGISTRO_INDICADORES:
        if filtros["dominio"] != "Todos" and indicador["dominio"] != filtros["dominio"]:
            continue
        if filtros["indicador"] != "Todos" and indicador["id"] != filtros["indicador"]:
            continue
        if filtros["status"] != "Todos" and indicador["status"] != filtros["status"]:
            continue
        if filtros["unidade"] != "Todas" and indicador["unidade"] != filtros["unidade"]:
            continue
        indicadores.append(indicador)
    return indicadores


def resolver_indicador(definicao, data_inicio, data_fim, cache=None):
    if definicao.get("tipo_origem") == "bloqueado":
        return indicador_resultado(definicao, data_inicio, data_fim, None, status=definicao["status"], limitacao=definicao["descricao"])
    resolver = RESOLVERS[definicao["tipo_origem"]]
    try:
        return resolver(definicao, data_inicio, data_fim, cache)
    except Exception as erro:
        return indicador_resultado(definicao, data_inicio, data_fim, None, status=STATUS_SEM_DADOS, limitacao=f"Fonte sem dados ou indisponivel para o periodo: {erro}")


def montar_indicadores(filtros):
    cache = {}
    saida = [resolver_indicador(item, filtros["data_inicio"], filtros["data_fim"], cache) for item in filtrar_registro(filtros)]
    if filtros["status"] != "Todos":
        saida = [item for item in saida if item["status_dados"] == filtros["status"] or item["status"] == filtros["status"]]
    return saida


def comparar_indicador(indicador, data_inicio, data_fim, cache=None):
    anterior_inicio, anterior_fim = deslocar_periodo_anterior(data_inicio, data_fim)
    atual = resolver_indicador(indicador, data_inicio, data_fim, cache)
    anterior = resolver_indicador(indicador, anterior_inicio, anterior_fim, cache)
    valor_atual = atual.get("valor")
    valor_anterior = anterior.get("valor")
    comparavel = valor_atual is not None and valor_anterior is not None and atual["unidade"] == anterior["unidade"]
    variacao_abs = None
    variacao_pct = None
    leitura = "Sem base comparavel"
    if comparavel:
        variacao_abs = round(valor_atual - valor_anterior, 4)
        if valor_anterior:
            variacao_pct = round((variacao_abs / valor_anterior) * 100, 2)
        leitura = "Permaneceu estavel" if abs(variacao_abs) <= abs(valor_anterior or 1) * TOLERANCIA_ESTAVEL else ("Aumentou" if variacao_abs > 0 else "Reduziu")
    return {
        **atual,
        "valor_atual": valor_atual,
        "valor_anterior": valor_anterior,
        "periodo_atual": atual["periodo"],
        "periodo_anterior": anterior["periodo"],
        "variacao_abs": variacao_abs,
        "variacao_pct": variacao_pct,
        "comparabilidade": "Comparavel" if comparavel else "Sem base comparavel",
        "leitura": leitura,
    }


def montar_cache_producao_periodos(indicadores, periodos):
    cache = {}
    slugs = sorted({
        item.get("slug_origem")
        for item in indicadores
        if item.get("tipo_origem") == "producao" and item.get("slug_origem")
    })
    if not slugs or not periodos:
        return cache

    data_inicio = min(inicio for _, inicio, _ in periodos)
    data_fim = max(fim for _, _, fim in periodos)
    for slug in slugs:
        try:
            cache[slug] = montar_resumos_gerenciais_producao_periodos(
                slug,
                arg_periodo(data_inicio, data_fim),
                periodos,
            )
        except Exception:
            cache[slug] = None
    return cache


def comparar_indicador_com_periodos(indicador, data_inicio, data_fim, anterior_inicio, anterior_fim, cache=None, producao_periodos=None):
    if indicador.get("tipo_origem") == "producao" and producao_periodos and producao_periodos.get(indicador.get("slug_origem")):
        atual = resolver_producao_contexto(
            indicador,
            data_inicio,
            data_fim,
            producao_periodos[indicador["slug_origem"]].get("atual"),
        )
        anterior = resolver_producao_contexto(
            indicador,
            anterior_inicio,
            anterior_fim,
            producao_periodos[indicador["slug_origem"]].get("anterior"),
        )
    else:
        atual = resolver_indicador(indicador, data_inicio, data_fim, cache)
        anterior = resolver_indicador(indicador, anterior_inicio, anterior_fim, cache)

    valor_atual = atual.get("valor")
    valor_anterior = anterior.get("valor")
    comparavel = valor_atual is not None and valor_anterior is not None and atual["unidade"] == anterior["unidade"]
    variacao_abs = None
    variacao_pct = None
    leitura = "Sem base comparavel"
    if comparavel:
        variacao_abs = round(valor_atual - valor_anterior, 4)
        if valor_anterior:
            variacao_pct = round((variacao_abs / valor_anterior) * 100, 2)
        leitura = "Permaneceu estavel" if abs(variacao_abs) <= abs(valor_anterior or 1) * TOLERANCIA_ESTAVEL else ("Aumentou" if variacao_abs > 0 else "Reduziu")
    return {
        **atual,
        "valor_atual": valor_atual,
        "valor_anterior": valor_anterior,
        "periodo_atual": atual["periodo"],
        "periodo_anterior": anterior["periodo"],
        "variacao_abs": variacao_abs,
        "variacao_pct": variacao_pct,
        "comparabilidade": "Comparavel" if comparavel else "Sem base comparavel",
        "leitura": leitura,
    }


def montar_comparativos(filtros):
    cache = {}
    indicadores = [item for item in filtrar_registro(filtros) if item["status"] in [STATUS_DISPONIVEL, STATUS_EVOLUCAO]]
    anterior_inicio, anterior_fim = deslocar_periodo_anterior(filtros["data_inicio"], filtros["data_fim"])
    periodos = [
        ("atual", filtros["data_inicio"], filtros["data_fim"]),
        ("anterior", anterior_inicio, anterior_fim),
    ]
    producao_periodos = montar_cache_producao_periodos(indicadores, periodos)
    return [
        comparar_indicador_com_periodos(
            item,
            filtros["data_inicio"],
            filtros["data_fim"],
            anterior_inicio,
            anterior_fim,
            cache,
            producao_periodos,
        )
        for item in indicadores
    ]


def tendencia_sem_granularidade(indicador):
    return {
        **indicador,
        "serie": [],
        "direcao_tendencia": "Historico insuficiente",
        "cobertura": 0,
        "status_dados": STATUS_SEM_DADOS,
        "limitacao_resultado": "Granularidade nao sustentada pela fonte.",
    }


def montar_tendencia_saida(indicador, series):
    validos = [item["valor"] for item in series if item.get("valor") is not None]
    direcao = "Historico insuficiente"
    if len(validos) >= 3:
        primeiro, ultimo = validos[0], validos[-1]
        delta = ultimo - primeiro
        base = abs(primeiro) if primeiro else 1
        if abs(delta) <= base * TOLERANCIA_ESTAVEL:
            direcao = "Estavel"
        else:
            sinais = []
            for a, b in zip(validos, validos[1:]):
                if abs(b - a) <= (abs(a) if a else 1) * TOLERANCIA_ESTAVEL:
                    sinais.append(0)
                else:
                    sinais.append(1 if b > a else -1)
            sinais_reais = [s for s in sinais if s != 0]
            if sinais_reais and all(s > 0 for s in sinais_reais):
                direcao = "Crescente"
            elif sinais_reais and all(s < 0 for s in sinais_reais):
                direcao = "Decrescente"
            else:
                direcao = "Oscilante"
    return {
        **indicador,
        "serie": series,
        "direcao_tendencia": direcao,
        "cobertura": len(validos),
        "status_dados": STATUS_DISPONIVEL if validos else STATUS_SEM_DADOS,
        "limitacao_resultado": indicador.get("limitacoes", ""),
    }


def tendencia_indicador(indicador, filtros, cache=None):
    series = []
    granularidade = filtros["granularidade"]
    if granularidade not in indicador.get("granularidades", []):
        return tendencia_sem_granularidade(indicador)
    for periodo, inicio, fim in periodos_serie(filtros["data_inicio"], filtros["data_fim"], granularidade):
        valor = resolver_indicador(indicador, inicio, fim, cache)
        series.append({"periodo": periodo, "valor": valor.get("valor"), "status": valor.get("status_dados")})
    return montar_tendencia_saida(indicador, series)


def tendencia_indicador_producao_lote(indicador, filtros, periodos, producao_periodos):
    if filtros["granularidade"] not in indicador.get("granularidades", []):
        return tendencia_sem_granularidade(indicador)
    dados_slug = producao_periodos.get(indicador.get("slug_origem")) or {}
    series = []
    for periodo, inicio, fim in periodos:
        valor = resolver_producao_contexto(indicador, inicio, fim, dados_slug.get(periodo))
        series.append({"periodo": periodo, "valor": valor.get("valor"), "status": valor.get("status_dados")})
    return montar_tendencia_saida(indicador, series)


def montar_tendencias(filtros):
    cache = {}
    indicadores = [item for item in filtrar_registro(filtros) if item["status"] in [STATUS_DISPONIVEL, STATUS_EVOLUCAO]]
    periodos = periodos_serie(filtros["data_inicio"], filtros["data_fim"], filtros["granularidade"])
    producao_periodos = montar_cache_producao_periodos(
        [item for item in indicadores if filtros["granularidade"] in item.get("granularidades", [])],
        periodos,
    )
    saida = []
    for item in indicadores:
        if item.get("tipo_origem") == "producao" and producao_periodos.get(item.get("slug_origem")):
            saida.append(tendencia_indicador_producao_lote(item, filtros, periodos, producao_periodos))
        else:
            saida.append(tendencia_indicador(item, filtros, cache))
    return saida


def comparativos_todos_sob_demanda(filtros):
    return (
        filtros["dominio"] == "Todos"
        and filtros["indicador"] == "Todos"
        and not filtros.get("carregar_todos")
    )


def montar_resumo_dominios_comparativos(filtros):
    anterior_inicio, anterior_fim = deslocar_periodo_anterior(filtros["data_inicio"], filtros["data_fim"])
    saida = []
    for dominio in sorted({item["dominio"] for item in REGISTRO_INDICADORES}):
        filtros_dominio = dict(filtros)
        filtros_dominio["dominio"] = dominio
        indicadores = [
            item for item in filtrar_registro(filtros_dominio)
            if item["status"] in [STATUS_DISPONIVEL, STATUS_EVOLUCAO]
        ]
        saida.append({
            "dominio": dominio,
            "indicadores": len(indicadores),
            "disponiveis": sum(1 for item in indicadores if item["status"] == STATUS_DISPONIVEL),
            "evolucao": sum(1 for item in indicadores if item["status"] == STATUS_EVOLUCAO),
            "periodo_atual": f"{filtros['data_inicio']} a {filtros['data_fim']}",
            "periodo_anterior": f"{anterior_inicio} a {anterior_fim}",
            "query": urlencode({
                **{k: v for k, v in filtros.items() if k != "carregar_todos" and v not in ["", "Todos", "Todas", False]},
                "dominio": dominio,
            }),
        })
    return saida


def opcoes_filtro():
    return {
        "dominios": sorted({item["dominio"] for item in REGISTRO_INDICADORES}),
        "indicadores": sorted([(item["id"], item["nome"], item["dominio"]) for item in REGISTRO_INDICADORES], key=lambda x: (x[2], x[1])),
        "status": [STATUS_DISPONIVEL, STATUS_EVOLUCAO, STATUS_SEM_DADOS, STATUS_ESTRUTURACAO, STATUS_CONGELADO, STATUS_FUTURO],
        "unidades": sorted({item["unidade"] for item in REGISTRO_INDICADORES if item.get("unidade")}),
        "granularidades": [("dia", "Dia"), ("semana", "Semana"), ("mes", "Mes")],
    }


def montar_contexto_relatorio_gerencial(slug, args):
    filtros = normalizar_filtros(args)
    resumo_dominios = []
    sob_demanda = False
    if slug == "indicadores":
        linhas = montar_indicadores(filtros)
        titulo = "Indicadores"
        objetivo = "Consultar indicadores oficiais sem criar nova fonte de verdade."
    elif slug == "comparativos":
        sob_demanda = comparativos_todos_sob_demanda(filtros)
        if sob_demanda:
            linhas = []
            resumo_dominios = montar_resumo_dominios_comparativos(filtros)
        else:
            linhas = montar_comparativos(filtros)
        titulo = "Comparativos"
        objetivo = "Comparar periodo atual com periodo anterior equivalente."
    elif slug == "tendencias":
        linhas = montar_tendencias(filtros)
        titulo = "Tendencias"
        objetivo = "Evidenciar direcao historica sem previsao ou meta inventada."
    else:
        raise KeyError(slug)

    if sob_demanda:
        resumo = {
            "total": sum(item["indicadores"] for item in resumo_dominios),
            "disponiveis": sum(item["disponiveis"] for item in resumo_dominios),
            "evolucao": sum(item["evolucao"] for item in resumo_dominios),
            "sem_dados": 0,
            "bloqueados": 0,
        }
    else:
        resumo = {
            "total": len(linhas),
            "disponiveis": sum(1 for i in linhas if i.get("status_dados") == STATUS_DISPONIVEL),
            "evolucao": sum(1 for i in linhas if i.get("status_dados") == STATUS_EVOLUCAO or i.get("status") == STATUS_EVOLUCAO),
            "sem_dados": sum(1 for i in linhas if i.get("status_dados") == STATUS_SEM_DADOS),
            "bloqueados": sum(1 for i in linhas if i.get("status") in [STATUS_ESTRUTURACAO, STATUS_CONGELADO, STATUS_FUTURO]),
        }
    return {
        "slug": slug,
        "titulo": titulo,
        "objetivo": objetivo,
        "filtros": filtros,
        "opcoes": opcoes_filtro(),
        "linhas": linhas,
        "resumo": resumo,
        "query_string": urlencode({k: v for k, v in filtros.items() if v not in ["", "Todos", "Todas", False]}),
        "comparativos_sob_demanda": sob_demanda,
        "resumo_dominios": resumo_dominios,
        "query_carregar_todos": urlencode({
            **{k: v for k, v in filtros.items() if k != "carregar_todos" and v not in ["", "Todos", "Todas", False]},
            "carregar_todos": 1,
        }),
        "limitacoes": [
            "A camada gerencial consome services oficiais e nao grava resultados.",
            "Ausencia de dados, metricas congeladas e metricas futuras nao sao convertidas em zero.",
            "Variacoes sao neutras; favorabilidade depende de regra oficial por indicador.",
        ],
    }


def gerar_excel_relatorio_gerencial(contexto):
    wb = Workbook()
    ws = wb.active
    ws.title = contexto["titulo"][:31]
    ws.append([contexto["titulo"]])
    ws.append([contexto["objetivo"]])
    ws.append([])
    ws.append(["Filtros"])
    for chave, valor in contexto["filtros"].items():
        ws.append([chave, valor])
    ws.append([])

    if contexto["slug"] == "comparativos":
        colunas = ["id", "nome", "dominio", "unidade", "periodo_atual", "valor_atual", "periodo_anterior", "valor_anterior", "variacao_abs", "variacao_pct", "comparabilidade", "leitura", "origem", "status_dados"]
        ws.append(colunas)
        for item in contexto["linhas"]:
            ws.append([item.get(c, "") for c in colunas])
    elif contexto["slug"] == "tendencias":
        colunas = ["id", "nome", "dominio", "unidade", "referencia_temporal", "direcao_tendencia", "cobertura", "status_dados", "origem"]
        ws.append(colunas)
        for item in contexto["linhas"]:
            ws.append([item.get(c, "") for c in colunas])
            for ponto in item.get("serie", []):
                ws.append(["", "serie", "", item.get("unidade"), ponto.get("periodo"), ponto.get("valor"), ponto.get("status"), "", ""])
    else:
        colunas = ["id", "nome", "dominio", "valor", "unidade", "periodo", "referencia_temporal", "status_dados", "origem", "limitacao_resultado"]
        ws.append(colunas)
        for item in contexto["linhas"]:
            ws.append([item.get(c, "") for c in colunas])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3B4D")
    for coluna in range(1, 16):
        ws.column_dimensions[get_column_letter(coluna)].width = 22
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo

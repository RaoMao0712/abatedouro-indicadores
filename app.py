from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from werkzeug.security import generate_password_hash
from datetime import datetime, timedelta
from functools import wraps
import calendar
from io import BytesIO
import os
import uuid
import sqlite3
import threading
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import DATABASE_URL, DB_NAME, conectar, inicializar_schema_uma_vez, q
from database.migrations import executar_alteracao_segura
from modules.auth import register_auth_routes
from modules.auth.decorators import perfil_permitido
from modules.auth.services import nome_usuario_atual, usuario_eh_admin
from modules.dashboard.routes import register_dashboard_routes
from modules.producao.routes import register_producao_routes
from modules.qualidade.routes import register_qualidade_routes
from modules.almoxarifado.routes import register_almoxarifado_routes
from modules.producao.services import buscar_op_por_id, gerar_producao_automatica_setores
from modules.almoxarifado.services import (
    buscar_insumos_almoxarifado,
    criar_tabelas_almoxarifado,
    criar_tabelas_estoque_almoxarifado,
)
from services import manutencao_service
from utils import calcular_horas_programadas, calcular_produtividade, setores_padrao, normalizar_chave_setor

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "segredo")


# ============================================================
# FORMATAÇÃO BR / APRESENTAÇÃO EXECUTIVA
# ============================================================

def formatar_numero_br(valor, casas=2):
    try:
        numero = float(valor or 0)
    except Exception:
        numero = 0

    texto = f"{numero:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_moeda_br(valor):
    return f"R$ {formatar_numero_br(valor, 2)}"


def formatar_percentual_br(valor):
    return f"{formatar_numero_br(valor, 2)}%"


@app.template_filter("br_numero")
def filtro_br_numero(valor, casas=2):
    return formatar_numero_br(valor, int(casas))


@app.template_filter("br_moeda")
def filtro_br_moeda(valor):
    return formatar_moeda_br(valor)


@app.template_filter("br_percentual")
def filtro_br_percentual(valor):
    return formatar_percentual_br(valor)


def preparar_grafico_despesas_operacionais(linhas_custos, receita_bruta, limite=6):
    """
    Prepara os dados visuais do gráfico de rosca da DRE.
    O tamanho das fatias usa participação dentro das despesas operacionais.
    O rótulo exibido usa % da receita, para manter leitura gerencial.
    """
    itens_validos = [
        {
            "categoria": item["categoria"],
            "valor": float(item["valor"] or 0),
            "percentual_receita": float(item["percentual"] or 0)
        }
        for item in linhas_custos
        if float(item["valor"] or 0) > 0
    ]

    itens_validos = sorted(itens_validos, key=lambda item: item["valor"], reverse=True)
    total_despesas = sum(item["valor"] for item in itens_validos)

    if total_despesas <= 0:
        return {
            "itens": [],
            "gradiente": "#e5e7eb 0deg 360deg",
            "total": 0,
            "percentual_receita_total": 0
        }

    cores = [
        "#2563eb",
        "#16a34a",
        "#f97316",
        "#8b5cf6",
        "#0891b2",
        "#64748b",
        "#dc2626"
    ]

    limite_principal = max(1, limite - 1)
    principais = itens_validos[:limite_principal]
    restantes = itens_validos[limite_principal:]

    itens_grafico = []

    for item in principais:
        itens_grafico.append(item)

    if restantes:
        valor_outras = sum(item["valor"] for item in restantes)
        percentual_receita_outras = sum(item["percentual_receita"] for item in restantes)
        itens_grafico.append({
            "categoria": f"Outras categorias ({len(restantes)})",
            "valor": valor_outras,
            "percentual_receita": percentual_receita_outras
        })

    segmentos = []
    angulo_atual = 0
    itens_saida = []

    for indice, item in enumerate(itens_grafico):
        fatia = item["valor"] / total_despesas
        graus = fatia * 360
        inicio = angulo_atual
        fim = 360 if indice == len(itens_grafico) - 1 else angulo_atual + graus
        meio = (inicio + fim) / 2
        cor = cores[indice % len(cores)]
        segmentos.append(f"{cor} {inicio:.2f}deg {fim:.2f}deg")

        # Coordenadas dos rótulos ao redor da rosca.
        # Regra visual validada: só exibimos rótulos externos para categorias
        # com pelo menos 5% da receita. As demais continuam na tabela lateral.
        #
        # Também limitamos a distância dos rótulos para impedir invasão da tabela
        # lateral e evitar balões excessivamente afastados do gráfico.
        import math
        rad = math.radians(meio - 90)
        x = 50 + (32 * math.cos(rad))
        y = 50 + (31 * math.sin(rad))
        x = max(28, min(72, x))
        y = max(22, min(78, y))
        alinhamento = "right" if x < 50 else "left"
        mostrar_rotulo = float(item["percentual_receita"] or 0) >= 5

        itens_saida.append({
            "categoria": item["categoria"],
            "valor": round(item["valor"], 2),
            "valor_formatado": formatar_moeda_br(item["valor"]),
            "percentual_receita": round(item["percentual_receita"], 2),
            "percentual_receita_formatado": formatar_percentual_br(item["percentual_receita"]),
            "percentual_despesa": round(fatia * 100, 2),
            "percentual_despesa_formatado": formatar_percentual_br(fatia * 100),
            "cor": cor,
            "x": round(x, 2),
            "y": round(y, 2),
            "alinhamento": alinhamento,
            "mostrar_rotulo": mostrar_rotulo
        })

        angulo_atual = fim

    return {
        "itens": itens_saida,
        "gradiente": ", ".join(segmentos),
        "total": round(total_despesas, 2),
        "total_formatado": formatar_moeda_br(total_despesas),
        "percentual_receita_total": round((total_despesas / receita_bruta * 100) if receita_bruta > 0 else 0, 2),
        "percentual_receita_total_formatado": formatar_percentual_br((total_despesas / receita_bruta * 100) if receita_bruta > 0 else 0)
    }


def preparar_linhas_custos_executivas(linhas_custos, limite=6):
    itens = [item for item in linhas_custos if float(item["valor"] or 0) > 0]
    itens = sorted(itens, key=lambda item: float(item["valor"] or 0), reverse=True)

    if len(itens) <= limite:
        return itens

    principais = itens[:limite - 1]
    restantes = itens[limite - 1:]
    principais.append({
        "categoria": f"Outras categorias ({len(restantes)})",
        "valor": round(sum(float(item["valor"] or 0) for item in restantes), 2),
        "percentual": round(sum(float(item["percentual"] or 0) for item in restantes), 2)
    })
    return principais

ROTINAS_ESTRUTURAIS_EXECUTADAS = set()
ROTINAS_ESTRUTURAIS_LOCK = threading.RLock()


def executar_rotina_estrutural_uma_vez(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        destino_banco = DATABASE_URL or os.path.abspath(DB_NAME)
        chave = (func.__name__, destino_banco)

        with ROTINAS_ESTRUTURAIS_LOCK:
            if chave in ROTINAS_ESTRUTURAIS_EXECUTADAS:
                return None

            resultado = func(*args, **kwargs)
            ROTINAS_ESTRUTURAIS_EXECUTADAS.add(chave)
            return resultado

    return wrapper


CATEGORIAS_CUSTOS = [
    "Mão de obra",
    "Energia",
    "Água",
    "Lenha",
    "Combustível",

    "Manutenção de Equipamentos",
    "Manutenção Predial",
    "Manutenção de Veículos",

    "Material de Limpeza",
    "Material de Escritório",
    "Serviços",

    "EPIs",

    "Marketing",
    "Cursos e Treinamentos",
    "Despesas com Viagens",

    "Consultoria e Responsabilidade Técnica",

    "Contratos com Clientes",

    "Insumos para Produção",

    "Impostos e Taxas",

    "Despesas Financeiras",


    "Outros"
]


def tentar_alter_table(cursor, conn, comando):
    executar_alteracao_segura(cursor, conn, comando)


@executar_rotina_estrutural_uma_vez
def criar_banco():
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            perfil TEXT DEFAULT 'admin'
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS fornecedores (
            id SERIAL PRIMARY KEY,
            nome TEXT UNIQUE NOT NULL
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS ordens_producao (
            id SERIAL PRIMARY KEY,
            data TEXT NOT NULL,
            fornecedor TEXT NOT NULL,
            gta TEXT,
            nota_fiscal TEXT,
            quantidade_aves INTEGER NOT NULL,
            mortes_antes_pendura INTEGER DEFAULT 0,
            peso_vivo REAL NOT NULL,
            peso_medio REAL NOT NULL,
            observacoes TEXT,
            status TEXT DEFAULT 'Aberta',
            sku TEXT DEFAULT 'Galinha Cortada'
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_setor (
            id SERIAL PRIMARY KEY,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            colaboradores INTEGER NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fim TEXT NOT NULL,
            horas_programadas REAL NOT NULL,
            horas_paradas REAL DEFAULT 0,
            motivo_parada TEXT,
            quantidade_produzida REAL NOT NULL,
            unidade TEXT NOT NULL,
            condenacoes REAL DEFAULT 0,
            perdas REAL DEFAULT 0,
            produtividade REAL NOT NULL,
            observacoes TEXT
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_producao (
            id SERIAL PRIMARY KEY,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            quantidade REAL NOT NULL,
            unidade TEXT NOT NULL,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_mao_obra (
            id SERIAL PRIMARY KEY,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            colaborador TEXT NOT NULL,
            funcao TEXT NOT NULL,
            setor TEXT NOT NULL,
            turno TEXT,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_paradas (
            id SERIAL PRIMARY KEY,
            evento_id TEXT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            motivo TEXT NOT NULL,
            hora_inicio TEXT,
            hora_fim TEXT,
            horas_paradas REAL NOT NULL,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_descartes (
            id SERIAL PRIMARY KEY,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            categoria TEXT NOT NULL,
            motivo TEXT NOT NULL,
            quantidade REAL NOT NULL,
            unidade TEXT NOT NULL,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_tempos_setor (
            id SERIAL PRIMARY KEY,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fim TEXT NOT NULL,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        tentar_alter_table(cursor, conn, "ALTER TABLE usuarios ADD COLUMN perfil TEXT DEFAULT 'admin'")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE ordens_producao ADD COLUMN status TEXT DEFAULT 'Aberta'")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE ordens_producao ADD COLUMN sku TEXT DEFAULT 'Galinha Cortada'")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE apontamentos_paradas ADD COLUMN evento_id TEXT")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE apontamentos_paradas ADD COLUMN manutencao_ordem_id INTEGER")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE apontamentos_paradas ADD COLUMN manutencao_aberta TEXT DEFAULT 'Nao'")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE apontamentos_paradas ADD COLUMN encerrada_por_manutencao TEXT DEFAULT 'Nao'")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE apontamentos_paradas ADD COLUMN data_fim TEXT")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE apontamentos_paradas ADD COLUMN equipamento TEXT")
        conn = conectar()
        cursor = conn.cursor()
        tentar_alter_table(cursor, conn, "ALTER TABLE apontamentos_paradas ADD COLUMN equipamento_id INTEGER")
        conn = conectar()
        cursor = conn.cursor()

    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            perfil TEXT DEFAULT 'admin'
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS fornecedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS ordens_producao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            fornecedor TEXT NOT NULL,
            gta TEXT,
            nota_fiscal TEXT,
            quantidade_aves INTEGER NOT NULL,
            mortes_antes_pendura INTEGER DEFAULT 0,
            peso_vivo REAL NOT NULL,
            peso_medio REAL NOT NULL,
            observacoes TEXT,
            status TEXT DEFAULT 'Aberta',
            sku TEXT DEFAULT 'Galinha Cortada'
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_setor (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            colaboradores INTEGER NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fim TEXT NOT NULL,
            horas_programadas REAL NOT NULL,
            horas_paradas REAL DEFAULT 0,
            motivo_parada TEXT,
            quantidade_produzida REAL NOT NULL,
            unidade TEXT NOT NULL,
            condenacoes REAL DEFAULT 0,
            perdas REAL DEFAULT 0,
            produtividade REAL NOT NULL,
            observacoes TEXT
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_producao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            quantidade REAL NOT NULL,
            unidade TEXT NOT NULL,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_mao_obra (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            colaborador TEXT NOT NULL,
            funcao TEXT NOT NULL,
            setor TEXT NOT NULL,
            turno TEXT,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_paradas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evento_id TEXT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            motivo TEXT NOT NULL,
            hora_inicio TEXT,
            hora_fim TEXT,
            horas_paradas REAL NOT NULL,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_descartes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            categoria TEXT NOT NULL,
            motivo TEXT NOT NULL,
            quantidade REAL NOT NULL,
            unidade TEXT NOT NULL,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_tempos_setor (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fim TEXT NOT NULL,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        try:
            cursor.execute("ALTER TABLE usuarios ADD COLUMN perfil TEXT DEFAULT 'admin'")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE ordens_producao ADD COLUMN status TEXT DEFAULT 'Aberta'")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE ordens_producao ADD COLUMN sku TEXT DEFAULT 'Galinha Cortada'")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE apontamentos_paradas ADD COLUMN evento_id TEXT")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE apontamentos_paradas ADD COLUMN manutencao_ordem_id INTEGER")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE apontamentos_paradas ADD COLUMN manutencao_aberta TEXT DEFAULT 'Nao'")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE apontamentos_paradas ADD COLUMN encerrada_por_manutencao TEXT DEFAULT 'Nao'")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE apontamentos_paradas ADD COLUMN data_fim TEXT")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE apontamentos_paradas ADD COLUMN equipamento TEXT")
        except sqlite3.OperationalError:
            pass

        try:
            cursor.execute("ALTER TABLE apontamentos_paradas ADD COLUMN equipamento_id INTEGER")
        except sqlite3.OperationalError:
            pass

    cursor.execute("SELECT COUNT(*) as total FROM usuarios")
    total = cursor.fetchone()["total"]

    if total == 0:
        cursor.execute(q("""
        INSERT INTO usuarios (nome, email, senha_hash, perfil)
        VALUES (?, ?, ?, ?)
        """), (
            "Administrador",
            "admin@app.com",
            generate_password_hash("admin123"),
            "admin"
        ))
    else:
        cursor.execute(q("""
        UPDATE usuarios
        SET perfil = 'admin'
        WHERE perfil IS NULL OR perfil = ''
        """))

    conn.commit()
    conn.close()


register_auth_routes(app, criar_banco)
register_dashboard_routes(app)


@executar_rotina_estrutural_uma_vez
def criar_tabela_tempos_setor():
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_tempos_setor (
            id SERIAL PRIMARY KEY,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fim TEXT NOT NULL,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS apontamentos_tempos_setor (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            setor TEXT NOT NULL,
            hora_inicio TEXT NOT NULL,
            hora_fim TEXT NOT NULL,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

    conn.commit()
    conn.close()





@executar_rotina_estrutural_uma_vez
def criar_tabelas_custos():
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS parametros_custos (
            id SERIAL PRIMARY KEY,
            sku TEXT UNIQUE NOT NULL,
            custo_ave REAL DEFAULT 0,
            unidade_custo_ave TEXT DEFAULT '',
            custo_embalagem REAL DEFAULT 0,
            unidade_custo_embalagem TEXT DEFAULT '',
            atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS custos_mensais (
            id SERIAL PRIMARY KEY,
            competencia TEXT NOT NULL,
            categoria TEXT NOT NULL,
            valor REAL NOT NULL,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS parametros_custos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT UNIQUE NOT NULL,
            custo_ave REAL DEFAULT 0,
            unidade_custo_ave TEXT DEFAULT '',
            custo_embalagem REAL DEFAULT 0,
            unidade_custo_embalagem TEXT DEFAULT '',
            atualizado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS custos_mensais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competencia TEXT NOT NULL,
            categoria TEXT NOT NULL,
            valor REAL NOT NULL,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

    conn.commit()

    parametros_padrao = [
        ("Galinha Cortada", 0, "R$/ave", 0, "R$/bandeja"),
        ("Galinha Inteira", 0, "R$/ave", 0, "R$/unidade")
    ]

    for item in parametros_padrao:
        try:
            cursor.execute(q("""
            INSERT INTO parametros_custos (
                sku,
                custo_ave,
                unidade_custo_ave,
                custo_embalagem,
                unidade_custo_embalagem
            ) VALUES (?, ?, ?, ?, ?)
            """), item)
            conn.commit()
        except Exception:
            conn.rollback()

    conn.close()


def buscar_parametros_custos():
    criar_tabelas_custos()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM parametros_custos
    ORDER BY sku
    """)

    parametros = cursor.fetchall()
    conn.close()

    return parametros


def buscar_custos_mensais(competencia_inicio=None, competencia_fim=None, categoria=None):
    criar_tabelas_custos()

    conn = conectar()
    cursor = conn.cursor()

    filtros = []
    parametros = []

    if competencia_inicio:
        filtros.append("competencia >= ?")
        parametros.append(competencia_inicio)

    if competencia_fim:
        filtros.append("competencia <= ?")
        parametros.append(competencia_fim)

    if categoria and categoria != "Todas":
        filtros.append("categoria = ?")
        parametros.append(categoria)

    where_sql = f"WHERE {' AND '.join(filtros)}" if filtros else ""

    cursor.execute(q(f"""
    SELECT *
    FROM custos_mensais
    {where_sql}
    ORDER BY competencia DESC, categoria ASC
    """), parametros)

    custos = cursor.fetchall()
    conn.close()

    return custos


def chave_sku_custo(sku):
    return (
        sku.lower()
        .replace(" ", "_")
        .replace("ã", "a")
    )


def salvar_parametros_custos(form):
    criar_tabelas_custos()

    conn = conectar()
    cursor = conn.cursor()

    skus = ["Galinha Cortada", "Galinha Inteira"]

    for sku in skus:
        chave = chave_sku_custo(sku)

        custo_ave = float(form.get(f"custo_ave_{chave}") or 0)
        custo_embalagem = float(form.get(f"custo_embalagem_{chave}") or 0)

        if sku == "Galinha Cortada":
            unidade_custo_ave = "R$/ave"
            unidade_custo_embalagem = "R$/bandeja"
        else:
            unidade_custo_ave = "R$/ave"
            unidade_custo_embalagem = "R$/unidade"

        cursor.execute(q("""
        UPDATE parametros_custos
        SET custo_ave = ?,
            unidade_custo_ave = ?,
            custo_embalagem = ?,
            unidade_custo_embalagem = ?
        WHERE sku = ?
        """), (
            custo_ave,
            unidade_custo_ave,
            custo_embalagem,
            unidade_custo_embalagem,
            sku
        ))

    conn.commit()
    conn.close()


def salvar_custo_mensal(form):
    criar_tabelas_custos()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    INSERT INTO custos_mensais (
        competencia,
        categoria,
        valor,
        observacoes
    ) VALUES (?, ?, ?, ?)
    """), (
        form["competencia"],
        form["categoria"],
        float(form["valor"]),
        form.get("observacoes", "")
    ))

    conn.commit()
    conn.close()


def salvar_custos_mensais_lote(form):
    criar_tabelas_custos()

    competencia = form["competencia"]
    observacoes_gerais = form.get("observacoes_gerais", "")
    categorias = form.getlist("categoria[]")
    valores = form.getlist("valor[]")
    observacoes = form.getlist("observacoes[]")

    if not categorias:
        raise ValueError("Adicione pelo menos uma linha de custo antes de confirmar.")

    if not (len(categorias) == len(valores) == len(observacoes)):
        raise ValueError("As linhas de custo estao incompletas. Revise categorias, valores e observacoes.")

    linhas = []
    for indice, valor_raw in enumerate(valores, start=1):
        categoria = categorias[indice - 1]
        observacao = observacoes[indice - 1].strip()

        if not categoria:
            raise ValueError(f"Selecione uma categoria na linha {indice}.")

        if categoria not in CATEGORIAS_CUSTOS:
            raise ValueError(f"A categoria da linha {indice} nao e valida.")

        try:
            valor = float(str(valor_raw).replace(",", "."))
        except (TypeError, ValueError):
            raise ValueError(f"Informe um valor valido na linha {indice}.")

        if valor <= 0:
            raise ValueError(f"O valor da linha {indice} precisa ser maior que zero.")

        if observacoes_gerais and observacao:
            observacao_final = f"{observacao} | {observacoes_gerais}"
        else:
            observacao_final = observacao or observacoes_gerais

        linhas.append((competencia, categoria, valor, observacao_final))

    conn = conectar()
    cursor = conn.cursor()

    cursor.executemany(q("""
    INSERT INTO custos_mensais (
        competencia,
        categoria,
        valor,
        observacoes
    ) VALUES (?, ?, ?, ?)
    """), linhas)

    conn.commit()
    conn.close()

    return len(linhas)





@executar_rotina_estrutural_uma_vez
def criar_tabela_vendas():
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS vendas_diarias (
            id SERIAL PRIMARY KEY,
            data TEXT NOT NULL,
            sku TEXT NOT NULL,
            quantidade REAL NOT NULL,
            unidade TEXT NOT NULL,
            quantidade_unidades REAL DEFAULT 0,
            quantidade_kg REAL DEFAULT 0,
            receita REAL NOT NULL,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS vendas_diarias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            sku TEXT NOT NULL,
            quantidade REAL NOT NULL,
            unidade TEXT NOT NULL,
            quantidade_unidades REAL DEFAULT 0,
            quantidade_kg REAL DEFAULT 0,
            receita REAL NOT NULL,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

    tentar_alter_table(cursor, conn, "ALTER TABLE vendas_diarias ADD COLUMN quantidade_unidades REAL DEFAULT 0")
    tentar_alter_table(cursor, conn, "ALTER TABLE vendas_diarias ADD COLUMN quantidade_kg REAL DEFAULT 0")

    conn.commit()
    conn.close()



def preparar_quantidades_venda(sku, form, quantidade_atual=None, unidade_atual=None):
    """
    Padroniza as quantidades de venda para suportar a lógica correta da DRE.

    Galinha Cortada:
    - quantidade_unidades = bandejas vendidas, usada para CMV;
    - quantidade_kg = kg vendidos, usado para receita/kg e CMV/kg;
    - quantidade legado = kg, para compatibilidade com telas antigas.

    Galinha Inteira:
    - quantidade_unidades = unidades vendidas;
    - quantidade_kg = 0;
    - quantidade legado = unidades.
    """
    quantidade_legacy = form.get("quantidade")
    quantidade_unidades_raw = (
        form.get("quantidade_unidades")
        or form.get("unidades_vendidas")
        or form.get("bandejas_vendidas")
        or ""
    )
    quantidade_kg_raw = (
        form.get("quantidade_kg")
        or form.get("kg_vendidos")
        or ""
    )

    if sku == "Galinha Cortada":
        quantidade_unidades = float(quantidade_unidades_raw or 0)
        quantidade_kg = float(quantidade_kg_raw or quantidade_legacy or quantidade_atual or 0)

        if quantidade_unidades <= 0:
            raise ValueError("Informe a quantidade de bandejas/unidades vendidas para Galinha Cortada.")

        if quantidade_kg <= 0:
            raise ValueError("Informe a quantidade em kg vendida para Galinha Cortada.")

        return {
            "quantidade": quantidade_kg,
            "unidade": "kg",
            "quantidade_unidades": quantidade_unidades,
            "quantidade_kg": quantidade_kg
        }

    quantidade_unidades = float(quantidade_unidades_raw or quantidade_legacy or quantidade_atual or 0)

    if quantidade_unidades <= 0:
        raise ValueError("Informe a quantidade de unidades vendidas.")

    return {
        "quantidade": quantidade_unidades,
        "unidade": "unidades",
        "quantidade_unidades": quantidade_unidades,
        "quantidade_kg": 0
    }


def buscar_venda_diaria_por_data_sku(data, sku, ignorar_id=None):
    """
    Busca lançamento de venda já existente para a mesma data e SKU.

    Regra de negócio:
    - A tela de Vendas Diárias registra a receita consolidada do dia por SKU.
    - Portanto, deve existir no máximo um lançamento para cada combinação Data + SKU.
    - Se o usuário precisar corrigir valores, deve editar o lançamento existente.
    """
    criar_tabela_vendas()

    conn = conectar()
    cursor = conn.cursor()

    if ignorar_id:
        cursor.execute(q("""
        SELECT *
        FROM vendas_diarias
        WHERE data = ?
          AND sku = ?
          AND id <> ?
        ORDER BY id DESC
        LIMIT 1
        """), (data, sku, ignorar_id))
    else:
        cursor.execute(q("""
        SELECT *
        FROM vendas_diarias
        WHERE data = ?
          AND sku = ?
        ORDER BY id DESC
        LIMIT 1
        """), (data, sku))

    venda = cursor.fetchone()
    conn.close()

    return venda


def salvar_venda_diaria(form):
    criar_tabela_vendas()

    data_venda = form["data"]
    sku = form["sku"]
    receita = float(form["receita"])

    if receita < 0:
        raise ValueError("A receita não pode ser negativa.")

    quantidades = preparar_quantidades_venda(sku, form)

    venda_existente = buscar_venda_diaria_por_data_sku(data_venda, sku)

    if venda_existente:
        raise ValueError(
            "Já existe uma venda lançada para esta data e SKU. "
            "Edite o lançamento existente em vez de cadastrar novamente."
        )

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    INSERT INTO vendas_diarias (
        data,
        sku,
        quantidade,
        unidade,
        quantidade_unidades,
        quantidade_kg,
        receita,
        observacoes
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """), (
        data_venda,
        sku,
        quantidades["quantidade"],
        quantidades["unidade"],
        quantidades["quantidade_unidades"],
        quantidades["quantidade_kg"],
        receita,
        form.get("observacoes", "")
    ))

    conn.commit()
    conn.close()


def buscar_vendas_diarias():
    criar_tabela_vendas()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM vendas_diarias
    ORDER BY data DESC, id DESC
    """)

    vendas = cursor.fetchall()
    conn.close()

    return vendas



def buscar_custo_mensal_por_id(custo_id):
    criar_tabelas_custos()
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q("""
    SELECT *
    FROM custos_mensais
    WHERE id = ?
    """), (custo_id,))
    custo = cursor.fetchone()
    conn.close()
    return custo


def buscar_venda_diaria_por_id(venda_id):
    criar_tabela_vendas()
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q("""
    SELECT *
    FROM vendas_diarias
    WHERE id = ?
    """), (venda_id,))
    venda = cursor.fetchone()
    conn.close()
    return venda



# ============================================================
# EXPEDIÇÃO - SPRINT 1.0
# ============================================================

@executar_rotina_estrutural_uma_vez
def criar_tabelas_expedicao():
    """
    Cria a fundação do módulo de Expedição.

    Sprint 1.0:
    - Apenas estrutura, listagem e ponto de entrada no menu.
    - Não baixa estoque.
    - Não gera venda.
    - Não interfere na DRE nem no Financeiro.
    """
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS expedicoes (
            id SERIAL PRIMARY KEY,
            numero_romaneio TEXT UNIQUE NOT NULL,
            data TEXT NOT NULL,
            tipo_movimentacao TEXT NOT NULL DEFAULT 'TRANSFERENCIA',
            destino TEXT NOT NULL,
            responsavel TEXT,
            observacoes TEXT,
            status TEXT NOT NULL DEFAULT 'Aberto',
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS expedicao_itens (
            id SERIAL PRIMARY KEY,
            expedicao_id INTEGER NOT NULL,
            op_id INTEGER,
            sku TEXT NOT NULL,
            quantidade_unidades REAL DEFAULT 0,
            quantidade_kg REAL DEFAULT 0,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS expedicoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_romaneio TEXT UNIQUE NOT NULL,
            data TEXT NOT NULL,
            tipo_movimentacao TEXT NOT NULL DEFAULT 'TRANSFERENCIA',
            destino TEXT NOT NULL,
            responsavel TEXT,
            observacoes TEXT,
            status TEXT NOT NULL DEFAULT 'Aberto',
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS expedicao_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expedicao_id INTEGER NOT NULL,
            op_id INTEGER,
            sku TEXT NOT NULL,
            quantidade_unidades REAL DEFAULT 0,
            quantidade_kg REAL DEFAULT 0,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

    conn.commit()
    conn.close()



# ============================================================
# ESTOQUE PI / PA - ARQUITETURA RASTREÁVEL
# ============================================================

BANDEJAS_POR_CAIXA = 12


def sku_sem_embalagem_secundaria(sku):
    return (sku or "").strip().lower() == "galinha inteira"


@executar_rotina_estrutural_uma_vez
def criar_tabelas_estoque_pi_pa():
    """
    Cria a estrutura de estoque em duas etapas:

    1) Produto Intermediário (PI)
       - nasce no encerramento da OP;
       - controla bandejas por OP/lote;
       - nenhuma bandeja fica sem vínculo com OP.

    2) Produto Acabado (PA)
       - nasce na Embalagem Secundária;
       - controla caixas físicas;
       - permite composição da caixa por uma ou mais OPs;
       - preparado para integração futura com balança, Zebra e leitor.
    """
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS estoque_produto_intermediario (
            id SERIAL PRIMARY KEY,
            data_movimentacao TEXT NOT NULL,
            tipo TEXT NOT NULL,
            op_id INTEGER,
            sku TEXT NOT NULL,
            quantidade_bandejas REAL DEFAULT 0,
            origem TEXT,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS embalagem_primaria_apontamentos (
            id SERIAL PRIMARY KEY,
            op_id INTEGER NOT NULL,
            data_apontamento TEXT NOT NULL,
            sku TEXT NOT NULL,
            quantidade_bandejas REAL DEFAULT 0,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS pa_caixas (
            id SERIAL PRIMARY KEY,
            codigo_caixa TEXT UNIQUE NOT NULL,
            sku TEXT NOT NULL,
            data_fabricacao TEXT,
            data_validade TEXT,
            peso_bruto REAL DEFAULT 0,
            peso_liquido REAL DEFAULT 0,
            quantidade_bandejas REAL DEFAULT 0,
            status TEXT DEFAULT 'Em estoque',
            origem TEXT,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS pa_caixa_composicao (
            id SERIAL PRIMARY KEY,
            caixa_id INTEGER NOT NULL,
            op_id INTEGER NOT NULL,
            quantidade_bandejas REAL DEFAULT 0,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS estoque_produto_intermediario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_movimentacao TEXT NOT NULL,
            tipo TEXT NOT NULL,
            op_id INTEGER,
            sku TEXT NOT NULL,
            quantidade_bandejas REAL DEFAULT 0,
            origem TEXT,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS embalagem_primaria_apontamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            data_apontamento TEXT NOT NULL,
            sku TEXT NOT NULL,
            quantidade_bandejas REAL DEFAULT 0,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS pa_caixas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_caixa TEXT UNIQUE NOT NULL,
            sku TEXT NOT NULL,
            data_fabricacao TEXT,
            data_validade TEXT,
            peso_bruto REAL DEFAULT 0,
            peso_liquido REAL DEFAULT 0,
            quantidade_bandejas REAL DEFAULT 0,
            status TEXT DEFAULT 'Em estoque',
            origem TEXT,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS pa_caixa_composicao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            caixa_id INTEGER NOT NULL,
            op_id INTEGER NOT NULL,
            quantidade_bandejas REAL DEFAULT 0,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

    conn.commit()
    conn.close()


def remover_movimentacoes_estoque_pi_por_op(op_id):
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    DELETE FROM estoque_produto_intermediario
    WHERE op_id = ?
      AND tipo IN (?, ?)
    """), (op_id, "ENTRADA_OP", "ENTRADA_EMBALAGEM_PRIMARIA"))

    cursor.execute(q("""
    DELETE FROM embalagem_primaria_apontamentos
    WHERE op_id = ?
    """), (op_id,))

    conn.commit()
    conn.close()


def op_possui_caixa_pa(op_id):
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT COUNT(*) AS total
    FROM pa_caixa_composicao
    WHERE op_id = ?
    """), (op_id,))

    linha = cursor.fetchone()
    conn.close()

    return float(linha["total"] or 0) > 0


def registrar_entrada_estoque_pi_op(op, unidades_produzidas, origem="Embalagem Primária", observacoes=None):
    criar_tabelas_estoque_pi_pa()

    op_id = op["id"]
    sku = op["sku"] or "Galinha Cortada"
    bandejas = float(unidades_produzidas or 0)

    # Segurança contra duplicidade: se a OP for reapontada, a entrada anterior é substituída.
    remover_movimentacoes_estoque_pi_por_op(op_id)

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    INSERT INTO estoque_produto_intermediario (
        data_movimentacao,
        tipo,
        op_id,
        sku,
        quantidade_bandejas,
        origem,
        observacoes
    ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """), (
        op["data"],
        "ENTRADA_EMBALAGEM_PRIMARIA",
        op_id,
        sku,
        bandejas,
        origem,
        observacoes or "Entrada automática de PI gerada no apontamento da Embalagem Primária."
    ))

    conn.commit()
    conn.close()


def calcular_perdas_aves_op(cursor, op_id):
    cursor.execute(q("""
    SELECT
        COALESCE(SUM(CASE
            WHEN LOWER(COALESCE(categoria, '')) LIKE '%%conden%%' THEN quantidade
            ELSE 0
        END), 0) AS condenacoes,
        COALESCE(SUM(CASE
            WHEN LOWER(COALESCE(categoria, '')) NOT LIKE '%%conden%%'
             AND LOWER(TRIM(COALESCE(motivo, ''))) <> 'morte na gaiola' THEN quantidade
            ELSE 0
        END), 0) AS descartes,
        COALESCE(SUM(CASE
            WHEN LOWER(TRIM(COALESCE(motivo, ''))) = 'morte na gaiola' THEN quantidade
            ELSE 0
        END), 0) AS mortes_na_gaiola
    FROM apontamentos_descartes
    WHERE op_id = ?
      AND LOWER(unidade) IN ('aves', 'ave', 'unidade', 'unidades')
    """), (op_id,))

    perdas = cursor.fetchone()
    return {
        "condenacoes": float(perdas["condenacoes"] or 0),
        "descartes": float(perdas["descartes"] or 0),
        "mortes_na_gaiola": float(perdas["mortes_na_gaiola"] or 0),
    }


def validar_balanco_aves_op(op, aves_produzidas, perdas):
    aves_vivas = float(op["quantidade_aves"] or 0)
    mortes_na_gaiola = float(op["mortes_antes_pendura"] or 0) + float(perdas["mortes_na_gaiola"] or 0)
    total_fechamento = float(aves_produzidas or 0) + float(perdas["descartes"] or 0) + float(perdas["condenacoes"] or 0) + mortes_na_gaiola
    saldo_aves = aves_vivas - total_fechamento

    if abs(saldo_aves) > 0.0001:
        raise ValueError(
            f"Balanco de aves divergente em {saldo_aves:g} aves. "
            "Revise Embalagem Primaria, descartes, condenacoes ou mortes na gaiola."
        )

    return saldo_aves


def registrar_lote_pa_galinha_inteira(cursor, op, unidades_vendaveis, aves_embaladas, kg_produzidos, observacoes):
    codigo_lote = f"GI-OP-{int(op['id']):05d}"
    data_fabricacao = op["data"] or datetime.now().strftime("%Y-%m-%d")
    data_validade = calcular_validade_padrao(data_fabricacao)
    observacao_lote = observacoes or f"Lote Galinha Inteira. Aves embaladas: {aves_embaladas:g}."

    if DATABASE_URL:
        cursor.execute(q("""
        INSERT INTO pa_caixas (
            codigo_caixa,
            sku,
            data_fabricacao,
            data_validade,
            peso_bruto,
            peso_liquido,
            quantidade_bandejas,
            status,
            origem,
            observacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        RETURNING id
        """), (
            codigo_lote,
            op["sku"] or "Galinha Inteira",
            data_fabricacao,
            data_validade,
            kg_produzidos,
            kg_produzidos,
            unidades_vendaveis,
            "Em estoque",
            "Embalagem Primaria",
            observacao_lote
        ))
        caixa_id = cursor.fetchone()["id"]
    else:
        cursor.execute(q("""
        INSERT INTO pa_caixas (
            codigo_caixa,
            sku,
            data_fabricacao,
            data_validade,
            peso_bruto,
            peso_liquido,
            quantidade_bandejas,
            status,
            origem,
            observacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """), (
            codigo_lote,
            op["sku"] or "Galinha Inteira",
            data_fabricacao,
            data_validade,
            kg_produzidos,
            kg_produzidos,
            unidades_vendaveis,
            "Em estoque",
            "Embalagem Primaria",
            observacao_lote
        ))
        caixa_id = cursor.lastrowid

    cursor.execute(q("""
    INSERT INTO pa_caixa_composicao (
        caixa_id,
        op_id,
        quantidade_bandejas
    ) VALUES (?, ?, ?)
    """), (caixa_id, op["id"], unidades_vendaveis))

    return codigo_lote


def registrar_apontamento_embalagem_primaria(op, quantidade_bandejas, observacoes="", kg_produzidos=None, pacotes_1_ave=0, pacotes_2_aves=0):
    criar_tabelas_estoque_pi_pa()

    if not op:
        raise ValueError("OP não encontrada.")

    if op["status"] == "Encerrada":
        raise ValueError("Esta OP já está encerrada.")

    if op_possui_caixa_pa(op["id"]):
        raise ValueError("Esta OP já possui caixas PA vinculadas. Não é possível reapontar a Embalagem Primária.")

    sku = op["sku"] or "Galinha Cortada"
    bandejas = float(quantidade_bandejas or 0)
    if bandejas <= 0 and not sku_sem_embalagem_secundaria(sku):
        raise ValueError("Informe uma quantidade válida de bandejas produzidas.")

    if sku_sem_embalagem_secundaria(sku):
        pacotes_1 = parse_numero_form(pacotes_1_ave or 0)
        pacotes_2 = parse_numero_form(pacotes_2_aves or 0)
        aves_embaladas = pacotes_1 + (pacotes_2 * 2)
        unidades_vendaveis = pacotes_1 + pacotes_2
        kg_final = parse_numero_form(kg_produzidos or 0)

        if aves_embaladas <= 0:
            aves_embaladas = bandejas
            unidades_vendaveis = bandejas

        if aves_embaladas <= 0:
            raise ValueError("Informe as aves embaladas ou os pacotes de Galinha Inteira.")

        if unidades_vendaveis <= 0:
            raise ValueError("Informe uma quantidade valida de unidades vendaveis.")

        if kg_final <= 0:
            raise ValueError("Informe o peso liquido produzido em kg para calcular o rendimento da OP.")

        conn = conectar()
        cursor = conn.cursor()
        perdas = calcular_perdas_aves_op(cursor, op["id"])
        conn.close()

        validar_balanco_aves_op(op, aves_embaladas, perdas)
        remover_movimentacoes_estoque_pi_por_op(op["id"])

        observacao_final = observacoes or f"Pacotes 1 ave: {pacotes_1:g} | Pacotes 2 aves: {pacotes_2:g}"

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute(q("""
        INSERT INTO embalagem_primaria_apontamentos (
            op_id,
            data_apontamento,
            sku,
            quantidade_bandejas,
            observacoes
        ) VALUES (?, ?, ?, ?, ?)
        """), (
            op["id"],
            op["data"],
            sku,
            aves_embaladas,
            observacao_final
        ))

        codigo_lote = registrar_lote_pa_galinha_inteira(
            cursor,
            op,
            unidades_vendaveis,
            aves_embaladas,
            kg_final,
            observacao_final
        )

        cursor.execute(q("""
        UPDATE ordens_producao
        SET status = ?
        WHERE id = ?
        """), ("Encerrada", op["id"]))

        conn.commit()
        conn.close()

        gerar_producao_automatica_setores(
            op=op,
            data_lancamento=op["data"],
            hora_inicio="N/A",
            hora_fim="N/A",
            unidades_produzidas=unidades_vendaveis,
            kg_produzidos=kg_final,
            descontar_almoco=False
        )

        return {
            "tipo": "encerramento_primaria",
            "codigo_lote": codigo_lote,
            "aves_embaladas": aves_embaladas,
            "unidades_vendaveis": unidades_vendaveis,
            "kg_produzidos": kg_final,
        }

    # Substitui o apontamento anterior da mesma OP, se existir.
    remover_movimentacoes_estoque_pi_por_op(op["id"])

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    INSERT INTO embalagem_primaria_apontamentos (
        op_id,
        data_apontamento,
        sku,
        quantidade_bandejas,
        observacoes
    ) VALUES (?, ?, ?, ?, ?)
    """), (
        op["id"],
        op["data"],
        sku,
        bandejas,
        observacoes or ""
    ))

    cursor.execute(q("""
    INSERT INTO estoque_produto_intermediario (
        data_movimentacao,
        tipo,
        op_id,
        sku,
        quantidade_bandejas,
        origem,
        observacoes
    ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """), (
        op["data"],
        "ENTRADA_EMBALAGEM_PRIMARIA",
        op["id"],
        sku,
        bandejas,
        "Embalagem Primária",
        "PI gerado pelo apontamento da Embalagem Primária. A OP permanece aberta até a validação da Embalagem Secundária."
    ))

    # A OP ainda não está encerrada. O status apenas sinaliza que há PI aguardando embalagem secundária.
    cursor.execute(q("""
    UPDATE ordens_producao
    SET status = ?
    WHERE id = ?
    """), ("Aguardando Embalagem Secundária", op["id"]))

    conn.commit()
    conn.close()

    return {
        "tipo": "pi",
        "bandejas": bandejas,
    }


def buscar_apontamentos_embalagem_primaria(limite=50):
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT ep.*, op.data AS data_op, op.fornecedor, op.quantidade_aves
    FROM embalagem_primaria_apontamentos ep
    LEFT JOIN ordens_producao op ON op.id = ep.op_id
    ORDER BY ep.id DESC
    LIMIT ?
    """), (limite,))

    linhas = cursor.fetchall()
    conn.close()
    return linhas


def buscar_apontamento_embalagem_primaria_por_op(op_id):
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT ep.*, op.data AS data_op, op.fornecedor, op.quantidade_aves
    FROM embalagem_primaria_apontamentos ep
    LEFT JOIN ordens_producao op ON op.id = ep.op_id
    WHERE ep.op_id = ?
    ORDER BY ep.id DESC
    LIMIT 1
    """), (op_id,))

    linha = cursor.fetchone()
    conn.close()
    return linha


def buscar_ops_para_embalagem_primaria():
    criar_banco()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT *
    FROM ordens_producao
    WHERE status <> ?
    ORDER BY data DESC, id DESC
    """), ("Encerrada",))

    ops = cursor.fetchall()
    conn.close()
    return ops


def buscar_saldos_estoque_pi():
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT
        pi.op_id,
        op.data AS data_op,
        pi.sku,
        COALESCE(SUM(
            CASE
                WHEN pi.tipo LIKE 'ENTRADA%' THEN pi.quantidade_bandejas
                WHEN pi.tipo LIKE 'SAIDA%' THEN -pi.quantidade_bandejas
                ELSE pi.quantidade_bandejas
            END
        ), 0) AS saldo_bandejas
    FROM estoque_produto_intermediario pi
    LEFT JOIN ordens_producao op ON op.id = pi.op_id
    GROUP BY pi.op_id, op.data, pi.sku
    HAVING COALESCE(SUM(
            CASE
                WHEN pi.tipo LIKE 'ENTRADA%' THEN pi.quantidade_bandejas
                WHEN pi.tipo LIKE 'SAIDA%' THEN -pi.quantidade_bandejas
                ELSE pi.quantidade_bandejas
            END
        ), 0) <> 0
    ORDER BY op.data ASC, pi.op_id ASC
    """))

    saldos = cursor.fetchall()
    conn.close()
    return saldos


def buscar_movimentacoes_estoque_pi(limite=50):
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT *
    FROM estoque_produto_intermediario
    ORDER BY data_movimentacao DESC, id DESC
    LIMIT ?
    """), (limite,))

    movimentacoes = cursor.fetchall()
    conn.close()
    return movimentacoes


def buscar_caixas_pa(limite=80):
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT *
    FROM pa_caixas
    ORDER BY id DESC
    LIMIT ?
    """), (limite,))

    caixas = cursor.fetchall()
    conn.close()
    return caixas


def buscar_ops_com_saldo_pi():
    return buscar_saldos_estoque_pi()


def gerar_codigo_caixa():
    criar_tabelas_estoque_pi_pa()

    hoje = datetime.now().strftime("%Y%m%d")
    prefixo = f"CX-{hoje}-"

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT codigo_caixa
    FROM pa_caixas
    WHERE codigo_caixa LIKE ?
    ORDER BY codigo_caixa DESC
    LIMIT 1
    """), (f"{prefixo}%",))

    ultima = cursor.fetchone()
    conn.close()

    if not ultima:
        return f"{prefixo}001"

    try:
        sequencia = int(str(ultima["codigo_caixa"]).split("-")[-1]) + 1
    except Exception:
        sequencia = 1

    return f"{prefixo}{sequencia:03d}"


def obter_sku_op(op_id):
    op = buscar_op_por_id(op_id)
    if not op:
        return None
    return op["sku"] or "Galinha Cortada"


def registrar_saida_pi_por_caixa(cursor, op_id, sku, bandejas, caixa_id):
    cursor.execute(q("""
    INSERT INTO estoque_produto_intermediario (
        data_movimentacao,
        tipo,
        op_id,
        sku,
        quantidade_bandejas,
        origem,
        observacoes
    ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """), (
        datetime.now().strftime("%Y-%m-%d"),
        "SAIDA_EMBALAGEM_SECUNDARIA",
        op_id,
        sku,
        float(bandejas or 0),
        "Embalagem Secundária",
        f"Bandejas consumidas na formação da caixa PA #{caixa_id}."
    ))


def parse_numero_form(valor):
    if valor is None:
        return 0.0

    texto = str(valor).strip()
    if not texto:
        return 0.0

    texto = texto.replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        raise ValueError(f"Valor numérico inválido: {valor}")


def calcular_validade_padrao(data_fabricacao):
    if not data_fabricacao:
        return ""

    try:
        data_base = datetime.strptime(data_fabricacao, "%Y-%m-%d")
        try:
            validade = data_base.replace(year=data_base.year + 1)
        except ValueError:
            validade = data_base + timedelta(days=365)
        return validade.strftime("%Y-%m-%d")
    except Exception:
        return ""


def preparar_composicao_caixa(form):
    op_principal = int(form.get("op_principal") or 0)
    bandejas_principal = parse_numero_form(form.get("bandejas_principal") or 0)
    op_complementar_raw = form.get("op_complementar") or ""
    bandejas_complementar = parse_numero_form(form.get("bandejas_complementar") or 0)

    composicao = []

    if op_principal and bandejas_principal > 0:
        composicao.append((op_principal, bandejas_principal))

    if op_complementar_raw and bandejas_complementar > 0:
        op_complementar = int(op_complementar_raw)
        if op_complementar == op_principal:
            raise ValueError("A OP complementar deve ser diferente da OP principal.")
        composicao.append((op_complementar, bandejas_complementar))

    if not composicao:
        raise ValueError("Informe ao menos uma OP e a quantidade de bandejas utilizadas.")

    total_bandejas = sum(qtd for _, qtd in composicao)

    if total_bandejas <= 0 or total_bandejas > BANDEJAS_POR_CAIXA:
        raise ValueError(f"A caixa deve conter entre 1 e {BANDEJAS_POR_CAIXA} bandejas.")

    skus = {obter_sku_op(op_id) for op_id, _ in composicao}
    skus.discard(None)

    if not skus:
        raise ValueError("Não foi possível identificar o SKU das OPs informadas.")

    if len(skus) > 1:
        raise ValueError("Não é permitido formar uma caixa com SKUs diferentes.")

    return composicao, total_bandejas, list(skus)[0]


def validar_saldo_pi_para_composicoes(composicoes):
    consumo_por_op = {}
    for composicao in composicoes:
        for op_id, qtd in composicao:
            consumo_por_op[op_id] = consumo_por_op.get(op_id, 0) + float(qtd or 0)

    saldos = {int(item["op_id"]): float(item["saldo_bandejas"] or 0) for item in buscar_saldos_estoque_pi()}
    for op_id, qtd_total in consumo_por_op.items():
        saldo_disponivel = saldos.get(op_id, 0)
        if qtd_total > saldo_disponivel:
            raise ValueError(
                f"A OP #{op_id} possui apenas {saldo_disponivel:g} bandejas disponíveis em PI. "
                f"O lançamento tentaria consumir {qtd_total:g}."
            )


def inserir_caixa_pa(cursor, codigo_caixa, sku, data_fabricacao, data_validade, peso_bruto, peso_liquido, total_bandejas, observacoes, composicao):
    if DATABASE_URL:
        cursor.execute(q("""
        INSERT INTO pa_caixas (
            codigo_caixa,
            sku,
            data_fabricacao,
            data_validade,
            peso_bruto,
            peso_liquido,
            quantidade_bandejas,
            status,
            origem,
            observacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        RETURNING id
        """), (
            codigo_caixa,
            sku,
            data_fabricacao or "",
            data_validade or "",
            peso_bruto,
            peso_liquido,
            total_bandejas,
            "Em estoque",
            "Embalagem Secundária",
            observacoes or ""
        ))
        caixa_id = cursor.fetchone()["id"]
    else:
        cursor.execute(q("""
        INSERT INTO pa_caixas (
            codigo_caixa,
            sku,
            data_fabricacao,
            data_validade,
            peso_bruto,
            peso_liquido,
            quantidade_bandejas,
            status,
            origem,
            observacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """), (
            codigo_caixa,
            sku,
            data_fabricacao or "",
            data_validade or "",
            peso_bruto,
            peso_liquido,
            total_bandejas,
            "Em estoque",
            "Embalagem Secundária",
            observacoes or ""
        ))
        caixa_id = cursor.lastrowid

    for op_id, qtd in composicao:
        cursor.execute(q("""
        INSERT INTO pa_caixa_composicao (
            caixa_id,
            op_id,
            quantidade_bandejas
        ) VALUES (?, ?, ?)
        """), (caixa_id, op_id, qtd))

        registrar_saida_pi_por_caixa(cursor, op_id, sku, qtd, caixa_id)

    return caixa_id


def registrar_caixa_pa_manual(form):
    criar_tabelas_estoque_pi_pa()

    composicao, total_bandejas, sku = preparar_composicao_caixa(form)
    validar_saldo_pi_para_composicoes([composicao])

    peso_bruto = parse_numero_form(form.get("peso_bruto") or 0)
    peso_liquido = parse_numero_form(form.get("peso_liquido") or 0)

    if peso_bruto <= 0:
        raise ValueError("Informe o peso bruto da caixa.")

    if peso_liquido <= 0:
        peso_liquido = peso_bruto - 0.5

    if peso_liquido <= 0:
        raise ValueError("O peso líquido calculado precisa ser maior que zero.")

    data_fabricacao = form.get("data_fabricacao") or datetime.now().strftime("%Y-%m-%d")
    data_validade = form.get("data_validade") or calcular_validade_padrao(data_fabricacao)
    codigo_caixa = form.get("codigo_caixa") or gerar_codigo_caixa()

    conn = conectar()
    cursor = conn.cursor()

    inserir_caixa_pa(
        cursor,
        codigo_caixa,
        sku,
        data_fabricacao,
        data_validade,
        peso_bruto,
        peso_liquido,
        total_bandejas,
        form.get("observacoes") or "",
        composicao
    )

    conn.commit()
    conn.close()

    return codigo_caixa


def registrar_caixas_pa_lote(form):
    criar_tabelas_estoque_pi_pa()

    composicao, total_bandejas, sku = preparar_composicao_caixa(form)
    linhas = [linha.strip() for linha in (form.get("pesos_brutos_lote") or "").splitlines() if linha.strip()]

    if not linhas:
        raise ValueError("Informe ao menos um peso bruto no lançamento em lote.")

    pesos_brutos = [parse_numero_form(linha) for linha in linhas]
    pesos_liquidos_raw = [linha.strip() for linha in (form.get("pesos_liquidos_lote") or "").splitlines() if linha.strip()]

    if pesos_liquidos_raw and len(pesos_liquidos_raw) != len(pesos_brutos):
        raise ValueError("A lista de pesos líquidos editados deve ter a mesma quantidade de linhas da lista de pesos brutos.")

    pesos_liquidos = []
    for indice, peso_bruto in enumerate(pesos_brutos):
        if peso_bruto <= 0:
            raise ValueError("Todos os pesos brutos do lote precisam ser maiores que zero.")

        if pesos_liquidos_raw:
            peso_liquido = parse_numero_form(pesos_liquidos_raw[indice])
        else:
            peso_liquido = peso_bruto - 0.5

        if peso_liquido <= 0:
            raise ValueError("Todos os pesos líquidos do lote precisam ser maiores que zero.")
        pesos_liquidos.append(peso_liquido)

    validar_saldo_pi_para_composicoes([composicao for _ in pesos_brutos])

    data_fabricacao = form.get("data_fabricacao") or datetime.now().strftime("%Y-%m-%d")
    data_validade = form.get("data_validade") or calcular_validade_padrao(data_fabricacao)
    observacoes = form.get("observacoes") or "Lançamento em lote na Embalagem Secundária"

    codigos = []
    primeiro_codigo = gerar_codigo_caixa()
    try:
        prefixo_codigo = "-".join(primeiro_codigo.split("-")[:-1])
        sequencia_codigo = int(primeiro_codigo.split("-")[-1])
    except Exception:
        prefixo_codigo = f"CX-{datetime.now().strftime('%Y%m%d')}"
        sequencia_codigo = 1

    conn = conectar()
    cursor = conn.cursor()

    for peso_bruto, peso_liquido in zip(pesos_brutos, pesos_liquidos):
        codigo_caixa = f"{prefixo_codigo}-{sequencia_codigo:03d}"
        sequencia_codigo += 1
        inserir_caixa_pa(
            cursor,
            codigo_caixa,
            sku,
            data_fabricacao,
            data_validade,
            peso_bruto,
            peso_liquido,
            total_bandejas,
            observacoes,
            composicao
        )
        codigos.append(codigo_caixa)

    conn.commit()
    conn.close()

    return codigos


def calcular_fechamento_industrial_op(op_id):
    """
    Consolida as validações necessárias para finalizar a Embalagem Secundária
    e encerrar oficialmente a OP.

    Regras:
    1) Aves vivas = bandejas apontadas na Embalagem Primária + descartes/condenações + mortes na gaiola.
    2) Todas as bandejas apontadas na Embalagem Primária precisam ter sido pesadas na Embalagem Secundária.
    3) O peso oficial da OP nasce da soma dos pesos líquidos das caixas PA vinculadas à OP.
    """
    criar_banco()
    criar_tabelas_estoque_pi_pa()

    op = buscar_op_por_id(op_id)
    if not op:
        raise ValueError("OP não encontrada.")

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT COALESCE(SUM(quantidade_bandejas), 0) AS total
    FROM embalagem_primaria_apontamentos
    WHERE op_id = ?
    """), (op_id,))
    bandejas_primaria = float(cursor.fetchone()["total"] or 0)

    cursor.execute(q("""
    SELECT
        COALESCE(SUM(CASE
            WHEN LOWER(COALESCE(categoria, '')) LIKE '%%conden%%' THEN quantidade
            ELSE 0
        END), 0) AS condenacoes,
        COALESCE(SUM(CASE
            WHEN LOWER(COALESCE(categoria, '')) NOT LIKE '%%conden%%'
             AND LOWER(TRIM(COALESCE(motivo, ''))) <> 'morte na gaiola' THEN quantidade
            ELSE 0
        END), 0) AS descartes,
        COALESCE(SUM(CASE
            WHEN LOWER(TRIM(COALESCE(motivo, ''))) = 'morte na gaiola' THEN quantidade
            ELSE 0
        END), 0) AS mortes_na_gaiola
    FROM apontamentos_descartes
    WHERE op_id = ?
      AND LOWER(unidade) IN ('aves', 'ave', 'unidade', 'unidades')
    """), (op_id,))
    perdas = cursor.fetchone()
    condenacoes = float(perdas["condenacoes"] or 0)
    descartes = float(perdas["descartes"] or 0)
    mortes_na_gaiola_descartes = float(perdas["mortes_na_gaiola"] or 0)

    cursor.execute(q("""
    SELECT
        COUNT(DISTINCT cx.id) AS caixas,
        COALESCE(SUM(comp.quantidade_bandejas), 0) AS bandejas_consumidas,
        COALESCE(SUM(cx.peso_liquido), 0) AS peso_liquido_total,
        COALESCE(SUM(cx.peso_bruto), 0) AS peso_bruto_total
    FROM pa_caixa_composicao comp
    INNER JOIN pa_caixas cx ON cx.id = comp.caixa_id
    WHERE comp.op_id = ?
    """), (op_id,))
    caixas = cursor.fetchone()

    conn.close()

    aves_vivas = float(op["quantidade_aves"] or 0)
    mortes_antes_pendura = float(op["mortes_antes_pendura"] or 0) + mortes_na_gaiola_descartes
    caixas_qtd = int(caixas["caixas"] or 0)
    bandejas_consumidas = float(caixas["bandejas_consumidas"] or 0)
    peso_liquido_total = float(caixas["peso_liquido_total"] or 0)
    peso_bruto_total = float(caixas["peso_bruto_total"] or 0)

    total_fechamento_aves = bandejas_primaria + descartes + condenacoes + mortes_antes_pendura
    saldo_aves = aves_vivas - total_fechamento_aves
    saldo_pi = bandejas_primaria - bandejas_consumidas

    tolerancia = 0.0001
    aves_ok = abs(saldo_aves) <= tolerancia
    pi_ok = abs(saldo_pi) <= tolerancia
    possui_peso = peso_liquido_total > 0 and caixas_qtd > 0
    pode_encerrar = aves_ok and pi_ok and possui_peso and op["status"] != "Encerrada"

    pendencias = []
    if not aves_ok:
        pendencias.append(
            f"Balanço de aves divergente em {saldo_aves:g} aves. Revise Embalagem Primária, descartes, condenações ou mortes na gaiola."
        )
    if not pi_ok:
        if saldo_pi > 0:
            pendencias.append(
                f"Existem {saldo_pi:g} bandejas sem pesagem. Lance uma caixa parcial antes de encerrar."
            )
        else:
            pendencias.append(
                f"A Embalagem Secundária consumiu {-saldo_pi:g} bandejas a mais que o PI produzido. Revise as caixas lançadas."
            )
    if not possui_peso:
        pendencias.append("Nenhuma caixa com peso líquido foi registrada para esta OP.")
    if op["status"] == "Encerrada":
        pendencias.append("Esta OP já está encerrada.")

    return {
        "op": op,
        "aves_vivas": aves_vivas,
        "mortes_antes_pendura": mortes_antes_pendura,
        "bandejas_primaria": bandejas_primaria,
        "descartes": descartes,
        "condenacoes": condenacoes,
        "total_fechamento_aves": total_fechamento_aves,
        "saldo_aves": saldo_aves,
        "aves_ok": aves_ok,
        "caixas": caixas_qtd,
        "bandejas_consumidas": bandejas_consumidas,
        "saldo_pi": saldo_pi,
        "pi_ok": pi_ok,
        "peso_liquido_total": peso_liquido_total,
        "peso_bruto_total": peso_bruto_total,
        "possui_peso": possui_peso,
        "pode_encerrar": pode_encerrar,
        "pendencias": pendencias,
    }


def finalizar_embalagem_secundaria_op(op_id):
    fechamento = calcular_fechamento_industrial_op(op_id)

    if not fechamento["pode_encerrar"]:
        raise ValueError("Não foi possível encerrar a OP: " + " ".join(fechamento["pendencias"]))

    op = fechamento["op"]
    unidades_produzidas = fechamento["bandejas_consumidas"]
    kg_produzidos = fechamento["peso_liquido_total"]

    gerar_producao_automatica_setores(
        op=op,
        data_lancamento=op["data"],
        hora_inicio="N/A",
        hora_fim="N/A",
        unidades_produzidas=unidades_produzidas,
        kg_produzidos=kg_produzidos,
        descontar_almoco=False
    )

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    UPDATE ordens_producao
    SET status = ?
    WHERE id = ?
    """), ("Encerrada", op_id))

    conn.commit()
    conn.close()

    return fechamento


def validar_reset_processamento_op(op_id):
    """
    Valida se a OP pode voltar ao estado anterior à Embalagem Primária.

    Esta rotina não é uma reabertura de OP encerrada. Ela serve para desfazer
    processamento operacional parcial: Embalagem Primária, PI, Embalagem
    Secundária e caixas PA ainda não expedidas.
    """
    criar_banco()
    criar_tabelas_estoque_pi_pa()

    op = buscar_op_por_id(op_id)
    if not op:
        raise ValueError("OP não encontrada.")

    if (op["status"] or "") == "Encerrada":
        raise ValueError("Esta OP já está encerrada. O reset operacional só é permitido antes do encerramento.")

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT DISTINCT caixa_id
    FROM pa_caixa_composicao
    WHERE op_id = ?
    """), (op_id,))
    caixas_ids = [linha["caixa_id"] for linha in cursor.fetchall()]

    if caixas_ids:
        placeholders = ",".join(["?"] * len(caixas_ids))

        cursor.execute(q(f"""
        SELECT caixa_id, COUNT(DISTINCT op_id) AS total_ops
        FROM pa_caixa_composicao
        WHERE caixa_id IN ({placeholders})
        GROUP BY caixa_id
        HAVING COUNT(DISTINCT op_id) > 1
        """), tuple(caixas_ids))
        caixas_mistas = cursor.fetchall()
        if caixas_mistas:
            conn.close()
            raise ValueError("Não é possível resetar esta OP porque existem caixas mistas vinculadas a outras OPs.")

        cursor.execute(q(f"""
        SELECT codigo_caixa, status
        FROM pa_caixas
        WHERE id IN ({placeholders})
          AND COALESCE(status, '') <> ?
        """), tuple(caixas_ids) + ("Em estoque",))
        caixas_indisponiveis = cursor.fetchall()
        if caixas_indisponiveis:
            conn.close()
            raise ValueError("Não é possível resetar esta OP porque existem caixas que não estão mais em estoque.")

    conn.close()
    return op, caixas_ids


def resetar_processamento_op(op_id, confirmacao):
    if str(confirmacao or "").strip().upper() != "RESETAR":
        raise ValueError("Digite RESETAR para confirmar o reset da OP.")

    op, caixas_ids = validar_reset_processamento_op(op_id)

    conn = conectar()
    cursor = conn.cursor()

    try:
        if caixas_ids:
            placeholders = ",".join(["?"] * len(caixas_ids))

            cursor.execute(q(f"""
            DELETE FROM pa_caixa_composicao
            WHERE caixa_id IN ({placeholders})
            """), tuple(caixas_ids))

            cursor.execute(q(f"""
            DELETE FROM pa_caixas
            WHERE id IN ({placeholders})
            """), tuple(caixas_ids))

        cursor.execute(q("""
        DELETE FROM estoque_produto_intermediario
        WHERE op_id = ?
        """), (op_id,))

        cursor.execute(q("""
        DELETE FROM embalagem_primaria_apontamentos
        WHERE op_id = ?
        """), (op_id,))

        cursor.execute(q("""
        UPDATE ordens_producao
        SET status = ?
        WHERE id = ?
        """), ("Aberta", op_id))

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "op": op,
        "caixas_removidas": len(caixas_ids),
    }


def buscar_resumo_pa_completo():
    """
    Calcula o resumo de Produto Acabado usando a base completa de caixas.

    Importante: a listagem visual de caixas pode continuar limitada, mas os
    cards de saldo não podem depender da lista exibida na tela. Por isso o
    resumo é apurado diretamente em pa_caixas, sem LIMIT.
    """
    criar_tabelas_estoque_pi_pa()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT
        COALESCE(COUNT(CASE WHEN status = ? THEN 1 END), 0) AS saldo_pa_caixas,
        COALESCE(SUM(CASE WHEN status = ? THEN quantidade_bandejas ELSE 0 END), 0) AS saldo_pa_bandejas,
        COALESCE(SUM(CASE WHEN status = ? THEN peso_liquido ELSE 0 END), 0) AS saldo_pa_kg
    FROM pa_caixas
    """), ("Em estoque", "Em estoque", "Em estoque"))

    resumo = cursor.fetchone()
    conn.close()

    return {
        "saldo_pa_caixas": int(resumo["saldo_pa_caixas"] or 0),
        "saldo_pa_bandejas": float(resumo["saldo_pa_bandejas"] or 0),
        "saldo_pa_kg": float(resumo["saldo_pa_kg"] or 0),
    }


def calcular_resumo_estoques_pi_pa(saldos_pi, caixas_pa=None):
    saldo_pi_bandejas = sum(float(item["saldo_bandejas"] or 0) for item in saldos_pi)
    resumo_pa = buscar_resumo_pa_completo()

    return {
        "saldo_pi_bandejas": saldo_pi_bandejas,
        "ops_com_pi": len({item["op_id"] for item in saldos_pi}),
        "saldo_pa_caixas": resumo_pa["saldo_pa_caixas"],
        "saldo_pa_bandejas": resumo_pa["saldo_pa_bandejas"],
        "saldo_pa_kg": resumo_pa["saldo_pa_kg"],
    }


# Compatibilidade temporária com chamadas antigas do Sprint PA-1.
@executar_rotina_estrutural_uma_vez
def criar_tabela_estoque_produto_acabado():
    criar_tabelas_estoque_pi_pa()


def remover_movimentacoes_estoque_pa_por_op(op_id):
    remover_movimentacoes_estoque_pi_por_op(op_id)


def registrar_entrada_estoque_pa_op(op, unidades_produzidas, kg_produzidos=None):
    registrar_entrada_estoque_pi_op(op, unidades_produzidas)


def buscar_saldos_estoque_pa():
    return buscar_saldos_estoque_pi()


def buscar_movimentacoes_estoque_pa(limite=50):
    return buscar_movimentacoes_estoque_pi(limite)


def calcular_resumo_estoque_pa(saldos):
    caixas = buscar_caixas_pa()
    return calcular_resumo_estoques_pi_pa(saldos, caixas)


def buscar_expedicoes(data_inicio=None, data_fim=None, status=None):
    criar_tabelas_expedicao()

    filtros = []
    parametros = []

    if data_inicio:
        filtros.append("data >= ?")
        parametros.append(data_inicio)

    if data_fim:
        filtros.append("data <= ?")
        parametros.append(data_fim)

    if status and status != "Todos":
        filtros.append("status = ?")
        parametros.append(status)

    where = ""
    if filtros:
        where = "WHERE " + " AND ".join(filtros)

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(f"""
    SELECT
        e.*,
        COALESCE(COUNT(i.id), 0) as total_itens,
        COALESCE(SUM(i.quantidade_unidades), 0) as total_unidades,
        COALESCE(SUM(i.quantidade_kg), 0) as total_kg
    FROM expedicoes e
    LEFT JOIN expedicao_itens i ON i.expedicao_id = e.id
    {where}
    GROUP BY e.id, e.numero_romaneio, e.data, e.tipo_movimentacao, e.destino, e.responsavel, e.observacoes, e.status, e.criado_em
    ORDER BY e.data DESC, e.id DESC
    """), tuple(parametros))

    expedicoes = cursor.fetchall()
    conn.close()
    return expedicoes


def calcular_resumo_expedicao(expedicoes):
    total_romaneios = len(expedicoes)
    abertos = sum(1 for item in expedicoes if item["status"] == "Aberto")
    concluidos = sum(1 for item in expedicoes if item["status"] == "Concluído")
    cancelados = sum(1 for item in expedicoes if item["status"] == "Cancelado")

    total_unidades = sum(float(item["total_unidades"] or 0) for item in expedicoes)
    total_kg = sum(float(item["total_kg"] or 0) for item in expedicoes)

    return {
        "total_romaneios": total_romaneios,
        "abertos": abertos,
        "concluidos": concluidos,
        "cancelados": cancelados,
        "total_unidades": round(total_unidades, 2),
        "total_kg": round(total_kg, 2)
    }




def normalizar_competencia(competencia):
    if not competencia:
        return ""

    competencia = str(competencia).strip()

    if len(competencia) >= 7:
        return competencia[:7]

    return competencia


def listar_competencias_periodo(competencia_inicio, competencia_fim):
    inicio = datetime.strptime(competencia_inicio + "-01", "%Y-%m-%d")
    fim = datetime.strptime(competencia_fim + "-01", "%Y-%m-%d")

    competencias = []
    atual = inicio

    while atual <= fim:
        competencias.append(atual.strftime("%Y-%m"))

        if atual.month == 12:
            atual = atual.replace(year=atual.year + 1, month=1)
        else:
            atual = atual.replace(month=atual.month + 1)

    return competencias


def buscar_dados_relatorio_custos(competencia_inicio, competencia_fim, categoria_filtro="Todas"):
    criar_tabelas_custos()

    competencias = listar_competencias_periodo(
        competencia_inicio,
        competencia_fim
    )

    categoria_filtro = categoria_filtro or "Todas"
    categorias_padrao = CATEGORIAS_CUSTOS

    dados_por_categoria_completo = {
        categoria: {competencia: 0 for competencia in competencias}
        for categoria in categorias_padrao
    }

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT
        competencia,
        categoria,
        COALESCE(SUM(valor), 0) as total
    FROM custos_mensais
    WHERE competencia BETWEEN ? AND ?
    GROUP BY competencia, categoria
    ORDER BY competencia, categoria
    """), (
        competencia_inicio,
        competencia_fim
    ))

    registros = cursor.fetchall()
    conn.close()

    categorias_encontradas = set()

    for item in registros:
        competencia = normalizar_competencia(item["competencia"])
        categoria = item["categoria"]
        categorias_encontradas.add(categoria)

        if categoria not in dados_por_categoria_completo:
            dados_por_categoria_completo[categoria] = {
                comp: 0 for comp in competencias
            }

        if competencia in dados_por_categoria_completo[categoria]:
            dados_por_categoria_completo[categoria][competencia] = float(item["total"] or 0)

    categorias_disponiveis = list(categorias_padrao)

    for categoria in sorted(categorias_encontradas):
        if categoria not in categorias_disponiveis:
            categorias_disponiveis.append(categoria)

    if categoria_filtro != "Todas":
        if categoria_filtro not in dados_por_categoria_completo:
            dados_por_categoria_completo[categoria_filtro] = {
                comp: 0 for comp in competencias
            }

        dados_por_categoria = {
            categoria_filtro: dados_por_categoria_completo[categoria_filtro]
        }
    else:
        dados_por_categoria = {
            categoria: dados_por_categoria_completo.get(
                categoria,
                {comp: 0 for comp in competencias}
            )
            for categoria in categorias_disponiveis
        }

    totais_por_categoria = {
        categoria: sum(valores.values())
        for categoria, valores in dados_por_categoria.items()
    }

    totais_por_competencia = {
        competencia: sum(
            dados_por_categoria[categoria].get(competencia, 0)
            for categoria in dados_por_categoria
        )
        for competencia in competencias
    }

    custo_total = sum(totais_por_categoria.values())
    media_mensal = custo_total / len(competencias) if competencias else 0

    maior_categoria = "Sem dados"
    valor_maior_categoria = 0

    if totais_por_categoria:
        maior_categoria = max(
            totais_por_categoria,
            key=lambda categoria: totais_por_categoria[categoria]
        )
        valor_maior_categoria = totais_por_categoria.get(maior_categoria, 0)

        if valor_maior_categoria == 0:
            maior_categoria = "Sem dados"

    maior_crescimento_categoria = "Sem dados"
    maior_crescimento_valor = 0

    for categoria, valores in dados_por_categoria.items():
        lista_valores = [valores.get(comp, 0) for comp in competencias]

        if len(lista_valores) < 2:
            continue

        crescimento = lista_valores[-1] - lista_valores[0]

        if crescimento > maior_crescimento_valor:
            maior_crescimento_valor = crescimento
            maior_crescimento_categoria = categoria

    # Gráfico executivo: exibe apenas as 5 maiores categorias do período
    # e agrupa as demais em "Outras Categorias".
    categorias_com_movimento = [
        categoria
        for categoria, total in sorted(
            totais_por_categoria.items(),
            key=lambda item: item[1],
            reverse=True
        )
        if float(total or 0) > 0
    ]

    categorias_principais = categorias_com_movimento[:5]
    categorias_restantes = categorias_com_movimento[5:]

    datasets = []

    for categoria in categorias_principais:
        valores = dados_por_categoria.get(categoria, {})
        datasets.append({
            "label": categoria,
            "data": [
                round(valores.get(competencia, 0), 2)
                for competencia in competencias
            ]
        })

    if categorias_restantes:
        datasets.append({
            "label": f"Outras Categorias ({len(categorias_restantes)})",
            "data": [
                round(
                    sum(
                        dados_por_categoria.get(categoria, {}).get(competencia, 0)
                        for categoria in categorias_restantes
                    ),
                    2
                )
                for competencia in competencias
            ]
        })

    resumo_categorias = []

    for categoria, total in sorted(
        totais_por_categoria.items(),
        key=lambda item: item[1],
        reverse=True
    ):
        percentual = 0

        if custo_total > 0:
            percentual = (total / custo_total) * 100

        resumo_categorias.append({
            "categoria": categoria,
            "total": round(total, 2),
            "percentual": round(percentual, 2)
        })

    return {
        "competencias": competencias,
        "datasets": datasets,
        "custo_total": round(custo_total, 2),
        "media_mensal": round(media_mensal, 2),
        "maior_categoria": maior_categoria,
        "valor_maior_categoria": round(valor_maior_categoria, 2),
        "maior_crescimento_categoria": maior_crescimento_categoria,
        "maior_crescimento_valor": round(maior_crescimento_valor, 2),
        "totais_por_competencia": totais_por_competencia,
        "resumo_categorias": resumo_categorias,
        "categorias_disponiveis": categorias_disponiveis,
        "categoria_filtro": categoria_filtro
    }




def valor_linha_venda(item, campo, padrao=0):
    try:
        valor = item[campo]
    except Exception:
        valor = padrao

    if valor is None:
        return padrao

    return float(valor or 0)


def normalizar_venda_para_dre(item):
    sku = item["sku"]
    quantidade = valor_linha_venda(item, "quantidade")
    unidade = (item["unidade"] or "").lower()
    quantidade_unidades = valor_linha_venda(item, "quantidade_unidades")
    quantidade_kg = valor_linha_venda(item, "quantidade_kg")
    receita = valor_linha_venda(item, "receita")

    # Compatibilidade com registros antigos: antes a Galinha Cortada era lançada só em kg.
    if sku == "Galinha Cortada":
        if quantidade_kg <= 0 and unidade == "kg":
            quantidade_kg = quantidade

        if quantidade_unidades <= 0 and unidade in ["unidades", "unidade", "aves", "ave"]:
            quantidade_unidades = quantidade
    else:
        if quantidade_unidades <= 0:
            quantidade_unidades = quantidade

        quantidade_kg = 0

    return {
        "sku": sku,
        "receita": receita,
        "quantidade": quantidade,
        "unidade": unidade,
        "quantidade_unidades": quantidade_unidades,
        "quantidade_kg": quantidade_kg
    }


def buscar_dados_dre_gerencial(competencia):
    criar_tabelas_custos()
    criar_tabela_vendas()

    ano, mes = competencia.split("-")
    ultimo_dia = calendar.monthrange(int(ano), int(mes))[1]
    data_inicio = f"{competencia}-01"
    data_fim = f"{competencia}-{ultimo_dia:02d}"

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT *
    FROM vendas_diarias
    WHERE data BETWEEN ? AND ?
    ORDER BY sku ASC, data ASC, id ASC
    """), (data_inicio, data_fim))

    vendas_linhas = [normalizar_venda_para_dre(item) for item in cursor.fetchall()]

    vendas_por_sku_dict = {}
    receita_bruta = 0

    for item in vendas_linhas:
        sku = item["sku"]

        if sku not in vendas_por_sku_dict:
            vendas_por_sku_dict[sku] = {
                "receita": 0,
                "quantidade": 0,
                "quantidade_unidades": 0,
                "quantidade_kg": 0
            }

        vendas_por_sku_dict[sku]["receita"] += item["receita"]
        vendas_por_sku_dict[sku]["quantidade_unidades"] += item["quantidade_unidades"]
        vendas_por_sku_dict[sku]["quantidade_kg"] += item["quantidade_kg"]

        if sku == "Galinha Cortada":
            vendas_por_sku_dict[sku]["quantidade"] += item["quantidade_kg"]
        else:
            vendas_por_sku_dict[sku]["quantidade"] += item["quantidade_unidades"]

        receita_bruta += item["receita"]

    vendas_por_sku = []

    for sku, venda in vendas_por_sku_dict.items():
        quantidade_base = venda["quantidade_kg"] if sku == "Galinha Cortada" else venda["quantidade_unidades"]
        unidade_base = "kg" if sku == "Galinha Cortada" else "unidades"
        preco_medio = venda["receita"] / quantidade_base if quantidade_base > 0 else 0

        vendas_por_sku.append({
            "sku": sku,
            "receita": round(venda["receita"], 2),
            "quantidade": round(quantidade_base, 2),
            "unidade": unidade_base,
            "quantidade_unidades": round(venda["quantidade_unidades"], 2),
            "quantidade_kg": round(venda["quantidade_kg"], 2),
            "preco_medio": round(preco_medio, 4)
        })

    cursor.execute("SELECT * FROM parametros_custos")

    parametros = {
        item["sku"]: item
        for item in cursor.fetchall()
    }

    cmv_por_sku = []
    cmv_total = 0

    for sku, venda in vendas_por_sku_dict.items():
        parametros_sku = parametros.get(sku)

        custo_ave = 0
        custo_embalagem = 0

        if parametros_sku:
            custo_ave = float(parametros_sku["custo_ave"] or 0)
            custo_embalagem = float(parametros_sku["custo_embalagem"] or 0)

        quantidade_unidades = float(venda["quantidade_unidades"] or 0)
        quantidade_kg = float(venda["quantidade_kg"] or 0)

        # Regra atual validada:
        # Galinha Cortada: (ave viva + embalagem) x bandejas vendidas.
        # CMV por kg = CMV total / kg vendidos.
        # Galinha Inteira: 1 x 1 por unidade vendida.
        custo_materia_prima_unitario = custo_ave
        custo_embalagem_unitario = custo_embalagem
        quantidade_cmv = quantidade_unidades

        custo_materia_prima = quantidade_cmv * custo_materia_prima_unitario
        custo_embalagens = quantidade_cmv * custo_embalagem_unitario
        cmv_sku = custo_materia_prima + custo_embalagens
        cmv_total += cmv_sku

        cmv_por_kg = 0
        if sku == "Galinha Cortada" and quantidade_kg > 0:
            cmv_por_kg = cmv_sku / quantidade_kg

        cmv_por_unidade = 0
        if quantidade_cmv > 0:
            cmv_por_unidade = cmv_sku / quantidade_cmv

        cmv_por_sku.append({
            "sku": sku,
            "quantidade_vendida": round(quantidade_kg if sku == "Galinha Cortada" else quantidade_unidades, 2),
            "quantidade_unidades": round(quantidade_unidades, 2),
            "quantidade_kg": round(quantidade_kg, 2),
            "custo_materia_prima_unitario": round(custo_materia_prima_unitario, 4),
            "custo_embalagem_unitario": round(custo_embalagem_unitario, 4),
            "materia_prima": round(custo_materia_prima, 2),
            "embalagem": round(custo_embalagens, 2),
            "cmv": round(cmv_sku, 2),
            "cmv_por_kg": round(cmv_por_kg, 4),
            "cmv_por_unidade": round(cmv_por_unidade, 4),
            "observacao_calculo": "CMV por bandeja vendida; CMV/kg calculado pelos kg vendidos." if sku == "Galinha Cortada" else "CMV 1 x 1 por unidade vendida."
        })

    cursor.execute(q("""
    SELECT
        categoria,
        COALESCE(SUM(valor), 0) as total
    FROM custos_mensais
    WHERE competencia = ?
    GROUP BY categoria
    ORDER BY categoria
    """), (competencia,))

    custos_raw = cursor.fetchall()
    conn.close()

    categorias = CATEGORIAS_CUSTOS
    custos = {
        categoria: 0
        for categoria in categorias
    }

    for item in custos_raw:
        categoria = item["categoria"]
        valor = float(item["total"] or 0)
        custos[categoria] = custos.get(categoria, 0) + valor

    custos_operacionais_total = sum(custos.values())
    margem_bruta = receita_bruta - cmv_total
    resultado_operacional = margem_bruta - custos_operacionais_total

    def perc(valor):
        if receita_bruta > 0:
            return (valor / receita_bruta) * 100

        return 0

    linhas_custos = [
        {
            "categoria": categoria,
            "valor": round(valor, 2),
            "percentual": round(perc(valor), 2)
        }
        for categoria, valor in custos.items()
    ]

    linhas_custos_executivas = preparar_linhas_custos_executivas(linhas_custos)
    despesas_grafico = preparar_grafico_despesas_operacionais(linhas_custos, receita_bruta)

    return {
        "receita_bruta": round(receita_bruta, 2),
        "vendas_por_sku": vendas_por_sku,
        "cmv_total": round(cmv_total, 2),
        "cmv_percentual": round(perc(cmv_total), 2),
        "cmv_por_sku": cmv_por_sku,
        "margem_bruta": round(margem_bruta, 2),
        "margem_bruta_percentual": round(perc(margem_bruta), 2),
        "custos_operacionais_total": round(custos_operacionais_total, 2),
        "custos_operacionais_percentual": round(perc(custos_operacionais_total), 2),
        "linhas_custos": linhas_custos,
        "linhas_custos_executivas": linhas_custos_executivas,
        "despesas_grafico": despesas_grafico,
        "resultado_operacional": round(resultado_operacional, 2),
        "margem_operacional_percentual": round(perc(resultado_operacional), 2)
    }

def gerar_excel_dre_gerencial(competencia, dados):
    wb = Workbook()
    ws = wb.active
    ws.title = "DRE Gerencial"

    azul = "1F3B4D"
    laranja = "F97316"
    cinza = "F8FAFC"
    branco = "FFFFFF"
    azul_resultado = "2563EB"
    vermelho = "DC2626"

    fill_topo = PatternFill("solid", fgColor=azul)
    fill_laranja = PatternFill("solid", fgColor=laranja)
    fill_cinza = PatternFill("solid", fgColor=cinza)
    fill_resultado = PatternFill(
        "solid",
        fgColor=azul_resultado if dados["resultado_operacional"] >= 0 else vermelho
    )

    fonte_titulo = Font(color=branco, bold=True, size=16)
    fonte_subtitulo = Font(color=branco, bold=True, size=11)
    fonte_header = Font(color=branco, bold=True)
    fonte_negrito = Font(bold=True, color=azul)
    fonte_resultado = Font(color=branco, bold=True, size=13)

    borda = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = "FRIGODATTA — DRE Gerencial Industrial"
    ws["A1"].fill = fill_topo
    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Competência: {competencia}"
    ws["A2"].fill = fill_topo
    ws["A2"].font = fonte_subtitulo
    ws["A2"].alignment = Alignment(horizontal="center")

    linha = 4

    ws[f"A{linha}"] = "Indicador"
    ws[f"B{linha}"] = "Valor"
    ws[f"C{linha}"] = "% Receita"
    ws[f"D{linha}"] = "Observação"

    for col in range(1, 5):
        celula = ws.cell(row=linha, column=col)
        celula.fill = fill_laranja
        celula.font = fonte_header
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda

    kpis = [
        ("Receita Bruta", dados["receita_bruta"], 100 if dados["receita_bruta"] > 0 else 0, "Venda de galinhas"),
        ("CMV", dados["cmv_total"], dados["cmv_percentual"], "Custo das vendas"),
        ("Margem Bruta", dados["margem_bruta"], dados["margem_bruta_percentual"], "Receita - CMV"),
        ("Custos Operacionais", dados["custos_operacionais_total"], dados["custos_operacionais_percentual"], "Custos mensais"),
        ("Resultado Operacional", dados["resultado_operacional"], dados["margem_operacional_percentual"], "Margem Bruta - Custos")
    ]

    for item in kpis:
        linha += 1
        ws[f"A{linha}"] = item[0]
        ws[f"B{linha}"] = item[1]
        ws[f"C{linha}"] = item[2] / 100
        ws[f"D{linha}"] = item[3]

        for col in range(1, 5):
            celula = ws.cell(row=linha, column=col)
            celula.border = borda
            celula.fill = fill_cinza
            if col == 1:
                celula.font = fonte_negrito

        ws[f"B{linha}"].number_format = 'R$ #,##0.00'
        ws[f"C{linha}"].number_format = '0.00%'

    linha += 3
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    ws.cell(row=linha, column=1).value = "DRE"
    ws.cell(row=linha, column=1).fill = fill_topo
    ws.cell(row=linha, column=1).font = fonte_header
    ws.cell(row=linha, column=1).alignment = Alignment(horizontal="center")

    linhas_dre = []

    linhas_dre.append(("Receita Bruta", dados["receita_bruta"], "receita"))

    for venda in dados["vendas_por_sku"]:
        linhas_dre.append((f"  Venda — {venda['sku']}", venda["receita"], "subitem"))

    linhas_dre.append(("(-) CMV", dados["cmv_total"], "normal"))

    for cmv in dados["cmv_por_sku"]:
        linhas_dre.append((f"  CMV — {cmv['sku']}", cmv["cmv"], "subitem"))

    linhas_dre.append(("= Margem Bruta", dados["margem_bruta"], "total"))
    linhas_dre.append(("Custos Operacionais", None, "grupo"))

    for custo in dados["linhas_custos"]:
        linhas_dre.append((f"(-) {custo['categoria']}", custo["valor"], "normal"))

    linhas_dre.append(("Total de Custos Operacionais", dados["custos_operacionais_total"], "total"))
    linhas_dre.append(("= Resultado Operacional", dados["resultado_operacional"], "resultado"))

    for descricao, valor, tipo in linhas_dre:
        linha += 1

        ws[f"A{linha}"] = descricao

        if valor is not None:
            ws[f"B{linha}"] = valor
            ws[f"B{linha}"].number_format = 'R$ #,##0.00'

        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=4)

        for col in range(1, 5):
            celula = ws.cell(row=linha, column=col)
            celula.border = borda

        if tipo == "grupo":
            for col in range(1, 5):
                ws.cell(row=linha, column=col).fill = fill_topo
                ws.cell(row=linha, column=col).font = fonte_header
        elif tipo == "resultado":
            for col in range(1, 5):
                ws.cell(row=linha, column=col).fill = fill_resultado
                ws.cell(row=linha, column=col).font = fonte_resultado
        elif tipo == "total":
            for col in range(1, 5):
                ws.cell(row=linha, column=col).fill = fill_cinza
                ws.cell(row=linha, column=col).font = fonte_negrito
        elif tipo == "subitem":
            ws[f"A{linha}"].font = Font(color="64748B")
            ws[f"B{linha}"].font = Font(color="64748B")
        else:
            ws[f"A{linha}"].font = fonte_negrito

    linha += 3
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    ws.cell(row=linha, column=1).value = (
        "Leitura executiva: "
        + ("A operação apresentou resultado operacional positivo." if dados["resultado_operacional"] >= 0 else "A operação apresentou resultado operacional negativo.")
    )
    ws.cell(row=linha, column=1).alignment = Alignment(wrap_text=True)
    ws.cell(row=linha, column=1).fill = PatternFill("solid", fgColor="FFF7ED")

    larguras = {
        "A": 38,
        "B": 18,
        "C": 14,
        "D": 32
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="center",
                horizontal="right" if cell.column >= 2 else "left",
                wrap_text=True
            )

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = "A4"

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return arquivo






# ============================================================
# MÓDULO ALMOXARIFADO
# ============================================================

UNIDADES_VENDA_SKU = [
    "Kg",
    "Un"
]

TIPOS_CONSUMO_RECEITA = [
    "Matéria-prima",
    "Embalagem primária - 1 pra 1",
    "Embalagem secundária - proporcional",
    "Outro"
]


@executar_rotina_estrutural_uma_vez
def criar_tabelas_receitas_sku():
    criar_tabelas_estoque_almoxarifado()

    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS skus (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL UNIQUE,
            unidade_venda TEXT NOT NULL DEFAULT 'Kg',
            ativo TEXT DEFAULT 'Sim',
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS receitas_sku (
            id SERIAL PRIMARY KEY,
            sku_id INTEGER NOT NULL,
            insumo_id INTEGER NOT NULL,
            quantidade_por_unidade REAL NOT NULL,
            tipo_consumo TEXT DEFAULT '',
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS skus (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            unidade_venda TEXT NOT NULL DEFAULT 'Kg',
            ativo TEXT DEFAULT 'Sim',
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS receitas_sku (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku_id INTEGER NOT NULL,
            insumo_id INTEGER NOT NULL,
            quantidade_por_unidade REAL NOT NULL,
            tipo_consumo TEXT DEFAULT '',
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

    conn.commit()

    skus_padrao = [
        ("Galinha Cortada", "Kg", "Produto vendido por kg. A OP consome insumos por unidade produzida e forma estoque em kg."),
        ("Galinha Inteira", "Un", "Produto vendido por unidade.")
    ]

    for sku in skus_padrao:
        try:
            cursor.execute(q("""
            INSERT INTO skus (nome, unidade_venda, ativo, observacoes)
            VALUES (?, ?, ?, ?)
            """), (sku[0], sku[1], "Sim", sku[2]))
            conn.commit()
        except Exception:
            conn.rollback()

    conn.close()


def buscar_skus(filtro_status="Todos"):
    criar_tabelas_receitas_sku()

    condicoes = ["1 = 1"]
    parametros = []

    if filtro_status and filtro_status != "Todos":
        condicoes.append("ativo = ?")
        parametros.append(filtro_status)

    where_sql = " AND ".join(condicoes)

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q(f"""
    SELECT *
    FROM skus
    WHERE {where_sql}
    ORDER BY ativo DESC, nome ASC
    """), tuple(parametros))

    skus = cursor.fetchall()
    conn.close()
    return skus


def salvar_sku(form):
    criar_tabelas_receitas_sku()

    nome = form.get("nome", "").strip()
    unidade_venda = form.get("unidade_venda", "Kg").strip()
    ativo = form.get("ativo", "Sim").strip()
    observacoes = form.get("observacoes", "").strip()

    if not nome:
        raise ValueError("Informe o nome do SKU.")

    if unidade_venda not in UNIDADES_VENDA_SKU:
        raise ValueError("Unidade de venda inválida.")

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    INSERT INTO skus (nome, unidade_venda, ativo, observacoes)
    VALUES (?, ?, ?, ?)
    """), (nome, unidade_venda, ativo, observacoes))

    conn.commit()
    conn.close()


def salvar_item_receita_sku(form):
    criar_tabelas_receitas_sku()

    sku_id = int(form.get("sku_id") or 0)
    insumo_id = int(form.get("insumo_id") or 0)
    quantidade_por_unidade = float(form.get("quantidade_por_unidade") or 0)
    tipo_consumo = form.get("tipo_consumo", "").strip()
    observacoes = form.get("observacoes", "").strip()

    if quantidade_por_unidade <= 0:
        raise ValueError("A quantidade por unidade precisa ser maior que zero.")

    if tipo_consumo and tipo_consumo not in TIPOS_CONSUMO_RECEITA:
        raise ValueError("Tipo de consumo inválido.")

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("SELECT id FROM skus WHERE id = ?"), (sku_id,))
    if not cursor.fetchone():
        conn.close()
        raise ValueError("Selecione um SKU válido.")

    cursor.execute(q("SELECT id FROM almoxarifado_insumos WHERE id = ?"), (insumo_id,))
    if not cursor.fetchone():
        conn.close()
        raise ValueError("Selecione um insumo válido.")

    cursor.execute(q("""
    INSERT INTO receitas_sku (
        sku_id,
        insumo_id,
        quantidade_por_unidade,
        tipo_consumo,
        observacoes
    ) VALUES (?, ?, ?, ?, ?)
    """), (
        sku_id,
        insumo_id,
        quantidade_por_unidade,
        tipo_consumo,
        observacoes
    ))

    conn.commit()
    conn.close()


def excluir_item_receita_sku(item_id):
    criar_tabelas_receitas_sku()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("DELETE FROM receitas_sku WHERE id = ?"), (item_id,))

    conn.commit()
    conn.close()


def buscar_receitas_sku():
    criar_tabelas_receitas_sku()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        r.*,
        s.nome as sku,
        s.unidade_venda as unidade_venda,
        i.descricao as insumo,
        i.categoria as categoria_insumo,
        i.unidade as unidade_insumo
    FROM receitas_sku r
    JOIN skus s ON s.id = r.sku_id
    JOIN almoxarifado_insumos i ON i.id = r.insumo_id
    ORDER BY s.nome ASC, r.id ASC
    """)

    itens = cursor.fetchall()
    conn.close()

    receitas = {}

    for item in itens:
        sku = item["sku"]
        if sku not in receitas:
            receitas[sku] = {
                "sku": sku,
                "unidade_venda": item["unidade_venda"],
                "itens": [],
                "total_itens": 0
            }

        receitas[sku]["itens"].append(item)
        receitas[sku]["total_itens"] += 1

    return list(receitas.values())


def calcular_resumo_receitas_sku(skus, receitas):
    total_skus = len(skus)
    skus_ativos = sum(1 for item in skus if item["ativo"] == "Sim")
    total_itens = sum(receita["total_itens"] for receita in receitas)
    skus_com_receita = len(receitas)

    return {
        "total_skus": total_skus,
        "skus_ativos": skus_ativos,
        "skus_com_receita": skus_com_receita,
        "total_itens": total_itens
    }


@app.route("/receitas-sku", methods=["GET", "POST"])
@perfil_permitido("pcp")
def receitas_sku():
    if request.method == "POST":
        acao = request.form.get("acao")

        try:
            if acao == "salvar_sku":
                salvar_sku(request.form)
                flash("SKU cadastrado com sucesso.")

            elif acao == "salvar_item_receita":
                salvar_item_receita_sku(request.form)
                flash("Item adicionado à receita com sucesso.")

            else:
                flash("Ação inválida.")

        except Exception as erro:
            flash(f"Erro ao salvar receita/SKU: {erro}")

        return redirect(url_for("receitas_sku"))

    skus = buscar_skus()
    skus_ativos = buscar_skus("Sim")
    insumos = buscar_insumos_almoxarifado("Todas", "Sim", "")
    receitas = buscar_receitas_sku()
    resumo = calcular_resumo_receitas_sku(skus, receitas)

    return render_template(
        "receitas_sku.html",
        skus=skus,
        skus_ativos=skus_ativos,
        insumos=insumos,
        receitas=receitas,
        resumo=resumo,
        unidades_venda=UNIDADES_VENDA_SKU,
        tipos_consumo=TIPOS_CONSUMO_RECEITA
    )


@app.route("/receitas-sku/item/<int:item_id>/excluir", methods=["POST"])
@perfil_permitido("pcp")
def excluir_item_receita_sku_rota(item_id):
    excluir_item_receita_sku(item_id)
    flash("Item removido da receita com sucesso.")
    return redirect(url_for("receitas_sku"))


# ============================================================
# MÓDULO FINANCEIRO - MOVIMENTAÇÃO DE CAIXA
# ============================================================

CATEGORIAS_FINANCEIRAS_ENTRADA = [
    "Venda de produtos",
    "Recebimento de cliente",
    "Aporte",
    "Empréstimo recebido",
    "Outras entradas"
]

CATEGORIAS_FINANCEIRAS_SAIDA = [
    "Fornecedor",
    "Mão de obra",
    "Energia",
    "Água",
    "Lenha",
    "Combustível",
    "Manutenção",
    "Embalagens",
    "Impostos",
    "Marketing",
    "Serviços terceiros",
    "Empréstimos e financiamentos",
    "Outras saídas"
]

FORMAS_PAGAMENTO_FINANCEIRO = [
    "Pix",
    "Dinheiro",
    "Boleto",
    "Cartão",
    "Transferência",
    "Cheque",
    "Outro"
]

# Status que o usuário escolhe no lançamento.
# "Em atraso" não deve ser escolhido manualmente; o sistema calcula pela data de vencimento.
STATUS_FINANCEIRO = [
    "Pendente",
    "Realizado",
    "Cancelado"
]

# Opções usadas nos filtros e na leitura gerencial da tela.
STATUS_FINANCEIRO_FILTRO = [
    "Todos",
    "A vencer",
    "Em atraso",
    "Realizado",
    "Cancelado"
]


def calcular_status_financeiro_visual(item, data_referencia=None):
    """
    Calcula o status visual/gerencial sem alterar o status gravado no banco.

    Regra:
    - Cancelado permanece Cancelado;
    - Realizado vira Liquidado;
    - Pendente com vencimento anterior à data de referência vira Em atraso;
    - Pendente com vencimento igual ou posterior vira A vencer.

    Observação de segurança:
    Se existir algum registro antigo com status "Atrasado", ele será exibido como "Em atraso"
    para manter compatibilidade com dados já lançados.
    """
    if data_referencia is None:
        data_referencia = datetime.now().date()

    status = (item.get("status") if hasattr(item, "get") else item["status"]) or "Pendente"
    data_vencimento = (item.get("data_vencimento") if hasattr(item, "get") else item["data_vencimento"]) or ""

    if status == "Cancelado":
        return "Cancelado"

    if status == "Realizado":
        return "Liquidado"

    if status == "Atrasado":
        return "Em atraso"

    try:
        vencimento = datetime.strptime(data_vencimento, "%Y-%m-%d").date()
    except Exception:
        return "A vencer"

    if vencimento < data_referencia:
        return "Em atraso"

    return "A vencer"


def preparar_movimentacoes_financeiras_para_tela(movimentacoes, status_filtro="Todos"):
    """
    Converte as linhas retornadas do banco em dicionários e adiciona campos calculados.
    Isso evita alterar a estrutura do banco e protege a lógica já existente.
    """
    hoje_data = datetime.now().date()
    resultado = []

    for item in movimentacoes:
        item_dict = dict(item)
        status_visual = calcular_status_financeiro_visual(item_dict, hoje_data)
        item_dict["status_original"] = item_dict.get("status", "Pendente")
        item_dict["status_visual"] = status_visual

        # Classe CSS simples para futura estilização do badge, sem obrigar mudança no template.
        item_dict["status_classe"] = (
            status_visual.lower()
            .replace(" ", "-")
            .replace("ç", "c")
            .replace("ã", "a")
        )

        if status_filtro == "Todos" or status_visual == status_filtro:
            resultado.append(item_dict)

    return resultado


@executar_rotina_estrutural_uma_vez
def criar_tabela_movimentacoes_financeiras():
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS movimentacoes_financeiras (
            id SERIAL PRIMARY KEY,
            data_vencimento TEXT NOT NULL,
            data_realizacao TEXT,
            tipo TEXT NOT NULL,
            categoria TEXT NOT NULL,
            descricao TEXT NOT NULL,
            valor REAL NOT NULL,
            forma_pagamento TEXT,
            status TEXT DEFAULT 'Pendente',
            parcelas INTEGER DEFAULT 1,
            parcela_atual INTEGER DEFAULT 1,
            intervalo_dias INTEGER DEFAULT 30,
            documento_id TEXT,
            data_documento TEXT,
            valor_documento REAL DEFAULT 0,
            prazo_medio_dias REAL DEFAULT 0,
            observacoes TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS movimentacoes_financeiras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_vencimento TEXT NOT NULL,
            data_realizacao TEXT,
            tipo TEXT NOT NULL,
            categoria TEXT NOT NULL,
            descricao TEXT NOT NULL,
            valor REAL NOT NULL,
            forma_pagamento TEXT,
            status TEXT DEFAULT 'Pendente',
            parcelas INTEGER DEFAULT 1,
            parcela_atual INTEGER DEFAULT 1,
            intervalo_dias INTEGER DEFAULT 30,
            documento_id TEXT,
            data_documento TEXT,
            valor_documento REAL DEFAULT 0,
            prazo_medio_dias REAL DEFAULT 0,
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """)

    tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN intervalo_dias INTEGER DEFAULT 30")
    tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN documento_id TEXT")
    tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN data_documento TEXT")
    tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN valor_documento REAL DEFAULT 0")
    tentar_alter_table(cursor, conn, "ALTER TABLE movimentacoes_financeiras ADD COLUMN prazo_medio_dias REAL DEFAULT 0")

    conn.commit()
    conn.close()


def adicionar_meses(data_base, meses):
    ano = data_base.year
    mes = data_base.month + meses

    while mes > 12:
        mes -= 12
        ano += 1

    while mes < 1:
        mes += 12
        ano -= 1

    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dia = min(data_base.day, ultimo_dia)

    return data_base.replace(year=ano, month=mes, day=dia)


def salvar_movimentacao_financeira(form):
    criar_tabela_movimentacoes_financeiras()

    tipo = form.get("tipo", "").strip()
    categoria = form.get("categoria", "").strip()
    descricao = form.get("descricao", "").strip()
    data_documento = form.get("data_documento") or datetime.now().strftime("%Y-%m-%d")
    data_realizacao = form.get("data_realizacao", "")
    forma_pagamento = form.get("forma_pagamento", "")
    status = form.get("status", "Pendente")
    observacoes = form.get("observacoes", "")
    valor_documento = float(form.get("valor") or 0)

    vencimentos = form.getlist("parcela_vencimento[]")
    valores = form.getlist("parcela_valor[]")

    if tipo not in ["Entrada", "Saída"]:
        raise ValueError("Tipo de movimentação inválido.")

    if not descricao:
        raise ValueError("Informe uma descrição para a movimentação.")

    if valor_documento <= 0:
        raise ValueError("O valor total do documento deve ser maior que zero.")

    parcelas_validas = []

    for vencimento, valor in zip(vencimentos, valores):
        vencimento = (vencimento or "").strip()
        valor = float(valor or 0)

        if vencimento and valor > 0:
            parcelas_validas.append({
                "vencimento": vencimento,
                "valor": round(valor, 2)
            })

    if not parcelas_validas:
        parcelas_validas.append({
            "vencimento": data_documento,
            "valor": round(valor_documento, 2)
        })

    soma_parcelas = round(sum(item["valor"] for item in parcelas_validas), 2)

    if abs(soma_parcelas - round(valor_documento, 2)) > 0.02:
        raise ValueError(
            f"A soma das parcelas (R$ {soma_parcelas:.2f}) precisa bater com o valor total do documento (R$ {valor_documento:.2f})."
        )

    data_base = datetime.strptime(data_documento, "%Y-%m-%d")
    prazo_ponderado = 0

    for item in parcelas_validas:
        data_vencimento = datetime.strptime(item["vencimento"], "%Y-%m-%d")
        dias = (data_vencimento - data_base).days
        prazo_ponderado += item["valor"] * dias

    prazo_medio_dias = prazo_ponderado / valor_documento if valor_documento > 0 else 0
    documento_id = uuid.uuid4().hex
    total_parcelas = len(parcelas_validas)

    conn = conectar()
    cursor = conn.cursor()

    for indice, parcela in enumerate(parcelas_validas, start=1):
        descricao_parcela = descricao

        if total_parcelas > 1:
            descricao_parcela = f"{descricao} ({indice}/{total_parcelas})"

        cursor.execute(q("""
        INSERT INTO movimentacoes_financeiras (
            data_vencimento,
            data_realizacao,
            tipo,
            categoria,
            descricao,
            valor,
            forma_pagamento,
            status,
            parcelas,
            parcela_atual,
            intervalo_dias,
            documento_id,
            data_documento,
            valor_documento,
            prazo_medio_dias,
            observacoes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """), (
            parcela["vencimento"],
            data_realizacao if status == "Realizado" else "",
            tipo,
            categoria,
            descricao_parcela,
            parcela["valor"],
            forma_pagamento,
            status,
            total_parcelas,
            indice,
            0,
            documento_id,
            data_documento,
            valor_documento,
            round(prazo_medio_dias, 2),
            observacoes
        ))

    conn.commit()
    conn.close()


def buscar_movimentacao_financeira_por_id(movimentacao_id):
    criar_tabela_movimentacoes_financeiras()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT *
    FROM movimentacoes_financeiras
    WHERE id = ?
    """), (movimentacao_id,))

    movimentacao = cursor.fetchone()
    conn.close()

    return movimentacao


def atualizar_movimentacao_financeira(movimentacao_id, form):
    criar_tabela_movimentacoes_financeiras()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    UPDATE movimentacoes_financeiras
    SET data_vencimento = ?,
        data_realizacao = ?,
        tipo = ?,
        categoria = ?,
        descricao = ?,
        valor = ?,
        forma_pagamento = ?,
        status = ?,
        intervalo_dias = ?,
        observacoes = ?
    WHERE id = ?
    """), (
        form.get("data_vencimento", ""),
        form.get("data_realizacao", ""),
        form.get("tipo", ""),
        form.get("categoria", ""),
        form.get("descricao", ""),
        float(form.get("valor") or 0),
        form.get("forma_pagamento", ""),
        form.get("status", ""),
        int(form.get("intervalo_dias") or 30),
        form.get("observacoes", ""),
        movimentacao_id
    ))

    conn.commit()
    conn.close()


def excluir_movimentacao_financeira(movimentacao_id):
    criar_tabela_movimentacoes_financeiras()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    DELETE FROM movimentacoes_financeiras
    WHERE id = ?
    """), (movimentacao_id,))

    conn.commit()
    conn.close()


def buscar_movimentacoes_financeiras(data_inicio, data_fim, tipo_filtro, status_filtro):
    criar_tabela_movimentacoes_financeiras()

    condicoes = ["data_vencimento BETWEEN ? AND ?"]
    parametros = [data_inicio, data_fim]

    if tipo_filtro in ["Entrada", "Saída"]:
        condicoes.append("tipo = ?")
        parametros.append(tipo_filtro)

    # O filtro por status visual é aplicado depois da consulta, porque "A vencer" e
    # "Em atraso" são calculados pela data de vencimento, não gravados no banco.
    if status_filtro == "Realizado":
        condicoes.append("status = ?")
        parametros.append("Realizado")
    elif status_filtro == "Cancelado":
        condicoes.append("status = ?")
        parametros.append("Cancelado")

    where_sql = " AND ".join(condicoes)

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q(f"""
    SELECT *
    FROM movimentacoes_financeiras
    WHERE {where_sql}
    ORDER BY data_vencimento ASC, id ASC
    """), tuple(parametros))

    movimentacoes = cursor.fetchall()
    conn.close()

    return preparar_movimentacoes_financeiras_para_tela(movimentacoes, status_filtro)


def calcular_resumo_financeiro(movimentacoes):
    entradas_previstas = 0
    saidas_previstas = 0
    entradas_realizadas = 0
    saidas_realizadas = 0

    for item in movimentacoes:
        valor = float(item["valor"] or 0)
        tipo = item["tipo"]
        status = item.get("status_original", item.get("status", "Pendente")) if hasattr(item, "get") else item["status"]

        # Cancelados não entram no previsto nem no realizado.
        if status == "Cancelado":
            continue

        if tipo == "Entrada":
            entradas_previstas += valor

            if status == "Realizado":
                entradas_realizadas += valor

        elif tipo == "Saída":
            saidas_previstas += valor

            if status == "Realizado":
                saidas_realizadas += valor

    saldo_previsto = entradas_previstas - saidas_previstas
    saldo_realizado = entradas_realizadas - saidas_realizadas

    return {
        "entradas_previstas": round(entradas_previstas, 2),
        "saidas_previstas": round(saidas_previstas, 2),
        "saldo_previsto": round(saldo_previsto, 2),
        "entradas_realizadas": round(entradas_realizadas, 2),
        "saidas_realizadas": round(saidas_realizadas, 2),
        "saldo_realizado": round(saldo_realizado, 2)
    }


def agrupar_fluxo_por_dia(movimentacoes):
    fluxo = {}

    for item in movimentacoes:
        data = item["data_vencimento"]
        valor = float(item["valor"] or 0)

        if data not in fluxo:
            fluxo[data] = {
                "data": data,
                "entradas": 0,
                "saidas": 0,
                "saldo": 0
            }

        status = item.get("status_original", item.get("status", "Pendente")) if hasattr(item, "get") else item["status"]

        # Fluxo diário ignora documentos cancelados.
        if status == "Cancelado":
            continue

        if item["tipo"] == "Entrada":
            fluxo[data]["entradas"] += valor
        else:
            fluxo[data]["saidas"] += valor

        fluxo[data]["saldo"] = fluxo[data]["entradas"] - fluxo[data]["saidas"]

    return [
        {
            "data": item["data"],
            "entradas": round(item["entradas"], 2),
            "saidas": round(item["saidas"], 2),
            "saldo": round(item["saldo"], 2)
        }
        for item in fluxo.values()
    ]



@executar_rotina_estrutural_uma_vez
def criar_tabela_fornecedores():
    conn = conectar()
    cursor = conn.cursor()

    if DATABASE_URL:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS fornecedores (
            id SERIAL PRIMARY KEY,
            nome TEXT UNIQUE NOT NULL
        )
        """)
    else:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS fornecedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        )
        """)

    conn.commit()
    conn.close()


def buscar_fornecedores():
    criar_tabela_fornecedores()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM fornecedores
    ORDER BY nome
    """)

    fornecedores = cursor.fetchall()
    conn.close()

    return fornecedores


@app.route("/embalagem-primaria", methods=["GET", "POST"])
@perfil_permitido("pcp", "producao")
def embalagem_primaria():
    if request.method == "POST":
        try:
            op_id = int(request.form.get("op_id") or 0)
            op = buscar_op_por_id(op_id)
            resultado = registrar_apontamento_embalagem_primaria(
                op=op,
                quantidade_bandejas=request.form.get("quantidade_bandejas"),
                observacoes=request.form.get("observacoes") or "",
                kg_produzidos=request.form.get("kg_produzidos"),
                pacotes_1_ave=request.form.get("pacotes_1_ave"),
                pacotes_2_aves=request.form.get("pacotes_2_aves")
            )
            if resultado.get("tipo") == "encerramento_primaria":
                flash(
                    "Galinha Inteira encerrada na Embalagem Primaria. "
                    f"Lote PA: {resultado['codigo_lote']} | "
                    f"Unidades vendaveis: {resultado['unidades_vendaveis']:.0f} | "
                    f"Peso produzido: {resultado['kg_produzidos']:.3f} kg."
                )
            else:
                flash("Embalagem Primária apontada com sucesso. O Estoque PI foi atualizado e a OP permanece pendente para Embalagem Secundária.")
        except ValueError as erro:
            flash(str(erro))

        return redirect(url_for("embalagem_primaria", op_id=request.form.get("op_id") or ""))

    op_id_selecionada = request.args.get("op_id", "")
    modo_edicao = request.args.get("editar") == "1"
    apontamento_edicao = None

    if modo_edicao and op_id_selecionada:
        try:
            apontamento_edicao = buscar_apontamento_embalagem_primaria_por_op(int(op_id_selecionada))
        except (TypeError, ValueError):
            apontamento_edicao = None

    ops = buscar_ops_para_embalagem_primaria()
    apontamentos = buscar_apontamentos_embalagem_primaria()
    saldos_pi = buscar_saldos_estoque_pi()
    caixas_pa = buscar_caixas_pa()
    resumo = calcular_resumo_estoques_pi_pa(saldos_pi, caixas_pa)

    return render_template(
        "embalagem_primaria.html",
        ops=ops,
        apontamentos=apontamentos,
        saldos_pi=saldos_pi,
        resumo=resumo,
        op_id_selecionada=str(op_id_selecionada),
        apontamento_edicao=apontamento_edicao,
        modo_edicao=modo_edicao
    )


@app.route("/estoque-produtos")
@perfil_permitido("pcp")
def estoque_produtos():
    saldos_pi = buscar_saldos_estoque_pi()
    movimentacoes_pi = buscar_movimentacoes_estoque_pi()
    caixas_pa = buscar_caixas_pa()
    resumo = calcular_resumo_estoques_pi_pa(saldos_pi, caixas_pa)

    return render_template(
        "estoque_produtos.html",
        saldos_pi=saldos_pi,
        movimentacoes_pi=movimentacoes_pi,
        caixas_pa=caixas_pa,
        resumo=resumo
    )


@app.route("/embalagem-secundaria/<int:op_id>/finalizar", methods=["POST"])
@perfil_permitido("pcp")
def finalizar_embalagem_secundaria(op_id):
    try:
        fechamento = finalizar_embalagem_secundaria_op(op_id)
        flash(
            "OP encerrada com sucesso. "
            f"Peso oficial: {fechamento['peso_liquido_total']:.3f} kg | "
            f"Caixas: {fechamento['caixas']} | "
            f"Bandejas: {fechamento['bandejas_consumidas']:.0f}."
        )
    except ValueError as erro:
        flash(str(erro))

    return redirect(url_for("embalagem_secundaria", op_id=op_id))


@app.route("/embalagem-secundaria/<int:op_id>/resetar", methods=["POST"])
@perfil_permitido("pcp")
def resetar_embalagem_secundaria_op(op_id):
    try:
        resultado = resetar_processamento_op(op_id, request.form.get("confirmacao_reset"))
        flash(
            "OP resetada com sucesso. "
            f"Caixas removidas: {resultado['caixas_removidas']}. "
            "A OP voltou para Aberta e pode ser reapontada desde a Embalagem Primária."
        )
        return redirect(url_for("embalagem_primaria", op_id=op_id))
    except ValueError as erro:
        flash(str(erro))
        return redirect(url_for("embalagem_secundaria", op_id=op_id))


@app.route("/embalagem-secundaria", methods=["GET", "POST"])

@perfil_permitido("pcp")
def embalagem_secundaria():
    if request.method == "POST":
        try:
            if request.form.get("modo_lancamento") == "lote":
                codigos = registrar_caixas_pa_lote(request.form)
                flash(f"{len(codigos)} caixas registradas no Estoque PA com sucesso.")
            else:
                codigo_caixa = registrar_caixa_pa_manual(request.form)
                flash(f"Caixa {codigo_caixa} registrada no Estoque PA com sucesso.")
        except ValueError as erro:
            flash(str(erro))

        return redirect(url_for("embalagem_secundaria", op_id=request.form.get("op_principal") or ""))

    saldos_pi = buscar_ops_com_saldo_pi()
    caixas_pa = buscar_caixas_pa()
    resumo = calcular_resumo_estoques_pi_pa(saldos_pi, caixas_pa)
    op_id_selecionada = request.args.get("op_id", "")
    op_selecionada = None
    caixas_op = []
    fechamento_op = None

    if op_id_selecionada:
        try:
            op_id_int = int(op_id_selecionada)
        except Exception:
            op_id_int = None

        if op_id_int:
            # A OP deve abrir para lançamento contínuo sempre que houver PI disponível.
            # O painel de fechamento é complementar e não pode impedir a abertura da tela de caixas.
            op_selecionada = next((item for item in saldos_pi if int(item["op_id"]) == op_id_int), None)

            try:
                fechamento_op = calcular_fechamento_industrial_op(op_id_int)
            except Exception:
                fechamento_op = None

            # Quando o saldo PI chega a zero, a OP deixa de aparecer em saldos_pi.
            # Ainda assim ela precisa permanecer carregada para conferência e encerramento.
            if op_selecionada is None and fechamento_op:
                op_base = fechamento_op["op"]
                op_selecionada = {
                    "op_id": op_id_int,
                    "data_op": op_base["data"],
                    "sku": op_base["sku"] or "Galinha Cortada",
                    "saldo_bandejas": fechamento_op["saldo_pi"],
                }

            try:
                conn = conectar()
                cursor = conn.cursor()
                cursor.execute(q("""
                SELECT cx.*
                FROM pa_caixas cx
                INNER JOIN pa_caixa_composicao comp ON comp.caixa_id = cx.id
                WHERE comp.op_id = ?
                ORDER BY cx.id DESC
                LIMIT 80
                """), (op_id_int,))
                caixas_op = cursor.fetchall()
                conn.close()
            except Exception:
                caixas_op = []

    return render_template(
        "embalagem_secundaria.html",
        saldos_pi=saldos_pi,
        caixas_pa=caixas_pa,
        resumo=resumo,
        hoje=datetime.now().strftime("%Y-%m-%d"),
        bandejas_por_caixa=BANDEJAS_POR_CAIXA,
        op_id_selecionada=str(op_id_selecionada),
        op_selecionada=op_selecionada,
        caixas_op=caixas_op,
        fechamento_op=fechamento_op
    )

@app.route("/expedicao")
@perfil_permitido("pcp")
def expedicao():
    hoje = datetime.now()
    primeiro_dia_mes = hoje.replace(day=1).strftime("%Y-%m-%d")
    data_inicio = request.args.get("data_inicio") or primeiro_dia_mes
    data_fim = request.args.get("data_fim") or hoje.strftime("%Y-%m-%d")
    status = request.args.get("status") or "Todos"

    expedicoes = buscar_expedicoes(data_inicio, data_fim, status)
    resumo = calcular_resumo_expedicao(expedicoes)

    return render_template(
        "expedicao.html",
        expedicoes=expedicoes,
        resumo=resumo,
        data_inicio=data_inicio,
        data_fim=data_fim,
        status=status,
        status_opcoes=["Todos", "Aberto", "Concluído", "Cancelado"]
    )



def gerar_numero_romaneio(data_romaneio):
    """
    Gera número sequencial diário do romaneio.
    Formato: ROM-AAAAMMDD-001

    Sprint 1.1:
    - Numeração simples e isolada.
    - Não baixa estoque.
    - Não gera venda.
    - Não interfere em DRE, Financeiro ou Almoxarifado.
    """
    criar_tabelas_expedicao()

    data_base = (data_romaneio or datetime.now().strftime("%Y-%m-%d")).strip()
    prefixo = "ROM-" + data_base.replace("-", "")

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT COUNT(*) as total
    FROM expedicoes
    WHERE numero_romaneio LIKE ?
    """), (f"{prefixo}-%",))

    resultado = cursor.fetchone()
    conn.close()

    total = int(resultado["total"] or 0)
    sequencial = total + 1

    return f"{prefixo}-{sequencial:03d}"


def salvar_romaneio_expedicao(form):
    """
    Salva apenas o cabeçalho do romaneio.

    Sprint 1.1:
    - Cria o romaneio em aberto.
    - Tipo fixo: TRANSFERENCIA.
    - Sem itens.
    - Sem baixa de estoque.
    """
    criar_tabelas_expedicao()

    data_romaneio = (form.get("data") or "").strip()
    destino = (form.get("destino") or "").strip()
    responsavel = (form.get("responsavel") or "").strip()
    observacoes = (form.get("observacoes") or "").strip()

    if not data_romaneio:
        raise ValueError("Informe a data do romaneio.")

    if not destino:
        raise ValueError("Informe o destino do romaneio.")

    numero_romaneio = gerar_numero_romaneio(data_romaneio)

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    INSERT INTO expedicoes (
        numero_romaneio,
        data,
        tipo_movimentacao,
        destino,
        responsavel,
        observacoes,
        status
    ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """), (
        numero_romaneio,
        data_romaneio,
        "TRANSFERENCIA",
        destino,
        responsavel,
        observacoes,
        "Aberto"
    ))

    conn.commit()
    conn.close()

    return numero_romaneio


@app.route("/expedicao/novo", methods=["GET", "POST"])
@perfil_permitido("pcp")
def novo_romaneio_expedicao():
    hoje = datetime.now().strftime("%Y-%m-%d")

    if request.method == "POST":
        try:
            numero_romaneio = salvar_romaneio_expedicao(request.form)
            flash(f"Romaneio {numero_romaneio} criado com sucesso.")
            return redirect(url_for("expedicao"))
        except Exception as erro:
            flash(f"Erro ao criar romaneio: {erro}")

    return render_template(
        "novo_romaneio.html",
        hoje=hoje
    )




def buscar_expedicao_por_id(expedicao_id):
    criar_tabelas_expedicao()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT *
    FROM expedicoes
    WHERE id = ?
    """), (expedicao_id,))

    expedicao = cursor.fetchone()
    conn.close()

    return expedicao


def buscar_itens_expedicao(expedicao_id):
    criar_tabelas_expedicao()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    SELECT *
    FROM expedicao_itens
    WHERE expedicao_id = ?
    ORDER BY id ASC
    """), (expedicao_id,))

    itens = cursor.fetchall()
    conn.close()

    return itens


def calcular_resumo_itens_expedicao(itens):
    total_itens = len(itens)
    total_unidades = sum(float(item["quantidade_unidades"] or 0) for item in itens)
    total_kg = sum(float(item["quantidade_kg"] or 0) for item in itens)

    return {
        "total_itens": total_itens,
        "total_unidades": round(total_unidades, 2),
        "total_kg": round(total_kg, 2)
    }


def salvar_item_expedicao(expedicao_id, form):
    """
    Salva item manual do romaneio.

    Sprint 1.2:
    - Itens manuais.
    - Sem baixa de estoque.
    - Sem vínculo obrigatório com OP.
    - Sem impacto na DRE, Financeiro ou Almoxarifado.
    """
    criar_tabelas_expedicao()

    expedicao = buscar_expedicao_por_id(expedicao_id)

    if not expedicao:
        raise ValueError("Romaneio não encontrado.")

    if expedicao["status"] != "Aberto":
        raise ValueError("Só é possível adicionar itens em romaneios abertos.")

    sku = (form.get("sku") or "").strip()
    quantidade_unidades = float(form.get("quantidade_unidades") or 0)
    quantidade_kg = float(form.get("quantidade_kg") or 0)

    if sku not in ["Galinha Cortada", "Galinha Inteira"]:
        raise ValueError("Selecione um SKU válido.")

    if quantidade_unidades < 0 or quantidade_kg < 0:
        raise ValueError("As quantidades não podem ser negativas.")

    if quantidade_unidades <= 0 and quantidade_kg <= 0:
        raise ValueError("Informe pelo menos uma quantidade para o item.")

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute(q("""
    INSERT INTO expedicao_itens (
        expedicao_id,
        op_id,
        sku,
        quantidade_unidades,
        quantidade_kg
    ) VALUES (?, ?, ?, ?, ?)
    """), (
        expedicao_id,
        None,
        sku,
        quantidade_unidades,
        quantidade_kg
    ))

    conn.commit()
    conn.close()


@app.route("/expedicao/<int:expedicao_id>", methods=["GET", "POST"])
@perfil_permitido("pcp")
def detalhe_romaneio_expedicao(expedicao_id):
    expedicao = buscar_expedicao_por_id(expedicao_id)

    if not expedicao:
        flash("Romaneio não encontrado.")
        return redirect(url_for("expedicao"))

    if request.method == "POST":
        try:
            salvar_item_expedicao(expedicao_id, request.form)
            flash("Item adicionado ao romaneio com sucesso.")
            return redirect(url_for("detalhe_romaneio_expedicao", expedicao_id=expedicao_id))
        except Exception as erro:
            flash(f"Erro ao adicionar item: {erro}")

    itens = buscar_itens_expedicao(expedicao_id)
    resumo_itens = calcular_resumo_itens_expedicao(itens)

    return render_template(
        "romaneio_detalhe.html",
        expedicao=expedicao,
        itens=itens,
        resumo_itens=resumo_itens,
        skus=["Galinha Cortada", "Galinha Inteira"]
    )

@app.route("/dre-gerencial/exportar-excel")
@perfil_permitido("pcp")
def exportar_dre_gerencial_excel():
    competencia = request.args.get("competencia") or datetime.now().strftime("%Y-%m")
    dados = buscar_dados_dre_gerencial(competencia)
    arquivo = gerar_excel_dre_gerencial(competencia, dados)

    nome_arquivo = f"DRE_Gerencial_{competencia}.xlsx"

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/dre-gerencial")
@perfil_permitido("pcp")
def dre_gerencial():
    competencia = request.args.get("competencia") or datetime.now().strftime("%Y-%m")
    dados = buscar_dados_dre_gerencial(competencia)

    return render_template(
        "dre_gerencial.html",
        competencia=competencia,
        receita_bruta=dados["receita_bruta"],
        vendas_por_sku=dados["vendas_por_sku"],
        cmv_total=dados["cmv_total"],
        cmv_percentual=dados["cmv_percentual"],
        cmv_por_sku=dados["cmv_por_sku"],
        margem_bruta=dados["margem_bruta"],
        margem_bruta_percentual=dados["margem_bruta_percentual"],
        custos_operacionais_total=dados["custos_operacionais_total"],
        custos_operacionais_percentual=dados["custos_operacionais_percentual"],
        linhas_custos=dados["linhas_custos"],
        linhas_custos_executivas=dados.get("linhas_custos_executivas", dados["linhas_custos"]),
        despesas_grafico=dados.get("despesas_grafico", {}),
        resultado_operacional=dados["resultado_operacional"],
        margem_operacional_percentual=dados["margem_operacional_percentual"]
    )



@app.route("/relatorio-custos")
@perfil_permitido("pcp")
def relatorio_custos():
    agora = datetime.now()
    competencia_fim = request.args.get("competencia_fim") or agora.strftime("%Y-%m")

    seis_meses_atras = agora

    for _ in range(5):
        if seis_meses_atras.month == 1:
            seis_meses_atras = seis_meses_atras.replace(
                year=seis_meses_atras.year - 1,
                month=12
            )
        else:
            seis_meses_atras = seis_meses_atras.replace(
                month=seis_meses_atras.month - 1
            )

    competencia_inicio = (
        request.args.get("competencia_inicio")
        or seis_meses_atras.strftime("%Y-%m")
    )

    if competencia_inicio > competencia_fim:
        competencia_inicio, competencia_fim = competencia_fim, competencia_inicio

    categoria_filtro = request.args.get("categoria") or "Todas"

    dados = buscar_dados_relatorio_custos(
        competencia_inicio,
        competencia_fim,
        categoria_filtro
    )

    return render_template(
        "relatorio_custos.html",
        competencia_inicio=competencia_inicio,
        competencia_fim=competencia_fim,
        categoria_filtro=categoria_filtro,
        categorias_custos=dados["categorias_disponiveis"],
        competencias=dados["competencias"],
        datasets=dados["datasets"],
        custo_total=dados["custo_total"],
        media_mensal=dados["media_mensal"],
        maior_categoria=dados["maior_categoria"],
        valor_maior_categoria=dados["valor_maior_categoria"],
        maior_crescimento_categoria=dados["maior_crescimento_categoria"],
        maior_crescimento_valor=dados["maior_crescimento_valor"],
        resumo_categorias=dados["resumo_categorias"]
    )



@app.route("/custos", methods=["GET", "POST"])
@perfil_permitido("pcp")
def custos():
    if request.method == "POST":
        acao = request.form.get("acao")

        try:
            if acao == "salvar_parametros":
                salvar_parametros_custos(request.form)
                flash("Parâmetros de CMV atualizados com sucesso.")

            elif acao == "salvar_custo_mensal":
                salvar_custo_mensal(request.form)
                flash("Custo mensal cadastrado com sucesso.")

            elif acao == "salvar_custos_lote":
                total_linhas = salvar_custos_mensais_lote(request.form)
                flash(f"{total_linhas} custos mensais cadastrados com sucesso.")

        except ValueError as erro:
            flash(str(erro) or "Verifique os valores informados. Use apenas números nos campos de custo.")

        return redirect(url_for("custos"))

    categorias_custos = CATEGORIAS_CUSTOS

    competencia_atual = datetime.now().strftime("%Y-%m")
    competencia_inicio = request.args.get("competencia_inicio") or competencia_atual
    competencia_fim = request.args.get("competencia_fim") or competencia_atual
    categoria_filtro = request.args.get("categoria") or "Todas"
    custos_filtrados = buscar_custos_mensais(
        competencia_inicio=competencia_inicio,
        competencia_fim=competencia_fim,
        categoria=categoria_filtro
    )
    total_custos_filtrados = sum(float(item["valor"] or 0) for item in custos_filtrados)

    return render_template(
        "custos.html",
        parametros=buscar_parametros_custos(),
        custos_mensais=custos_filtrados,
        categorias_custos=categorias_custos,
        competencia_atual=competencia_atual,
        competencia_inicio=competencia_inicio,
        competencia_fim=competencia_fim,
        categoria_filtro=categoria_filtro,
        total_custos_filtrados=total_custos_filtrados
    )




@app.route("/vendas", methods=["GET", "POST"])
@perfil_permitido("pcp")
def vendas():
    if request.method == "POST":
        try:
            salvar_venda_diaria(request.form)
            flash("Venda diária cadastrada com sucesso.")
        except ValueError as erro:
            flash(str(erro))

        return redirect(url_for("vendas"))

    hoje = datetime.now().strftime("%Y-%m-%d")

    return render_template(
        "vendas.html",
        hoje=hoje,
        vendas_diarias=buscar_vendas_diarias()
    )



@app.route("/custos/mensal/<int:custo_id>/editar", methods=["GET", "POST"])
@perfil_permitido("pcp")
def editar_custo_mensal(custo_id):
    custo = buscar_custo_mensal_por_id(custo_id)

    if not custo:
        flash("Custo mensal não encontrado.")
        return redirect(url_for("custos"))

    categorias_custos = CATEGORIAS_CUSTOS

    if request.method == "POST":
        try:
            conn = conectar()
            cursor = conn.cursor()
            cursor.execute(q("""
            UPDATE custos_mensais
            SET competencia = ?,
                categoria = ?,
                valor = ?,
                observacoes = ?
            WHERE id = ?
            """), (
                request.form["competencia"],
                request.form["categoria"],
                float(request.form["valor"]),
                request.form.get("observacoes", ""),
                custo_id
            ))
            conn.commit()
            conn.close()
            flash("Custo mensal atualizado com sucesso.")
            return redirect(url_for("custos"))
        except ValueError:
            flash("Verifique o valor informado. Use apenas números no campo de valor.")

    return render_template(
        "editar_custo_mensal.html",
        custo=custo,
        categorias_custos=categorias_custos
    )


@app.route("/custos/mensal/<int:custo_id>/excluir", methods=["POST"])
@perfil_permitido("pcp")
def excluir_custo_mensal(custo_id):
    if not buscar_custo_mensal_por_id(custo_id):
        flash("Custo mensal não encontrado.")
        return redirect(url_for("custos"))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q("""
    DELETE FROM custos_mensais
    WHERE id = ?
    """), (custo_id,))
    conn.commit()
    conn.close()

    flash("Custo mensal excluído com sucesso.")
    return redirect(url_for("custos"))


@app.route("/vendas/<int:venda_id>/editar", methods=["GET", "POST"])
@perfil_permitido("pcp")
def editar_venda_diaria(venda_id):
    venda = buscar_venda_diaria_por_id(venda_id)

    if not venda:
        flash("Venda diária não encontrada.")
        return redirect(url_for("vendas"))

    if request.method == "POST":
        try:
            sku = request.form["sku"]
            receita = float(request.form["receita"])

            if receita < 0:
                raise ValueError("A receita não pode ser negativa.")

            form_edicao = request.form.copy()

            try:
                quantidade_unidades_atual = float(venda["quantidade_unidades"] or 0)
            except Exception:
                quantidade_unidades_atual = 0

            try:
                quantidade_kg_atual = float(venda["quantidade_kg"] or 0)
            except Exception:
                quantidade_kg_atual = 0

            if sku == "Galinha Cortada":
                if not form_edicao.get("quantidade_unidades") and quantidade_unidades_atual > 0:
                    form_edicao["quantidade_unidades"] = str(quantidade_unidades_atual)

                if not form_edicao.get("quantidade_kg") and quantidade_kg_atual > 0:
                    form_edicao["quantidade_kg"] = str(quantidade_kg_atual)

            quantidades = preparar_quantidades_venda(
                sku,
                form_edicao,
                quantidade_atual=float(venda["quantidade"] or 0),
                unidade_atual=venda["unidade"]
            )

            venda_duplicada = buscar_venda_diaria_por_data_sku(
                request.form["data"],
                sku,
                ignorar_id=venda_id
            )

            if venda_duplicada:
                raise ValueError(
                    "Já existe outra venda lançada para esta data e SKU. "
                    "Ajuste o lançamento existente ou escolha outra data/SKU."
                )

            conn = conectar()
            cursor = conn.cursor()
            cursor.execute(q("""
            UPDATE vendas_diarias
            SET data = ?,
                sku = ?,
                quantidade = ?,
                unidade = ?,
                quantidade_unidades = ?,
                quantidade_kg = ?,
                receita = ?,
                observacoes = ?
            WHERE id = ?
            """), (
                request.form["data"],
                sku,
                quantidades["quantidade"],
                quantidades["unidade"],
                quantidades["quantidade_unidades"],
                quantidades["quantidade_kg"],
                receita,
                request.form.get("observacoes", ""),
                venda_id
            ))
            conn.commit()
            conn.close()

            flash("Venda diária atualizada com sucesso.")
            return redirect(url_for("vendas"))
        except ValueError as erro:
            flash(str(erro))

    return render_template("editar_venda_diaria.html", venda=venda)


@app.route("/vendas/<int:venda_id>/excluir", methods=["POST"])
@perfil_permitido("pcp")
def excluir_venda_diaria(venda_id):
    if not buscar_venda_diaria_por_id(venda_id):
        flash("Venda diária não encontrada.")
        return redirect(url_for("vendas"))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q("""
    DELETE FROM vendas_diarias
    WHERE id = ?
    """), (venda_id,))
    conn.commit()
    conn.close()

    flash("Venda diária excluída com sucesso.")
    return redirect(url_for("vendas"))




def destino_movimentacao_por_tipo(tipo):
    if tipo == "Saída":
        return "movimentacoes_despesas"
    return "movimentacoes_entradas"


def contexto_movimentacoes(visao, tipo_movimentacao=None):
    criar_tabela_movimentacoes_financeiras()

    agora = datetime.now()
    hoje = agora.strftime("%Y-%m-%d")
    primeiro_dia_mes = agora.replace(day=1).strftime("%Y-%m-%d")

    data_inicio = request.args.get("data_inicio") or primeiro_dia_mes
    data_fim = request.args.get("data_fim") or hoje
    status_filtro = request.args.get("status") or "Todos"
    tipo_filtro = tipo_movimentacao or request.args.get("tipo") or "Todos"

    movimentacoes = buscar_movimentacoes_financeiras(
        data_inicio,
        data_fim,
        tipo_filtro,
        status_filtro
    )

    return {
        "visao": visao,
        "hoje": hoje,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "tipo_filtro": tipo_filtro,
        "tipo_padrao": tipo_movimentacao or "Entrada",
        "status_filtro": status_filtro,
        "movimentacoes": movimentacoes,
        "resumo": calcular_resumo_financeiro(movimentacoes),
        "fluxo_diario": agrupar_fluxo_por_dia(movimentacoes),
        "categorias_entrada": CATEGORIAS_FINANCEIRAS_ENTRADA,
        "categorias_saida": CATEGORIAS_FINANCEIRAS_SAIDA,
        "categorias_lancamento": CATEGORIAS_FINANCEIRAS_ENTRADA if tipo_movimentacao == "Entrada" else CATEGORIAS_FINANCEIRAS_SAIDA,
        "formas_pagamento": FORMAS_PAGAMENTO_FINANCEIRO,
        "status_opcoes": STATUS_FINANCEIRO_FILTRO
    }


def salvar_movimentacao_por_visao(tipo_movimentacao, endpoint):
    form = request.form.copy()
    form["tipo"] = tipo_movimentacao

    try:
        salvar_movimentacao_financeira(form)
        flash("Movimentação lançada com sucesso.")
    except Exception as erro:
        flash(f"Erro ao salvar movimentação: {erro}")

    return redirect(url_for(endpoint))


@app.route("/financeiro")
@perfil_permitido("pcp")
def financeiro():
    return redirect(url_for("movimentacoes_entradas"))


@app.route("/movimentacoes")
@perfil_permitido("pcp")
def movimentacoes():
    return redirect(url_for("movimentacoes_entradas"))


@app.route("/movimentacoes/entradas", methods=["GET", "POST"])
@perfil_permitido("pcp")
def movimentacoes_entradas():
    if request.method == "POST":
        return salvar_movimentacao_por_visao("Entrada", "movimentacoes_entradas")

    return render_template(
        "financeiro.html",
        **contexto_movimentacoes("entradas", "Entrada")
    )


@app.route("/movimentacoes/despesas", methods=["GET", "POST"])
@perfil_permitido("pcp")
def movimentacoes_despesas():
    if request.method == "POST":
        return salvar_movimentacao_por_visao("Saída", "movimentacoes_despesas")

    return render_template(
        "financeiro.html",
        **contexto_movimentacoes("despesas", "Saída")
    )


@app.route("/movimentacoes/estoque")
@perfil_permitido("pcp")
def movimentacoes_estoque():
    return render_template(
        "financeiro.html",
        **contexto_movimentacoes("estoque")
    )


@app.route("/financeiro/editar/<int:movimentacao_id>", methods=["GET", "POST"])
@perfil_permitido("pcp")
def editar_movimentacao_financeira(movimentacao_id):
    movimentacao = buscar_movimentacao_financeira_por_id(movimentacao_id)

    if not movimentacao:
        flash("Movimentação financeira não encontrada.")
        return redirect(url_for("financeiro"))

    if request.method == "POST":
        try:
            atualizar_movimentacao_financeira(movimentacao_id, request.form)
            flash("Movimentação financeira atualizada com sucesso.")
            return redirect(url_for(destino_movimentacao_por_tipo(request.form.get("tipo", movimentacao["tipo"]))))
        except Exception as erro:
            flash(f"Erro ao atualizar movimentação: {erro}")

    return render_template(
        "financeiro_editar.html",
        movimentacao=movimentacao,
        categorias_entrada=CATEGORIAS_FINANCEIRAS_ENTRADA,
        categorias_saida=CATEGORIAS_FINANCEIRAS_SAIDA,
        formas_pagamento=FORMAS_PAGAMENTO_FINANCEIRO,
        status_opcoes=STATUS_FINANCEIRO,
        voltar_endpoint=destino_movimentacao_por_tipo(movimentacao["tipo"])
    )


@app.route("/financeiro/excluir/<int:movimentacao_id>", methods=["POST"])
@perfil_permitido("pcp")
def excluir_movimentacao_financeira_rota(movimentacao_id):
    movimentacao = buscar_movimentacao_financeira_por_id(movimentacao_id)
    excluir_movimentacao_financeira(movimentacao_id)
    flash("Movimentação financeira excluída com sucesso.")
    return redirect(url_for(destino_movimentacao_por_tipo(movimentacao["tipo"] if movimentacao else "Entrada")))


@app.route("/fornecedores", methods=["GET", "POST"])
@perfil_permitido("pcp")
def fornecedores():
    conn = conectar()
    cursor = conn.cursor()

    if request.method == "POST":
        nome = request.form["nome"].strip()

        if nome:
            try:
                cursor.execute(q("""
                INSERT INTO fornecedores (nome)
                VALUES (?)
                """), (nome,))

                conn.commit()
                flash("Fornecedor cadastrado com sucesso.")
            except Exception:
                conn.rollback()
                flash("Este fornecedor já está cadastrado.")

        conn.close()
        return redirect(url_for("fornecedores"))

    cursor.execute("""
    SELECT *
    FROM fornecedores
    ORDER BY nome
    """)

    fornecedores = cursor.fetchall()
    conn.close()

    return render_template(
        "fornecedores.html",
        fornecedores=fornecedores
    )


@app.route("/cadastros/equipamentos", methods=["GET", "POST"])
@perfil_permitido("pcp", "producao")
def cadastro_equipamentos_manutencao():
    if request.method == "POST":
        try:
            manutencao_service.salvar_equipamento_manutencao(request.form)
            flash("Equipamento cadastrado com sucesso.")
        except Exception as erro:
            flash(str(erro))

        return redirect(url_for("cadastro_equipamentos_manutencao"))

    return render_template(
        "cadastro_equipamentos.html",
        **manutencao_service.preparar_contexto_cadastro_equipamentos(request.args)
    )


@app.route("/cadastros/equipamentos/<int:equipamento_id>/editar", methods=["POST"])
@perfil_permitido("pcp", "producao")
def editar_equipamento_manutencao(equipamento_id):
    try:
        manutencao_service.atualizar_equipamento_manutencao(equipamento_id, request.form)
        flash("Equipamento atualizado com sucesso.")
    except Exception as erro:
        flash(str(erro))

    return redirect(url_for("cadastro_equipamentos_manutencao", busca=request.form.get("busca", "")))


@app.route("/cadastros/equipamentos/<int:equipamento_id>/excluir", methods=["POST"])
@perfil_permitido("pcp", "producao")
def excluir_equipamento_manutencao(equipamento_id):
    try:
        manutencao_service.excluir_equipamento_manutencao(equipamento_id)
        flash("Equipamento removido com sucesso.")
    except Exception as erro:
        flash(str(erro))

    return redirect(url_for("cadastro_equipamentos_manutencao", busca=request.form.get("busca", "")))


@app.route("/manutencao", methods=["GET", "POST"])
@perfil_permitido("pcp", "producao")
def manutencao():
    if request.method == "POST":
        try:
            manutencao_service.salvar_ordem_manutencao(request.form)
            flash("Ordem de manutencao aberta com sucesso.")
        except Exception as erro:
            flash(str(erro))

        return redirect(url_for("manutencao"))

    return render_template(
        "manutencao.html",
        **manutencao_service.preparar_contexto_manutencao(request.args)
    )


@app.route("/manutencao/ordem/<int:ordem_id>/atualizar", methods=["POST"])
@perfil_permitido("pcp", "producao")
def atualizar_ordem_manutencao_rota(ordem_id):
    try:
        manutencao_service.atualizar_ordem_manutencao(ordem_id, request.form)
        flash("Ordem de manutencao atualizada com sucesso.")
    except Exception as erro:
        flash(str(erro))

    return redirect(url_for("manutencao"))


@app.route("/manutencao/ordem/<int:ordem_id>/recursos", methods=["POST"])
@perfil_permitido("pcp", "producao")
def salvar_recursos_ordem_manutencao_rota(ordem_id):
    try:
        manutencao_service.salvar_recursos_ordem_manutencao(ordem_id, request.form)
        flash("Lista de materiais e terceiros atualizada com sucesso.")
    except Exception as erro:
        flash(str(erro))

    return redirect(url_for(
        "manutencao",
        status=request.form.get("status_filtro", "Todos"),
        equipamento_id=request.form.get("equipamento_filtro", ""),
    ))


# ============================================================
# IMPORTAÇÃO OFICIAL DE MAIO/2026
# Rota temporária para carregar OPs + descartes via Excel.
# ============================================================

MOTIVOS_DESCARTE_IMPORTACAO_MAIO = [
    "Hematomas",
    "Contaminação no processo",
    "Aspecto repugnante",
    "Doença",
    "Cozimento",
    "Caquexia",
    "Fratura",
    "Carcaça incompleta",
    "Outra",
]


def normalizar_cabecalho_importacao(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    mapa = str.maketrans("áàâãéêíóôõúç", "aaaaeeiooouc")
    return texto.translate(mapa).replace("_", " ")


def valor_excel_para_data_texto(valor):
    if not valor:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    for formato in ["%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"]:
        try:
            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except Exception:
            pass
    raise ValueError(f"Data inválida: {valor}")


def numero_importacao(valor, padrao=0):
    if valor is None or str(valor).strip() == "":
        return padrao
    if isinstance(valor, str):
        valor = valor.replace("R$", "").replace(".", "").replace(",", ".").strip()
    return float(valor)


def inteiro_importacao(valor, padrao=0):
    return int(round(numero_importacao(valor, padrao)))


def texto_importacao(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def mapa_colunas_por_cabecalho(ws):
    cabecalhos = {}
    for idx, celula in enumerate(ws[1], start=1):
        cabecalhos[normalizar_cabecalho_importacao(celula.value)] = idx
    return cabecalhos


def localizar_coluna(cabecalhos, nomes_possiveis):
    for nome in nomes_possiveis:
        chave = normalizar_cabecalho_importacao(nome)
        if chave in cabecalhos:
            return cabecalhos[chave]
    return None


def ler_planilha_importacao_maio(arquivo_excel):
    wb = load_workbook(arquivo_excel, data_only=True)
    erros = []
    avisos = []
    ops = []
    descartes = []

    if "OPs" not in wb.sheetnames:
        return {"erros": ["A planilha precisa ter uma aba chamada OPs."], "avisos": [], "ops": [], "descartes": []}

    ws_ops = wb["OPs"]
    cab = mapa_colunas_por_cabecalho(ws_ops)
    col = {
        "data": localizar_coluna(cab, ["data"]),
        "sku": localizar_coluna(cab, ["sku"]),
        "fornecedor": localizar_coluna(cab, ["fornecedor"]),
        "gta": localizar_coluna(cab, ["gta"]),
        "nota_fiscal": localizar_coluna(cab, ["nota fiscal", "nf", "nota_fiscal"]),
        "quantidade_aves": localizar_coluna(cab, ["quantidade de aves", "aves", "quantidade_aves"]),
        "mortes_antes_pendura": localizar_coluna(cab, ["mortes na gaiola", "mortes_antes_pendura"]),
        "peso_vivo": localizar_coluna(cab, ["peso vivo", "peso_vivo"]),
        "unidades_produzidas": localizar_coluna(cab, ["unidades produzidas", "unidades_produzidas"]),
        "kg_produzidos": localizar_coluna(cab, ["kg produzidos", "kg_produzidos"]),
        "observacoes": localizar_coluna(cab, ["observações", "observacoes", "obs"]),
    }

    for chave in ["data", "sku", "fornecedor", "quantidade_aves", "mortes_antes_pendura", "peso_vivo", "unidades_produzidas", "kg_produzidos"]:
        if not col.get(chave):
            erros.append(f"Aba OPs: coluna obrigatória ausente: {chave}.")

    if erros:
        return {"erros": erros, "avisos": avisos, "ops": ops, "descartes": descartes}

    for linha in range(2, ws_ops.max_row + 1):
        if not ws_ops.cell(linha, col["data"]).value:
            continue
        try:
            op = {
                "linha": linha,
                "data": valor_excel_para_data_texto(ws_ops.cell(linha, col["data"]).value),
                "sku": texto_importacao(ws_ops.cell(linha, col["sku"]).value) or "Galinha Cortada",
                "fornecedor": texto_importacao(ws_ops.cell(linha, col["fornecedor"]).value),
                "gta": texto_importacao(ws_ops.cell(linha, col["gta"]).value) if col.get("gta") else "",
                "nota_fiscal": texto_importacao(ws_ops.cell(linha, col["nota_fiscal"]).value) if col.get("nota_fiscal") else "",
                "quantidade_aves": inteiro_importacao(ws_ops.cell(linha, col["quantidade_aves"]).value),
                "mortes_antes_pendura": inteiro_importacao(ws_ops.cell(linha, col["mortes_antes_pendura"]).value),
                "peso_vivo": numero_importacao(ws_ops.cell(linha, col["peso_vivo"]).value),
                "unidades_produzidas": numero_importacao(ws_ops.cell(linha, col["unidades_produzidas"]).value),
                "kg_produzidos": numero_importacao(ws_ops.cell(linha, col["kg_produzidos"]).value),
                "observacoes": texto_importacao(ws_ops.cell(linha, col["observacoes"]).value) if col.get("observacoes") else "",
            }
            if not op["fornecedor"]:
                erros.append(f"Aba OPs linha {linha}: fornecedor vazio.")
            if op["quantidade_aves"] <= 0:
                erros.append(f"Aba OPs linha {linha}: quantidade de aves precisa ser maior que zero.")
            if op["peso_vivo"] <= 0:
                erros.append(f"Aba OPs linha {linha}: peso vivo precisa ser maior que zero.")
            if op["unidades_produzidas"] <= 0:
                erros.append(f"Aba OPs linha {linha}: unidades produzidas precisa ser maior que zero.")
            if op["kg_produzidos"] <= 0:
                erros.append(f"Aba OPs linha {linha}: kg produzidos está vazio ou zerado.")
            ops.append(op)
        except Exception as erro:
            erros.append(f"Aba OPs linha {linha}: {erro}")

    if "Descartes" in wb.sheetnames:
        ws_desc = wb["Descartes"]
        cab_desc = mapa_colunas_por_cabecalho(ws_desc)
        col_data = localizar_coluna(cab_desc, ["data"])
        if not col_data:
            erros.append("Aba Descartes: coluna Data ausente.")
        else:
            col_motivos = []
            for motivo in MOTIVOS_DESCARTE_IMPORTACAO_MAIO:
                c = localizar_coluna(cab_desc, [motivo])
                if c:
                    col_motivos.append((motivo, c))
            if not col_motivos:
                avisos.append("Aba Descartes existe, mas nenhum motivo conhecido foi encontrado.")
            for linha in range(2, ws_desc.max_row + 1):
                data_raw = ws_desc.cell(linha, col_data).value
                if not data_raw:
                    continue
                try:
                    data = valor_excel_para_data_texto(data_raw)
                    for motivo, c in col_motivos:
                        qtd = numero_importacao(ws_desc.cell(linha, c).value)
                        if qtd > 0:
                            descartes.append({
                                "linha": linha,
                                "data": data,
                                "setor": "Não informado",
                                "categoria": "Condenação / Descarte",
                                "motivo": motivo,
                                "quantidade": qtd,
                                "unidade": "aves",
                                "observacoes": "Importado da planilha oficial de maio/2026. Setor não informado na origem.",
                            })
                except Exception as erro:
                    erros.append(f"Aba Descartes linha {linha}: {erro}")
    else:
        avisos.append("A planilha não possui aba Descartes. Apenas OPs serão importadas.")

    if not ops:
        erros.append("Nenhuma OP válida encontrada para importação.")

    return {"erros": erros, "avisos": avisos, "ops": ops, "descartes": descartes}


def resumir_importacao_maio(dados):
    ops = dados.get("ops", [])
    descartes = dados.get("descartes", [])
    total_descartes = sum(item["quantidade"] for item in descartes)
    motivos = {}
    for item in descartes:
        motivos[item["motivo"]] = motivos.get(item["motivo"], 0) + item["quantidade"]
    return {
        "total_ops": len(ops),
        "total_aves": round(sum(op["quantidade_aves"] for op in ops), 2),
        "total_mortes": round(sum(op["mortes_antes_pendura"] for op in ops), 2),
        "total_peso_vivo": round(sum(op["peso_vivo"] for op in ops), 2),
        "total_unidades": round(sum(op["unidades_produzidas"] for op in ops), 2),
        "total_kg": round(sum(op["kg_produzidos"] for op in ops), 2),
        "total_descartes": round(total_descartes, 2),
        "motivos": [
            {"motivo": m, "quantidade": round(q, 2), "percentual": round((q / total_descartes * 100), 2) if total_descartes > 0 else 0}
            for m, q in sorted(motivos.items(), key=lambda par: par[1], reverse=True)
        ],
    }


def obter_id_inserido(cursor):
    if DATABASE_URL:
        cursor.execute("SELECT LASTVAL() as id")
        return cursor.fetchone()["id"]
    return cursor.lastrowid


def excluir_historico_maio_2026(cursor):
    cursor.execute(q("""
    SELECT id FROM ordens_producao
    WHERE data BETWEEN ? AND ?
      AND observacoes LIKE ?
    """), ("2026-05-01", "2026-05-31", "%Importação oficial Maio/2026%"))
    ids = [item["id"] for item in cursor.fetchall()]
    if not ids:
        return 0
    placeholders = ",".join(["?"] * len(ids))
    for tabela in ["apontamentos_setor", "apontamentos_producao", "apontamentos_mao_obra", "apontamentos_paradas", "apontamentos_descartes", "apontamentos_tempos_setor"]:
        cursor.execute(q(f"DELETE FROM {tabela} WHERE op_id IN ({placeholders})"), tuple(ids))
    cursor.execute(q(f"DELETE FROM ordens_producao WHERE id IN ({placeholders})"), tuple(ids))
    return len(ids)


def importar_dados_oficiais_maio(dados, substituir=True):
    criar_banco()
    criar_tabela_tempos_setor()
    if dados.get("erros"):
        raise ValueError("A planilha possui erros de validação e não pode ser importada.")
    conn = conectar()
    cursor = conn.cursor()
    try:
        removidas = excluir_historico_maio_2026(cursor) if substituir else 0
        op_por_data = {}
        ops_importadas = 0
        descartes_importados = 0
        for op in dados["ops"]:
            peso_medio = op["peso_vivo"] / op["quantidade_aves"] if op["quantidade_aves"] else 0
            obs = (op.get("observacoes") or "").strip()
            obs = (obs + " | " if obs else "") + "Importação oficial Maio/2026"
            cursor.execute(q("""
            INSERT INTO ordens_producao (
                data, sku, fornecedor, gta, nota_fiscal, quantidade_aves,
                mortes_antes_pendura, peso_vivo, peso_medio, observacoes, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """), (op["data"], op["sku"], op["fornecedor"], op["gta"], op["nota_fiscal"], op["quantidade_aves"], op["mortes_antes_pendura"], op["peso_vivo"], peso_medio, obs, "Encerrada"))
            op_id = obter_id_inserido(cursor)
            op_por_data[op["data"]] = op_id
            ops_importadas += 1
            cursor.execute(q("""
            INSERT INTO apontamentos_producao (op_id, data, setor, quantidade, unidade, observacoes)
            VALUES (?, ?, ?, ?, ?, ?)
            """), (op_id, op["data"], "Expedição", op["unidades_produzidas"], "unidades", "Produção final importada da planilha oficial de maio/2026."))
            cursor.execute(q("""
            INSERT INTO apontamentos_producao (op_id, data, setor, quantidade, unidade, observacoes)
            VALUES (?, ?, ?, ?, ?, ?)
            """), (op_id, op["data"], "Expedição", op["kg_produzidos"], "kg", "Kg final produzido importado da planilha oficial de maio/2026."))
        for descarte in dados["descartes"]:
            op_id = op_por_data.get(descarte["data"])
            if not op_id:
                continue
            cursor.execute(q("""
            INSERT INTO apontamentos_descartes (op_id, data, setor, categoria, motivo, quantidade, unidade, observacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """), (op_id, descarte["data"], descarte["setor"], descarte["categoria"], descarte["motivo"], descarte["quantidade"], descarte["unidade"], descarte["observacoes"]))
            descartes_importados += 1
        conn.commit()
        conn.close()
        return {"removidas": removidas, "ops_importadas": ops_importadas, "descartes_importados": descartes_importados}
    except Exception:
        conn.rollback()
        conn.close()
        raise


@app.route("/importar-maio", methods=["GET", "POST"])
@perfil_permitido("pcp")
def importar_maio():
    resultado = None
    resumo = None
    erros = []
    avisos = []
    importado = False
    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        acao = request.form.get("acao", "validar")
        substituir = request.form.get("substituir") == "sim"
        if not arquivo or not arquivo.filename:
            erros.append("Selecione a planilha de maio em Excel.")
        else:
            try:
                dados = ler_planilha_importacao_maio(arquivo)
                erros = dados.get("erros", [])
                avisos = dados.get("avisos", [])
                resumo = resumir_importacao_maio(dados)
                if acao == "importar" and not erros:
                    if not substituir:
                        erros.append("Marque a confirmação de substituição segura dos dados oficiais de Maio/2026 importados anteriormente.")
                    else:
                        resultado = importar_dados_oficiais_maio(dados, substituir=True)
                        importado = True
                        flash("Importação oficial de Maio/2026 concluída com sucesso.")
            except Exception as erro:
                erros.append(str(erro))
    return render_template("importar_maio.html", resumo=resumo, erros=erros, avisos=avisos, resultado=resultado, importado=importado)


register_producao_routes(app, {
    "op_possui_caixa_pa": op_possui_caixa_pa,
    "remover_movimentacoes_estoque_pi_por_op": remover_movimentacoes_estoque_pi_por_op,
})

register_qualidade_routes(app, {
    "criar_banco": criar_banco,
})

register_almoxarifado_routes(app)

def inicializar_schema_aplicacao():
    inicializar_schema_uma_vez([
        criar_banco,
        criar_tabela_tempos_setor,
        criar_tabelas_custos,
        criar_tabela_vendas,
        criar_tabelas_expedicao,
        criar_tabelas_estoque_pi_pa,
        criar_tabelas_almoxarifado,
        criar_tabelas_estoque_almoxarifado,
        criar_tabelas_receitas_sku,
        criar_tabela_movimentacoes_financeiras,
        criar_tabela_fornecedores,
        manutencao_service.criar_tabelas_manutencao,
    ])


inicializar_schema_aplicacao()


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)

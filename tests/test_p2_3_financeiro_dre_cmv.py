"""Cenarios de aceite da P2.3: FIFO, estorno, calculabilidade e DRE."""

import os
from pathlib import Path
import sys
import tempfile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BANCO = tempfile.NamedTemporaryFile(prefix="frigodatta-p23-", suffix=".db", delete=False)
BANCO.close()
os.environ["DB_NAME"] = BANCO.name
os.environ.pop("DATABASE_URL", None)
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database import conectar  # noqa: E402
from modules.cmv import services as cmv  # noqa: E402
from modules.dre.services import buscar_dados_dre_gerencial  # noqa: E402
from modules.movimentacoes.services import criar_tabela_movimentacoes_financeiras  # noqa: E402


def limpar():
    cmv.criar_tabelas_cmv()
    conn = conectar()
    cur = conn.cursor()
    for tabela in ("cmv_consumos", "cmv_eventos", "cmv_camadas", "cmv_auditoria"):
        cur.execute(f"DELETE FROM {tabela}")
    conn.commit(); conn.close()


def camada(chave, quantidade, custo, conhecida=True, data="2026-08-01"):
    return cmv.registrar_camada(produto="Frango", unidade="KG", data_entrada=data,
        quantidade=quantidade, custo_unitario=custo, custo_conhecido=conhecida,
        origem_tipo="ENTRADA_TESTE", idempotency_key=chave, usuario="Teste")


def venda(chave, quantidade=150):
    return cmv.registrar_saida(data_evento="2026-08-10", documento="VENDA-1",
        produto="Frango", unidade="KG", quantidade=quantidade,
        origem_tipo="VENDA_TESTE", idempotency_key=chave, usuario="Teste")


def test_fifo_100_a_10_mais_100_a_12_venda_150():
    limpar(); camada("L1", 100, 10); camada("L2", 100, 12, data="2026-08-02")
    evento, criado = venda("V1")
    assert criado and evento["estado_calculo"] == "CALCULAVEL"
    assert evento["custo_total"] == 1600
    resumo = cmv.resumo_periodo("2026-08-01", "2026-08-31")
    assert resumo["cmv_total"] == 1600 and resumo["cobertura_percentual"] == 100
    estoque = cmv.estoque_valorizado()[0]
    assert estoque["quantidade"] == 50 and estoque["valor_conhecido"] == 600


def test_estorno_restaura_as_mesmas_camadas_e_zera_cmv():
    limpar(); camada("L1", 100, 10); camada("L2", 100, 12, data="2026-08-02")
    evento, _ = venda("V1")
    estorno, criado = cmv.estornar_saida(evento["id"], data_evento="2026-08-11",
        idempotency_key="E1", justificativa="Cancelamento controlado", usuario="Teste")
    assert criado and estorno["custo_total"] == -1600
    resumo = cmv.resumo_periodo("2026-08-01", "2026-08-31")
    assert resumo["cmv_total"] == 0
    assert cmv.estoque_valorizado()[0]["quantidade"] == 200


def test_cobertura_parcial_e_ausente_nao_viram_zero():
    limpar(); camada("L1", 90, 10)
    evento, _ = venda("V1", 100)
    assert evento["estado_calculo"] == "PARCIAL"
    resumo = cmv.resumo_periodo("2026-08-01", "2026-08-31")
    assert resumo["cmv_total"] == 900 and resumo["cobertura_percentual"] == 90
    limpar(); evento, _ = venda("V2", 25)
    assert evento["estado_calculo"] == "NAO_CALCULAVEL" and evento["custo_total"] is None


def test_custo_zero_explicito_e_calculavel_e_idempotente():
    limpar(); _, criada = camada("GRATIS", 10, 0, conhecida=True)
    _, repetida = camada("GRATIS", 10, 0, conhecida=True)
    evento, _ = venda("V1", 10)
    assert criada and not repetida
    assert evento["estado_calculo"] == "CALCULAVEL" and evento["custo_total"] == 0


def test_descarte_fica_fora_do_cmv_de_venda_e_exportacoes_abrem():
    limpar(); camada("L1", 20, 5)
    evento, _ = cmv.registrar_descarte(data_evento="2026-08-12", documento="PNC-1",
        produto="Frango", unidade="KG", quantidade=5, origem_tipo="PNC_DESCARTE",
        idempotency_key="D1", usuario="Teste")
    assert evento["tipo"] == "DESCARTE"
    assert cmv.resumo_periodo("2026-08-01", "2026-08-31")["cmv_total"] is None
    arquivo = cmv.gerar_excel("2026-08-01", "2026-08-31")
    assert load_workbook(arquivo).sheetnames == ["Resumo CMV", "Consumos FIFO", "Estoque valorizado"]
    assert cmv.gerar_pdf("2026-08-01", "2026-08-31").read(4) == b"%PDF"


def test_cenario_dre_completo_usa_competencia_e_servico_fifo():
    limpar()
    conn = conectar(); cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS vendas_diarias(id INTEGER PRIMARY KEY,data TEXT,sku TEXT,quantidade REAL,unidade TEXT,quantidade_unidades REAL,quantidade_kg REAL,receita REAL)")
    cur.execute("DELETE FROM vendas_diarias")
    conn.close()
    criar_tabela_movimentacoes_financeiras()
    conn = conectar(); cur = conn.cursor()
    cur.execute("DELETE FROM movimentacoes_financeiras")
    linhas = [
        ("Entrada", "Receita Bruta", 10000, "Receita Bruta", "Entrada"),
        ("Saida", "Impostos sobre Vendas", 1000, "Deducoes da Receita", "Saida"),
        ("Saida", "Energia", 2000, "Despesas Operacionais", "Saida"),
        ("Saida", "Despesas Financeiras", 500, "Resultado Nao Operacional", "Saida"),
    ]
    for tipo, categoria, valor, linha, tipo_conta in linhas:
        cur.execute("INSERT INTO movimentacoes_financeiras(data_documento,data_vencimento,tipo,categoria,descricao,valor,status,linha_dre,tipo_conta,categoria_plano) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    ("2026-08-10", "2026-08-20", tipo, categoria, categoria, valor, "Pendente", linha, tipo_conta, categoria))
    conn.commit(); conn.close()
    camada("L1", 100, 10); camada("L2", 100, 12, data="2026-08-02"); venda("V1")
    dados = buscar_dados_dre_gerencial("2026-08")
    assert dados["receita_bruta"] == 10000
    assert dados["deducoes_receita"] == 1000
    assert dados["receita_operacional_liquida"] == 9000
    assert dados["cmv_total"] == 1600 and dados["cmv_estado"] == "CALCULAVEL"
    assert dados["margem_bruta"] == 7400
    assert dados["custos_operacionais_total"] == 2000
    assert dados["resultado_operacional"] == 5400
    assert dados["resultado_nao_operacional"] == -500
    assert dados["resultado_gerencial_periodo"] == 4900


def test_custo_por_op_usa_consumo_valorizado_e_peso_liquido_oficial():
    limpar(); conn = conectar(); cur = conn.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS ordens_producao(id INTEGER PRIMARY KEY,data TEXT,fornecedor TEXT NOT NULL,quantidade_aves INTEGER NOT NULL,peso_vivo REAL NOT NULL,peso_medio REAL NOT NULL,sku TEXT,status TEXT);
    CREATE TABLE IF NOT EXISTS almoxarifado_movimentacoes(id INTEGER PRIMARY KEY,data_movimentacao TEXT NOT NULL,insumo_id INTEGER NOT NULL,lote_id INTEGER,quantidade REAL,valor_unitario REAL,valor_total REAL,op_id INTEGER,tipo TEXT);
    CREATE TABLE IF NOT EXISTS pa_caixas(id INTEGER PRIMARY KEY,codigo_caixa TEXT NOT NULL,sku TEXT,peso_liquido REAL,quantidade_bandejas REAL,status TEXT);
    CREATE TABLE IF NOT EXISTS pa_caixa_composicao(id INTEGER PRIMARY KEY,caixa_id INTEGER,op_id INTEGER,quantidade_bandejas REAL);
    DELETE FROM ordens_producao; DELETE FROM almoxarifado_movimentacoes;
    DELETE FROM pa_caixas; DELETE FROM pa_caixa_composicao;
    INSERT INTO ordens_producao(id,data,fornecedor,quantidade_aves,peso_vivo,peso_medio,sku,status) VALUES(77,'2026-08-05','Fornecedor teste',100,200,2,'Frango Cortado','Encerrada');
    INSERT INTO almoxarifado_movimentacoes(id,data_movimentacao,insumo_id,lote_id,quantidade,valor_unitario,valor_total,op_id,tipo) VALUES(1,'2026-08-05',1,10,50,8,400,77,'SAIDA_OP');
    INSERT INTO almoxarifado_movimentacoes(id,data_movimentacao,insumo_id,lote_id,quantidade,valor_unitario,valor_total,op_id,tipo) VALUES(2,'2026-08-05',2,11,20,5,100,77,'SAIDA_OP');
    INSERT INTO pa_caixas(id,codigo_caixa,sku,peso_liquido,quantidade_bandejas,status) VALUES(1,'CX-P23','Frango Cortado',100,12,'Em estoque');
    INSERT INTO pa_caixa_composicao(id,caixa_id,op_id,quantidade_bandejas) VALUES(1,1,77,12);
    """)
    conn.commit(); conn.close()
    custo = cmv.calcular_custo_op(77)
    assert custo["estado_calculo"] == "CALCULAVEL"
    assert custo["custo_total"] == 500 and custo["quantidade_pa"] == 100
    assert custo["custo_unitario"] == 5
    camada_op, criada = cmv.registrar_camada_op(77, usuario="Teste")
    assert criada and camada_op["op_id"] == 77 and camada_op["custo_unitario"] == 5

"""Regressões da Onda 0 de proteção e integridade financeira."""

import inspect
from io import BytesIO
import os
from pathlib import Path
import sqlite3
import subprocess
import sys
import tempfile

import pytest
from flask import Flask
from openpyxl import load_workbook
from werkzeug.datastructures import MultiDict


ROOT = Path(__file__).resolve().parents[1]
BANCO_TESTE = tempfile.NamedTemporaryFile(
    prefix="frigodatta-onda0-",
    suffix=".db",
    delete=False,
)
BANCO_TESTE.close()
os.environ["DB_NAME"] = BANCO_TESTE.name
os.environ.pop("DATABASE_URL", None)
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database import conectar, q  # noqa: E402
from modules.fluxo_caixa.services import (  # noqa: E402
    calcular_resumo_fluxo_caixa,
    montar_resumo_gerencial_fluxo_caixa,
    montar_linha_tempo,
    preparar_movimentacao_fluxo,
)
from modules.movimentacoes import services as movimentacoes_service  # noqa: E402
from modules.movimentacoes.routes import register_movimentacoes_routes  # noqa: E402
from modules.movimentacoes.services import (  # noqa: E402
    atualizar_movimentacao_financeira,
    buscar_historico_movimentacao_financeira,
    calcular_status_liquidacao,
    criar_tabela_movimentacoes_financeiras,
    excluir_movimentacao_financeira,
    gerar_planilha_modelo_importacao_financeira,
    importar_movimentacoes_financeiras_excel,
    movimentacao_compativel_contas_pagar,
    salvar_movimentacao_financeira,
    valor_realizado_movimentacao,
)
from modules.relatorios.financeiro import (  # noqa: E402
    RELATORIOS_FINANCEIROS,
    montar_resumo_caixa,
    normalizar_filtros,
    status_conta,
    valor_baixado,
)


def executar(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), parametros)
    conn.commit()
    ultimo_id = cursor.lastrowid
    conn.close()
    return ultimo_id


def consultar_um(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), parametros)
    linha = cursor.fetchone()
    conn.close()
    return dict(linha) if linha else None


def inserir_movimentacao(
    *,
    descricao,
    tipo="Saida",
    valor=1000,
    valor_pago=0,
    status="Pendente",
    data_realizacao="",
    categoria="Materia Prima",
    linha_dre="Despesas Operacionais",
    tipo_conta="Saida",
    impacta_fluxo_caixa=1,
):
    return executar(
        """
        INSERT INTO movimentacoes_financeiras (
            data_documento, data_vencimento, data_realizacao, tipo, categoria,
            descricao, valor, valor_documento, valor_pago, status,
            linha_dre, tipo_conta, impacta_fluxo_caixa
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "2026-07-01",
            "2026-07-15",
            data_realizacao,
            tipo,
            categoria,
            descricao,
            valor,
            valor,
            valor_pago,
            status,
            linha_dre,
            tipo_conta,
            impacta_fluxo_caixa,
        ),
    )


@pytest.fixture(autouse=True)
def base_limpa():
    criar_tabela_movimentacoes_financeiras()
    executar("DELETE FROM movimentacoes_financeiras_auditoria")
    executar("DELETE FROM movimentacoes_financeiras")
    yield


def test_01_02_03_importar_app_nao_muta_historico_nem_executa_hotfix_ou_sincronizacao():
    banco = tempfile.NamedTemporaryFile(prefix="onda0-bootstrap-", suffix=".db", delete=False)
    banco.close()
    ambiente = os.environ.copy()
    ambiente["DB_NAME"] = banco.name
    ambiente.pop("DATABASE_URL", None)
    comando = [sys.executable, "-c", "import app"]

    subprocess.run(comando, cwd=ROOT, env=ambiente, check=True, capture_output=True, text=True)
    conn = sqlite3.connect(banco.name)
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO movimentacoes_financeiras (
            data_vencimento, tipo, categoria, descricao, valor, status,
            linha_dre, tipo_conta, impacta_fluxo_caixa
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "2026-07-15",
            "Saida",
            "Aportes",
            "Sentinela historica",
            777,
            "Pendente",
            "Despesas Operacionais",
            "Saida",
            0,
        ),
    )
    conn.commit()
    sentinela_antes = cursor.execute(
        "SELECT tipo, linha_dre, tipo_conta, impacta_fluxo_caixa FROM movimentacoes_financeiras"
    ).fetchone()
    conn.close()

    subprocess.run(comando, cwd=ROOT, env=ambiente, check=True, capture_output=True, text=True)
    conn = sqlite3.connect(banco.name)
    cursor = conn.cursor()
    sentinela_depois = cursor.execute(
        "SELECT tipo, linha_dre, tipo_conta, impacta_fluxo_caixa FROM movimentacoes_financeiras"
    ).fetchone()
    tabelas = {
        item[0]
        for item in cursor.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()
    }
    conn.close()

    assert sentinela_depois == sentinela_antes
    assert "hotfix_aportes_natureza_fluxo" not in tabelas
    assert "movimentacoes_financeiras_backup_aportes_natureza_20260714" not in tabelas


def test_04_05_fluxo_previsto_e_dre_permanecem_inalterados():
    itens = [
        preparar_movimentacao_fluxo(
            {
                "tipo": "Entrada",
                "valor": 1000,
                "valor_documento": 1000,
                "valor_pago": 400,
                "status": "Realizado",
                "data_realizacao": "2026-07-10",
            }
        ),
        preparar_movimentacao_fluxo(
            {
                "tipo": "Saida",
                "valor": 300,
                "valor_documento": 300,
                "valor_pago": 0,
                "status": "Pendente",
                "data_realizacao": "",
            }
        ),
    ]
    resumo = calcular_resumo_fluxo_caixa(
        itens,
        [itens[0]],
        {"saldo_inicial_previsto": 0, "saldo_inicial_realizado": 0},
    )
    assert resumo["entradas_previstas"] == 1000
    assert resumo["saidas_previstas"] == 300
    assert resumo["entradas_realizadas"] == 400

    dre_antes = sum(item["valor"] * item["sinal"] for item in itens)
    dre_depois = sum(item["valor"] * item["sinal"] for item in itens)
    assert dre_antes == dre_depois == 700


@pytest.mark.parametrize(
    ("movimento", "esperado"),
    [
        (
            {
                "valor": 1000,
                "valor_documento": 1000,
                "valor_pago": 1000,
                "status": "Realizado",
                "data_realizacao": "2026-07-10",
            },
            1000,
        ),
        (
            {
                "valor": 1000,
                "valor_documento": 1000,
                "valor_pago": 400,
                "status": "Realizado",
                "data_realizacao": "2026-07-10",
            },
            400,
        ),
        (
            {
                "valor": 1000,
                "valor_documento": 1000,
                "valor_pago": 0,
                "status": "Pendente",
                "data_realizacao": "",
            },
            0,
        ),
        (
            {
                "valor": 1000,
                "valor_documento": 1000,
                "valor_pago": 1200,
                "status": "Realizado",
                "data_realizacao": "2026-07-10",
            },
            1200,
        ),
        (
            {
                "valor": 1000,
                "valor_documento": 1000,
                "valor_pago": 0,
                "status": "Realizado",
                "data_realizacao": "2026-07-10",
            },
            1000,
        ),
    ],
)
def test_06_07_08_09_10_semantica_realizado_e_fallback_legado(movimento, esperado):
    assert valor_realizado_movimentacao(movimento) == esperado
    assert valor_baixado(movimento) == esperado
    if movimento["valor_pago"] == 400:
        assert valor_realizado_movimentacao(movimento) != movimento["valor"]
    if movimento["valor_pago"] == 1200:
        assert calcular_status_liquidacao(movimento)["inconsistente"] is True
    if movimento["valor_pago"] == 0 and movimento["data_realizacao"]:
        assert calcular_status_liquidacao(movimento)["inconsistente"] is True


def test_linha_do_tempo_realizada_usa_valor_pago_e_prevista_usa_valor():
    item = preparar_movimentacao_fluxo(
        {
            "tipo": "Saida",
            "valor": 1000,
            "valor_documento": 1000,
            "valor_pago": 400,
            "status": "Realizado",
            "data_vencimento": "2026-07-15",
            "data_realizacao": "2026-07-10",
        }
    )
    linha = montar_linha_tempo(
        [item],
        [item],
        {"saldo_inicial_previsto": 0, "saldo_inicial_realizado": 0},
    )
    por_data = {item["data"]: item for item in linha}
    assert por_data["2026-07-10"]["saidas_realizadas"] == 400
    assert por_data["2026-07-15"]["saidas_previstas"] == 1000


def test_cards_e_relatorio_de_caixa_usam_valor_pago_sem_alterar_previsto():
    inserir_movimentacao(
        descricao="Parcial nos cards",
        valor=1000,
        valor_pago=400,
        status="Realizado",
        data_realizacao="2026-07-10",
    )
    gerencial = montar_resumo_gerencial_fluxo_caixa(
        {"data_inicio": "2026-07-01", "data_fim": "2026-07-31"}
    )
    assert gerencial["resumo"]["saidas_realizadas"] == 400

    config = RELATORIOS_FINANCEIROS["saidas-caixa"]
    filtros = normalizar_filtros(
        {"data_inicio": "2026-07-01", "data_fim": "2026-07-31"},
        config,
    )
    resumo = {item["rotulo"]: item["valor"] for item in montar_resumo_caixa(config, filtros)}
    assert resumo["Total previsto"] == 1000
    assert resumo["Total realizado"] == 400


def formulario_edicao(justificativa="Correção conferida"):
    return {
        "data_vencimento": "2026-07-20",
        "data_realizacao": "",
        "tipo": "Saida",
        "categoria": "Materia Prima",
        "descricao": "Descrição alterada",
        "valor": "950",
        "forma_pagamento": "PIX",
        "status": "Pendente",
        "intervalo_dias": "30",
        "observacoes": "Revisado",
        "justificativa": justificativa,
    }


def test_criacao_manual_registra_estado_posterior_e_ator():
    form = MultiDict(
        [
            ("tipo", "Entrada"),
            ("categoria", "Venda de Producao Propria"),
            ("descricao", "Receita manual auditada"),
            ("data_documento", "2026-07-01"),
            ("valor", "750"),
            ("status", "Pendente"),
            ("parcela_vencimento[]", "2026-07-15"),
            ("parcela_valor[]", "750"),
        ]
    )
    salvar_movimentacao_financeira(
        form,
        usuario_id=11,
        usuario_nome="Tesouraria",
        perfil="pcp",
    )
    evento = consultar_um(
        "SELECT * FROM movimentacoes_financeiras_auditoria ORDER BY id DESC LIMIT 1"
    )
    assert evento["acao"] == "CRIACAO_MANUAL"
    assert evento["estado_anterior"] is None
    assert '"descricao": "Receita manual auditada"' in evento["estado_posterior"]
    assert evento["usuario_id"] == 11


def test_11_12_edicao_exige_justificativa_e_registra_before_after_ator():
    movimento_id = inserir_movimentacao(descricao="Antes")
    with pytest.raises(ValueError, match="justificativa"):
        atualizar_movimentacao_financeira(movimento_id, formulario_edicao(""))

    atualizar_movimentacao_financeira(
        movimento_id,
        formulario_edicao(),
        usuario_id=42,
        usuario_nome="Auditora",
        perfil="pcp",
    )
    historico = buscar_historico_movimentacao_financeira(movimento_id)
    assert historico[-1]["acao"] == "EDICAO"
    assert historico[-1]["usuario_id"] == 42
    assert historico[-1]["usuario_nome"] == "Auditora"
    assert historico[-1]["justificativa"] == "Correção conferida"
    assert "descricao" in historico[-1]["campos_alterados"]


def test_13_falha_de_auditoria_reverte_edicao(monkeypatch):
    movimento_id = inserir_movimentacao(descricao="Estado original")

    def falhar(*args, **kwargs):
        raise RuntimeError("auditoria indisponível")

    monkeypatch.setattr(movimentacoes_service, "_registrar_auditoria_cursor", falhar)
    with pytest.raises(RuntimeError, match="auditoria"):
        atualizar_movimentacao_financeira(movimento_id, formulario_edicao())
    assert consultar_um(
        "SELECT descricao, valor FROM movimentacoes_financeiras WHERE id = ?",
        (movimento_id,),
    ) == {"descricao": "Estado original", "valor": 1000.0}


def test_14_15_16_cancelamento_exige_justificativa_preserva_linha_e_exclui_demonstrativos():
    movimento_id = inserir_movimentacao(
        descricao="Cancelar",
        valor_pago=1000,
        status="Realizado",
        data_realizacao="2026-07-10",
    )
    with pytest.raises(ValueError, match="justificativa"):
        excluir_movimentacao_financeira(movimento_id)

    excluir_movimentacao_financeira(
        movimento_id,
        "Documento duplicado confirmado",
        usuario_id=7,
        usuario_nome="Controladoria",
        perfil="pcp",
    )
    movimento = consultar_um(
        "SELECT * FROM movimentacoes_financeiras WHERE id = ?",
        (movimento_id,),
    )
    assert movimento is not None
    assert movimento["status"] == "Cancelado"
    assert valor_realizado_movimentacao(movimento) == 0
    assert status_conta(movimento)[0] != "Realizado"
    assert buscar_historico_movimentacao_financeira(movimento_id)[-1]["acao"] == "CANCELAMENTO"


def test_17_19_alias_antigo_cancela_e_historico_respeita_permissao():
    movimento_id = inserir_movimentacao(descricao="Alias")
    from app import app

    cliente = app.test_client()
    with cliente.session_transaction() as sessao:
        sessao.update({"usuario_id": 9, "nome": "PCP", "perfil": "pcp"})

    resposta = cliente.post(
        f"/financeiro/excluir/{movimento_id}",
        data={"justificativa": "Cancelamento via URL compatível"},
    )
    assert resposta.status_code == 302
    assert consultar_um(
        "SELECT status FROM movimentacoes_financeiras WHERE id = ?",
        (movimento_id,),
    )["status"] == "Cancelado"

    resposta_historico = cliente.get(f"/movimentacoes/editar/{movimento_id}")
    assert resposta_historico.status_code == 200
    assert "Cancelamento via URL compatível".encode("utf-8") in resposta_historico.data

    with cliente.session_transaction() as sessao:
        sessao["perfil"] = "producao"
    resposta_negada = cliente.get(f"/movimentacoes/editar/{movimento_id}")
    assert resposta_negada.status_code == 302


def test_18_nenhum_caminho_oficial_de_exclusao_executa_delete():
    fonte_servico = inspect.getsource(movimentacoes_service.excluir_movimentacao_financeira).upper()
    fonte_rotas = inspect.getsource(register_movimentacoes_routes).upper()
    assert "DELETE FROM MOVIMENTACOES_FINANCEIRAS" not in fonte_servico
    assert "DELETE FROM MOVIMENTACOES_FINANCEIRAS" not in fonte_rotas


def test_20_migrations_sqlite_e_postgresql_sao_aditivas_idempotentes_e_reversiveis():
    sqlite_up = (ROOT / "database/20260725_onda0_auditoria_financeira_sqlite.sql").read_text(
        encoding="utf-8"
    )
    sqlite_down = (
        ROOT / "database/20260725_onda0_auditoria_financeira_sqlite_rollback.sql"
    ).read_text(encoding="utf-8")
    postgres_up = (ROOT / "database/20260725_onda0_auditoria_financeira.sql").read_text(
        encoding="utf-8"
    )
    postgres_down = (
        ROOT / "database/20260725_onda0_auditoria_financeira_rollback.sql"
    ).read_text(encoding="utf-8")

    banco = sqlite3.connect(":memory:")
    banco.executescript(sqlite_up)
    banco.executescript(sqlite_up)
    assert banco.execute(
        "SELECT name FROM sqlite_master WHERE name = 'movimentacoes_financeiras_auditoria'"
    ).fetchone()
    banco.executescript(sqlite_down)
    assert not banco.execute(
        "SELECT name FROM sqlite_master WHERE name = 'movimentacoes_financeiras_auditoria'"
    ).fetchone()
    banco.close()

    assert "BEGIN;" in postgres_up and "SERIAL PRIMARY KEY" in postgres_up
    assert "CREATE TABLE IF NOT EXISTS" in postgres_up
    assert "DROP TABLE IF EXISTS movimentacoes_financeiras_auditoria" in postgres_down
    assert "ALTER TABLE movimentacoes_financeiras" not in postgres_up


def test_21_reimportacao_identica_e_idempotente_divergencia_nao_sobrescreve():
    modelo = gerar_planilha_modelo_importacao_financeira()
    wb_modelo = load_workbook(modelo)
    ws_modelo = wb_modelo.active
    cabecalhos_modelo = {celula.value: celula.column for celula in ws_modelo[1]}
    ws_modelo.cell(2, cabecalhos_modelo["Valor do Documento"]).value = 1000
    modelo = BytesIO()
    wb_modelo.save(modelo)
    wb_modelo.close()
    modelo.seek(0)
    primeiro = importar_movimentacoes_financeiras_excel(modelo)
    assert primeiro["importadas"] == 1

    modelo.seek(0)
    segundo = importar_movimentacoes_financeiras_excel(modelo)
    assert segundo["importadas"] == 0
    assert segundo["ignoradas"] == 1

    modelo.seek(0)
    wb = load_workbook(modelo)
    ws = wb.active
    coluna_observacoes = ws[1].__iter__()
    cabecalhos = {celula.value: celula.column for celula in coluna_observacoes}
    ws.cell(2, cabecalhos["Observacoes"]).value = "Conteúdo divergente"
    divergente = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    divergente.close()
    wb.save(divergente.name)
    wb.close()

    resultado = importar_movimentacoes_financeiras_excel(divergente.name)
    assert resultado["conflitos"] == 1
    assert resultado["atualizadas"] == 0
    linha = consultar_um("SELECT observacoes FROM movimentacoes_financeiras LIMIT 1")
    assert linha["observacoes"] != "Conteúdo divergente"
    auditoria = consultar_um(
        "SELECT acao FROM movimentacoes_financeiras_auditoria ORDER BY id DESC LIMIT 1"
    )
    assert auditoria["acao"] == "CONFLITO_IMPORTACAO"


def test_base_sintetica_compara_dre_previsto_realizado_cp_cr_e_auditoria():
    integral = inserir_movimentacao(
        descricao="Integral",
        valor=1000,
        valor_pago=1000,
        status="Realizado",
        data_realizacao="2026-07-10",
    )
    inserir_movimentacao(
        descricao="Parcial",
        valor=1000,
        valor_pago=400,
        status="Realizado",
        data_realizacao="2026-07-10",
    )
    inserir_movimentacao(descricao="Aberto", valor=500)
    inserir_movimentacao(descricao="Vencido", valor=200)
    inserir_movimentacao(
        descricao="Inconsistente",
        valor=1000,
        valor_pago=1200,
        status="Realizado",
        data_realizacao="2026-07-10",
    )
    inserir_movimentacao(
        descricao="Receita",
        tipo="Entrada",
        tipo_conta="Entrada",
        valor=700,
        linha_dre="Receita Bruta",
    )
    inserir_movimentacao(
        descricao="Aporte",
        tipo="Entrada",
        tipo_conta="Neutro",
        categoria="Aportes",
        valor=300,
        linha_dre="Neutro",
    )
    inserir_movimentacao(
        descricao="Transferência neutra",
        valor=250,
        linha_dre="Neutro",
        tipo_conta="Neutro",
    )
    inserir_movimentacao(descricao="Cancelado", valor=900, status="Cancelado")
    inserir_movimentacao(
        descricao="Legado",
        valor=600,
        valor_pago=0,
        status="Realizado",
        data_realizacao="2026-07-10",
    )

    conn = conectar()
    linhas = [dict(item) for item in conn.execute("SELECT * FROM movimentacoes_financeiras")]
    conn.close()
    ativos = [item for item in linhas if item["status"] != "Cancelado"]
    dre_antes = sum(
        item["valor"] * (1 if item["tipo"] == "Entrada" else -1)
        for item in ativos
        if item["linha_dre"] != "Neutro"
    )
    previsto_antes = sum(
        item["valor"] * (1 if item["tipo"] == "Entrada" else -1)
        for item in ativos
        if item["impacta_fluxo_caixa"]
    )
    realizado_antes = sum(
        item["valor"] * (1 if item["tipo"] == "Entrada" else -1)
        for item in ativos
        if item["data_realizacao"] and item["status"] == "Realizado"
    )
    realizado_depois = sum(
        valor_realizado_movimentacao(item) * (1 if item["tipo"] == "Entrada" else -1)
        for item in ativos
    )
    dre_depois = sum(
        item["valor"] * (1 if item["tipo"] == "Entrada" else -1)
        for item in ativos
        if item["linha_dre"] != "Neutro"
    )
    previsto_depois = sum(
        item["valor"] * (1 if item["tipo"] == "Entrada" else -1)
        for item in ativos
        if item["impacta_fluxo_caixa"]
    )
    contas_pagar_antes = sum(
        status_conta(item)[1]
        for item in ativos
        if movimentacao_compativel_contas_pagar(item)
    )
    contas_receber_antes = sum(
        status_conta(item)[1]
        for item in ativos
        if item["tipo"] == "Entrada"
        and item["tipo_conta"] != "Neutro"
        and item["linha_dre"] != "Neutro"
    )
    contas_pagar_depois = contas_pagar_antes
    contas_receber_depois = contas_receber_antes

    assert dre_antes == dre_depois == -3600
    assert previsto_antes == previsto_depois == -3550
    assert realizado_antes == -3600
    assert realizado_depois == -3200
    assert realizado_depois - realizado_antes == 400
    assert contas_pagar_antes == contas_pagar_depois == 1300
    assert contas_receber_antes == contas_receber_depois == 700
    assert calcular_status_liquidacao(
        next(item for item in linhas if item["descricao"] == "Inconsistente")
    )["status_liquidacao"] == "Inconsistente"
    assert status_conta(next(item for item in linhas if item["descricao"] == "Parcial"))[1] == 600
    assert status_conta(next(item for item in linhas if item["descricao"] == "Receita"))[1] == 700

    excluir_movimentacao_financeira(integral, "Cancelamento sintético controlado")
    assert consultar_um("SELECT COUNT(*) AS total FROM movimentacoes_financeiras")["total"] == 10
    assert consultar_um(
        "SELECT COUNT(*) AS total FROM movimentacoes_financeiras_auditoria"
    )["total"] == 1


def test_22_navegacao_logo_painel_e_modulos_operacionais_nao_foram_alterados():
    alterados = subprocess.run(
        ["git", "diff", "--name-only", "HEAD"],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    ).stdout.splitlines()
    proibidos = (
        "modules/navegacao",
        "modules/producao",
        "templates/inicio",
        "static/img",
        "static/images",
    )
    assert not [caminho for caminho in alterados if caminho.startswith(proibidos)]

from io import BytesIO
from pathlib import Path
import sqlite3
import sys
import tempfile

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from modules.relatorios import financeiro


COLUNAS = [
    "id", "documento_id", "data_documento", "data_vencimento", "data_realizacao",
    "tipo", "categoria", "categoria_plano", "subcategoria", "favorecido", "parceiro",
    "origem_importacao", "valor", "valor_documento", "valor_pago", "valor_liquido",
    "status", "linha_dre", "tipo_conta", "impacta_fluxo_caixa", "descricao",
    "historico", "numero_documento",
]


def abrir_conexao(caminho):
    conn = sqlite3.connect(caminho)
    conn.row_factory = sqlite3.Row
    return conn


def criar_base(caminho):
    conn = abrir_conexao(caminho)
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE movimentacoes_financeiras (
            id INTEGER PRIMARY KEY,
            documento_id TEXT,
            data_documento TEXT,
            data_vencimento TEXT,
            data_realizacao TEXT,
            tipo TEXT,
            categoria TEXT,
            categoria_plano TEXT,
            subcategoria TEXT,
            favorecido TEXT,
            parceiro TEXT,
            origem_importacao TEXT,
            valor REAL,
            valor_documento REAL,
            valor_pago REAL,
            valor_liquido REAL,
            status TEXT,
            linha_dre TEXT,
            tipo_conta TEXT,
            impacta_fluxo_caixa INTEGER,
            descricao TEXT,
            historico TEXT,
            numero_documento TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def inserir(conn, **dados):
    valores = {coluna: dados.get(coluna) for coluna in COLUNAS}
    colunas = ", ".join(COLUNAS)
    placeholders = ", ".join(["?"] * len(COLUNAS))
    conn.execute(
        f"INSERT INTO movimentacoes_financeiras ({colunas}) VALUES ({placeholders})",
        [valores[coluna] for coluna in COLUNAS],
    )


def popular_base(caminho):
    conn = abrir_conexao(caminho)
    linhas_validas = [
        (1, "2026-01-10", "2026-03-02", 1000, "Fornecedor A", "Materia Prima", "DOC-001"),
        (2, "2026-02-05", "2026-03-03", 900, "Fornecedor B", "Embalagem", "DOC-002"),
        (3, "2026-03-01", "2026-03-04", 800, "Fornecedor C", "Energia", "DOC-003"),
        (4, "2026-01-20", "2026-03-05", 700, "Fornecedor D", "Materia Prima", "DOC-004"),
        (5, "2026-02-18", "2026-03-06", 600, "Fornecedor E", "Manutencao", "DOC-005"),
        (6, "2026-03-08", "2026-03-06", 600, "Fornecedor F", "Servicos", "DOC-006"),
        (7, "2026-02-22", "2026-03-07", 500, "Fornecedor G", "Embalagem", "DOC-007"),
        (8, "2026-03-12", "2026-03-08", 100, "Fornecedor H", "Agua", "DOC-008"),
        (9, "", "2026-03-09", 300, "Fornecedor I", "Frete", "DOC-009"),
    ]
    for id_, data_doc, baixa, valor, favorecido, categoria, numero in linhas_validas:
        inserir(
            conn,
            id=id_,
            documento_id=numero,
            data_documento=data_doc,
            data_vencimento="2026-03-15",
            data_realizacao=baixa,
            tipo="Saida",
            categoria=categoria,
            categoria_plano=categoria,
            subcategoria="Sub " + categoria,
            favorecido=favorecido,
            origem_importacao="Teste",
            valor=valor,
            valor_documento=valor,
            valor_pago=valor,
            valor_liquido=valor,
            status="Pago",
            linha_dre="Despesas Operacionais",
            tipo_conta="Operacional",
            impacta_fluxo_caixa=1,
            descricao="Pagamento teste",
            historico="Historico " + numero,
            numero_documento=numero,
        )

    exclusoes = [
        dict(id=20, tipo="Entrada", categoria="Venda", valor=9999, status="Recebido", impacta_fluxo_caixa=1),
        dict(id=21, tipo="Entrada", categoria="Aportes", valor=8888, status="Recebido", impacta_fluxo_caixa=1),
        dict(id=22, tipo="Saida", categoria="Transferencia", valor=7777, status="Pago", impacta_fluxo_caixa=0),
        dict(id=23, tipo="Saida", categoria="Aberto", valor=6666, status="Pendente", data_realizacao="", impacta_fluxo_caixa=1),
        dict(id=24, tipo="Saida", categoria="Fora", valor=5555, status="Pago", data_realizacao="2026-04-01", impacta_fluxo_caixa=1),
        dict(id=25, tipo="Saida", categoria="Cancelado", valor=4444, status="Cancelado", impacta_fluxo_caixa=1),
    ]
    for item in exclusoes:
        inserir(
            conn,
            id=item["id"],
            documento_id=f"DOC-{item['id']}",
            data_documento="2026-03-10",
            data_vencimento="2026-03-20",
            data_realizacao=item.get("data_realizacao", "2026-03-10"),
            tipo=item["tipo"],
            categoria=item["categoria"],
            categoria_plano=item["categoria"],
            subcategoria="Controle",
            favorecido="Controle",
            origem_importacao="Teste",
            valor=item["valor"],
            valor_documento=item["valor"],
            valor_pago=item["valor"],
            valor_liquido=item["valor"],
            status=item["status"],
            linha_dre="Neutro",
            tipo_conta="Controle",
            impacta_fluxo_caixa=item["impacta_fluxo_caixa"],
            descricao="Exclusao controle",
            historico="Exclusao controle",
            numero_documento=f"DOC-{item['id']}",
        )
    conn.commit()
    conn.close()


class BaseTemporaria:
    def __enter__(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        criar_base(self.tmp.name)
        popular_base(self.tmp.name)
        self.original_conectar = financeiro.conectar
        self.original_q = financeiro.q
        financeiro.conectar = lambda: abrir_conexao(self.tmp.name)
        financeiro.q = lambda sql: sql
        return self.tmp.name

    def __exit__(self, exc_type, exc, tb):
        financeiro.conectar = self.original_conectar
        financeiro.q = self.original_q
        Path(self.tmp.name).unlink(missing_ok=True)


def contexto_marco(natureza="Saida"):
    return financeiro.montar_contexto_relatorio_financeiro(
        "competencia-realizacao",
        {
            "data_inicio": "2026-03-01",
            "data_fim": "2026-03-31",
            "referencia_data": "realizacao",
            "natureza": natureza,
        },
    )


def substituir_linhas(caminho, linhas):
    conn = abrir_conexao(caminho)
    conn.execute("DELETE FROM movimentacoes_financeiras")
    for posicao, linha in enumerate(linhas, start=1):
        valor = linha.get("valor", 100)
        categoria = linha.get("categoria", "Controle")
        inserir(
            conn,
            id=posicao,
            documento_id=f"DOC-{posicao}",
            data_documento=linha.get("data_documento", "2026-03-01"),
            data_vencimento=linha.get("data_vencimento", "2026-03-15"),
            data_realizacao=linha.get("data_realizacao", "2026-03-20"),
            tipo=linha.get("tipo", "Saida"),
            categoria=categoria,
            categoria_plano=categoria,
            subcategoria=linha.get("subcategoria", "Controle"),
            favorecido=linha.get("favorecido", f"Fornecedor {posicao}"),
            origem_importacao="Teste",
            valor=valor,
            valor_documento=valor,
            valor_pago=linha.get("valor_pago", valor),
            valor_liquido=valor,
            status=linha.get("status", "Pago"),
            linha_dre=linha.get("linha_dre", "Despesas Operacionais"),
            tipo_conta="Operacional",
            impacta_fluxo_caixa=linha.get("impacta_fluxo_caixa", 1),
            descricao="Pagamento teste",
            historico=f"Historico {posicao}",
            numero_documento=f"DOC-{posicao}",
        )
    conn.commit()
    conn.close()


def leitura_do_contexto(contexto):
    return contexto["analise_pagamentos"]["leitura_gerencial"]["texto_exportacao"]


def test_pagamentos_realizados_reconciliam_por_competencia():
    with BaseTemporaria():
        contexto = contexto_marco()
        analise = contexto["analise_pagamentos"]
        resumo = analise["resumo"]

        assert analise["ativa"] is True
        assert resumo["total_pago"] == 5500
        assert resumo["quantidade_pagamentos"] == 9
        assert resumo["eventos_sem_data_documento"] == 1
        assert resumo["reconciliacao_ok"] is True
        assert round(sum(item["valor"] for item in analise["origem_competencia"]), 2) == resumo["total_pago"]
        assert round(sum(item["percentual"] for item in analise["origem_competencia"]), 1) == 100.0
        leitura = leitura_do_contexto(contexto)
        assert "R$ 3.700,00" in leitura
        assert "67,27%" in leitura
        assert "A maior parte dos pagamentos corresponde a documentos originados antes" in leitura
        assert "Fornecedor A" in leitura


def test_top5_individual_e_deterministico():
    with BaseTemporaria():
        top5 = contexto_marco()["analise_pagamentos"]["top5"]

        assert len(top5) == 5
        assert [item["id"] for item in top5] == [1, 2, 3, 4, 5]
        assert [item["valor_realizado"] for item in top5] == [1000, 900, 800, 700, 600]
        assert contexto_marco()["analise_pagamentos"]["resumo"]["valor_top5"] == 4000


def test_competencia_sem_data_documento_e_exclusoes():
    with BaseTemporaria():
        analise = contexto_marco()["analise_pagamentos"]
        ids = {item["id"] for item in analise["dados_detalhados"]}
        competencias = {item["competencia"]: item for item in analise["origem_competencia"]}

        assert {20, 21, 22, 23, 24, 25}.isdisjoint(ids)
        assert competencias["2026-01"]["valor"] == 1700
        assert competencias["2026-02"]["valor"] == 2000
        assert competencias["2026-03"]["valor"] == 1500
        assert competencias["Sem data do documento"]["valor"] == 300
        assert competencias["2026-02"]["principal"] is True
        leitura = leitura_do_contexto(contexto_marco())
        assert "1 eventos sem Data do Documento" in leitura
        assert "R$ 300,00" in leitura


def test_natureza_todas_e_entrada_preservam_comportamento():
    with BaseTemporaria():
        todas = contexto_marco("Todas")
        entrada = contexto_marco("Entrada")
        invalida = financeiro.montar_contexto_relatorio_financeiro(
            "competencia-realizacao",
            {
                "data_inicio": "2026-03-01",
                "data_fim": "2026-03-31",
                "referencia_data": "realizacao",
                "natureza": "Invalida",
            },
        )

        assert todas["analise_pagamentos"]["ativa"] is False
        assert entrada["analise_pagamentos"]["ativa"] is False
        assert invalida["filtros"]["natureza"] == "Todas"
        assert any(item["tipo"] == "Entrada" for item in todas["detalhes"])
        assert all(item["tipo"] == "Entrada" for item in entrada["detalhes"])
        assert "leitura dos pagamentos" in todas["analise_pagamentos"]["leitura_gerencial"]["texto_exportacao"]


def test_exportacao_excel_reconcilia_com_tela():
    with BaseTemporaria():
        contexto = contexto_marco()
        arquivo = financeiro.gerar_excel_relatorio_financeiro(contexto)
        wb = load_workbook(BytesIO(arquivo.getvalue()), data_only=True)

        assert wb.sheetnames == [
            "Resumo",
            "Top 5 Pagamentos",
            "Origem por Competencia",
            "Dados Detalhados",
            "Parametros",
        ]
        resumo = {linha[0].value: linha[1].value for linha in wb["Resumo"].iter_rows(min_row=2, max_col=2)}
        assert resumo["Total pago"] == contexto["analise_pagamentos"]["resumo"]["total_pago"]
        celulas_resumo = [cell.value for row in wb["Resumo"].iter_rows() for cell in row if cell.value]
        leitura_excel = celulas_resumo[celulas_resumo.index("LEITURA GERENCIAL") + 1]
        assert leitura_excel == contexto["analise_pagamentos"]["leitura_gerencial"]["texto_exportacao"]
        assert wb["Top 5 Pagamentos"].max_row == 6
        assert wb["Dados Detalhados"].max_row == 10


def test_leitura_maioria_no_proprio_periodo():
    with BaseTemporaria() as caminho:
        substituir_linhas(caminho, [
            {"data_documento": "2026-03-01", "valor": 700},
            {"data_documento": "2026-03-15", "valor": 100},
            {"data_documento": "2026-02-01", "valor": 200},
        ])
        leitura = leitura_do_contexto(contexto_marco())
        assert "R$ 800,00" in leitura
        assert "80,00%" in leitura
        assert "documentos originados no proprio periodo" in leitura


def test_leitura_ausencia_de_pagamentos():
    with BaseTemporaria() as caminho:
        substituir_linhas(caminho, [])
        contexto = contexto_marco()
        leitura = leitura_do_contexto(contexto)
        assert contexto["analise_pagamentos"]["leitura_gerencial"]["sem_dados"] is True
        assert "nao foram encontrados pagamentos realizados" in leitura


def test_leitura_competencia_posterior():
    with BaseTemporaria() as caminho:
        substituir_linhas(caminho, [
            {"data_documento": "2026-04-01", "valor": 300},
            {"data_documento": "2026-03-01", "valor": 100},
        ])
        leitura = leitura_do_contexto(contexto_marco())
        assert "competencia posterior" in leitura or "Documentos com competencia posterior" in leitura
        assert "R$ 300,00" in leitura
        assert "75,00%" in leitura


def test_leitura_empate_objetivo_entre_competencias():
    with BaseTemporaria() as caminho:
        substituir_linhas(caminho, [
            {"data_documento": "2026-02-01", "valor": 100},
            {"data_documento": "2026-03-01", "valor": 100},
        ])
        leitura = leitura_do_contexto(contexto_marco())
        assert "concentracao semelhante" in leitura


def test_leitura_marco_2026_valores_conhecidos():
    analise = {
        "ativa": True,
        "resumo": {
            "total_pago": 408319.70,
            "quantidade_pagamentos": 288,
            "valor_top5": 79947.47,
            "percentual_top5": 19.58,
            "eventos_sem_data_documento": 0,
            "reconciliacao_ok": True,
        },
        "origem_competencia": [
            {"competencia": "2026-01", "quantidade": 13, "valor": 67099.13, "percentual": 16.43},
            {"competencia": "2026-02", "quantidade": 60, "valor": 171433.83, "percentual": 41.99},
            {"competencia": "2026-03", "quantidade": 215, "valor": 169786.74, "percentual": 41.58},
        ],
        "top5": [
            {
                "favorecido": "FRANCISCO HELDER DE OLIVEIRA PEIXOTO",
                "valor_realizado": 31500,
                "categoria": "Materia Prima",
                "competencia_origem": "2026-01",
            }
        ],
    }
    leitura = financeiro.montar_leitura_gerencial_pagamentos(
        analise,
        {"data_inicio": "2026-03-01", "data_fim": "2026-03-31"},
    )["texto_exportacao"]

    assert "288 pagamentos" in leitura
    assert "R$ 408.319,70" in leitura
    assert "R$ 238.532,96" in leitura
    assert "58,42%" in leitura
    assert "2026-02 foi a principal competencia" in leitura
    assert "R$ 171.433,83" in leitura
    assert "41,99%" in leitura
    assert "R$ 169.786,74" in leitura
    assert "41,58%" in leitura
    assert "concentracao semelhante" in leitura
    assert "R$ 79.947,47" in leitura
    assert "19,58%" in leitura
    assert "Nao foram encontrados eventos sem Data do Documento" in leitura
    assert "reconciliada" in leitura


if __name__ == "__main__":
    falhas = []
    for nome, funcao in sorted(globals().items()):
        if nome.startswith("test_") and callable(funcao):
            try:
                funcao()
                print(f"OK {nome}")
            except Exception as erro:
                falhas.append((nome, erro))
                print(f"FAIL {nome}: {erro}")

    if falhas:
        raise SystemExit(1)

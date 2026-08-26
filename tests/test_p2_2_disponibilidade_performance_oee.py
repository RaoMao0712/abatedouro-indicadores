from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import sqlite3

import pytest
from openpyxl import load_workbook

from modules.producao import disponibilidade as disp
from modules.producao import performance as perf
from modules.producao import oee


@pytest.fixture()
def banco(tmp_path, monkeypatch):
    caminho = tmp_path / "p2-2.db"

    def conectar():
        conn = sqlite3.connect(caminho)
        conn.row_factory = sqlite3.Row
        return conn

    for modulo in (disp, perf, oee):
        monkeypatch.setattr(modulo, "conectar", conectar)
        monkeypatch.setattr(modulo, "DATABASE_URL", None)
    conn = conectar()
    conn.executescript("""
    CREATE TABLE ordens_producao(
      id INTEGER PRIMARY KEY,data TEXT,quantidade_aves INTEGER,
      mortes_antes_pendura INTEGER,status TEXT,sku TEXT,fornecedor TEXT,
      peso_vivo REAL,peso_medio REAL,gta TEXT,nota_fiscal TEXT,observacoes TEXT
    );
    CREATE TABLE apontamentos_paradas(
      id INTEGER PRIMARY KEY,op_id INTEGER,data TEXT,data_fim TEXT,setor TEXT,
      motivo TEXT,hora_inicio TEXT,hora_fim TEXT,horas_paradas REAL,
      manutencao_aberta TEXT,afeta_linha_abate INTEGER,natureza_disponibilidade TEXT
    );
    CREATE TABLE apontamentos_descartes(
      id INTEGER PRIMARY KEY,op_id INTEGER,data TEXT,setor TEXT,categoria TEXT,
      motivo TEXT,quantidade REAL,unidade TEXT,observacoes TEXT
    );
    INSERT INTO ordens_producao VALUES
      (1,'2026-08-10',1000,100,'Encerrada','Galinha Cortada','F1',2000,2,NULL,NULL,NULL),
      (2,'2026-08-11',500,0,'Encerrada','Galinha Inteira','F1',900,1.8,NULL,NULL,NULL),
      (3,'2026-08-12',2000,14,'Aberta','Galinha Cortada','F2',3320,1.66,NULL,NULL,NULL),
      (4,'2026-08-13',800,0,'Estornada','Galinha Cortada','F2',1400,1.75,NULL,NULL,NULL);
    """)
    conn.commit(); conn.close()
    disp.criar_tabelas_disponibilidade()
    perf.criar_tabelas_performance()
    oee.criar_tabelas_oee()
    return conectar


def preparar_op_1(banco, *, velocidade="1000", aves="900"):
    disp.salvar_programacao(
        1, "2026-08-10T08:00-04:00", "2026-08-10T17:00-04:00",
        [{"categoria": "ALMOCO", "inicio_previsto": "12:00", "fim_previsto": "13:00"}],
        perfil="pcp", usuario="PCP",
    )
    conn = banco()
    conn.execute("""UPDATE linha_abate_programacoes SET
      inicio_real='2026-08-10T08:00:00-04:00',fim_real='2026-08-10T17:00:00-04:00'
      WHERE op_id=1""")
    conn.executemany("""INSERT INTO apontamentos_paradas
      (id,op_id,data,setor,motivo,hora_inicio,hora_fim,horas_paradas,
       afeta_linha_abate,natureza_disponibilidade)
      VALUES (?,1,'2026-08-10','Linha','Falha',?,?,0,1,'NAO_PLANEJADA')""", [
        (1, "09:00", "09:20"), (2, "09:10", "09:30"),
    ])
    conn.execute("""INSERT INTO linha_performance_snapshots_op
      (op_id,velocidade_id,linha,configuracao,sku,velocidade_aves_hora,
       vigencia_inicio,resolvido_em,resolvido_por,versao,atual)
      VALUES (1,1,'LINHA_ABATE','PADRAO','Galinha Cortada',?,'2026-08-01',
              '2026-08-10T08:00:00-04:00','Prod',1,1)""", (velocidade,))
    conn.execute("""INSERT INTO linha_performance_contagens
      (op_id,aves_recebidas,mortes_antes_pendura,aves_processadas,origem_calculo,
       confirmado_por,confirmado_em,versao,atual)
      VALUES (1,'1000','100',?,'AVES_RECEBIDAS - MORTES_PRE_PENDURA','Prod',
              '2026-08-10T17:00:00-04:00',1,1)""", (aves,))
    conn.commit(); conn.close()


def filtros(**extras):
    base = {
        "data_inicio": "2026-08-01", "data_fim": "2026-08-31",
        "op_id": "1", "status": "Encerrada", "fornecedor": "Todos",
    }
    base.update(extras)
    return base


def test_cenario_a_disponibilidade_subtrai_pausa_e_uniao_de_paradas(banco):
    preparar_op_1(banco)
    r = disp.calcular_disponibilidade(1)
    assert r["tempo_planejado_liquido_minutos"] == Decimal("480")
    assert r["paradas_nao_planejadas_minutos"] == Decimal("30")
    assert r["tempo_operacional_minutos"] == Decimal("450")
    assert r["disponibilidade"] == Decimal("93.7500")


def test_cenario_b_performance_manual_usa_aves_consideradas(banco):
    preparar_op_1(banco)
    r = perf.calcular_performance(1)
    assert r["quantidade_total_considerada"] == Decimal("900")
    assert r["producao_teorica"] == Decimal("7500")
    assert r["performance"] == Decimal("12.00")


def test_performance_acima_de_cem_nao_e_truncada(banco):
    preparar_op_1(banco, velocidade="100")
    r = perf.calcular_performance(1)
    assert r["performance"] == Decimal("120.0")
    assert any("acima de 100%" in alerta for alerta in r["alertas"])


def test_qualidade_e_oee_permanecem_nao_calculaveis_sem_base(banco):
    preparar_op_1(banco)
    r = oee.calcular_oee(1)
    assert r["disponibilidade"]["situacao"] == "CALCULAVEL"
    assert r["performance"]["situacao"] == "CALCULAVEL"
    assert r["qualidade"]["situacao"] == "NAO_CALCULAVEL"
    assert r["qualidade"]["qualidade"] is None
    assert r["situacao"] == "NAO_CALCULAVEL" and r["oee"] is None


def test_formula_oee_decimal_quando_tres_componentes_futuros_foreem_validos(banco):
    preparar_op_1(banco)
    d = disp.calcular_disponibilidade(1)
    p = perf.calcular_performance(1, disponibilidade=d)
    q = {"situacao": "CALCULAVEL", "qualidade": Decimal("90"), "motivos": [], "alertas": [], "inconsistencias": []}
    r = oee.calcular_oee(1, disponibilidade=d, performance=p, qualidade=q)
    assert r["situacao"] == "CALCULAVEL"
    assert r["oee"] == Decimal("10.125000")


def test_op_aberta_sem_programacao_e_em_andamento(banco):
    assert disp.calcular_disponibilidade(3)["situacao"] == "EM_ANDAMENTO"
    assert perf.calcular_performance(3)["situacao"] == "EM_ANDAMENTO"
    r = oee.calcular_oee(3)
    assert r["situacao"] == "EM_ANDAMENTO" and r["oee"] is None


def test_parada_aberta_em_op_aberta_mantem_operacao_em_andamento(banco):
    conn = banco()
    conn.execute("UPDATE ordens_producao SET status='Aberta' WHERE id=1")
    conn.execute("""INSERT INTO apontamentos_paradas
      (op_id,data,setor,motivo,hora_inicio,hora_fim,horas_paradas,
       afeta_linha_abate,natureza_disponibilidade)
      VALUES (1,'2026-08-10','Corte','Falha aberta','10:00','',0,1,'NAO_PLANEJADA')""")
    conn.commit(); conn.close()
    r = disp.calcular_disponibilidade(1)
    assert r["situacao"] == "EM_ANDAMENTO" and r["disponibilidade"] is None


def test_op_estornada_e_historica_sem_indicador_vigente(banco):
    assert disp.calcular_disponibilidade(4)["situacao"] == "NAO_CALCULAVEL"
    assert perf.calcular_performance(4)["situacao"] == "NAO_CALCULAVEL"
    assert oee.calcular_oee(4)["situacao"] == "NAO_CALCULAVEL"


def test_historico_sem_programacao_nao_recebe_valores_atuais(banco):
    r = oee.calcular_oee(2)
    assert r["disponibilidade"]["disponibilidade"] is None
    assert r["performance"]["performance"] is None
    assert r["oee"] is None


def test_consolidacao_ponderada_nao_faz_media_simples(banco):
    preparar_op_1(banco)
    r = oee.consolidar_oee(filtros())
    assert r["totais"]["tempo_planejado_minutos"] == Decimal("480")
    assert r["totais"]["tempo_operacional_minutos"] == Decimal("450")
    assert r["totais"]["disponibilidade"] == Decimal("93.7500")
    assert r["totais"]["aves_consideradas"] == Decimal("900")
    assert r["totais"]["capacidade_teorica"] == Decimal("7500")
    assert r["totais"]["performance"] == Decimal("12.00")


def test_periodo_com_op_sem_base_fica_nao_calculavel_em_vez_de_parcial(banco):
    preparar_op_1(banco)
    r = oee.consolidar_oee(filtros(op_id=""))
    assert r["totais"]["situacao_disponibilidade"] == "NAO_CALCULAVEL"
    assert r["totais"]["disponibilidade"] is None
    assert r["totais"]["situacao_oee"] == "NAO_CALCULAVEL"


def test_filtros_status_aberta_e_historico(banco):
    abertas = oee.consolidar_oee(filtros(op_id="", status="Aberta"))
    historico = oee.consolidar_oee(filtros(op_id="", status="Historico"))
    assert [item["op_id"] for item in abertas["linhas"]] == [3]
    assert [item["op_id"] for item in historico["linhas"]] == [4]


def test_configuracao_de_ganchos_e_versionada_sem_default_oitenta_porcento(banco):
    conn = banco()
    colunas_parada = {item[1] for item in conn.execute("PRAGMA table_info(apontamentos_paradas)")}
    conn.close()
    assert {"registrado_por", "registrado_por_id"} <= colunas_parada
    assert oee.listar_configuracoes_fisicas() == []
    registro = oee.registrar_configuracao_fisica(
        "2026-08-01", None, 500, 401, 500, 398, "Contagem fisica",
        usuario="Admin", perfil="admin",
    )
    itens = oee.listar_configuracoes_fisicas()
    assert registro == itens[0]["id"]
    assert itens[0]["noria_1_ganchos_operacionais"] == 401
    assert itens[0]["noria_2_ganchos_operacionais"] == 398


def test_ganchos_nao_podem_ser_estimados_ou_exceder_instalados(banco):
    with pytest.raises(ValueError, match="exceder"):
        oee.registrar_configuracao_fisica(
            "2026-08-01", None, 500, 501, 500, 400, "Teste",
            usuario="Admin", perfil="admin",
        )
    assert oee.listar_configuracoes_fisicas() == []


def test_configuracao_fisica_exige_admin_e_nao_afeta_performance(banco):
    with pytest.raises(PermissionError):
        oee.registrar_configuracao_fisica(
            "2026-08-01", None, 500, 400, 500, 400, "Teste",
            usuario="PCP", perfil="pcp",
        )
    preparar_op_1(banco)
    antes = perf.calcular_performance(1)["performance"]
    oee.registrar_configuracao_fisica(
        "2026-08-01", None, 500, 300, 500, 350, "Contagem",
        usuario="Admin", perfil="admin",
    )
    assert perf.calcular_performance(1)["performance"] == antes


def test_excel_preserva_data_e_numeros(banco):
    preparar_op_1(banco)
    contexto = oee.montar_contexto_oee(filtros(), "oee")
    wb = load_workbook(BytesIO(oee.gerar_excel_oee(contexto).getvalue()))
    ws = wb.active
    assert isinstance(ws["B2"].value, date)
    assert isinstance(ws["E2"].value, (int, float))
    assert ws["L2"].value is None and ws["M2"].value is None


def test_tela_e_dashboard_referenciam_o_mesmo_servico_oficial():
    raiz = Path(__file__).parents[1]
    dashboard_service = (raiz / "modules/dashboard/services.py").read_text(encoding="utf-8")
    dashboard = (raiz / "templates/dashboard.html").read_text(encoding="utf-8")
    relatorio = (raiz / "templates/relatorio_oee.html").read_text(encoding="utf-8")
    assert "montar_contexto_oee" in dashboard_service
    assert "oee_oficial.totais" in dashboard
    assert "OEE parcial n&atilde;o permitido" in dashboard
    assert "Percentuais individuais n&atilde;o s&atilde;o promediados" in relatorio


def test_qualidade_nao_desconta_pnc_condenacao_ou_rendimento_por_conveniencia(banco):
    codigo = Path(oee.__file__).read_text(encoding="utf-8").lower()
    assert "nao converte rendimento, pnc ou condenacao" in codigo
    r = oee.calcular_qualidade(1)
    assert r["unidades_boas"] is None and r["unidades_processadas"] is None

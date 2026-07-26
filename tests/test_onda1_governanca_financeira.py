"""Regressões da Onda 1 de origem e governança financeira."""

from io import BytesIO
import os
from pathlib import Path
import sqlite3
import subprocess
import sys
import tempfile
import time

import pytest
from flask import Flask
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage, MultiDict


ROOT = Path(__file__).resolve().parents[1]
BANCO_TESTE = tempfile.NamedTemporaryFile(prefix="frigodatta-onda1-", suffix=".db", delete=False)
BANCO_TESTE.close()
os.environ["DB_NAME"] = BANCO_TESTE.name
os.environ.pop("DATABASE_URL", None)
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database import conectar, q  # noqa: E402
from modules.movimentacoes import services as service  # noqa: E402
from modules.movimentacoes import routes as routes_module  # noqa: E402
from modules.movimentacoes.origens import (  # noqa: E402
    IMPORTACAO_FINANCEIRA,
    LEGADO_IMPORTADO,
    MANUAL_CONTROLADO,
    OPERACIONAL,
    SANKHYA,
    buscar_configuracao_corte,
    buscar_origem_principal_movimentacao,
    criar_tabelas_governanca_financeira,
    montar_contexto_governanca_financeira,
    registrar_origem_principal_cursor,
)
from modules.movimentacoes.routes import register_movimentacoes_routes  # noqa: E402
from modules.movimentacoes.services import (  # noqa: E402
    alterar_origem_principal_movimentacao,
    atualizar_movimentacao_financeira,
    buscar_movimentacoes_financeiras,
    criar_tabela_movimentacoes_financeiras,
    excluir_movimentacao_financeira,
    gerar_planilha_modelo_importacao_financeira,
    importar_movimentacoes_financeiras_excel,
    salvar_movimentacao_financeira,
)


def executar(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), parametros)
    ultimo_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return ultimo_id


def consultar(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), parametros)
    itens = [dict(item) for item in cursor.fetchall()]
    conn.close()
    return itens


def contar(tabela):
    return consultar(f"SELECT COUNT(*) AS quantidade FROM {tabela}")[0]["quantidade"]


def inserir_legado(descricao="Legado", tipo="Entrada", valor=100):
    return executar(
        """
        INSERT INTO movimentacoes_financeiras (
            data_documento, data_vencimento, tipo, categoria, descricao,
            valor, valor_documento, valor_pago, status, impacta_fluxo_caixa
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        ("2026-07-01", "2026-07-15", tipo, "Outros", descricao, valor, valor, 0, "Pendente", 1),
    )


def formulario_manual(**alteracoes):
    dados = [
        ("tipo", "Entrada"),
        ("categoria", "Venda de Producao Propria"),
        ("descricao", "Manual controlado"),
        ("data_documento", "2026-07-01"),
        ("valor", "250"),
        ("status", "Pendente"),
        ("parcela_vencimento[]", "2026-07-15"),
        ("parcela_valor[]", "250"),
        ("justificativa", "Fato sem integração disponível."),
        ("referencia_evidencia", "NF TESTE 101"),
    ]
    form = MultiDict(dados)
    for chave, valor in alteracoes.items():
        form.setlist(chave, [valor])
    return form


def arquivo_importacao(*, valor=1000, observacoes="", data_documento="2026-01-01", filename="financeiro.xlsx"):
    modelo = gerar_planilha_modelo_importacao_financeira()
    wb = load_workbook(modelo)
    ws = wb.active
    colunas = {celula.value: celula.column for celula in ws[1]}
    ws.cell(2, colunas["Valor do Documento"]).value = valor
    ws.cell(2, colunas["Data do Documento"]).value = data_documento
    ws.cell(2, colunas["Observacoes"]).value = observacoes
    conteudo = BytesIO()
    wb.save(conteudo)
    wb.close()
    conteudo.seek(0)
    return FileStorage(stream=conteudo, filename=filename)


def importar(arquivo):
    return importar_movimentacoes_financeiras_excel(
        arquivo,
        usuario_id=42,
        usuario_nome="Controladoria",
        perfil="pcp",
    )


@pytest.fixture(autouse=True)
def base_limpa():
    criar_tabela_movimentacoes_financeiras()
    criar_tabelas_governanca_financeira()
    for tabela in (
        "movimentacoes_financeiras_importacao_linhas",
        "movimentacoes_financeiras_origens",
        "movimentacoes_financeiras_importacao_lotes",
        "movimentacoes_financeiras_configuracao_corte",
        "movimentacoes_financeiras_auditoria",
        "movimentacoes_financeiras",
    ):
        executar(f"DELETE FROM {tabela}")
    yield


def test_01_historico_sem_origem_resolve_legado():
    movimento_id = inserir_legado()
    origem = buscar_origem_principal_movimentacao(movimento_id)
    assert origem["modo_origem"] == LEGADO_IMPORTADO
    assert origem["origem_persistida"] is False


def test_02_leitura_legado_nao_altera_banco():
    inserir_legado()
    antes = contar("movimentacoes_financeiras_origens")
    buscar_movimentacoes_financeiras("2026-01-01", "2026-12-31", "Todos", "Todos")
    montar_contexto_governanca_financeira()
    assert contar("movimentacoes_financeiras_origens") == antes == 0


def test_03_manual_gera_origem_controlada():
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    origem = consultar("SELECT * FROM movimentacoes_financeiras_origens")[0]
    assert origem["modo"] == MANUAL_CONTROLADO
    assert origem["papel"] == "PRINCIPAL"
    assert origem["auditoria_id"]


def test_04_manual_exige_justificativa():
    with pytest.raises(ValueError, match="justificativa"):
        salvar_movimentacao_financeira(formulario_manual(justificativa="  "), 42, "Controladoria", "pcp")


def test_05_manual_exige_evidencia_util():
    with pytest.raises(ValueError, match="referência"):
        salvar_movimentacao_financeira(formulario_manual(referencia_evidencia=" "), 42, "Controladoria", "pcp")


def test_06_usuario_nao_escolhe_sankhya_ou_operacional():
    form = formulario_manual(modo_origem=SANKHYA)
    salvar_movimentacao_financeira(form, 42, "Controladoria", "pcp")
    assert consultar("SELECT modo FROM movimentacoes_financeiras_origens")[0]["modo"] == MANUAL_CONTROLADO
    html = (ROOT / "templates" / "financeiro.html").read_text(encoding="utf-8")
    assert 'name="modo_origem"' not in html


def test_07_movimentacao_e_origem_na_mesma_transacao():
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    movimento = consultar("SELECT id FROM movimentacoes_financeiras")[0]
    origem = consultar("SELECT movimentacao_id FROM movimentacoes_financeiras_origens")[0]
    assert origem["movimentacao_id"] == movimento["id"]


def test_08_falha_origem_reverte_movimentacao(monkeypatch):
    monkeypatch.setattr(service, "registrar_origem_principal_cursor", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("origem")))
    with pytest.raises(RuntimeError, match="origem"):
        salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    assert contar("movimentacoes_financeiras") == 0


def test_09_importacao_nova_gera_origem_financeira():
    resultado = importar(arquivo_importacao())
    origem = consultar("SELECT * FROM movimentacoes_financeiras_origens")[0]
    assert resultado["importadas"] == 1
    assert origem["modo"] == IMPORTACAO_FINANCEIRA
    assert origem["lote_importacao_id"] == resultado["lote_id"]


def test_10_lote_registra_hash_usuario_totais_e_nome_sanitizado():
    resultado = importar(arquivo_importacao(filename="../pasta/financeiro?.xlsx"))
    lote = consultar("SELECT * FROM movimentacoes_financeiras_importacao_lotes")[0]
    assert len(lote["arquivo_hash"]) == 64
    assert lote["usuario_id"] == 42 and lote["usuario_nome"] == "Controladoria"
    assert lote["quantidade_total"] == lote["importadas"] == 1
    assert "/" not in lote["arquivo_nome"] and "?" not in lote["arquivo_nome"]
    assert resultado["arquivo_hash"] == lote["arquivo_hash"]


def test_11_cada_linha_possui_resultado_persistido():
    resultado = importar(arquivo_importacao())
    linha = consultar("SELECT * FROM movimentacoes_financeiras_importacao_linhas WHERE lote_id = ?", (resultado["lote_id"],))[0]
    assert linha["numero_linha"] == 2
    assert linha["status"] == "IMPORTADA"
    assert len(linha["hash_normalizado"]) == 64
    assert linha["movimentacao_id"]


def test_12_reimportacao_identica_e_noop():
    primeiro = importar(arquivo_importacao())
    segundo = importar(arquivo_importacao())
    assert primeiro["importadas"] == 1
    assert segundo["importadas"] == 0 and segundo["identicas"] == 1
    assert contar("movimentacoes_financeiras") == 1


def test_13_reimportacao_identica_nao_duplica_origem():
    importar(arquivo_importacao())
    importar(arquivo_importacao())
    assert contar("movimentacoes_financeiras_origens") == 1
    assert contar("movimentacoes_financeiras_importacao_lotes") == 2


def test_14_divergencia_nao_sobrescreve():
    importar(arquivo_importacao(observacoes="Original"))
    importar(arquivo_importacao(observacoes="Divergente"))
    movimento = consultar("SELECT observacoes FROM movimentacoes_financeiras")[0]
    assert movimento["observacoes"] == "Original"


def test_15_divergencia_cria_conflito_persistido():
    importar(arquivo_importacao(observacoes="Original"))
    resultado = importar(arquivo_importacao(observacoes="Divergente"))
    linha = consultar(
        "SELECT * FROM movimentacoes_financeiras_importacao_linhas WHERE lote_id = ?",
        (resultado["lote_id"],),
    )[0]
    assert linha["status"] == "CONFLITANTE" and linha["auditoria_id"]
    assert consultar("SELECT acao FROM movimentacoes_financeiras_auditoria ORDER BY id DESC")[0]["acao"] == "CONFLITO_IMPORTACAO"


def test_16_linha_rejeitada_nao_cria_movimentacao():
    resultado = importar(arquivo_importacao(valor=0))
    assert resultado["rejeitadas"] == 1
    assert contar("movimentacoes_financeiras") == 0
    assert consultar("SELECT status FROM movimentacoes_financeiras_importacao_linhas")[0]["status"] == "REJEITADA"


def test_falha_integral_importador_reverte_movimento_e_registra_lote_falho(monkeypatch):
    monkeypatch.setattr(service, "registrar_origem_principal_cursor", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("origem")))
    with pytest.raises(RuntimeError, match="origem"):
        importar(arquivo_importacao())
    assert contar("movimentacoes_financeiras") == 0
    lote = consultar("SELECT status, mensagem_final FROM movimentacoes_financeiras_importacao_lotes")[0]
    assert lote["status"] == "FALHOU"
    assert "RuntimeError" in lote["mensagem_final"]


def test_17_mudanca_origem_exige_justificativa_e_auditoria():
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    movimento_id = consultar("SELECT id FROM movimentacoes_financeiras")[0]["id"]
    with pytest.raises(ValueError, match="justificativa"):
        alterar_origem_principal_movimentacao(movimento_id, IMPORTACAO_FINANCEIRA, "", 42, "Controladoria")
    alterar_origem_principal_movimentacao(
        movimento_id, IMPORTACAO_FINANCEIRA, "Correção controlada da classificação.", 42, "Controladoria"
    )
    evento = consultar("SELECT * FROM movimentacoes_financeiras_auditoria ORDER BY id DESC")[0]
    assert evento["acao"] == "ALTERACAO_ORIGEM"
    assert "MANUAL_CONTROLADO" in evento["estado_anterior"]
    assert "IMPORTACAO_FINANCEIRA" in evento["estado_posterior"]


def test_18_filtros_por_origem_funcionam():
    inserir_legado()
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    importar(arquivo_importacao())
    assert len(buscar_movimentacoes_financeiras("2026-01-01", "2026-12-31", "Todos", "Todos", LEGADO_IMPORTADO)) == 1
    assert len(buscar_movimentacoes_financeiras("2026-01-01", "2026-12-31", "Todos", "Todos", MANUAL_CONTROLADO)) == 1
    assert len(buscar_movimentacoes_financeiras("2026-01-01", "2026-12-31", "Todos", "Todos", IMPORTACAO_FINANCEIRA)) == 1


def test_19_governanca_respeita_permissoes(monkeypatch):
    monkeypatch.setattr(routes_module, "render_template", lambda *args, **kwargs: "governanca")
    app = Flask(__name__, template_folder=str(ROOT / "templates"))
    app.secret_key = "teste"
    app.add_url_rule("/login", "login", lambda: "login")
    app.add_url_rule("/inicio", "inicio", lambda: "inicio")
    register_movimentacoes_routes(app)
    cliente = app.test_client()
    assert cliente.get("/movimentacoes/governanca").status_code == 302
    with cliente.session_transaction() as sessao:
        sessao["usuario_id"] = 42
        sessao["nome"] = "PCP"
        sessao["perfil"] = "pcp"
    assert cliente.get("/movimentacoes/governanca").status_code == 200
    with cliente.session_transaction() as sessao:
        sessao["perfil"] = "admin"
    assert cliente.get("/movimentacoes/governanca").status_code == 200


def test_20_data_corte_permanece_desativada():
    configuracao = buscar_configuracao_corte()
    assert configuracao["ativo"] == 0
    assert configuracao["data_corte"] is None
    assert configuracao["situacao"].startswith("Não")
    assert contar("movimentacoes_financeiras_configuracao_corte") == 0


def test_21_nenhum_registro_sankhya_criado():
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    importar(arquivo_importacao())
    assert consultar("SELECT COUNT(*) AS quantidade FROM movimentacoes_financeiras_origens WHERE modo = ?", (SANKHYA,))[0]["quantidade"] == 0


def test_22_nenhum_registro_operacional_criado():
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    importar(arquivo_importacao())
    assert consultar("SELECT COUNT(*) AS quantidade FROM movimentacoes_financeiras_origens WHERE modo = ?", (OPERACIONAL,))[0]["quantidade"] == 0


def snapshot_financeiro():
    linha = consultar(
        """
        SELECT
          COUNT(*) AS quantidade,
          SUM(CASE WHEN tipo = 'Entrada' AND status != 'Cancelado' THEN valor ELSE 0 END) AS receitas,
          SUM(CASE WHEN tipo != 'Entrada' AND status != 'Cancelado' THEN valor ELSE 0 END) AS despesas,
          SUM(CASE WHEN tipo = 'Entrada' AND status != 'Cancelado' THEN valor ELSE -valor END) AS saldo,
          SUM(CASE WHEN status = 'Pendente' THEN valor ELSE 0 END) AS fluxo_previsto,
          SUM(CASE WHEN status = 'Realizado' THEN valor_pago ELSE 0 END) AS fluxo_realizado,
          SUM(CASE WHEN tipo != 'Entrada' AND status = 'Pendente' THEN valor - valor_pago ELSE 0 END) AS contas_pagar,
          SUM(CASE WHEN tipo = 'Entrada' AND status = 'Pendente' THEN valor - valor_pago ELSE 0 END) AS contas_receber,
          SUM(CASE WHEN status = 'Pendente' AND valor_pago = 0 THEN 1 ELSE 0 END) AS abertos,
          SUM(CASE WHEN status = 'Pendente' AND valor_pago > 0 AND valor_pago < valor THEN 1 ELSE 0 END) AS parciais,
          SUM(CASE WHEN status = 'Cancelado' THEN 1 ELSE 0 END) AS cancelados
        FROM movimentacoes_financeiras
        """
    )[0]
    return linha


def preparar_base_comparacao():
    ids = []
    casos = [
        ("Receita aberta", "Entrada", 1000, 0, "Pendente", "2026-07-15"),
        ("Despesa aberta", "Saída", 400, 0, "Pendente", "2026-07-15"),
        ("Receita realizada", "Entrada", 700, 700, "Realizado", "2026-07-10"),
        ("Despesa parcial", "Saída", 300, 100, "Pendente", "2026-07-01"),
        ("Cancelada", "Saída", 200, 0, "Cancelado", "2026-07-01"),
        ("Aporte", "Entrada", 500, 500, "Realizado", "2026-07-05"),
        ("Transferência neutra", "Entrada", 250, 250, "Realizado", "2026-07-05"),
    ]
    for descricao, tipo, valor, pago, status, vencimento in casos:
        ids.append(executar(
            """
            INSERT INTO movimentacoes_financeiras (
                data_documento, data_vencimento, tipo, categoria, descricao,
                valor, valor_documento, valor_pago, status, impacta_fluxo_caixa
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("2026-07-01", vencimento, tipo, "Outros", descricao, valor, valor, pago, status, 1),
        ))
    return ids


def adicionar_apenas_governanca(ids):
    conn = conectar()
    cursor = conn.cursor()
    for indice, movimento_id in enumerate(ids[:2], 1):
        registrar_origem_principal_cursor(
            cursor, movimento_id, MANUAL_CONTROLADO, "FRIGODATTA", "TESTE_SINTETICO",
            42, "Controladoria", chave_idempotente=f"teste-{indice}",
            metadados={"justificativa": "Comparação sintética", "referencia_evidencia": f"TESTE-{indice}"},
        )
    conn.commit()
    conn.close()


def test_23_dre_permanece_identica():
    ids = preparar_base_comparacao()
    antes = snapshot_financeiro()
    adicionar_apenas_governanca(ids)
    depois = snapshot_financeiro()
    assert (antes["receitas"], antes["despesas"], antes["saldo"]) == (depois["receitas"], depois["despesas"], depois["saldo"])


def test_24_fluxo_previsto_permanece_identico():
    ids = preparar_base_comparacao()
    antes = snapshot_financeiro()["fluxo_previsto"]
    adicionar_apenas_governanca(ids)
    assert snapshot_financeiro()["fluxo_previsto"] == antes


def test_25_fluxo_realizado_permanece_identico():
    ids = preparar_base_comparacao()
    antes = snapshot_financeiro()["fluxo_realizado"]
    adicionar_apenas_governanca(ids)
    assert snapshot_financeiro()["fluxo_realizado"] == antes


def test_26_contas_pagar_receber_permanecem_identicas():
    ids = preparar_base_comparacao()
    antes = snapshot_financeiro()
    adicionar_apenas_governanca(ids)
    depois = snapshot_financeiro()
    assert (antes["contas_pagar"], antes["contas_receber"]) == (depois["contas_pagar"], depois["contas_receber"])


def test_27_liquidacao_e_status_permanecem_identicos():
    ids = preparar_base_comparacao()
    antes = snapshot_financeiro()
    adicionar_apenas_governanca(ids)
    depois = snapshot_financeiro()
    assert (antes["abertos"], antes["parciais"], antes["cancelados"]) == (depois["abertos"], depois["parciais"], depois["cancelados"])


def test_28_cancelamento_preserva_origem():
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    movimento_id = consultar("SELECT id FROM movimentacoes_financeiras")[0]["id"]
    origem_antes = consultar("SELECT * FROM movimentacoes_financeiras_origens")[0]
    excluir_movimentacao_financeira(movimento_id, "Cancelamento controlado.", 42, "Controladoria", "pcp")
    origem_depois = consultar("SELECT * FROM movimentacoes_financeiras_origens")[0]
    assert origem_depois["id"] == origem_antes["id"]
    assert origem_depois["modo"] == MANUAL_CONTROLADO


def test_29_edicao_preserva_e_audita_origem():
    salvar_movimentacao_financeira(formulario_manual(), 42, "Controladoria", "pcp")
    movimento_id = consultar("SELECT id FROM movimentacoes_financeiras")[0]["id"]
    form = formulario_manual(descricao="Manual editado")
    form.setlist("justificativa", ["Correção descritiva auditada."])
    atualizar_movimentacao_financeira(movimento_id, form, 42, "Controladoria", "pcp")
    assert consultar("SELECT modo FROM movimentacoes_financeiras_origens")[0]["modo"] == MANUAL_CONTROLADO
    assert consultar("SELECT acao FROM movimentacoes_financeiras_auditoria ORDER BY id DESC")[0]["acao"] == "EDICAO"


def test_30_importar_app_nao_executa_backfill_hotfix_ou_reclassificacao():
    banco = tempfile.NamedTemporaryFile(prefix="onda1-bootstrap-", suffix=".db", delete=False)
    banco.close()
    ambiente = os.environ.copy()
    ambiente["DB_NAME"] = banco.name
    ambiente.pop("DATABASE_URL", None)
    codigo = (
        "import sqlite3, app; "
        f"c=sqlite3.connect(r'{banco.name}'); "
        "print(c.execute(\"SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='movimentacoes_financeiras_origens'\").fetchone()[0]); "
        "print(c.execute(\"SELECT COUNT(*) FROM movimentacoes_financeiras\").fetchone()[0]); c.close()"
    )
    processo = subprocess.run([sys.executable, "-c", codigo], cwd=ROOT, env=ambiente, text=True, capture_output=True, check=True)
    assert processo.stdout.strip().splitlines()[-2:] == ["0", "0"]


def test_31_migrations_sqlite_e_postgresql_idempotentes_e_com_rollback():
    banco = sqlite3.connect(":memory:")
    up_sqlite = (ROOT / "database" / "20260726_onda1_governanca_origens_financeiras_sqlite.sql").read_text(encoding="utf-8")
    rollback_sqlite = (ROOT / "database" / "20260726_onda1_governanca_origens_financeiras_sqlite_rollback.sql").read_text(encoding="utf-8")
    banco.executescript(up_sqlite)
    banco.executescript(up_sqlite)
    banco.executescript(rollback_sqlite)
    tabelas = banco.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'movimentacoes_financeiras_%'").fetchall()
    banco.close()
    assert tabelas == []
    up_pg = (ROOT / "database" / "20260726_onda1_governanca_origens_financeiras.sql").read_text(encoding="utf-8")
    rollback_pg = (ROOT / "database" / "20260726_onda1_governanca_origens_financeiras_rollback.sql").read_text(encoding="utf-8")
    assert "BEGIN;" in up_pg and "IF NOT EXISTS" in up_pg and "COMMIT;" in up_pg
    assert "DROP TABLE IF EXISTS" in rollback_pg and "BEGIN;" in rollback_pg and "COMMIT;" in rollback_pg


def test_32_listagem_sem_n1_e_sem_regressao_relevante():
    conn = conectar()
    cursor = conn.cursor()
    cursor.executemany(
        """
        INSERT INTO movimentacoes_financeiras (
            data_documento, data_vencimento, tipo, categoria, descricao,
            valor, valor_documento, valor_pago, status, impacta_fluxo_caixa
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            ("2026-07-01", "2026-07-15", "Entrada", "Outros", f"Legado {i}", 10, 10, 0, "Pendente", 1)
            for i in range(1000)
        ],
    )
    conn.commit()
    conn.close()
    inicio = time.perf_counter()
    itens = buscar_movimentacoes_financeiras("2026-01-01", "2026-12-31", "Todos", "Todos")
    duracao = time.perf_counter() - inicio
    fonte = (ROOT / "modules" / "movimentacoes" / "services.py").read_text(encoding="utf-8")
    trecho = fonte[fonte.index("def buscar_movimentacoes_financeiras"):fonte.index("def calcular_resumo_financeiro")]
    assert len(itens) == 1000 and duracao < 1.5
    assert "LEFT JOIN movimentacoes_financeiras_origens" in trecho
    assert trecho.count("cursor.execute") == 1

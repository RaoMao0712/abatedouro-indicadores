"""Regressões da exportação individualizada da Auditoria Financeira."""

from io import BytesIO
import os
from pathlib import Path
import sys
import tempfile
import time
import tracemalloc

from flask import Flask
from openpyxl import load_workbook
import pytest
from werkzeug.datastructures import MultiDict


ROOT = Path(__file__).resolve().parents[1]
BANCO_TESTE = tempfile.NamedTemporaryFile(
    prefix="frigodatta-exportacao-auditoria-",
    suffix=".db",
    delete=False,
)
BANCO_TESTE.close()
os.environ["DB_NAME"] = BANCO_TESTE.name
os.environ.pop("DATABASE_URL", None)
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database import conectar, q  # noqa: E402
from modules.movimentacoes import services as service  # noqa: E402
from modules.movimentacoes.origens import criar_tabelas_governanca_financeira  # noqa: E402
from modules.movimentacoes.routes import register_movimentacoes_routes  # noqa: E402
from modules.movimentacoes.services import (  # noqa: E402
    buscar_movimentacoes_exportacao_auditoria,
    criar_tabela_movimentacoes_financeiras,
    excluir_movimentacao_financeira,
    gerar_excel_auditoria_financeira,
    montar_contexto_auditoria_financeira,
    normalizar_filtros_auditoria,
)


def executar(sql, parametros=()):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(q(sql), parametros)
    ultimo_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return ultimo_id


def inserir_movimentacao(
    indice,
    *,
    data_documento="2026-01-10",
    data_vencimento="2026-02-10",
    data_realizacao="",
    tipo="Saida",
    categoria="Categoria teste",
    descricao="Movimento sintético",
    favorecido="Fornecedor sintético",
    historico="Histórico sintético",
    observacoes="Observação sintética",
    numero_documento=None,
    cnpj_cpf="00123456000199",
    valor=100,
    valor_pago=0,
    valor_liquido=None,
    status="Pendente",
    import_key=None,
    origem_importacao="excel_movimentacoes",
):
    valor_liquido = valor if valor_liquido is None else valor_liquido
    return executar(
        """
        INSERT INTO movimentacoes_financeiras (
            data_documento, data_vencimento, data_realizacao, tipo, categoria,
            descricao, valor, valor_documento, valor_pago, valor_liquido,
            forma_pagamento, status, parcelas, parcela_atual, documento_id,
            import_key, cnpj_cpf, numero_documento, favorecido, parceiro,
            historico, observacoes, origem_importacao, grupo_gerencial,
            categoria_plano, subcategoria, centro_analise, linha_dre,
            tipo_conta, impacta_fluxo_caixa
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            data_documento,
            data_vencimento,
            data_realizacao,
            tipo,
            categoria,
            descricao,
            valor,
            valor,
            valor_pago,
            valor_liquido,
            "Boleto",
            status,
            1,
            1,
            str(numero_documento or f"DOC-{indice}"),
            import_key or f"teste-exportacao-{indice}",
            cnpj_cpf,
            str(numero_documento or f"NF-{indice:05d}"),
            favorecido,
            favorecido,
            historico,
            observacoes,
            origem_importacao,
            "Despesas Operacionais",
            categoria,
            "Subcategoria teste",
            "Centro teste",
            "Despesas Operacionais",
            "Saida",
            1,
        ),
    )


def carregar_exportacao(args=None):
    contexto = montar_contexto_auditoria_financeira(
        MultiDict(args or {}),
        exportar=True,
    )
    arquivo = gerar_excel_auditoria_financeira(contexto)
    return contexto, load_workbook(arquivo, data_only=False)


@pytest.fixture(autouse=True)
def base_limpa():
    criar_tabela_movimentacoes_financeiras()
    criar_tabelas_governanca_financeira()
    for tabela in (
        "movimentacoes_financeiras_importacao_linhas",
        "movimentacoes_financeiras_origens",
        "movimentacoes_financeiras_importacao_lotes",
        "movimentacoes_financeiras_configuracao_corte",
        "movimentacoes_financeiras_auditoria",
        "movimentacoes_financeiras",
    ):
        executar(f"DELETE FROM {tabela}")
    yield


def test_exportacao_sem_filtros_preserva_abas_e_tem_uma_linha_por_id():
    ids = [
        inserir_movimentacao(1, tipo="Entrada", valor=125.5),
        inserir_movimentacao(2, valor=80.25),
        inserir_movimentacao(3, valor=19.75),
    ]

    contexto, wb = carregar_exportacao()

    assert wb.sheetnames == [
        "Visao geral",
        "Por categoria",
        "Por mes",
        "Pendencias",
        "Movimentações",
    ]
    ws = wb["Movimentações"]
    cabecalhos = {celula.value: celula.column for celula in ws[1]}
    ids_exportados = [
        ws.cell(linha, cabecalhos["ID movimentacao"]).value
        for linha in range(2, ws.max_row + 1)
    ]
    assert ids_exportados == ids
    assert len(ids_exportados) == len(set(ids_exportados)) == 3
    assert contexto["indicadores"]["total_movimentacoes"] == 3
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == f"A1:AU{ws.max_row}"
    wb.close()


@pytest.mark.parametrize(
    ("filtro", "valor", "documento_esperado"),
    [
        ("data_inicio", "2026-03-01", "NF-00003"),
        ("data_fim", "2026-01-31", "NF-00001"),
        ("categoria", "Categoria B", "NF-00002"),
        ("favorecido", "Fornecedor B", "NF-00002"),
        ("descricao", "Descrição B", "NF-00002"),
        ("historico", "Histórico B", "NF-00002"),
        ("mes", "2026-02", "NF-00002"),
        ("somente_nao_classificados", "1", "NF-00003"),
    ],
)
def test_exportacao_reutiliza_todos_os_filtros_existentes(
    filtro,
    valor,
    documento_esperado,
):
    inserir_movimentacao(
        1,
        data_documento="2026-01-10",
        categoria="Categoria A",
        favorecido="Fornecedor A",
        descricao="Descrição A",
        historico="Histórico A",
    )
    inserir_movimentacao(
        2,
        data_documento="2026-02-10",
        categoria="Categoria B",
        favorecido="Fornecedor B",
        descricao="Descrição B",
        historico="Histórico B",
    )
    inserir_movimentacao(
        3,
        data_documento="2026-03-10",
        categoria="Não Classificado",
        favorecido="Fornecedor C",
        descricao="Descrição C",
        historico="Histórico C",
    )

    contexto, wb = carregar_exportacao([(filtro, valor)])
    ws = wb["Movimentações"]
    cabecalhos = {celula.value: celula.column for celula in ws[1]}
    documentos = [
        ws.cell(linha, cabecalhos["Numero documento"]).value
        for linha in range(2, ws.max_row + 1)
    ]
    assert documentos == [documento_esperado]
    assert contexto["indicadores"]["total_movimentacoes"] == 1
    wb.close()


def test_populacao_vazia_mantem_cabecalho_e_consolidados():
    contexto, wb = carregar_exportacao()
    ws = wb["Movimentações"]
    assert ws.max_row == 1
    assert ws.max_column == 47
    assert contexto["indicadores"]["total_movimentacoes"] == 0
    assert wb["Visao geral"]["B2"].value == 0
    wb.close()


def test_origem_legada_persistida_lote_e_cancelamento_nao_duplicam_linhas():
    legado_id = inserir_movimentacao(1)
    persistido_id = inserir_movimentacao(2, status="Realizado", valor_pago=100)
    cancelado_id = inserir_movimentacao(3)

    lote_id = executar(
        """
        INSERT INTO movimentacoes_financeiras_importacao_lotes (
            arquivo_nome, arquivo_hash, tipo_importador, modo_origem,
            usuario_id, usuario_nome, status, quantidade_total,
            importadas, metadados_tecnicos
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "sintetico.xlsx",
            "a" * 64,
            "MOVIMENTACOES",
            "IMPORTACAO_FINANCEIRA",
            42,
            "Controladoria teste",
            "CONCLUIDO",
            1,
            1,
            "{}",
        ),
    )
    executar(
        """
        INSERT INTO movimentacoes_financeiras_origens (
            movimentacao_id, papel, modo, sistema_origem, modulo_origem,
            tipo_evento, chave_externa, chave_idempotente,
            lote_importacao_id, linha_arquivo, usuario_id, usuario_nome,
            metadados, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            persistido_id,
            "PRINCIPAL",
            "IMPORTACAO_FINANCEIRA",
            "Planilha sintética",
            "Financeiro",
            "IMPORTACAO",
            "EXT-2",
            "IDEM-2",
            lote_id,
            2,
            42,
            "Controladoria teste",
            "{}",
            "ATIVA",
        ),
    )
    executar(
        """
        INSERT INTO movimentacoes_financeiras_origens (
            movimentacao_id, papel, modo, sistema_origem, tipo_evento,
            usuario_nome, metadados, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            persistido_id,
            "SECUNDARIA",
            "MANUAL_CONTROLADO",
            "Teste",
            "REFERENCIA",
            "Teste",
            "{}",
            "INATIVA",
        ),
    )
    excluir_movimentacao_financeira(
        cancelado_id,
        justificativa="Cancelamento sintético auditado.",
        usuario_id=7,
        usuario_nome="Auditor teste",
        perfil="pcp",
    )

    _, wb = carregar_exportacao()
    ws = wb["Movimentações"]
    cabecalhos = {celula.value: celula.column for celula in ws[1]}
    linhas = {
        ws.cell(linha, cabecalhos["ID movimentacao"]).value: linha
        for linha in range(2, ws.max_row + 1)
    }
    assert set(linhas) == {legado_id, persistido_id, cancelado_id}
    assert ws.cell(linhas[legado_id], cabecalhos["Modo de origem resolvido"]).value == "LEGADO_IMPORTADO"
    assert ws.cell(linhas[legado_id], cabecalhos["Origem persistida"]).value == "Nao"
    assert ws.cell(linhas[persistido_id], cabecalhos["Origem persistida"]).value == "Sim"
    assert ws.cell(linhas[persistido_id], cabecalhos["ID lote"]).value == lote_id
    assert ws.cell(linhas[cancelado_id], cabecalhos["Cancelada"]).value == "Sim"
    assert ws.cell(linhas[cancelado_id], cabecalhos["Justificativa cancelamento"]).value == "Cancelamento sintético auditado."
    assert ws.cell(linhas[cancelado_id], cabecalhos["Saldo aberto"]).value is None
    wb.close()


def test_tipos_excel_valores_totais_e_protecao_contra_formula_injection():
    inserir_movimentacao(
        1,
        data_documento="2026-01-02",
        data_vencimento="2026-01-20",
        data_realizacao="2026-01-19",
        descricao="=2+2",
        favorecido="@usuario",
        historico="+COMANDO",
        observacoes="-ARGUMENTO",
        numero_documento="-0001",
        cnpj_cpf="00123456000199",
        valor=123.45,
        valor_pago=123.45,
        valor_liquido=123.45,
        status="Realizado",
    )
    contexto, wb = carregar_exportacao()
    ws = wb["Movimentações"]
    cabecalhos = {celula.value: celula.column for celula in ws[1]}

    for coluna in (
        "Favorecido/Cliente",
        "Descricao",
        "Historico",
        "Observacoes",
        "Numero documento",
    ):
        celula = ws.cell(2, cabecalhos[coluna])
        assert celula.data_type != "f"
        assert celula.value.startswith("'")

    assert ws.cell(2, cabecalhos["CNPJ/CPF"]).value == "00123456000199"
    assert ws.cell(2, cabecalhos["Data documento"]).is_date
    assert ws.cell(2, cabecalhos["Data vencimento"]).is_date
    assert ws.cell(2, cabecalhos["Data pagamento/realizacao"]).is_date
    assert isinstance(ws.cell(2, cabecalhos["Valor documental"]).value, (int, float))
    assert isinstance(ws.cell(2, cabecalhos["Valor pago"]).value, (int, float))
    assert isinstance(ws.cell(2, cabecalhos["Valor liquido"]).value, (int, float))
    assert contexto["indicadores"]["total_despesas"] == 123.45
    assert ws.cell(2, cabecalhos["Valor regra financeira"]).value == 123.45
    wb.close()


def test_totais_individualizados_reconciliam_com_agregados():
    inserir_movimentacao(1, tipo="Entrada", valor=150)
    inserir_movimentacao(2, tipo="Saida", valor=80)
    inserir_movimentacao(3, tipo="Saida", valor=20)
    contexto, wb = carregar_exportacao()
    ws = wb["Movimentações"]
    cabecalhos = {celula.value: celula.column for celula in ws[1]}
    receitas = 0
    despesas = 0
    for linha in range(2, ws.max_row + 1):
        valor = ws.cell(linha, cabecalhos["Valor regra financeira"]).value
        if ws.cell(linha, cabecalhos["Natureza"]).value == "Entrada":
            receitas += valor
        else:
            despesas += valor

    assert receitas == contexto["indicadores"]["total_receitas"] == 150
    assert despesas == contexto["indicadores"]["total_despesas"] == 100
    assert receitas - despesas == contexto["indicadores"]["saldo_liquido"] == 50
    wb.close()


def test_consulta_individualizada_e_unica_e_constante(monkeypatch):
    for indice in range(10):
        inserir_movimentacao(indice)

    sql_executados = []
    conectar_real = service.conectar

    def conectar_rastreado():
        conn = conectar_real()
        conn.set_trace_callback(
            lambda sql: sql_executados.append(sql)
            if sql.lstrip().upper().startswith(("SELECT", "WITH"))
            else None
        )
        return conn

    monkeypatch.setattr(service, "conectar", conectar_rastreado)
    itens = buscar_movimentacoes_exportacao_auditoria(
        normalizar_filtros_auditoria(MultiDict())
    )
    assert len(itens) == 10
    assert len(sql_executados) == 1


def test_permissoes_e_nome_do_download_sao_preservados():
    inserir_movimentacao(1)
    app = Flask(__name__, template_folder=str(ROOT / "templates"))
    app.secret_key = "teste"
    app.add_url_rule("/login", "login", lambda: "login")
    app.add_url_rule("/inicio", "inicio", lambda: "inicio")
    register_movimentacoes_routes(app)
    cliente = app.test_client()

    assert cliente.get("/movimentacoes/auditoria/exportar").status_code == 302
    with cliente.session_transaction() as sessao:
        sessao.update({"usuario_id": 42, "nome": "PCP", "perfil": "producao"})
    assert cliente.get("/movimentacoes/auditoria/exportar").status_code == 302

    with cliente.session_transaction() as sessao:
        sessao["perfil"] = "pcp"
    resposta = cliente.get("/movimentacoes/auditoria/exportar")
    assert resposta.status_code == 200
    assert resposta.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "auditoria_financeira_" in resposta.headers["Content-Disposition"]
    assert resposta.headers["Content-Disposition"].rstrip('"').endswith(".xlsx")
    wb = load_workbook(BytesIO(resposta.data), read_only=True)
    assert "Movimentações" in wb.sheetnames
    wb.close()


def test_exportacao_5000_registros_sem_truncagem_e_memoria_compativel():
    conn = conectar()
    cursor = conn.cursor()
    cursor.executemany(
        """
        INSERT INTO movimentacoes_financeiras (
            data_documento, data_vencimento, tipo, categoria, descricao,
            valor, valor_documento, valor_pago, valor_liquido, status,
            import_key, origem_importacao, linha_dre, tipo_conta,
            impacta_fluxo_caixa
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                f"2026-{(indice % 12) + 1:02d}-01",
                f"2026-{(indice % 12) + 1:02d}-20",
                "Saida",
                "Categoria desempenho",
                f"Movimento sintético {indice}",
                10,
                10,
                0,
                10,
                "Pendente",
                f"desempenho-{indice}",
                "excel_movimentacoes",
                "Despesas Operacionais",
                "Saida",
                1,
            )
            for indice in range(5000)
        ],
    )
    conn.commit()
    conn.close()

    tracemalloc.start()
    inicio = time.perf_counter()
    contexto = montar_contexto_auditoria_financeira(MultiDict(), exportar=True)
    arquivo = gerar_excel_auditoria_financeira(contexto)
    duracao = time.perf_counter() - inicio
    _, pico = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb["Movimentações"]
    assert contexto["indicadores"]["total_movimentacoes"] == 5000
    assert ws.max_row == 5001
    assert duracao < 45
    assert pico < 256 * 1024 * 1024
    assert len(arquivo.getvalue()) > 100_000
    wb.close()

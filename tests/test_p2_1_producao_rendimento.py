from datetime import date
from io import BytesIO
import sqlite3

import pytest
from openpyxl import load_workbook

from modules.relatorios import producao
from modules.dashboard import repositories as dashboard_repositories


@pytest.fixture()
def banco(tmp_path, monkeypatch):
    caminho = tmp_path / "p2-1.db"

    def conectar():
        conn = sqlite3.connect(caminho)
        conn.row_factory = sqlite3.Row
        return conn

    monkeypatch.setattr(producao, "conectar", conectar)
    monkeypatch.setattr(producao, "DATABASE_URL", None)
    conn = conectar()
    conn.executescript("""
        CREATE TABLE ordens_producao(
            id INTEGER PRIMARY KEY, data TEXT, fornecedor TEXT, sku TEXT, status TEXT,
            quantidade_aves INTEGER, mortes_antes_pendura INTEGER, peso_vivo REAL,
            peso_medio REAL, versao_operacional INTEGER DEFAULT 0
        );
        CREATE TABLE apontamentos_producao(
            id INTEGER PRIMARY KEY, op_id INTEGER, data TEXT, setor TEXT,
            quantidade REAL, unidade TEXT, vigente INTEGER DEFAULT 1
        );
        CREATE TABLE apontamentos_descartes(
            id INTEGER PRIMARY KEY, op_id INTEGER, data TEXT, setor TEXT, categoria TEXT,
            motivo TEXT, quantidade REAL, unidade TEXT
        );
        CREATE TABLE embalagem_primaria_apontamentos(
            id INTEGER PRIMARY KEY, op_id INTEGER, data_apontamento TEXT, sku TEXT,
            quantidade_bandejas REAL
        );
        CREATE TABLE estoque_produto_intermediario(
            id INTEGER PRIMARY KEY, op_id INTEGER, tipo TEXT, quantidade_bandejas REAL
        );
        CREATE TABLE pa_caixas(
            id INTEGER PRIMARY KEY, codigo_caixa TEXT, quantidade_bandejas REAL,
            peso_liquido REAL, peso_bruto REAL, status TEXT
        );
        CREATE TABLE pa_caixa_composicao(
            id INTEGER PRIMARY KEY, caixa_id INTEGER, op_id INTEGER, quantidade_bandejas REAL
        );
        CREATE TABLE op_operacoes_auditoria(
            id INTEGER PRIMARY KEY, op_id INTEGER, tipo TEXT
        );
        CREATE TABLE pa_nao_conformes(
            id INTEGER PRIMARY KEY, op_id INTEGER, quantidade REAL, peso REAL, status TEXT
        );

        INSERT INTO ordens_producao VALUES
          (1,'2026-08-01','Fornecedor A','Galinha Cortada','Encerrada',100,5,200,2,0),
          (2,'2026-08-01','Fornecedor A','Galinha Cortada','Aberta',60,0,120,2,1),
          (3,'2026-08-01','Fornecedor B','Galinha Cortada','Encerrada',50,0,100,2,2),
          (4,'2026-08-02','Fornecedor C','Galinha Cortada','Estornada',40,0,80,2,1);

        INSERT INTO apontamentos_descartes VALUES
          (1,1,'2026-08-01','Recepcao','Morte','Morte na gaiola',2,'aves'),
          (2,1,'2026-08-01','Inspecao','Condenacao','Condenacao sanitaria',3,'aves'),
          (3,1,'2026-08-01','Corte','Descarte','Perda operacional',5,'aves');

        INSERT INTO embalagem_primaria_apontamentos VALUES
          (1,1,'2026-08-01','Galinha Cortada',85),
          (2,2,'2026-08-05','Galinha Cortada',60),
          (3,3,'2026-08-10','Galinha Cortada',50),
          (4,4,'2026-08-02','Galinha Cortada',40);

        INSERT INTO estoque_produto_intermediario VALUES
          (1,1,'ENTRADA_EMBALAGEM_PRIMARIA',85),(2,1,'SAIDA_EMBALAGEM_SECUNDARIA',85),
          (3,2,'ENTRADA_EMBALAGEM_PRIMARIA',60),(4,2,'SAIDA_EMBALAGEM_SECUNDARIA',12),
          (5,3,'ENTRADA_EMBALAGEM_PRIMARIA',50),(6,3,'SAIDA_EMBALAGEM_SECUNDARIA',50),
          (7,4,'ENTRADA_EMBALAGEM_PRIMARIA',40),(8,4,'SAIDA_EMBALAGEM_SECUNDARIA',40),
          (9,4,'ENTRADA_ESTORNO_CAIXA',40),(10,4,'SAIDA_ESTORNO_OP',40);

        INSERT INTO pa_caixas VALUES
          (1,'CX-1',40,50,52,'Em estoque'),
          (2,'CX-2',45,55,57,'Em estoque'),
          (3,'CX-EST',12,20,21,'Estornada'),
          (4,'CX-PARCIAL',12,15,16,'Em estoque'),
          (5,'CX-REABERTA',50,60,62,'Em estoque'),
          (6,'CX-OP-EST',40,48,50,'Estornada');
        INSERT INTO pa_caixa_composicao VALUES
          (1,1,1,40),(2,2,1,45),(3,3,1,12),(4,4,2,12),(5,5,3,50),(6,6,4,40);

        INSERT INTO apontamentos_producao VALUES
          (1,1,'2026-08-01','Expedicao',999,'kg',1),
          (2,1,'2026-08-09','Expedicao',777,'kg',0),
          (3,3,'2026-08-13','Expedicao',999,'kg',1);
        INSERT INTO op_operacoes_auditoria VALUES
          (1,2,'RETOMADA_EMBALAGEM_SECUNDARIA'),
          (2,3,'REABERTURA'),(3,3,'RETOMADA_EMBALAGEM_SECUNDARIA'),
          (4,4,'ESTORNO_INTEGRAL');
        INSERT INTO pa_nao_conformes VALUES(1,1,12,15,'DESCARTADO');
    """)
    conn.commit()
    conn.close()
    return conectar


def filtros(**extras):
    dados = {
        "data_inicio": "2026-08-01", "data_fim": "2026-08-31",
        "status": "Todos", "sku": "Todos", "fornecedor": "Todos",
        "op_id": "", "lote": "", "causa": "Todos", "setor": "Todos",
        "situacao": "Todas", "escopo": "producao_valida", "granularidade": "dia",
    }
    dados.update(extras)
    return dados


def test_producao_valida_usa_pa_liquido_ativo_e_exclui_estornos(banco):
    linhas = producao.buscar_ops_agregadas(filtros())
    assert [item["op_id"] for item in linhas] == [3, 1]
    op1 = next(item for item in linhas if item["op_id"] == 1)
    assert op1["caixas"] == 2
    assert op1["producao_secundaria"] == 85
    assert op1["peso_produzido"] == pytest.approx(105)
    assert op1["peso_produzido"] != 999
    assert op1["pnc_registros"] == 1


def test_aves_consideradas_perdas_e_rendimentos_tem_denominador_oficial(banco):
    op1 = next(item for item in producao.buscar_ops_agregadas(filtros()) if item["op_id"] == 1)
    assert op1["mortes_total"] == 7
    assert op1["aves_consideradas"] == 93
    assert op1["condenacoes_aves"] == 3
    assert op1["descartes_aves"] == 5
    assert op1["perdas_aves"] == 15
    assert op1["rendimento"] == 52.5
    assert op1["rendimento_aves"] == pytest.approx(91.4, abs=.01)


def test_pi_reconcilia_producao_consumo_saldo_e_estorno_integral(banco):
    op1 = next(item for item in producao.buscar_ops_agregadas(filtros()) if item["op_id"] == 1)
    assert (op1["pi_produzido_valido"], op1["pi_consumido_valido"], op1["pi_saldo"]) == (85, 85, 0)
    assert op1["pi_reconciliado"] is True
    historico = producao.buscar_ops_agregadas(filtros(escopo="historico"))
    assert [item["op_id"] for item in historico] == [4]
    assert historico[0]["pi_produzido_valido"] == 0
    assert historico[0]["pi_reconciliado"] is True


def test_aberta_retomada_reaberta_e_estornada_nao_duplicam_indicadores(banco):
    andamento = producao.buscar_ops_agregadas(filtros(escopo="andamento"))
    assert [item["op_id"] for item in andamento] == [2]
    assert andamento[0]["indicador_retomada"] is True
    validas = producao.buscar_ops_agregadas(filtros())
    op3 = next(item for item in validas if item["op_id"] == 3)
    assert op3["indicador_reabertura"] is True
    assert op3["indicador_retomada"] is True
    assert op3["caixas"] == 1
    historico = producao.buscar_ops_agregadas(filtros(escopo="historico"))[0]
    assert historico["indicador_estorno"] is True
    assert historico["caixas"] == 0 and historico["peso_produzido"] == 0


def test_data_da_op_prevalece_sobre_datas_dos_apontamentos(banco):
    contexto = producao.montar_contexto_relatorio_producao("producao-por-periodo", filtros())
    assert [grupo["grupo"] for grupo in contexto["agrupamentos"]] == ["2026-08-01"]
    assert contexto["totais"]["ops"] == 2
    assert contexto["totais"]["peso_produzido"] == 165


def test_agregacao_de_rendimento_e_ponderada(banco):
    contexto = producao.montar_contexto_relatorio_producao("rendimento", filtros())
    assert contexto["totais"]["peso_produzido"] == 165
    assert contexto["totais"]["peso_vivo"] == 300
    assert contexto["totais"]["rendimento"] == 55


def test_todos_os_estados_mantem_cards_apenas_da_producao_valida(banco):
    contexto = producao.montar_contexto_relatorio_producao(
        "producao-por-op", filtros(escopo="todas")
    )
    assert {item["op_id"] for item in contexto["detalhes"]} == {1, 2, 3, 4}
    assert contexto["totais"]["ops"] == 2
    assert contexto["totais"]["peso_produzido"] == 165


def test_excel_preserva_datas_e_numeros(banco):
    contexto = producao.montar_contexto_relatorio_producao("producao-por-op", filtros())
    arquivo = producao.gerar_excel_relatorio_producao(contexto)
    planilha = load_workbook(BytesIO(arquivo.read()), data_only=True).active
    cabecalho = next(
        linha for linha in planilha.iter_rows(values_only=True)
        if linha[0] == "op_id" and "data_op" in linha
    )
    inicio = next(
        i for i, linha in enumerate(planilha.iter_rows(values_only=True), start=1)
        if linha[0] == "op_id" and "data_op" in linha
    )
    primeira = [celula.value for celula in planilha[inicio + 1]]
    dados = dict(zip(cabecalho, primeira))
    assert isinstance(dados["data_op"], (date,))
    assert isinstance(dados["peso_produzido"], (int, float))
    assert isinstance(dados["rendimento"], (int, float))


def test_caixa_mista_aloca_peso_e_conta_caixa_fisica_uma_vez(banco):
    conn = banco()
    conn.execute("INSERT INTO pa_caixas VALUES(7,'CX-MISTA',10,20,21,'Em estoque')")
    conn.execute("INSERT INTO pa_caixa_composicao VALUES(7,7,1,6)")
    conn.execute("INSERT INTO pa_caixa_composicao VALUES(8,7,3,4)")
    conn.commit()
    conn.close()
    linhas = {item["op_id"]: item for item in producao.buscar_ops_agregadas(filtros())}
    assert linhas[1]["peso_produzido"] == 117
    assert linhas[3]["peso_produzido"] == 68
    assert linhas[1]["caixas"] + linhas[3]["caixas"] == 4
    assert linhas[1]["caixas_compartilhadas"] == 1
    assert linhas[3]["caixas_compartilhadas"] == 1


def test_rendimento_em_peso_nao_inventa_peso_para_produto_sem_base_pa(banco):
    conn = banco()
    conn.execute("""INSERT INTO ordens_producao VALUES
        (5,'2026-08-03','Fornecedor D','Galinha Inteira','Encerrada',25,0,50,2,0)""")
    conn.execute("INSERT INTO embalagem_primaria_apontamentos VALUES(5,5,'2026-08-03','Galinha Inteira',25)")
    conn.execute("INSERT INTO pa_caixas VALUES(8,'LOTE-INTEIRA',NULL,NULL,NULL,'Em estoque')")
    conn.execute("INSERT INTO pa_caixa_composicao VALUES(9,8,5,25)")
    conn.commit()
    conn.close()
    contexto = producao.montar_contexto_relatorio_producao("rendimento", filtros())
    op5 = next(item for item in contexto["detalhes"] if item["op_id"] == 5)
    assert op5["rendimento_aplicavel"] is False
    assert op5["peso_base_rendimento"] == 0
    assert contexto["totais"]["rendimento"] == 55


def test_consultas_e_exportacao_nao_alteram_base_fisica(banco):
    conn = banco()
    antes = {
        tabela: conn.execute(f"SELECT COUNT(*) FROM {tabela}").fetchone()[0]
        for tabela in (
            "ordens_producao", "embalagem_primaria_apontamentos",
            "estoque_produto_intermediario", "pa_caixas", "pa_caixa_composicao",
            "apontamentos_producao", "apontamentos_descartes", "pa_nao_conformes",
        )
    }
    conn.close()
    contexto = producao.montar_contexto_relatorio_producao("producao-por-op", filtros(escopo="todas"))
    producao.gerar_excel_relatorio_producao(contexto)
    conn = banco()
    depois = {
        tabela: conn.execute(f"SELECT COUNT(*) FROM {tabela}").fetchone()[0]
        for tabela in antes
    }
    conn.close()
    assert depois == antes


def test_dashboard_usa_a_mesma_base_fisica_de_pa_ativo(banco, monkeypatch):
    conn = banco()
    conn.executescript("""
        CREATE TABLE apontamentos_paradas(
            id INTEGER, evento_id TEXT, op_id INTEGER, data TEXT, setor TEXT,
            motivo TEXT, horas_paradas REAL, observacoes TEXT
        );
        CREATE TABLE apontamentos_mao_obra(
            id INTEGER, op_id INTEGER, data TEXT, setor TEXT, colaborador TEXT
        );
    """)
    conn.commit()
    conn.close()
    monkeypatch.setattr(dashboard_repositories, "conectar", banco)
    dados = dashboard_repositories.buscar_dados_dashboard(
        "2026-08-01", "2026-08-31", "Encerrada", "Todos"
    )
    assert dados["kg_produzidos_rendimento"] == 165
    assert dados["peso_entrada_rendimento"] == 300
    assert dados["kg_produzidos_rendimento"] != 1998
